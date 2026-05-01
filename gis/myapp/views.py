from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.hashers import check_password, make_password
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_exempt
from django.db import transaction
from django.db.models import Avg, Count, F, Prefetch, Q, Sum
from django.db.models.functions import Lower
from io import BytesIO
from functools import wraps
from datetime import date, datetime, timedelta
from urllib.parse import quote
import base64
import json
import re
import unicodedata
import logging
from secrets import randbelow
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import uuid4
from xml.sax.saxutils import escape

from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.core.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.core.paginator import Paginator
from django.utils import timezone
from django.utils.html import escapejs, format_html, strip_tags
from django.utils.encoding import force_bytes, force_str
from django.utils.http import url_has_allowed_host_and_scheme, urlsafe_base64_decode, urlsafe_base64_encode
from django.contrib.auth.models import User
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile

from .storage import find_legacy_media_path, normalize_db_media_name

from .forms import (
    AccountProfileForm,
    AccountRecoveryRequestForm,
    AboutBuiltinSectionFormSet,
    AboutCustomBlockFormSet,
    AboutPageContentForm,
    AboutFeaturedBranchItemFormSet,
    AboutPageSlideFormSet,
    CheckoutForm,
    PaymentProofUploadForm,
    HomeCategorySpotlightItemFormSet,
    HomeHeroSlideFormSet,
    HomePageContentForm,
    HomeServiceCommitmentItemFormSet,
    RegisterForm,
    LoginForm,
    NewsArticleAdminForm,
    PharmacyAdminForm,
    MedicineAdminForm,
    MedicineReviewForm,
    PasswordResetOtpVerificationForm,
    ProfilePasswordChangeForm,
    PromotionAdminForm,
    OrderStatusUpdateForm,
    CustomUserCreateForm,
    CustomUserUpdateForm,
    PharmacyReviewForm,
    ReturnRefundRequestForm,
    ReturnRefundRequestAdminUpdateForm,
    PurchaseImportExcelForm,
    StockExportBatchForm,
    StockExportItemFormSet,
    UsernameRecoveryOtpVerificationForm,
)
from .emails import (
    send_account_recovery_otp_email,
    send_order_cancelled_email,
    send_order_confirmation_email,
    send_order_invoice_email,
    send_order_payment_confirmed_email,
    send_order_status_update_email,
    send_return_request_received_email,
    send_return_request_status_update_email,
    send_account_profile_updated_email,
    send_password_changed_email,
    send_registration_confirmation_email,
)
from .models import (
    AboutPageContent,
    AboutCustomBlock,
    AboutFeaturedBranchItem,
    AboutPageSlide,
    Cart,
    CartItem,
    HomeCategorySpotlightItem,
    HomeHeroSlide,
    HomePageContent,
    HomeServiceCommitmentItem,
    Medicine,
    MEDICINE_PRODUCT_TYPE_MEDICINE,
    MEDICINE_PRODUCT_TYPE_SUPPLEMENT,
    MEDICINE_PRODUCT_TYPE_CHOICES,
    MEDICINE_SHARED_SYNC_FIELDS,
    MedicinePromotion,
    MedicineReview,
    NewsArticle,
    Order,
    OrderItem,
    OrderPrescriptionProof,
    Pharmacy,
    PharmacyReview,
    ReturnRefundEvidence,
    ReturnRefundRequest,
    PurchaseImportBatch,
    PurchaseImportItem,
    MedicineLot,
    OrderItemLotAllocation,
    StockExportBatch,
    StockExportItem,
    StockExportLotAllocation,
    StoredMediaFile,
    AccountOtpChallenge,
    UserProfile,
    build_medicine_catalog_key,
    fold_text_for_match,
    sync_medicine_catalog_metadata,
)
from .tools.calculations import (
    calculate_air_distance_km,
    estimate_road_distance_km,
    normalize_departure_time_str,
)
from .tools.geocode import reverse_geocode_coordinates, search_address_candidates
from .tools.routing import DeliveryRoutingService
from .tokens import email_activation_token

try:
    import qrcode # type: ignore
except ImportError:  # pragma: no cover
    qrcode = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except ImportError:  # pragma: no cover
    colors = None
    A4 = None
    ParagraphStyle = None
    getSampleStyleSheet = None
    mm = None
    pdfmetrics = None
    TTFont = None
    Paragraph = None
    SimpleDocTemplate = None
    Spacer = None
    Table = None
    TableStyle = None


logger = logging.getLogger(__name__)
delivery_service = DeliveryRoutingService()
PHARMACY_FALLBACK_NAME = 'pharmacies/pm.jpg'
MEDICINE_FALLBACK_NAME = 'medicines/oresol.jpg'
PDF_FONT_NAME = "GISArial"
PDF_FONT_PATH = Path(r"C:\Windows\Fonts\arial.ttf")
DEFAULT_HOME_SLIDES = [
    {
        "legacy_static_path": "images/20260305065905-0-event-banner-slide-2560.webp",
        "alt_text": "Banner trang chủ 1",
        "link_url": "/about/",
    },
    {
        "legacy_static_path": "images/20260305094937-0-Banner.png",
        "alt_text": "Banner trang chủ 2",
        "link_url": "/about/",
    },
    {
        "legacy_static_path": "images/Banner_Hero_Full_PC_1216x280_4217b59127.webp",
        "alt_text": "Banner trang chủ 3",
        "link_url": "/about/",
    },
]
DEFAULT_ABOUT_SLIDES = [
    {
        "legacy_static_path": "images/Banner_Hero_Full_PC_1216x280_4217b59127.webp",
        "alt_text": "Banner giới thiệu 1",
        "link_url": "/products/",
    },
    {
        "legacy_static_path": "images/20260305065905-0-event-banner-slide-2560.webp",
        "alt_text": "Banner giới thiệu 2",
        "link_url": "/map/",
    },
    {
        "legacy_static_path": "images/20260305094937-0-Banner.png",
        "alt_text": "Banner giới thiệu 3",
        "link_url": "/about/",
    },
]
DEFAULT_HOME_CATEGORY_ITEMS = [
    {"title": "Chăm sóc cơ bản", "subtitle": "10 sản phẩm", "icon_class": "fas fa-capsules", "link_url": "/products/?category=Cham%20soc%20co%20ban"},
    {"title": "Dạ dày - tiêu hóa", "subtitle": "10 sản phẩm", "icon_class": "fas fa-heartbeat", "link_url": "/products/?category=Da%20day%20-%20tieu%20hoa"},
    {"title": "Dị ứng - hô hấp", "subtitle": "10 sản phẩm", "icon_class": "fas fa-prescription-bottle-alt", "link_url": "/products/?category=Di%20ung%20-%20ho%20hap"},
    {"title": "Giảm đau - hạ sốt", "subtitle": "10 sản phẩm", "icon_class": "fas fa-syringe", "link_url": "/products/?category=Giam%20dau%20-%20ha%20sot"},
    {"title": "Ho - cảm", "subtitle": "10 sản phẩm", "icon_class": "fas fa-shield-virus", "link_url": "/products/?category=Ho%20-%20cam"},
    {"title": "Kháng sinh", "subtitle": "10 sản phẩm", "icon_class": "fas fa-leaf", "link_url": "/products/?category=Khang%20sinh"},
]
DEFAULT_HOME_COMMITMENT_ITEMS = [
    {
        "title": "Thuốc chính hãng",
        "body": "Hiển thị rõ nhà sản xuất, chi nhánh bán và tình trạng còn hàng.",
        "icon_class": "fas fa-certificate",
    },
    {
        "title": "Định vị nhà thuốc",
        "body": "Kết hợp bản đồ GIS để chọn nhà thuốc và điểm giao phù hợp.",
        "icon_class": "fas fa-map-marker-alt",
    },
    {
        "title": "Phí giao hàng minh bạch",
        "body": "Khoảng cách và phí ship được cập nhật trực tiếp khi chọn vị trí nhận hàng.",
        "icon_class": "fas fa-shipping-fast",
    },
    {
        "title": "Theo dõi đơn hàng",
        "body": "Lưu lịch sử mua hàng để người dùng dễ dàng tra cứu lại.",
        "icon_class": "fas fa-history",
    },
]

ADMIN_PERMISSION_DEFINITIONS = [
    {"key": "dashboard", "label": "Tổng quan", "description": "Xem tổng quan vận hành", "actions": ("view",)},
    {"key": "home_page", "label": "Trang chủ", "description": "Quản lý slider và các khối nội dung trang chủ", "actions": ("view", "update")},
    {"key": "about_page", "label": "Trang giới thiệu", "description": "Quản lý nội dung hiển thị ở trang Giới thiệu", "actions": ("view", "update")},
    {"key": "news", "label": "Tin tức", "description": "Quản lý bài viết tin tức hiển thị ngoài website", "actions": ("view", "create", "update", "delete")},
    {"key": "order", "label": "Đơn hàng", "description": "Xem, cập nhật trạng thái và xóa đơn hàng", "actions": ("view", "update", "delete")},
    {"key": "return_request", "label": "Trả hàng / hoàn tiền", "description": "Xem, xử lý và xóa yêu cầu hoàn tiền", "actions": ("view", "update", "delete")},
    {"key": "medicine", "label": "Sản phẩm", "description": "Xem, thêm, sửa và xóa sản phẩm", "actions": ("view", "create", "update", "delete")},
    {"key": "purchase_import", "label": "Nhập kho", "description": "Xem và tạo phiếu nhập", "actions": ("view", "create", "delete")},
    {"key": "stock_export", "label": "Xuất kho", "description": "Xem và tạo phiếu xuất", "actions": ("view", "create", "delete")},
    {"key": "inventory_lot", "label": "Lô tồn kho", "description": "Theo dõi tồn theo lô FEFO", "actions": ("view",)},
    {"key": "promotion", "label": "Khuyến mãi", "description": "Xem, thêm và sửa khuyến mãi", "actions": ("view", "create", "update", "delete")},
    {"key": "reports", "label": "Báo cáo", "description": "Xem báo cáo và thống kê", "actions": ("view",)},
    {"key": "pharmacy", "label": "Chi nhánh", "description": "Quản lý chi nhánh", "actions": ("view", "create", "update", "delete")},
    {"key": "user", "label": "Tài khoản", "description": "Quản lý tài khoản người dùng", "actions": ("view", "create", "update", "delete")},
    {"key": "permission", "label": "Phân quyền", "description": "Thiết lập quyền chi tiết cho tài khoản", "actions": ("view", "update")},
]
ADMIN_PERMISSION_MAP = {item["key"]: item for item in ADMIN_PERMISSION_DEFINITIONS}
BRANCH_SCOPED_ADMIN_MODELS = {"medicine", "order", "return_request", "purchase_import", "stock_export", "promotion", "inventory_lot", "reports"}


ACCOUNT_OTP_EXPIRE_MINUTES = 10
ACCOUNT_OTP_MAX_ATTEMPTS = 5
ACCOUNT_RECOVERY_RESULT_SESSION_KEY = "account_recovery_username_result"


def is_username_recovery_request(request):
    recovery_mode = (request.GET.get("recovery") or request.POST.get("recovery") or "").strip().lower()
    return recovery_mode == "username"


def mask_email_address(email):
    local_part, _, domain = str(email or "").partition("@")
    if not domain:
        return email or ""
    safe_local = local_part[:2] + "*" * max(len(local_part) - 2, 1)
    return f"{safe_local}@{domain}"


def build_account_recovery_context(*, request, is_username_recovery, extra=None):
    context = {
        "is_username_recovery": is_username_recovery,
        "recovery_title": "Quên tên đăng nhập" if is_username_recovery else "Quên mật khẩu",
        "recovery_subtitle": (
            "Nhập email đã đăng ký để nhận mã OTP lấy lại tên đăng nhập."
            if is_username_recovery
            else "Nhập email đã đăng ký để nhận mã OTP đặt lại mật khẩu."
        ),
    }
    if extra:
        context.update(extra)
    return context


def expire_active_account_otp_challenges(*, user, purpose):
    AccountOtpChallenge.objects.filter(
        user=user,
        purpose=purpose,
        consumed_at__isnull=True,
        expires_at__gt=timezone.now(),
    ).update(consumed_at=timezone.now(), updated_at=timezone.now())


def create_account_otp_challenge(*, user, purpose):
    expire_active_account_otp_challenges(user=user, purpose=purpose)
    otp_code = f"{randbelow(1000000):06d}"
    challenge = AccountOtpChallenge.objects.create(
        user=user,
        purpose=purpose,
        email=(user.email or "").strip(),
        otp_hash=make_password(otp_code),
        username_snapshot=user.get_username(),
        expires_at=timezone.now() + timedelta(minutes=ACCOUNT_OTP_EXPIRE_MINUTES),
    )
    return challenge, otp_code


def account_recovery_request_view(request):
    is_username_recovery = is_username_recovery_request(request)
    recovery_mode = "username" if is_username_recovery else "password"
    purpose = (
        AccountOtpChallenge.PURPOSE_USERNAME_RECOVERY
        if is_username_recovery
        else AccountOtpChallenge.PURPOSE_PASSWORD_RESET
    )

    if request.method == "POST":
        form = AccountRecoveryRequestForm(request.POST, recovery_mode=recovery_mode)
        if form.is_valid():
            matched_user = form.matched_user
            challenge, otp_code = create_account_otp_challenge(user=matched_user, purpose=purpose)
            sent = send_account_recovery_otp_email(
                user=matched_user,
                challenge=challenge,
                otp_code=otp_code,
                request=request,
            )
            if sent:
                messages.success(
                    request,
                    "Hệ thống đã gửi mã OTP vào email của bạn. Vui lòng kiểm tra hộp thư để tiếp tục.",
                )
                return redirect("account_recovery_verify", token=challenge.public_token)
            form.add_error(None, "Không thể gửi email OTP lúc này. Vui lòng thử lại sau ít phút.")
    else:
        form = AccountRecoveryRequestForm(recovery_mode=recovery_mode)

    return render(
        request,
        "registration/password_reset_form.html",
        build_account_recovery_context(
            request=request,
            is_username_recovery=is_username_recovery,
            extra={"form": form},
        ),
    )


def account_recovery_verify_view(request, token):
    challenge = get_object_or_404(
        AccountOtpChallenge.objects.select_related("user"),
        public_token=token,
    )
    is_username_recovery = challenge.purpose == AccountOtpChallenge.PURPOSE_USERNAME_RECOVERY
    challenge_locked = challenge.attempts >= ACCOUNT_OTP_MAX_ATTEMPTS
    challenge_available = bool(
        getattr(challenge, "user", None)
        and challenge.user.is_active
        and challenge.is_active
        and not challenge_locked
    )

    if is_username_recovery:
        form = UsernameRecoveryOtpVerificationForm(request.POST or None)
    else:
        form = PasswordResetOtpVerificationForm(challenge.user, request.POST or None)

    if request.method == "POST" and challenge_available and form.is_valid():
        otp_code = form.cleaned_data["otp_code"]
        if not check_password(otp_code, challenge.otp_hash):
            challenge.attempts += 1
            if challenge.attempts >= ACCOUNT_OTP_MAX_ATTEMPTS:
                challenge.consumed_at = timezone.now()
            challenge.save(update_fields=["attempts", "consumed_at", "updated_at"])
            remaining_attempts = max(ACCOUNT_OTP_MAX_ATTEMPTS - challenge.attempts, 0)
            if remaining_attempts:
                form.add_error("otp_code", f"Mã OTP không đúng. Bạn còn {remaining_attempts} lần thử.")
            else:
                form.add_error("otp_code", "Mã OTP đã bị khóa do nhập sai quá nhiều lần. Vui lòng yêu cầu mã mới.")
        else:
            challenge.consumed_at = timezone.now()
            challenge.save(update_fields=["consumed_at", "updated_at"])
            if is_username_recovery:
                request.session[ACCOUNT_RECOVERY_RESULT_SESSION_KEY] = {
                    "username": challenge.username_snapshot or challenge.user.get_username(),
                    "email": challenge.email,
                }
                messages.success(request, "Đã xác thực OTP thành công. Bạn có thể xem lại tên đăng nhập ngay bên dưới.")
                return redirect("username_recovery_complete")

            challenge.user.set_password(form.cleaned_data["new_password1"])
            challenge.user.save(update_fields=["password"])
            send_password_changed_email(challenge.user, request=request, change_source='recovery')
            messages.success(request, "Mật khẩu đã được cập nhật thành công.")
            return redirect("password_reset_complete")

    if request.method == "POST" and not challenge_available:
        messages.error(request, "Mã OTP này đã hết hạn hoặc không còn hiệu lực. Vui lòng yêu cầu mã mới.")

    challenge_state = "active"
    if challenge_locked:
        challenge_state = "locked"
    elif challenge.is_consumed:
        challenge_state = "used"
    elif challenge.is_expired:
        challenge_state = "expired"
    elif not getattr(challenge, "user", None) or not challenge.user.is_active:
        challenge_state = "invalid"

    return render(
        request,
        "registration/password_reset_confirm.html",
        build_account_recovery_context(
            request=request,
            is_username_recovery=is_username_recovery,
            extra={
                "form": form,
                "challenge": challenge,
                "challenge_state": challenge_state,
                "challenge_available": challenge_available,
                "masked_email": mask_email_address(challenge.email),
                "otp_expires_minutes": ACCOUNT_OTP_EXPIRE_MINUTES,
            },
        ),
    )


def account_recovery_password_complete_view(request):
    return render(request, "registration/password_reset_complete.html")


def username_recovery_complete_view(request):
    result = request.session.pop(ACCOUNT_RECOVERY_RESULT_SESSION_KEY, None)
    if not result:
        messages.info(request, "Bạn chưa hoàn tất bước xác thực OTP để lấy lại tên đăng nhập.")
        return redirect(f"{reverse('password_reset')}?recovery=username")
    return render(
        request,
        "registration/username_recovery_complete.html",
        {
            "recovered_username": result.get("username", ""),
            "recovery_email": result.get("email", ""),
        },
    )


def legacy_password_reset_redirect_view(request, uidb64=None, token=None):
    messages.info(
        request,
        "Liên kết khôi phục cũ không còn được sử dụng. Vui lòng yêu cầu mã OTP mới để tiếp tục.",
    )
    return redirect("password_reset")


def get_media_url_or_fallback(file_name, legacy_url=''):
    try:
        return default_storage.url(file_name)
    except Exception:
        return legacy_url or '/db-media/' + file_name.lstrip('/')


PHARMACY_FALLBACK_IMAGE = get_media_url_or_fallback(PHARMACY_FALLBACK_NAME, '/db-media/pharmacies/pm.jpg')
MEDICINE_FALLBACK_IMAGE = get_media_url_or_fallback(MEDICINE_FALLBACK_NAME, '/db-media/medicines/oresol.jpg')




def get_excel_workbook_loader():
    try:
        from openpyxl import load_workbook as _load_workbook
        return _load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Máy đang thiếu thư viện openpyxl nên chưa thể đọc file Excel .xlsx. "
            "Hãy cài openpyxl hoặc dùng đúng bản môi trường đã có thư viện này."
        ) from exc


def get_excel_workbook_builder():
    try:
        from openpyxl import Workbook as _Workbook
        return _Workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Máy đang thiếu thư viện openpyxl nên chưa thể xuất file Excel .xlsx. "
            "Hãy cài openpyxl hoặc dùng đúng bản môi trường đã có thư viện này."
        ) from exc




def build_promotion_group_key(name, unit, manufacturer=''):
    return (
        fold_text_for_match(name),
        fold_text_for_match(unit),
        fold_text_for_match(manufacturer),
    )


def build_active_promotion_queryset(queryset=None, *, on_date=None):
    today = on_date or timezone.localdate()
    queryset = queryset if queryset is not None else MedicinePromotion.objects.all()
    return queryset.filter(
        is_active=True,
    ).filter(
        Q(start_date__isnull=True) | Q(start_date__lte=today),
        Q(end_date__isnull=True) | Q(end_date__gte=today),
    )


def build_active_promotion_map(medicines, *, on_date=None):
    medicine_list = [medicine for medicine in medicines if getattr(medicine, 'id', None)]
    if not medicine_list:
        return {}

    requested_keys = {
        build_promotion_group_key(medicine.name, medicine.unit, medicine.manufacturer)
        for medicine in medicine_list
    }
    promotions = list(
        build_active_promotion_queryset(
            MedicinePromotion.objects.select_related('medicine'),
            on_date=on_date,
        ).order_by('-discount_percent', '-id')
    )

    group_mapping = {}
    for promotion in promotions:
        medicine = getattr(promotion, 'medicine', None)
        if medicine is None:
            continue
        group_key = build_promotion_group_key(medicine.name, medicine.unit, medicine.manufacturer)
        if group_key not in requested_keys or group_key in group_mapping:
            continue
        group_mapping[group_key] = promotion

    mapping = {}
    for medicine in medicine_list:
        promotion = group_mapping.get(build_promotion_group_key(medicine.name, medicine.unit, medicine.manufacturer))
        medicine._catalog_group_active_promotion = promotion
        mapping[medicine.id] = promotion
    return mapping


def get_discounted_price(base_price, discount_percent):
    try:
        price_value = int(base_price or 0)
    except (TypeError, ValueError):
        price_value = 0
    try:
        discount_value = int(discount_percent or 0)
    except (TypeError, ValueError):
        discount_value = 0
    discounted = int(price_value * max(0, 100 - discount_value) / 100)
    return max(discounted, 0)


CUSTOMER_TIER_RULES = [
    {
        'code': 'diamond',
        'label': 'Kim cương',
        'threshold': 8_000_000,
        'discount_percent': 8,
    },
    {
        'code': 'gold',
        'label': 'Vàng',
        'threshold': 5_000_000,
        'discount_percent': 6,
    },
    {
        'code': 'silver',
        'label': 'Bạc',
        'threshold': 3_000_000,
        'discount_percent': 4,
    },
    {
        'code': 'bronze',
        'label': 'Đồng',
        'threshold': 1_000_000,
        'discount_percent': 2,
    },
]


def apply_customer_tier_discount(base_price, discount_percent):
    return get_discounted_price(base_price, discount_percent)


def calculate_loyalty_discount_total(base_total, discount_percent):
    try:
        base_value = int(base_total or 0)
    except (TypeError, ValueError):
        base_value = 0
    return max(base_value - get_discounted_price(base_value, discount_percent), 0)


def resolve_customer_tier(total_spent):
    spent_value = int(total_spent or 0)
    for tier in CUSTOMER_TIER_RULES:
        if spent_value >= tier['threshold']:
            return dict(tier)
    return {
        'code': 'standard',
        'label': 'Chưa xếp hạng',
        'threshold': 0,
        'discount_percent': 0,
    }


def build_customer_loyalty_context(user):
    if not user or not user.is_authenticated:
        return {
            'code': 'guest',
            'label': 'Khách lẻ',
            'discount_percent': 0,
            'total_spent': 0,
            'next_tier': CUSTOMER_TIER_RULES[-1],
            'amount_to_next_tier': CUSTOMER_TIER_RULES[-1]['threshold'],
            'is_customer': False,
            'show_program_hint': True,
        }

    if user.is_staff or user.is_superuser:
        return {
            'code': 'internal',
            'label': 'Không áp dụng',
            'discount_percent': 0,
            'total_spent': 0,
            'next_tier': None,
            'amount_to_next_tier': 0,
            'is_customer': False,
            'show_program_hint': False,
        }

    totals = Order.objects.filter(
        user=user,
        status=Order.STATUS_COMPLETED,
    ).aggregate(total_spent=Sum('total_product_price'))
    total_spent = int(totals.get('total_spent') or 0)
    tier = resolve_customer_tier(total_spent)

    ordered_rules = list(reversed(CUSTOMER_TIER_RULES))
    next_tier = None
    amount_to_next_tier = 0
    for candidate in ordered_rules:
        if total_spent < candidate['threshold']:
            next_tier = candidate
            amount_to_next_tier = candidate['threshold'] - total_spent
            break

    return {
        'code': tier['code'],
        'label': tier['label'],
        'discount_percent': tier['discount_percent'],
        'total_spent': total_spent,
        'next_tier': next_tier,
        'amount_to_next_tier': amount_to_next_tier,
        'is_customer': True,
        'show_program_hint': True,
    }


def build_cart_pricing_snapshot(cart, user=None):
    loyalty = build_customer_loyalty_context(user)
    cart_items = list(cart.items.select_related('medicine', 'medicine__pharmacy').order_by('id'))
    system_subtotal = 0
    line_item_count = len(cart_items)
    total_quantity = 0

    for item in cart_items:
        system_unit_price = int(item.medicine.current_price)
        line_system_total = system_unit_price * item.quantity

        item.system_unit_price = system_unit_price
        item.loyalty_unit_price = system_unit_price
        item.loyalty_unit_discount = 0
        item.line_system_total = line_system_total
        item.line_loyalty_discount = 0
        item.line_final_total = line_system_total
        item.customer_tier_discount_percent = loyalty['discount_percent']
        item.display_has_tier_discount = False

        system_subtotal += line_system_total
        total_quantity += item.quantity

    loyalty_discount_total = calculate_loyalty_discount_total(system_subtotal, loyalty['discount_percent'])
    final_product_total = max(system_subtotal - loyalty_discount_total, 0)
    return {
        'items': cart_items,
        'line_item_count': line_item_count,
        'total_quantity': total_quantity,
        'system_subtotal': system_subtotal,
        'loyalty_discount_total': loyalty_discount_total,
        'final_product_total': final_product_total,
        'loyalty': loyalty,
    }


def build_medicine_discount_payload(medicine, promotion=None):
    promotion = promotion or getattr(medicine, 'active_promotion', None)
    original_price = int(getattr(medicine, 'price', 0) or 0)
    if not promotion:
        return {
            'has_discount': False,
            'promotion': None,
            'original_price': original_price,
            'discounted_price': original_price,
            'discount_percent': 0,
            'label': '',
        }
    discounted_price = get_discounted_price(original_price, promotion.discount_percent)
    return {
        'has_discount': discounted_price < original_price,
        'promotion': promotion,
        'original_price': original_price,
        'discounted_price': discounted_price,
        'discount_percent': promotion.discount_percent,
        'label': promotion.resolved_title,
    }


def attach_discount_payloads(medicines, *, on_date=None):
    medicine_list = list(medicines)
    promotion_map = build_active_promotion_map(medicine_list, on_date=on_date)
    for medicine in medicine_list:
        payload = build_medicine_discount_payload(medicine, promotion_map.get(medicine.id))
        medicine.active_discount_payload = payload
        medicine._prefetched_active_promotions = [payload['promotion']] if payload['promotion'] else []
        medicine.current_price_value = payload['discounted_price']
        medicine.original_price_value = payload['original_price']
        medicine.discount_percent_value = payload['discount_percent']
        medicine.has_discount_value = payload['has_discount']
    return medicine_list


def deduplicate_catalog_medicines(medicines, *, on_date=None, limit=None):
    medicine_list = attach_discount_payloads(list(medicines), on_date=on_date)
    grouped = {}
    for medicine in medicine_list:
        key = normalize_catalog_key(medicine.name, medicine.unit, medicine.manufacturer)
        rank = (
            0 if medicine.quantity > 0 else 1,
            0 if getattr(medicine, 'has_discount_value', False) else 1,
            getattr(medicine, 'current_price_value', medicine.price),
            medicine.id,
        )
        current = grouped.get(key)
        if current is None or rank < current[0]:
            grouped[key] = (rank, medicine)
    unique_medicines = [entry[1] for entry in grouped.values()]
    unique_medicines.sort(
        key=lambda item: (
            0 if item.quantity > 0 else 1,
            0 if getattr(item, 'has_discount_value', False) else 1,
            getattr(item, 'current_price_value', item.price),
            item.name.casefold(),
            item.id,
        )
    )
    if limit is not None:
        unique_medicines = unique_medicines[:limit]
    return unique_medicines

def serve_db_media_file(request, file_name):
    try:
        normalized_name = normalize_db_media_name(file_name)
    except ValueError:
        raise Http404('Tên tệp không hợp lệ.')

    media_obj = StoredMediaFile.objects.filter(file_name=normalized_name).only('file_data', 'content_type', 'file_size').first()
    if media_obj:
        file_bytes = media_obj.file_data.tobytes() if hasattr(media_obj.file_data, 'tobytes') else bytes(media_obj.file_data)
        response = FileResponse(BytesIO(file_bytes), content_type=media_obj.content_type or 'application/octet-stream')
        if media_obj.file_size:
            response['Content-Length'] = str(media_obj.file_size)
        response['Cache-Control'] = 'public, max-age=86400'
        return response

    legacy_path = find_legacy_media_path(normalized_name)
    if legacy_path and legacy_path.is_file():
        response = FileResponse(open(legacy_path, 'rb'), content_type=None)
        response['Cache-Control'] = 'public, max-age=3600'
        return response

    raise Http404('Không tìm thấy tệp media.')


def get_or_create_cart(request):
    """
    Lấy giỏ hàng hiện tại của người dùng.
    """
    if request.user.is_authenticated:
        cart, _ = Cart.objects.get_or_create(user=request.user)
        return cart

    if not request.session.session_key:
        request.session.create()

    cart, _ = Cart.objects.get_or_create(session_key=request.session.session_key)
    return cart


def get_cart_items_count(cart):
    return sum(item.quantity for item in cart.items.all())


def request_expects_json(request):
    accept_header = request.headers.get('Accept', '')
    requested_format = (request.POST.get('response_format') or request.GET.get('response_format') or request.POST.get('format') or request.GET.get('format') or '').strip().lower()
    return (
        request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        or 'application/json' in accept_header
        or requested_format == 'json'
    )


def get_safe_redirect_url(request, default='home'):
    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url
    return reverse(default)


def admin_panel_required(view_func):
    @wraps(view_func)
    def wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            login_url = reverse('login')
            return redirect(f"{login_url}?next={request.get_full_path()}")
        if not request.user.is_staff:
            raise PermissionDenied('Chỉ tài khoản quản trị hoặc nhân viên được cấp quyền mới được truy cập khu vực quản trị.')
        return view_func(request, *args, **kwargs)

    return wrapped


def build_review_summary(review_queryset):
    aggregate = review_queryset.aggregate(avg=Avg('rating'), count=Count('id'))
    average = float(aggregate.get('avg') or 0)
    count = int(aggregate.get('count') or 0)
    distribution_map = {
        row['rating']: row['total']
        for row in review_queryset.order_by().values('rating').annotate(total=Count('id')).order_by('-rating')
    }
    distribution = []
    for rating in range(5, 0, -1):
        rating_count = distribution_map.get(rating, 0)
        distribution.append({
            'rating': rating,
            'count': rating_count,
            'percent': round((rating_count / count) * 100, 1) if count else 0,
        })

    return {
        'average': round(average, 1) if count else 0,
        'count': count,
        'rounded': int(round(average)) if count else 0,
        'distribution': distribution,
    }


def parse_review_rating(value):
    try:
        rating = int(value)
    except (TypeError, ValueError):
        return None
    return rating if 1 <= rating <= 5 else None


def get_review_author_name(user):
    first_name = (getattr(user, 'first_name', '') or '').strip()
    return first_name or user.username


def serialize_review_item(review):
    was_updated = bool(getattr(review, 'was_updated_by_user', False))
    author_name = get_review_author_name(review.user)
    return {
        'id': review.id,
        'author_name': author_name,
        'rating': review.rating,
        'comment': review.comment or 'Khách hàng chưa để lại cảm nhận chi tiết.',
        'updated_at': review.updated_at.strftime('%d/%m/%Y %H:%M'),
        'was_updated': was_updated,
        'update_note': f'Đánh giá do người dùng "{author_name}" cập nhật lại.' if was_updated else '',
    }


def build_review_panel(review_queryset, active_rating=None, page_number=1, per_page=5):
    if active_rating:
        review_queryset = review_queryset.filter(rating=active_rating)

    paginator = Paginator(review_queryset, per_page)
    page_obj = paginator.get_page(page_number)
    items = list(page_obj.object_list)

    return {
        'items': items,
        'serialized_items': [serialize_review_item(item) for item in items],
        'active_rating': active_rating,
        'page': page_obj.number,
        'has_next': page_obj.has_next(),
        'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
        'total_count': paginator.count,
        'per_page': per_page,
    }


def user_has_completed_purchase_for_medicine(user, medicine):
    if not user.is_authenticated:
        return False
    return OrderItem.objects.filter(
        order__user=user,
        order__status=Order.STATUS_COMPLETED,
        medicine=medicine,
    ).exists()


def user_has_completed_purchase_for_pharmacy(user, pharmacy):
    if not user.is_authenticated:
        return False
    return Order.objects.filter(
        user=user,
        pharmacy=pharmacy,
        status=Order.STATUS_COMPLETED,
    ).exists()


def can_user_review_medicine(user, medicine, existing_review=None):
    return bool(existing_review) or user_has_completed_purchase_for_medicine(user, medicine)


def can_user_review_pharmacy(user, pharmacy, existing_review=None):
    return bool(existing_review) or user_has_completed_purchase_for_pharmacy(user, pharmacy)


def build_review_access_context(is_authenticated, can_review, entity_label):
    if not is_authenticated:
        return {
            'can_review': False,
            'message': f'Bạn cần đăng nhập để gửi đánh giá cho {entity_label}.',
        }
    if can_review:
        return {
            'can_review': True,
            'message': '',
        }
    return {
        'can_review': False,
        'message': f'Chỉ khách hàng đã mua và có đơn hàng hoàn thành mới được đánh giá {entity_label}.',
    }


def build_review_api_payload(review_panel):
    return {
        'items': review_panel['serialized_items'],
        'has_next': review_panel['has_next'],
        'next_page': review_panel['next_page'],
        'page': review_panel['page'],
        'active_rating': review_panel['active_rating'],
        'total_count': review_panel['total_count'],
        'per_page': review_panel['per_page'],
    }




def build_order_estimated_delivery_at(created_at, departure_time_str, duration_minutes):
    base_dt = timezone.localtime(created_at or timezone.now())
    normalized_time = normalize_departure_time_str(departure_time_str)

    try:
        departure_hour, departure_minute = [int(part) for part in normalized_time.split(':', 1)]
    except (TypeError, ValueError):
        departure_hour = base_dt.hour
        departure_minute = base_dt.minute

    estimated_start = base_dt.replace(hour=departure_hour, minute=departure_minute, second=0, microsecond=0)
    if estimated_start < base_dt:
        estimated_start = base_dt

    try:
        duration_value = max(int(round(float(duration_minutes or 0))), 0)
    except (TypeError, ValueError):
        duration_value = 0

    return estimated_start + timedelta(minutes=duration_value)


def complete_order_workflow(order, *, completed_by_customer=False, auto_completed=False):
    if order.status == Order.STATUS_COMPLETED:
        return order

    now = timezone.now()

    with transaction.atomic():
        if order.pk:
            order = Order.objects.select_for_update().get(pk=order.pk)

        order.status = Order.STATUS_COMPLETED
        order.completed_at = now

        if completed_by_customer:
            order.received_confirmed_at = now
        if auto_completed:
            order.auto_completed_at = now

        if order.payment_method == Order.PAYMENT_COD:
            order.payment_status = Order.PAYMENT_STATUS_PAID
            if order.payment_confirmed_at is None:
                order.payment_confirmed_at = now

        order.save()

    return order


def auto_complete_order_if_due(order, now=None):
    now = now or timezone.now()
    if order.status != Order.STATUS_SHIPPING:
        return False

    if order.requires_payment_confirmation and order.payment_status != Order.PAYMENT_STATUS_PAID:
        return False

    deadline = order.auto_complete_deadline_at
    if deadline is None or deadline > now:
        return False

    complete_order_workflow(order, auto_completed=True)
    return True


def auto_complete_overdue_shipping_orders(order_queryset):
    now = timezone.now()
    candidate_orders = list(order_queryset.filter(status=Order.STATUS_SHIPPING))
    completed_count = 0
    for order in candidate_orders:
        if auto_complete_order_if_due(order, now=now):
            completed_count += 1
    return completed_count


def get_customer_order_status_meta(order):
    return_request = getattr(order, "user_return_request", None) or getattr(order, "return_request", None)
    if return_request is not None:
        if return_request.status == ReturnRefundRequest.STATUS_PROCESSING:
            return {
                "label": "Trả hàng / hoàn tiền: Đang xử lý",
                "badge_class": "history-status--refund-processing",
                "short_label": "Đang xử lý hoàn tiền",
            }
        if return_request.status == ReturnRefundRequest.STATUS_APPROVED:
            return {
                "label": "Trả hàng / hoàn tiền: Chấp nhận hoàn tiền",
                "badge_class": "history-status--refund-approved",
                "short_label": "Đã chấp nhận hoàn tiền",
            }
        return {
            "label": "Trả hàng / hoàn tiền: Từ chối hoàn tiền",
            "badge_class": "history-status--refund-rejected",
            "short_label": "Đã từ chối hoàn tiền",
        }

    mapping = {
        Order.STATUS_PENDING: {
            "label": order.get_status_display(),
            "badge_class": "history-status--pending",
            "short_label": order.get_status_display(),
        },
        Order.STATUS_CONFIRMED: {
            "label": order.get_status_display(),
            "badge_class": "history-status--confirmed",
            "short_label": order.get_status_display(),
        },
        Order.STATUS_PACKING: {
            "label": order.get_status_display(),
            "badge_class": "history-status--packing",
            "short_label": order.get_status_display(),
        },
        Order.STATUS_SHIPPING: {
            "label": order.get_status_display(),
            "badge_class": "history-status--shipping",
            "short_label": order.get_status_display(),
        },
        Order.STATUS_COMPLETED: {
            "label": order.get_status_display(),
            "badge_class": "history-status--completed",
            "short_label": order.get_status_display(),
        },
        Order.STATUS_CANCELLED: {
            "label": order.get_status_display(),
            "badge_class": "history-status--cancelled",
            "short_label": order.get_status_display(),
        },
        Order.STATUS_FAILED_DELIVERY: {
            "label": order.get_status_display(),
            "badge_class": "history-status--failed-delivery",
            "short_label": order.get_status_display(),
        },
    }
    return mapping.get(order.status, {
        "label": order.get_status_display(),
        "badge_class": "",
        "short_label": order.get_status_display(),
    })


def decorate_order_for_customer_display(order):
    order.user_return_request = getattr(order, "return_request", None)
    status_meta = get_customer_order_status_meta(order)
    order.customer_status_label = status_meta["label"]
    order.customer_status_short_label = status_meta["short_label"]
    order.customer_status_badge_class = status_meta["badge_class"]
    return order


def get_order_for_customer_or_404(user, order_id):
    order = get_object_or_404(
        Order.objects.select_related('pharmacy', 'user').prefetch_related('items__medicine', 'return_request', 'prescription_proof_images'),
        pk=order_id,
        user=user,
    )
    return decorate_order_for_customer_display(order)


def build_order_history_queryset_for_user(user, filter_state=None):
    base_queryset = Order.objects.filter(user=user).select_related('pharmacy').prefetch_related('items__medicine', 'return_request', 'prescription_proof_images').order_by('-created_at', '-id')
    auto_complete_overdue_shipping_orders(base_queryset)
    queryset = Order.objects.filter(user=user).select_related('pharmacy').prefetch_related('items__medicine', 'return_request', 'prescription_proof_images').order_by('-created_at', '-id')
    if filter_state:
        queryset = apply_order_history_filters(queryset, filter_state)
    orders = list(queryset)
    return [decorate_order_for_customer_display(order) for order in orders]


def build_return_request_initial(order, existing_request=None):
    initial = {
        'contact_phone': (existing_request.contact_phone if existing_request else '') or order.phone,
        'contact_email': (existing_request.contact_email if existing_request else '') or getattr(order.user, 'email', ''),
    }
    return initial


def get_order_history_filter_state(request):
    return {
        "status": (request.GET.get("status") or "").strip(),
        "payment_method": (request.GET.get("payment_method") or "").strip(),
        "refund_status": (request.GET.get("refund_status") or "").strip(),
    }


def apply_order_history_filters(queryset, filter_state):
    status_value = filter_state.get("status") or ""
    payment_method = filter_state.get("payment_method") or ""
    refund_status = filter_state.get("refund_status") or ""

    if status_value in {choice[0] for choice in Order.STATUS_CHOICES}:
        queryset = queryset.filter(status=status_value)

    if payment_method in {choice[0] for choice in Order.PAYMENT_METHOD_CHOICES}:
        queryset = queryset.filter(payment_method=payment_method)

    if refund_status == "none":
        queryset = queryset.filter(return_request__isnull=True)
    elif refund_status in {choice[0] for choice in ReturnRefundRequest.STATUS_CHOICES}:
        queryset = queryset.filter(return_request__status=refund_status)

    return queryset


def get_invoice_staff_display_name(order):
    return order.resolved_invoice_staff_name


def get_user_full_name_or_username(user):
    if not user:
        return ""
    return get_user_display_name(user) if getattr(user, "is_authenticated", False) else str(user)


def parse_excel_date_value(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Không đọc được hạn sử dụng '{raw}'.")


def parse_excel_int(value, field_label):
    if value in (None, ""):
        raise ValueError(f"Thiếu giá trị cho cột {field_label}.")
    try:
        decimal_value = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f"Giá trị '{value}' ở cột {field_label} không hợp lệ.")
    integer_value = int(decimal_value)
    if integer_value < 0:
        raise ValueError(f"Giá trị ở cột {field_label} không được âm.")
    return integer_value


def normalize_excel_header(value):
    normalized = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^a-zA-Z0-9]+", "_", normalized).strip("_").lower()
    return normalized


def get_purchase_import_header_map(header_row):
    aliases = {
        "medicine_id": {"medicine_id", "ma_thuoc", "id_thuoc", "product_id", "id_san_pham"},
        "medicine_name": {"medicine_name", "ten_thuoc", "ten_san_pham", "product_name", "name"},
        "product_type": {"product_type", "loai_san_pham", "loai_thuoc", "type"},
        "manufacturer": {"manufacturer", "nha_san_xuat", "hang_san_xuat"},
        "unit": {"unit", "don_vi", "don_vi_tinh"},
        "quantity": {"quantity", "so_luong", "qty"},
        "expiry_date": {"expiry_date", "han_su_dung", "expiry", "hsd"},
        "import_price": {"import_price", "gia_nhap", "don_gia_nhap"},
        "sale_price": {"sale_price", "gia_ban", "don_gia_ban", "price"},
        "category": {"category", "danh_muc", "nhom_hang"},
        "origin": {"origin", "xuat_xu"},
        "description": {"description", "mo_ta"},
        "usage": {"usage", "cong_dung"},
        "ingredients": {"ingredients", "thanh_phan"},
        "dosage": {"dosage", "cach_dung", "lieu_dung"},
        "note": {"note", "ghi_chu"},
    }
    normalized_headers = [normalize_excel_header(cell) for cell in header_row]
    header_map = {}
    for index, normalized in enumerate(normalized_headers):
        for canonical, names in aliases.items():
            if normalized in names and canonical not in header_map:
                header_map[canonical] = index
    return header_map


def parse_optional_excel_int(value, field_label):
    if value in (None, ""):
        return None
    return parse_excel_int(value, field_label)


def choose_richer_text_variant(*values):
    candidates = []
    for index, raw_value in enumerate(values):
        text = str(raw_value or "").strip()
        if not text:
            continue
        accent_score = sum(1 for character in text if ord(character) > 127)
        alpha_score = sum(1 for character in text if character.isalpha())
        candidates.append(((accent_score, alpha_score, len(text), -index), text))

    if not candidates:
        return ""
    return max(candidates, key=lambda item: item[0])[1]


def build_medicine_short_description(*values):
    for raw_value in values:
        text = strip_tags(str(raw_value or "")).strip()
        if text:
            return text
    return ""


def normalize_product_type_value(raw_value, *, fallback=MEDICINE_PRODUCT_TYPE_MEDICINE):
    folded_value = fold_text_for_match(raw_value)
    if folded_value in {"thuoc", "medicine", "drug", "duoc pham"}:
        return MEDICINE_PRODUCT_TYPE_MEDICINE
    if folded_value in {"thuc pham chuc nang", "supplement", "functional food", "tpcn"}:
        return MEDICINE_PRODUCT_TYPE_SUPPLEMENT
    return fallback


def choose_representative_medicine(candidates):
    def medicine_rank(medicine):
        metadata_score = sum(
            1
            for field_name in ("description", "usage", "ingredients", "dosage", "origin", "category")
            if str(getattr(medicine, field_name, "") or "").strip()
        )
        has_media = bool(getattr(getattr(medicine, "image", None), "name", "") or str(getattr(medicine, "gallery_urls", "") or "").strip())
        return (
            0 if has_media else 1,
            -metadata_score,
            0 if int(getattr(medicine, "quantity", 0) or 0) > 0 else 1,
            -int(getattr(medicine, "quantity", 0) or 0),
            medicine.id,
        )

    return min(candidates, key=medicine_rank)


def harmonize_imported_medicine_catalog_metadata(medicine, row_payload, *, template_medicine=None):
    if medicine is None:
        return []

    previous_catalog_key = build_medicine_catalog_key(medicine.name, medicine.unit, medicine.manufacturer)
    updated_fields = []

    preferred_name = choose_richer_text_variant(
        row_payload.get("medicine_name"),
        getattr(template_medicine, "name", ""),
        medicine.name,
    )
    if preferred_name and preferred_name != medicine.name:
        medicine.name = preferred_name
        updated_fields.append("name")

    preferred_unit = choose_richer_text_variant(
        row_payload.get("unit"),
        getattr(template_medicine, "unit", ""),
        medicine.unit,
    )
    if not preferred_unit:
        preferred_unit = "Hộp"
    if preferred_unit and preferred_unit != medicine.unit:
        medicine.unit = preferred_unit
        updated_fields.append("unit")

    preferred_manufacturer = choose_richer_text_variant(
        row_payload.get("manufacturer"),
        getattr(template_medicine, "manufacturer", ""),
        medicine.manufacturer,
    )
    if preferred_manufacturer != medicine.manufacturer:
        medicine.manufacturer = preferred_manufacturer
        updated_fields.append("manufacturer")

    preferred_product_type = normalize_product_type_value(
        row_payload.get("product_type"),
        fallback=getattr(template_medicine, "product_type", getattr(medicine, "product_type", MEDICINE_PRODUCT_TYPE_MEDICINE)),
    )
    if preferred_product_type != medicine.product_type:
        medicine.product_type = preferred_product_type
        updated_fields.append("product_type")

    short_description = build_medicine_short_description(
        row_payload.get("short_description"),
        row_payload.get("description"),
        getattr(template_medicine, "short_description", ""),
        getattr(medicine, "short_description", ""),
    )
    if short_description != getattr(medicine, "short_description", ""):
        medicine.short_description = short_description
        updated_fields.append("short_description")

    text_priority_fields = ("category", "origin", "description", "usage", "ingredients", "dosage")
    for field_name in text_priority_fields:
        preferred_value = (
            str(row_payload.get(field_name) or "").strip()
            or str(getattr(template_medicine, field_name, "") or "").strip()
            or str(getattr(medicine, field_name, "") or "").strip()
        )
        if preferred_value != getattr(medicine, field_name):
            setattr(medicine, field_name, preferred_value)
            updated_fields.append(field_name)

    if template_medicine and not medicine.image and template_medicine.image:
        medicine.image = template_medicine.image
        updated_fields.append("image")

    if template_medicine and not medicine.gallery_urls and template_medicine.gallery_urls:
        medicine.gallery_urls = template_medicine.gallery_urls
        updated_fields.append("gallery_urls")

    if template_medicine and template_medicine.prescription_required and not medicine.prescription_required:
        medicine.prescription_required = True
        updated_fields.append("prescription_required")

    if updated_fields:
        medicine.save(update_fields=list(dict.fromkeys(updated_fields)))
        sync_medicine_catalog_metadata(
            medicine,
            previous_catalog_key=previous_catalog_key,
            field_names=updated_fields,
        )

    return updated_fields


def find_global_medicine_template(row_payload):
    medicine_id = str(row_payload.get("medicine_id") or "").strip()
    if medicine_id.isdigit():
        template = Medicine.objects.filter(pk=int(medicine_id)).first()
        if template:
            return template

    medicine_name = (row_payload.get("medicine_name") or "").strip()
    if not medicine_name:
        return None

    manufacturer = (row_payload.get("manufacturer") or "").strip()
    unit = (row_payload.get("unit") or "").strip()
    normalized_name = fold_text_for_match(medicine_name)
    normalized_manufacturer = fold_text_for_match(manufacturer)
    normalized_unit = fold_text_for_match(unit)

    queryset = Medicine.objects.all().order_by("id")
    for candidate in queryset:
        if fold_text_for_match(candidate.name) != normalized_name:
            continue
        if normalized_manufacturer and fold_text_for_match(candidate.manufacturer) != normalized_manufacturer:
            continue
        if normalized_unit and fold_text_for_match(candidate.unit) != normalized_unit:
            continue
        return candidate
    return None


def get_or_create_medicine_for_import(pharmacy, row_payload, *, row_number=None):
    medicine_id = str(row_payload.get("medicine_id") or "").strip()
    manufacturer = (row_payload.get("manufacturer") or "").strip()
    unit = (row_payload.get("unit") or "").strip()
    template_medicine = find_global_medicine_template(row_payload)

    if medicine_id.isdigit():
        direct_match = Medicine.objects.select_for_update().filter(pk=int(medicine_id), pharmacy=pharmacy).first()
        if direct_match:
            harmonize_imported_medicine_catalog_metadata(
                direct_match,
                row_payload,
                template_medicine=template_medicine,
            )
            return direct_match, False

    medicine_name = (row_payload.get("medicine_name") or "").strip()
    if not medicine_name and template_medicine:
        medicine_name = template_medicine.name
    if not manufacturer and template_medicine:
        manufacturer = template_medicine.manufacturer or ""
    if not unit and template_medicine:
        unit = template_medicine.unit or ""

    if not medicine_name:
        row_suffix = f" ở dòng {row_number}" if row_number else ""
        raise ValueError(f"Mỗi dòng nhập hàng phải có mã thuốc hợp lệ hoặc tên thuốc{row_suffix}.")

    normalized_name = fold_text_for_match(medicine_name)
    normalized_manufacturer = fold_text_for_match(manufacturer)
    normalized_unit = fold_text_for_match(unit)

    queryset = Medicine.objects.select_for_update().filter(pharmacy=pharmacy).order_by("id")
    matches = []
    for candidate in queryset:
        if fold_text_for_match(candidate.name) != normalized_name:
            continue
        if normalized_manufacturer and fold_text_for_match(candidate.manufacturer) != normalized_manufacturer:
            continue
        if normalized_unit and fold_text_for_match(candidate.unit) != normalized_unit:
            continue
        matches.append(candidate)
        if len(matches) > 1:
            break

    if matches:
        medicine = choose_representative_medicine(matches)
        harmonize_imported_medicine_catalog_metadata(
            medicine,
            row_payload,
            template_medicine=template_medicine,
        )
        return medicine, False

    import_price = parse_optional_excel_int(
        row_payload.get("import_price"),
        f"Giá nhập (dòng {row_number})" if row_number else "Giá nhập",
    ) or 0
    sale_price = parse_optional_excel_int(
        row_payload.get("sale_price"),
        f"Giá bán (dòng {row_number})" if row_number else "Giá bán",
    )
    selling_price = sale_price if sale_price is not None else (template_medicine.price if template_medicine and template_medicine.price else import_price)

    medicine = Medicine.objects.create(
        pharmacy=pharmacy,
        name=medicine_name if not template_medicine else template_medicine.name if fold_text_for_match(template_medicine.name)==normalized_name else medicine_name,
        product_type=normalize_product_type_value(
            row_payload.get("product_type"),
            fallback=getattr(template_medicine, "product_type", MEDICINE_PRODUCT_TYPE_MEDICINE),
        ),
        category=(row_payload.get("category") or (template_medicine.category if template_medicine else "") or "").strip(),
        unit=unit or (template_medicine.unit if template_medicine else "") or "Hộp",
        manufacturer=manufacturer or (template_medicine.manufacturer if template_medicine else ""),
        origin=(row_payload.get("origin") or (template_medicine.origin if template_medicine else "") or "").strip(),
        price=selling_price or 0,
        quantity=0,
        image=(template_medicine.image if template_medicine and template_medicine.image else None),
        gallery_urls=(template_medicine.gallery_urls if template_medicine and template_medicine.gallery_urls else ""),
        short_description=build_medicine_short_description(
            row_payload.get("short_description"),
            row_payload.get("description"),
            getattr(template_medicine, "short_description", ""),
        ),
        description=(row_payload.get("description") or (template_medicine.description if template_medicine else "") or "").strip(),
        usage=(row_payload.get("usage") or (template_medicine.usage if template_medicine else "") or "").strip(),
        ingredients=(row_payload.get("ingredients") or (template_medicine.ingredients if template_medicine else "") or "").strip(),
        dosage=(row_payload.get("dosage") or (template_medicine.dosage if template_medicine else "") or "").strip(),
        prescription_required=(template_medicine.prescription_required if template_medicine else False),
    )
    harmonize_imported_medicine_catalog_metadata(
        medicine,
        row_payload,
        template_medicine=template_medicine,
    )
    sync_medicine_catalog_metadata(
        medicine,
        field_names=MEDICINE_SHARED_SYNC_FIELDS,
    )
    return medicine, True


def build_purchase_import_preview_payload(user):
    managed_pharmacy = get_admin_scope_pharmacy(user)
    pharmacy_queryset = Pharmacy.objects.order_by("name")
    if managed_pharmacy is not None:
        pharmacy_queryset = pharmacy_queryset.filter(pk=managed_pharmacy.pk)

    pharmacy_ids = list(pharmacy_queryset.values_list("id", flat=True))
    medicine_queryset = Medicine.objects.filter(pharmacy_id__in=pharmacy_ids).only("id", "pharmacy_id", "name", "manufacturer", "unit")

    pharmacy_catalog = {str(pharmacy_id): [] for pharmacy_id in pharmacy_ids}
    template_catalog = {}

    for medicine in medicine_queryset:
        catalog_item = {
            "id": str(medicine.id),
            "match_name": fold_text_for_match(medicine.name),
            "match_manufacturer": fold_text_for_match(medicine.manufacturer),
            "match_unit": fold_text_for_match(medicine.unit),
            "display_name": medicine.name,
            "display_manufacturer": medicine.manufacturer or "",
            "display_unit": medicine.unit or "",
        }
        pharmacy_catalog.setdefault(str(medicine.pharmacy_id), []).append(catalog_item)
        template_catalog[str(medicine.id)] = catalog_item

    return {
        "all_branches_value": PurchaseImportExcelForm.ALL_BRANCHES_VALUE,
        "pharmacy_catalog": pharmacy_catalog,
        "template_catalog": template_catalog,
    }


def get_active_lots_queryset(queryset=None):
    queryset = queryset if queryset is not None else MedicineLot.objects.all()
    return queryset.filter(remaining_quantity__gt=0)


def get_sellable_lots_queryset(queryset=None, *, today=None):
    today = today or timezone.localdate()
    queryset = get_active_lots_queryset(queryset)
    return queryset.filter(Q(expiry_date__isnull=True) | Q(expiry_date__gte=today))


def ensure_medicine_has_inventory_lot_snapshot(medicine):
    if medicine is None or not getattr(medicine, "pk", None):
        return None
    if MedicineLot.objects.filter(medicine=medicine).exists():
        return None

    current_quantity = int(getattr(medicine, "quantity", 0) or 0)
    if current_quantity <= 0:
        return None

    return MedicineLot.objects.create(
        medicine=medicine,
        pharmacy=medicine.pharmacy,
        source_type=MedicineLot.SOURCE_MANUAL,
        source_label="Ton kho dong bo",
        import_price=0,
        expiry_date=medicine.expiry_date,
        received_quantity=current_quantity,
        remaining_quantity=current_quantity,
        note="Dong bo ton kho tu du lieu cu chua co lich su lo.",
    )


def get_fefo_lot_queryset_for_medicine(medicine, *, lock=False, include_expired=False):
    queryset = MedicineLot.objects.filter(medicine=medicine, pharmacy=medicine.pharmacy, remaining_quantity__gt=0)
    if lock:
        queryset = queryset.select_for_update()
    if not include_expired:
        queryset = get_sellable_lots_queryset(queryset)
    return queryset.order_by(F('expiry_date').asc(nulls_last=True), 'created_at', 'id')


def get_stock_export_lot_queryset_for_medicine(medicine, export_scope, *, lock=False):
    if export_scope == StockExportBatch.EXPORT_SCOPE_EXPIRED:
        queryset = MedicineLot.objects.filter(
            medicine=medicine,
            pharmacy=medicine.pharmacy,
            remaining_quantity__gt=0,
            expiry_date__isnull=False,
            expiry_date__lt=timezone.localdate(),
        )
        if lock:
            queryset = queryset.select_for_update()
        return queryset.order_by(F('expiry_date').asc(nulls_last=True), 'created_at', 'id')
    return get_fefo_lot_queryset_for_medicine(medicine, lock=lock)


def recalculate_medicine_inventory_snapshot(medicine):
    if medicine is None or not getattr(medicine, 'pk', None):
        return False

    today = timezone.localdate()
    remaining_queryset = MedicineLot.objects.filter(medicine=medicine, remaining_quantity__gt=0)
    sellable_total = (
        remaining_queryset.filter(Q(expiry_date__isnull=True) | Q(expiry_date__gte=today))
        .aggregate(total=Sum('remaining_quantity'))
        .get('total')
        or 0
    )
    next_expiry = remaining_queryset.exclude(expiry_date__isnull=True).order_by('expiry_date', 'id').values_list('expiry_date', flat=True).first()

    update_fields = []
    if medicine.quantity != sellable_total:
        medicine.quantity = sellable_total
        update_fields.append('quantity')
    if medicine.expiry_date != next_expiry:
        medicine.expiry_date = next_expiry
        update_fields.append('expiry_date')

    if update_fields:
        medicine.save(update_fields=update_fields)
        return True
    return False


def update_medicine_expiry_from_import(medicine, imported_expiry_date):
    return recalculate_medicine_inventory_snapshot(medicine)


def recalculate_medicine_expiry_from_import_history(medicine, *, exclude_batch_id=None):
    return recalculate_medicine_inventory_snapshot(medicine)


def build_purchase_import_row_payload(header_map, row_values):
    payload = {}
    for canonical, index in header_map.items():
        if index < len(row_values):
            payload[canonical] = row_values[index]
    return payload


def create_import_lot_for_item(batch, purchase_item):
    return MedicineLot.objects.create(
        medicine=purchase_item.medicine,
        pharmacy=batch.pharmacy,
        purchase_batch=batch,
        purchase_item=purchase_item,
        source_type=MedicineLot.SOURCE_IMPORT,
        source_label=batch.resolved_invoice_code,
        import_price=purchase_item.import_price or 0,
        expiry_date=purchase_item.expiry_date,
        received_quantity=purchase_item.imported_quantity,
        remaining_quantity=purchase_item.imported_quantity,
        note=purchase_item.note or '',
    )


_cached_pdf_font_name = None


def get_receipt_font_name():
    global _cached_pdf_font_name
    if _cached_pdf_font_name:
        return _cached_pdf_font_name
    if pdfmetrics is None or TTFont is None:
        raise RuntimeError("Máy đang thiếu thư viện reportlab nên chưa thể tạo PDF phiếu kho.")
    if PDF_FONT_PATH.exists():
        try:
            pdfmetrics.registerFont(TTFont(PDF_FONT_NAME, str(PDF_FONT_PATH)))
            _cached_pdf_font_name = PDF_FONT_NAME
            return _cached_pdf_font_name
        except Exception:
            pass
    _cached_pdf_font_name = "Helvetica"
    return _cached_pdf_font_name


def build_receipt_styles():
    font_name = get_receipt_font_name()
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "GISReceiptTitle",
            parent=styles["Title"],
            fontName=font_name,
            fontSize=18,
            leading=24,
            alignment=1,
            spaceAfter=10,
        ),
        "subtitle": ParagraphStyle(
            "GISReceiptSubtitle",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=10,
            leading=14,
            alignment=1,
            textColor=colors.HexColor("#4f5f7d"),
        ),
        "body": ParagraphStyle(
            "GISReceiptBody",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=9,
            leading=13,
        ),
        "body_bold": ParagraphStyle(
            "GISReceiptBodyBold",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=9,
            leading=13,
        ),
        "table_header": ParagraphStyle(
            "GISReceiptTableHeader",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=8.5,
            leading=11,
            textColor=colors.white,
            alignment=1,
        ),
        "small": ParagraphStyle(
            "GISReceiptSmall",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=8,
            leading=11,
            textColor=colors.HexColor("#627394"),
        ),
    }


def to_receipt_paragraph(value, style, *, allow_markup=False):
    raw_value = str(value or "").strip() or "-"
    text = raw_value.replace("\n", "<br/>") if allow_markup else escape(raw_value).replace("\n", "<br/>")
    return Paragraph(text, style)


def sanitize_receipt_filename(value, prefix):
    raw_value = re.sub(r"[^A-Za-z0-9._-]+", "-", str(value or "").strip()).strip("-") or prefix
    return f"{raw_value}.pdf"


def get_admin_user_role_label(admin_user):
    if not getattr(admin_user, "is_authenticated", False):
        return ""
    if admin_user.is_superuser:
        return "Quản trị viên hệ thống"
    if admin_user.is_staff:
        return "Nhân viên chi nhánh"
    return "Người dùng"


def build_admin_receipt_identity(admin_user, explicit_name=""):
    if not getattr(admin_user, "is_authenticated", False):
        return {
            "name": explicit_name or "Chưa cập nhật",
            "email": "",
            "role": "",
        }
    profile = getattr(admin_user, "profile", None) or get_or_create_user_profile(admin_user)
    name = (
        (explicit_name or "").strip()
        or (getattr(profile, "full_name", "") or "").strip()
        or (admin_user.get_full_name() or "").strip()
        or admin_user.username
    )
    return {
        "name": name,
        "email": (admin_user.email or "").strip(),
        "role": get_admin_user_role_label(admin_user),
    }


def build_receipt_pdf_bytes(
    *,
    title,
    subtitle,
    info_pairs,
    table_headers,
    table_rows,
    totals_pairs=None,
    footer_note="",
    signature_left=None,
    signature_right=None,
):
    if SimpleDocTemplate is None:
        raise RuntimeError("Máy đang thiếu thư viện reportlab nên chưa thể tạo PDF phiếu kho.")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    styles = build_receipt_styles()
    story = [
        Paragraph(escape(title), styles["title"]),
        Paragraph(escape(subtitle), styles["subtitle"]),
        Spacer(1, 8),
    ]

    info_data = []
    current_row = []
    for label, value in info_pairs:
        current_row.extend(
            [
                to_receipt_paragraph(f"<b>{escape(label)}:</b>", styles["body_bold"], allow_markup=True),
                to_receipt_paragraph(value, styles["body"]),
            ]
        )
        if len(current_row) == 4:
            info_data.append(current_row)
            current_row = []
    if current_row:
        while len(current_row) < 4:
            current_row.append("")
        info_data.append(current_row)

    info_table = Table(info_data, colWidths=[30 * mm, 57 * mm, 30 * mm, 57 * mm])
    info_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f6f8fc")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#d7deea")),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d7deea")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.extend([info_table, Spacer(1, 10)])

    rendered_rows = [
        [Paragraph(escape(str(header)), styles["table_header"]) for header in table_headers]
    ]
    for row in table_rows:
        rendered_rows.append([to_receipt_paragraph(value, styles["body"]) for value in row])

    line_table = Table(rendered_rows, repeatRows=1)
    line_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f64e0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d7deea")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fbff")]),
            ]
        )
    )
    story.extend([line_table, Spacer(1, 10)])

    if totals_pairs:
        totals_data = [
            [
                to_receipt_paragraph(f"<b>{escape(label)}</b>", styles["body_bold"], allow_markup=True),
                to_receipt_paragraph(value, styles["body"]),
            ]
            for label, value in totals_pairs
        ]
        totals_table = Table(totals_data, colWidths=[55 * mm, 125 * mm], hAlign="RIGHT")
        totals_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f6f8fc")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#d7deea")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d7deea")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.extend([totals_table, Spacer(1, 10)])

    if footer_note:
        story.append(Paragraph(escape(footer_note), styles["small"]))

    if signature_left or signature_right:
        signature_left = signature_left or {}
        signature_right = signature_right or {}
        story.extend([Spacer(1, 18)])
        signature_table = Table(
            [
                [
                    Paragraph(escape(signature_left.get("title", "Người lập phiếu")), styles["body_bold"]),
                    Paragraph(escape(signature_right.get("title", "Xác nhận")), styles["body_bold"]),
                ],
                [
                    Paragraph("", styles["body"]),
                    Paragraph("", styles["body"]),
                ],
                [
                    to_receipt_paragraph(
                        f"<b>{escape(signature_left.get('name', ''))}</b><br/>{escape(signature_left.get('role', ''))}",
                        styles["body"],
                        allow_markup=True,
                    ),
                    to_receipt_paragraph(
                        f"<b>{escape(signature_right.get('name', ''))}</b><br/>{escape(signature_right.get('role', ''))}",
                        styles["body"],
                        allow_markup=True,
                    ),
                ],
            ],
            colWidths=[90 * mm, 90 * mm],
        )
        signature_table.setStyle(
            TableStyle(
                [
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#20395f")),
                ]
            )
        )
        story.append(signature_table)

    doc.build(story)
    return buffer.getvalue()


def ensure_purchase_import_receipt_pdf(batch, *, force=False):
    if getattr(batch, "receipt_pdf", None) and getattr(batch.receipt_pdf, "name", "") and not force:
        return batch.receipt_pdf
    items = list(batch.items.all())
    total_amount = sum((item.import_price or 0) * (item.imported_quantity or 0) for item in items)
    pdf_bytes = build_receipt_pdf_bytes(
        title="PHIẾU NHẬP KHO",
        subtitle="Hệ thống GIS Pharma",
        info_pairs=[
            ("Mã phiếu", batch.resolved_invoice_code),
            ("Chi nhánh", getattr(batch.pharmacy, "name", "")),
            ("Ngày lập", timezone.localtime(batch.created_at).strftime("%d/%m/%Y %H:%M") if batch.created_at else "-"),
            ("Người lập phiếu", batch.resolved_imported_by_name),
            ("Email", batch.imported_by_email or "-"),
            ("Chức vụ", batch.imported_by_role or "-"),
            ("Số dòng", str(batch.total_lines)),
            ("File nguồn", getattr(batch.source_file, "name", "") or "-"),
            ("Ghi chú", batch.note or "-"),
        ],
        table_headers=["STT", "Sản phẩm", "Đơn vị", "Tồn trước", "Nhập", "Tồn sau", "Giá nhập", "HSD"],
        table_rows=[
            [
                str(index),
                item.medicine_name,
                item.unit or "-",
                str(item.previous_quantity),
                str(item.imported_quantity),
                str(item.new_quantity),
                format_vnd(item.import_price),
                item.expiry_date.strftime("%d/%m/%Y") if item.expiry_date else "-",
            ]
            for index, item in enumerate(items, start=1)
        ],
        totals_pairs=[
            ("Tổng số lượng nhập", str(batch.total_quantity)),
            ("Tổng giá trị nhập", format_vnd(total_amount)),
        ],
        footer_note="Phiếu PDF được hệ thống tạo tự động và lưu trực tiếp trong PostgreSQL.",
        signature_left={
            "title": "Người lập phiếu",
            "name": batch.resolved_imported_by_name,
            "role": batch.imported_by_role or "",
        },
    )
    file_name = sanitize_receipt_filename(batch.resolved_invoice_code, "phieu-nhap-kho")
    batch.receipt_pdf.save(file_name, ContentFile(pdf_bytes), save=False)
    batch.save(update_fields=["receipt_pdf"])
    return batch.receipt_pdf


def ensure_stock_export_receipt_pdf(batch, *, force=False):
    if getattr(batch, "receipt_pdf", None) and getattr(batch.receipt_pdf, "name", "") and not force:
        return batch.receipt_pdf
    items = list(batch.items.all())
    pdf_bytes = build_receipt_pdf_bytes(
        title="PHIẾU XUẤT KHO",
        subtitle="Hệ thống GIS Pharma",
        info_pairs=[
            ("Mã phiếu", batch.resolved_export_code),
            ("Chi nhánh", getattr(batch.pharmacy, "name", "")),
            ("Ngày lập", timezone.localtime(batch.created_at).strftime("%d/%m/%Y %H:%M") if batch.created_at else "-"),
            ("Người lập phiếu", batch.resolved_exported_by_name),
            ("Email", batch.exported_by_email or "-"),
            ("Chức vụ", batch.exported_by_role or "-"),
            ("Nơi nhận / mục đích", batch.destination_name or "-"),
            ("Số dòng", str(batch.total_lines)),
            ("Ghi chú", batch.note or "-"),
        ],
        table_headers=["STT", "Sản phẩm", "Đơn vị", "Tồn trước", "Xuất", "Tồn sau", "Ghi chú"],
        table_rows=[
            [
                str(index),
                item.medicine_name,
                item.unit or "-",
                str(item.previous_quantity),
                str(item.exported_quantity),
                str(item.remaining_quantity),
                item.note or "-",
            ]
            for index, item in enumerate(items, start=1)
        ],
        totals_pairs=[
            ("Tổng số lượng xuất", str(batch.total_quantity)),
            ("Tổng dòng hàng", str(batch.total_lines)),
        ],
        footer_note="Phiếu PDF được hệ thống tạo tự động và lưu trực tiếp trong PostgreSQL.",
        signature_left={
            "title": "Người lập phiếu",
            "name": batch.resolved_exported_by_name,
            "role": batch.exported_by_role or "",
        },
    )
    file_name = sanitize_receipt_filename(batch.resolved_export_code, "phieu-xuat-kho")
    batch.receipt_pdf.save(file_name, ContentFile(pdf_bytes), save=False)
    batch.save(update_fields=["receipt_pdf"])
    return batch.receipt_pdf


def allocate_order_item_inventory(order_item, *, prefer_existing_allocations=True):
    medicine = getattr(order_item, 'medicine', None)
    if medicine is None:
        raise ValueError(f"Không thể xử lý tồn kho cho dòng đơn #{order_item.pk} vì thuốc gốc không còn tồn tại.")

    ensure_medicine_has_inventory_lot_snapshot(medicine)
    existing_allocations = list(order_item.lot_allocations.select_related('lot').order_by('id'))
    if prefer_existing_allocations and existing_allocations:
        lot_ids = [allocation.lot_id for allocation in existing_allocations if allocation.lot_id]
        locked_lots = {
            lot.id: lot
            for lot in MedicineLot.objects.select_for_update().filter(id__in=lot_ids)
        }
        can_reuse = True
        for allocation in existing_allocations:
            locked_lot = locked_lots.get(allocation.lot_id)
            if locked_lot is None or locked_lot.remaining_quantity < allocation.quantity:
                can_reuse = False
                break
        if can_reuse:
            for allocation in existing_allocations:
                locked_lot = locked_lots[allocation.lot_id]
                locked_lot.remaining_quantity -= allocation.quantity
                if locked_lot.remaining_quantity < 0:
                    locked_lot.remaining_quantity = 0
                locked_lot.save(update_fields=['remaining_quantity'])
            recalculate_medicine_inventory_snapshot(medicine)
            return existing_allocations
        order_item.lot_allocations.all().delete()

    fefo_lots = list(get_fefo_lot_queryset_for_medicine(medicine, lock=True))
    available_total = sum(lot.remaining_quantity for lot in fefo_lots)
    if available_total < order_item.quantity:
        raise ValueError(
            f"Thuốc '{medicine.name}' tại {medicine.pharmacy.name} không đủ tồn theo FEFO. "
            f"Hiện chỉ còn {available_total} {medicine.unit.lower()} có thể xuất bán."
        )

    remaining_to_allocate = order_item.quantity
    created_allocations = []
    for lot in fefo_lots:
        if remaining_to_allocate <= 0:
            break
        taken_quantity = min(lot.remaining_quantity, remaining_to_allocate)
        if taken_quantity <= 0:
            continue
        created_allocations.append(
            OrderItemLotAllocation.objects.create(
                order_item=order_item,
                lot=lot,
                quantity=taken_quantity,
                lot_expiry_date=lot.expiry_date,
                lot_import_price=lot.import_price,
                lot_source_label=lot.source_label or '',
            )
        )
        lot.remaining_quantity -= taken_quantity
        lot.save(update_fields=['remaining_quantity'])
        remaining_to_allocate -= taken_quantity

    recalculate_medicine_inventory_snapshot(medicine)
    return created_allocations


def restore_order_item_inventory(order_item):
    medicine = getattr(order_item, 'medicine', None)
    if medicine is None:
        return []

    ensure_medicine_has_inventory_lot_snapshot(medicine)
    allocations = list(order_item.lot_allocations.select_related('lot').order_by('id'))
    restored_lots = []
    if not allocations:
        fallback_lot = MedicineLot.objects.create(
            medicine=medicine,
            pharmacy=order_item.order.pharmacy or medicine.pharmacy,
            source_type=MedicineLot.SOURCE_RETURN,
            source_label=f"Hoàn kho {order_item.order.order_code}",
            import_price=0,
            expiry_date=medicine.expiry_date,
            received_quantity=order_item.quantity,
            remaining_quantity=order_item.quantity,
            note='Hoàn kho bổ sung do đơn cũ chưa có dữ liệu phân bổ lô.',
        )
        restored_lots.append(fallback_lot)
        recalculate_medicine_inventory_snapshot(medicine)
        return restored_lots

    locked_lots = {
        lot.id: lot
        for lot in MedicineLot.objects.select_for_update().filter(id__in=[allocation.lot_id for allocation in allocations if allocation.lot_id])
    }

    for allocation in allocations:
        locked_lot = locked_lots.get(allocation.lot_id)
        if locked_lot is None:
            locked_lot = MedicineLot.objects.create(
                medicine=medicine,
                pharmacy=order_item.order.pharmacy or medicine.pharmacy,
                source_type=MedicineLot.SOURCE_RETURN,
                source_label=allocation.lot_source_label or f"Hoàn kho {order_item.order.order_code}",
                import_price=allocation.lot_import_price or 0,
                expiry_date=allocation.lot_expiry_date,
                received_quantity=allocation.quantity,
                remaining_quantity=allocation.quantity,
                note='Khôi phục lô thay thế vì lô gốc không còn khả dụng.',
            )
            restored_lots.append(locked_lot)
            continue

        locked_lot.remaining_quantity += allocation.quantity
        if locked_lot.remaining_quantity > locked_lot.received_quantity:
            locked_lot.received_quantity = locked_lot.remaining_quantity
            locked_lot.save(update_fields=['remaining_quantity', 'received_quantity'])
        else:
            locked_lot.save(update_fields=['remaining_quantity'])
        restored_lots.append(locked_lot)

    recalculate_medicine_inventory_snapshot(medicine)
    return restored_lots


def deduct_inventory_for_order(order, *, prefer_existing_allocations=True):
    order_items = list(order.items.select_related('medicine', 'order__pharmacy').all())
    for order_item in order_items:
        allocate_order_item_inventory(order_item, prefer_existing_allocations=prefer_existing_allocations)
    return order


def restore_inventory_for_order(order):
    order_items = list(order.items.select_related('medicine', 'order__pharmacy').all())
    for order_item in order_items:
        restore_order_item_inventory(order_item)
    return order


def create_purchase_import_batch_for_pharmacy(*, pharmacy, source_file_name, source_file_bytes, invoice_code, imported_by_name, admin_user, note, rows, header_map):
    admin_identity = build_admin_receipt_identity(admin_user, explicit_name=imported_by_name)
    batch = PurchaseImportBatch.objects.create(
        pharmacy=pharmacy,
        invoice_code=invoice_code,
        source_file=ContentFile(source_file_bytes, name=source_file_name),
        imported_by=admin_user if getattr(admin_user, "is_authenticated", False) else None,
        imported_by_name=admin_identity["name"],
        imported_by_email=admin_identity["email"],
        imported_by_role=admin_identity["role"],
        note=note,
    )

    created_items = []
    total_quantity = 0
    for row_number, row_values in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in row_values):
            continue
        payload = build_purchase_import_row_payload(header_map, row_values)
        medicine, was_created = get_or_create_medicine_for_import(pharmacy, payload, row_number=row_number)
        imported_quantity = parse_excel_int(payload.get("quantity"), f"Số lượng (dòng {row_number})")
        if imported_quantity <= 0:
            raise ValueError(f"Số lượng nhập ở dòng {row_number} phải lớn hơn 0.")
        import_price = 0
        if payload.get("import_price") not in (None, ""):
            import_price = parse_excel_int(payload.get("import_price"), f"Giá nhập (dòng {row_number})")
        expiry_date = None
        if payload.get("expiry_date") not in (None, ""):
            expiry_date = parse_excel_date_value(payload.get("expiry_date"))

        previous_quantity = medicine.quantity
        sale_price = parse_optional_excel_int(payload.get("sale_price"), f"Giá bán (dòng {row_number})")
        update_fields = []
        if sale_price is not None and sale_price != medicine.price:
            medicine.price = sale_price
            update_fields.append('price')
        elif was_created and import_price and not medicine.price:
            medicine.price = import_price
            update_fields.append('price')
        if update_fields:
            medicine.save(update_fields=update_fields)
            sync_medicine_catalog_metadata(medicine, field_names=update_fields)

        created_item = PurchaseImportItem.objects.create(
            batch=batch,
            medicine=medicine,
            medicine_name=medicine.name,
            manufacturer=medicine.manufacturer or "",
            unit=medicine.unit or "",
            previous_quantity=previous_quantity,
            imported_quantity=imported_quantity,
            new_quantity=previous_quantity,
            import_price=import_price,
            expiry_date=expiry_date,
            note=(str(payload.get("note") or "").strip())[:255],
        )
        create_import_lot_for_item(batch, created_item)
        recalculate_medicine_inventory_snapshot(medicine)
        created_item.new_quantity = medicine.quantity
        created_item.save(update_fields=['new_quantity'])
        created_items.append(created_item)
        total_quantity += imported_quantity

    if not created_items:
        raise ValueError("File Excel chưa có dòng dữ liệu hợp lệ để nhập hàng.")

    batch.total_lines = len(created_items)
    batch.total_quantity = total_quantity
    if not batch.invoice_code:
        batch.invoice_code = batch.resolved_invoice_code
    batch.save(update_fields=["total_lines", "total_quantity", "invoice_code"])
    ensure_purchase_import_receipt_pdf(batch, force=True)
    return batch


def process_purchase_import_excel(*, form, admin_user):
    target_pharmacies = form.get_target_pharmacies()
    if not target_pharmacies:
        raise ValueError("Không tìm thấy chi nhánh phù hợp để nhập hàng.")

    excel_file = form.cleaned_data["excel_file"]
    load_workbook = get_excel_workbook_loader()
    workbook = load_workbook(excel_file, data_only=True)
    if hasattr(excel_file, "seek"):
        excel_file.seek(0)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    source_file_name = getattr(excel_file, "name", "purchase_import.xlsx") or "purchase_import.xlsx"
    source_file_bytes = excel_file.read() if hasattr(excel_file, "read") else b""
    if hasattr(excel_file, "seek"):
        excel_file.seek(0)
    if not rows:
        raise ValueError("File Excel không có dữ liệu.")

    header_map = get_purchase_import_header_map(rows[0])
    if "quantity" not in header_map or ("medicine_id" not in header_map and "medicine_name" not in header_map):
        raise ValueError("File Excel cần có ít nhất cột số lượng và mã thuốc hoặc tên thuốc.")

    imported_by_name = get_user_display_name(admin_user) or getattr(admin_user, "username", "") or "Nhân viên nhập hàng"
    raw_note = (form.cleaned_data.get("note") or "").strip()
    invoice_code = (form.cleaned_data.get("invoice_code") or "").strip()

    created_batches = []
    with transaction.atomic():
        for pharmacy in target_pharmacies:
            batch_note = raw_note
            if form.is_all_branches_selected():
                scope_note = f"Nhập đồng loạt cho toàn bộ chi nhánh • Chi nhánh nhận: {pharmacy.name}"
                batch_note = "\n".join(part for part in [scope_note, raw_note] if part).strip()
            created_batches.append(
                create_purchase_import_batch_for_pharmacy(
                    pharmacy=pharmacy,
                    source_file_name=source_file_name,
                    source_file_bytes=source_file_bytes,
                    invoice_code=invoice_code,
                    imported_by_name=imported_by_name,
                    admin_user=admin_user,
                    note=batch_note,
                    rows=rows,
                    header_map=header_map,
                )
            )

    return created_batches


def allocate_stock_export_item_inventory(export_item, *, quantity, export_scope=StockExportBatch.EXPORT_SCOPE_STANDARD):
    medicine = getattr(export_item, "medicine", None)
    if medicine is None:
        raise ValueError("Không thể xuất kho vì sản phẩm nguồn không còn tồn tại.")

    selected_lots = list(get_stock_export_lot_queryset_for_medicine(medicine, export_scope, lock=True))
    available_total = sum(lot.remaining_quantity for lot in selected_lots)
    if available_total < quantity:
        if export_scope == StockExportBatch.EXPORT_SCOPE_EXPIRED:
            raise ValueError(
                f"Sản phẩm '{medicine.name}' chỉ còn {available_total} {medicine.unit.lower()} ở các lô đã hết hạn, không đủ để xử lý {quantity}."
            )
        raise ValueError(
            f"Sản phẩm '{medicine.name}' chỉ còn {available_total} {medicine.unit.lower()} khả dụng, không đủ để xuất {quantity}."
        )

    remaining_to_allocate = quantity
    allocations = []
    for lot in selected_lots:
        if remaining_to_allocate <= 0:
            break
        taken_quantity = min(lot.remaining_quantity, remaining_to_allocate)
        if taken_quantity <= 0:
            continue
        allocations.append(
            StockExportLotAllocation.objects.create(
                export_item=export_item,
                lot=lot,
                quantity=taken_quantity,
                lot_expiry_date=lot.expiry_date,
                lot_import_price=lot.import_price,
                lot_source_label=lot.source_label or "",
            )
        )
        lot.remaining_quantity -= taken_quantity
        lot.save(update_fields=["remaining_quantity"])
        remaining_to_allocate -= taken_quantity

    recalculate_medicine_inventory_snapshot(medicine)
    return allocations


def create_stock_export_batch(*, batch_form, item_formset, admin_user):
    pharmacy = batch_form.cleaned_data["pharmacy"]
    export_scope = batch_form.cleaned_data["export_scope"]
    admin_identity = build_admin_receipt_identity(admin_user)
    export_code = (batch_form.cleaned_data.get("export_code") or "").strip()
    note = (batch_form.cleaned_data.get("note") or "").strip()
    destination_name = (batch_form.cleaned_data.get("destination_name") or "").strip()

    batch = StockExportBatch.objects.create(
        pharmacy=pharmacy,
        export_scope=export_scope,
        export_code=export_code,
        exported_by=admin_user if getattr(admin_user, "is_authenticated", False) else None,
        exported_by_name=admin_identity["name"],
        exported_by_email=admin_identity["email"],
        exported_by_role=admin_identity["role"],
        destination_name=destination_name,
        note=note,
    )

    created_items = []
    total_quantity = 0
    for form in item_formset:
        if not getattr(form, "cleaned_data", None):
            continue
        if form.cleaned_data.get("DELETE"):
            continue
        medicine = form.cleaned_data.get("medicine")
        quantity = int(form.cleaned_data.get("quantity") or 0)
        line_note = (form.cleaned_data.get("note") or "").strip()[:255]
        if medicine is None or quantity <= 0:
            continue
        if medicine.pharmacy_id != pharmacy.id:
            raise ValueError(f"Sản phẩm '{medicine.name}' không thuộc chi nhánh {pharmacy.name}.")

        previous_quantity = int(medicine.quantity or 0)
        export_item = StockExportItem.objects.create(
            batch=batch,
            medicine=medicine,
            medicine_name=medicine.name,
            manufacturer=medicine.manufacturer or "",
            unit=medicine.unit or "",
            previous_quantity=previous_quantity,
            exported_quantity=quantity,
            remaining_quantity=previous_quantity,
            note=line_note,
        )
        allocate_stock_export_item_inventory(export_item, quantity=quantity, export_scope=export_scope)
        medicine.refresh_from_db(fields=["quantity"])
        export_item.remaining_quantity = int(medicine.quantity or 0)
        export_item.save(update_fields=["remaining_quantity"])
        created_items.append(export_item)
        total_quantity += quantity

    if not created_items:
        raise ValueError("Phiếu xuất cần có ít nhất một dòng sản phẩm hợp lệ.")

    if not batch.export_code:
        batch.export_code = batch.resolved_export_code
    batch.total_lines = len(created_items)
    batch.total_quantity = total_quantity
    batch.save(update_fields=["export_code", "total_lines", "total_quantity"])
    ensure_stock_export_receipt_pdf(batch, force=True)
    return batch


def get_expired_medicines_queryset(queryset):
    today = timezone.localdate()
    return queryset.filter(expiry_date__isnull=False, expiry_date__lt=today)


def get_expiring_soon_medicines_queryset(queryset, days=183):
    today = timezone.localdate()
    deadline = today + timedelta(days=days)
    return queryset.filter(expiry_date__isnull=False, expiry_date__gte=today, expiry_date__lte=deadline)


def get_expiry_badge_payload(medicine):
    if not medicine.expiry_date:
        return ("Chưa có HSD", "secondary")
    today = timezone.localdate()
    if medicine.expiry_date < today:
        return ("Đã hết hạn", "danger")
    if medicine.expiry_date <= today + timedelta(days=183):
        return ("Cần chú ý HSD", "warning")
    return (medicine.expiry_date.strftime("%d/%m/%Y"), "success")


def format_vnd(amount):
    try:
        numeric_value = int(amount or 0)
    except (TypeError, ValueError):
        numeric_value = 0
    return f"{numeric_value:,} đ".replace(',', '.')



TIME_RANGE_PATTERN = re.compile(
    r'(\d{1,2})(?:\s*[:hg]\s*(\d{1,2}))?\s*(?:-|–|—|đến|to)\s*(\d{1,2})(?:\s*[:hg]\s*(\d{1,2}))?',
    re.IGNORECASE,
)


def get_default_departure_time_value(raw_value=''):
    return normalize_departure_time_str(raw_value)


def parse_opening_hours_range(raw_value):
    text = (raw_value or '').strip()
    if not text:
        return None

    match = TIME_RANGE_PATTERN.search(text)
    if not match:
        return None

    start_hour = int(match.group(1) or 0)
    start_minute = int(match.group(2) or 0)
    end_hour = int(match.group(3) or 0)
    end_minute = int(match.group(4) or 0)

    if start_hour > 23 or end_hour > 23 or start_minute > 59 or end_minute > 59:
        return None

    start_total = start_hour * 60 + start_minute
    end_total = end_hour * 60 + end_minute

    return {
        'start_minutes': start_total,
        'end_minutes': end_total,
        'start_text': f'{start_hour:02d}:{start_minute:02d}',
        'end_text': f'{end_hour:02d}:{end_minute:02d}',
        'overnight': end_total <= start_total,
    }


def build_opening_status(opening_hours_text):
    schedule = parse_opening_hours_range(opening_hours_text)
    if not schedule:
        return {
            'is_open': None,
            'label': 'Chưa rõ giờ hoạt động',
            'detail': 'Thông tin giờ mở cửa đang được cập nhật.',
            'badge_class': 'is-neutral',
            'icon': 'fas fa-clock',
        }

    now = timezone.localtime()
    current_minutes = now.hour * 60 + now.minute
    is_overnight = schedule['overnight']

    if is_overnight:
        is_open = current_minutes >= schedule['start_minutes'] or current_minutes < schedule['end_minutes']
    else:
        is_open = schedule['start_minutes'] <= current_minutes < schedule['end_minutes']

    if is_open:
        detail = f"Đang phục vụ khách, dự kiến đến {schedule['end_text']}."
        badge_class = 'is-open'
        label = 'Đang mở cửa'
        icon = 'fas fa-store-alt'
    else:
        detail = f"Hiện đang đóng, mở lại lúc {schedule['start_text']}."
        badge_class = 'is-closed'
        label = 'Đang đóng cửa'
        icon = 'fas fa-store-slash'

    return {
        'is_open': is_open,
        'label': label,
        'detail': detail,
        'badge_class': badge_class,
        'icon': icon,
        'start_text': schedule['start_text'],
        'end_text': schedule['end_text'],
    }


def is_customer_user(user):
    return user.is_authenticated and not user.is_staff and not user.is_superuser


def get_or_create_user_profile(user):
    profile, _ = UserProfile.objects.get_or_create(
        user=user,
        defaults={
            'full_name': user.get_full_name() or user.username,
            'phone': '',
            'address_text': '',
        },
    )
    return profile


def get_managed_pharmacy_for_user(user):
    if not user.is_authenticated or not user.is_staff or user.is_superuser:
        return None
    managed_pharmacy = get_or_create_user_profile(user).managed_pharmacy
    if managed_pharmacy is not None:
        return managed_pharmacy

    # In single-branch deployments, allow staff accounts without an explicit assignment
    # to operate within the only available branch instead of failing every branch-scoped action.
    if Pharmacy.objects.count() == 1:
        return Pharmacy.objects.order_by("id").first()
    return None


def get_admin_scope_pharmacy(user):
    return get_managed_pharmacy_for_user(user)


def build_default_admin_permissions_for_user(user):
    permissions = {
        item["key"]: {action: False for action in item["actions"]}
        for item in ADMIN_PERMISSION_DEFINITIONS
    }
    if not user or not getattr(user, "is_authenticated", False) or not user.is_staff or user.is_superuser:
        return permissions

    defaults = {
        "dashboard": {"view": True},
        "order": {"view": True, "update": True},
        "return_request": {"view": True, "update": True},
        "medicine": {"view": True, "create": True, "update": True},
        "purchase_import": {"view": True, "create": True},
        "stock_export": {"view": True, "create": True},
        "inventory_lot": {"view": True},
        "promotion": {"view": True, "create": True, "update": True},
        "reports": {"view": True},
    }
    for module_key, actions in defaults.items():
        if module_key not in permissions:
            continue
        permissions[module_key].update(actions)
    return permissions


def normalize_admin_permissions_payload(raw_payload, *, user=None):
    normalized = build_default_admin_permissions_for_user(user)
    if not isinstance(raw_payload, dict):
        return normalized
    for module_key, module_permissions in raw_payload.items():
        if module_key not in normalized or not isinstance(module_permissions, dict):
            continue
        for action in normalized[module_key]:
            normalized[module_key][action] = bool(module_permissions.get(action, normalized[module_key][action]))
    for module_permissions in normalized.values():
        if any(module_permissions.get(action) for action in module_permissions if action != "view"):
            module_permissions["view"] = True
    return normalized


def get_user_admin_permissions(user):
    if not user or not getattr(user, "is_authenticated", False):
        return {}
    if user.is_superuser:
        return {
            item["key"]: {action: True for action in item["actions"]}
            for item in ADMIN_PERMISSION_DEFINITIONS
        }
    profile = getattr(user, "profile", None) or get_or_create_user_profile(user)
    return normalize_admin_permissions_payload(getattr(profile, "admin_permissions", {}) or {}, user=user)


def user_has_admin_permission(user, module_key, action="view"):
    if not user or not getattr(user, "is_authenticated", False) or not user.is_staff:
        return False
    if user.is_superuser:
        return True
    permissions = get_user_admin_permissions(user)
    module_permissions = permissions.get(module_key, {})
    return bool(module_permissions.get(action, False))


def can_create_admin_model(user, model_key):
    return user_has_admin_permission(user, model_key, "create")


def can_update_admin_model(user, model_key):
    return user_has_admin_permission(user, model_key, "update")


def disable_form_fields(form):
    for field in form.fields.values():
        field.disabled = True
    return form


def filter_queryset_by_admin_scope(queryset, user, model_key):
    managed_pharmacy = get_admin_scope_pharmacy(user)
    if not managed_pharmacy:
        return queryset
    if model_key == 'medicine':
        return queryset.filter(pharmacy=managed_pharmacy)
    if model_key == 'order':
        return queryset.filter(pharmacy=managed_pharmacy)
    if model_key == 'return_request':
        return queryset.filter(order__pharmacy=managed_pharmacy)
    if model_key == 'purchase_import':
        return queryset.filter(pharmacy=managed_pharmacy)
    if model_key == 'stock_export':
        return queryset.filter(pharmacy=managed_pharmacy)
    if model_key == 'promotion':
        return queryset.filter(medicine__pharmacy=managed_pharmacy)
    if model_key == 'inventory_lot':
        return queryset.filter(pharmacy=managed_pharmacy)
    return queryset


def ensure_object_is_within_admin_scope(user, model_key, obj):
    managed_pharmacy = get_admin_scope_pharmacy(user)
    if not managed_pharmacy:
        return

    if model_key == 'medicine':
        object_pharmacy_id = getattr(obj, 'pharmacy_id', None)
    elif model_key == 'order':
        object_pharmacy_id = getattr(obj, 'pharmacy_id', None)
    elif model_key == 'return_request':
        object_pharmacy_id = getattr(getattr(obj, 'order', None), 'pharmacy_id', None)
    elif model_key == 'purchase_import':
        object_pharmacy_id = getattr(obj, 'pharmacy_id', None)
    elif model_key == 'stock_export':
        object_pharmacy_id = getattr(obj, 'pharmacy_id', None)
    elif model_key == 'promotion':
        object_pharmacy_id = getattr(getattr(obj, 'medicine', None), 'pharmacy_id', None)
    elif model_key == 'inventory_lot':
        object_pharmacy_id = getattr(obj, 'pharmacy_id', None)
    else:
        object_pharmacy_id = None

    if object_pharmacy_id != managed_pharmacy.pk:
        raise PermissionDenied('Bạn chỉ được thao tác với dữ liệu thuộc chi nhánh mình đang phụ trách.')


def get_user_display_name(user):
    if not user.is_authenticated:
        return ''

    profile = get_or_create_user_profile(user)
    return (
        profile.full_name.strip()
        or user.get_full_name().strip()
        or user.first_name.strip()
        or user.username
    )


def get_entity_gallery_urls(instance, fallback_url):
    urls = list(getattr(instance, 'gallery_image_list', []) or [])
    if not urls and fallback_url:
        urls.append(fallback_url)

    return urls


def build_saved_address_payload(profile):
    if not profile or not profile.address_text:
        return None

    return {
        'address_text': profile.address_text,
        'lat': profile.address_lat,
        'lng': profile.address_lng,
    }


def get_requested_stock_export_pharmacy(request, *, batch_form=None):
    if batch_form is not None and getattr(batch_form, "cleaned_data", None):
        cleaned_pharmacy = batch_form.cleaned_data.get("pharmacy")
        if cleaned_pharmacy is not None:
            return cleaned_pharmacy

    managed_pharmacy = get_admin_scope_pharmacy(request.user)
    if managed_pharmacy is not None:
        return managed_pharmacy

    raw_pharmacy_id = (request.POST.get("pharmacy") or request.GET.get("pharmacy") or "").strip()
    if raw_pharmacy_id.isdigit():
        return Pharmacy.objects.filter(pk=int(raw_pharmacy_id)).first()
    return None


def get_requested_stock_export_scope(request, *, batch_form=None):
    if batch_form is not None and getattr(batch_form, "cleaned_data", None):
        cleaned_scope = (batch_form.cleaned_data.get("export_scope") or "").strip()
        if cleaned_scope in dict(StockExportBatch.EXPORT_SCOPE_CHOICES):
            return cleaned_scope
    raw_scope = (request.POST.get("export_scope") or request.GET.get("export_scope") or "").strip()
    if raw_scope in dict(StockExportBatch.EXPORT_SCOPE_CHOICES):
        return raw_scope
    return StockExportBatch.EXPORT_SCOPE_STANDARD


def build_stock_export_medicine_insights(pharmacy, *, export_scope):
    if pharmacy is None:
        return {
            "medicine_options": [],
            "summary_cards": [],
            "alert_groups": [],
            "mode_label": dict(StockExportBatch.EXPORT_SCOPE_CHOICES).get(export_scope, "Xuất kho"),
        }

    today = timezone.localdate()
    warning_deadline = today + timedelta(days=183)
    medicines = list(
        Medicine.objects.filter(pharmacy=pharmacy)
        .prefetch_related("lots")
        .order_by(Lower("name"), "id")
    )

    medicine_options = []
    expired_alerts = []
    warning_alerts = []
    discrepancy_alerts = []

    for medicine in medicines:
        active_lots = [lot for lot in medicine.lots.all() if int(lot.remaining_quantity or 0) > 0]
        sellable_lots = [lot for lot in active_lots if not lot.expiry_date or lot.expiry_date >= today]
        expired_lots = [lot for lot in active_lots if lot.expiry_date and lot.expiry_date < today]
        warning_lots = [
            lot for lot in sellable_lots
            if lot.expiry_date and today <= lot.expiry_date <= warning_deadline
        ]

        sellable_quantity = sum(int(lot.remaining_quantity or 0) for lot in sellable_lots)
        expired_quantity = sum(int(lot.remaining_quantity or 0) for lot in expired_lots)
        warning_quantity = sum(int(lot.remaining_quantity or 0) for lot in warning_lots)
        total_remaining = sum(int(lot.remaining_quantity or 0) for lot in active_lots)
        discrepancy_quantity = int(medicine.quantity or 0) - int(sellable_quantity or 0)
        next_sellable_expiry = next((lot.expiry_date for lot in sellable_lots if lot.expiry_date), None)
        next_expired_expiry = next((lot.expiry_date for lot in expired_lots if lot.expiry_date), None)

        if export_scope == StockExportBatch.EXPORT_SCOPE_EXPIRED and expired_quantity <= 0:
            continue
        if export_scope != StockExportBatch.EXPORT_SCOPE_EXPIRED and sellable_quantity <= 0:
            continue

        item = {
            "id": medicine.id,
            "name": medicine.name,
            "unit": medicine.unit or "Đơn vị",
            "manufacturer": medicine.manufacturer or "Chưa cập nhật nhà sản xuất",
            "category": medicine.category or "Chưa phân loại",
            "product_type_label": dict(MEDICINE_PRODUCT_TYPE_CHOICES).get(medicine.product_type, "Sản phẩm"),
            "sellable_quantity": sellable_quantity,
            "expired_quantity": expired_quantity,
            "warning_quantity": warning_quantity,
            "snapshot_quantity": int(medicine.quantity or 0),
            "total_remaining": total_remaining,
            "discrepancy_quantity": discrepancy_quantity,
            "available_lot_count": len(sellable_lots),
            "expired_lot_count": len(expired_lots),
            "warning_lot_count": len(warning_lots),
            "next_sellable_expiry_label": next_sellable_expiry.strftime("%d/%m/%Y") if next_sellable_expiry else "Chưa có HSD",
            "next_expired_expiry_label": next_expired_expiry.strftime("%d/%m/%Y") if next_expired_expiry else "Không có lô hết hạn",
            "preferred_quantity": expired_quantity if export_scope == StockExportBatch.EXPORT_SCOPE_EXPIRED else sellable_quantity,
            "preferred_quantity_label": "Tồn xử lý" if export_scope == StockExportBatch.EXPORT_SCOPE_EXPIRED else "Tồn bán được",
        }
        medicine_options.append(item)

        if expired_quantity > 0:
            expired_alerts.append(item)
        if warning_quantity > 0:
            warning_alerts.append(item)
        if discrepancy_quantity != 0:
            discrepancy_alerts.append(item)

    medicine_options.sort(
        key=lambda item: (
            0 if item["expired_quantity"] > 0 else 1,
            0 if item["warning_quantity"] > 0 else 1,
            -item["preferred_quantity"],
            item["name"].casefold(),
        )
    )

    summary_cards = [
        {
            "label": "Sản phẩm khả dụng",
            "value": sum(1 for item in medicine_options if item["sellable_quantity"] > 0),
            "tone": "success",
            "hint": "Có thể xuất cho bán hàng, chuyển kho hoặc đối soát.",
        },
        {
            "label": "Sản phẩm cận hạn",
            "value": sum(1 for item in warning_alerts if item["warning_quantity"] > 0),
            "tone": "warning",
            "hint": "Nên ưu tiên theo dõi hoặc xử lý trước khi quá hạn.",
        },
        {
            "label": "Sản phẩm hết hạn",
            "value": sum(1 for item in expired_alerts if item["expired_quantity"] > 0),
            "tone": "danger",
            "hint": "Có thể chuyển sang chế độ xử lý hàng hết hạn để xuất loại bỏ.",
        },
        {
            "label": "Mặt hàng chênh lệch",
            "value": sum(1 for item in discrepancy_alerts if item["discrepancy_quantity"] != 0),
            "tone": "info",
            "hint": "Chênh giữa tồn hiển thị và tồn bán được từ các lô.",
        },
    ]

    alert_groups = [
        {
            "key": "expired",
            "title": "Hàng đã hết hạn",
            "tone": "danger",
            "description": "Các lô đã quá hạn nhưng vẫn còn tồn thực tế trong kho dữ liệu.",
            "items": expired_alerts[:6],
            "count": len(expired_alerts),
        },
        {
            "key": "warning",
            "title": "Hàng cận hạn ≤ 6 tháng",
            "tone": "warning",
            "description": "Nên ưu tiên bán, chuyển kho hoặc tạo kế hoạch xả hàng.",
            "items": warning_alerts[:6],
            "count": len(warning_alerts),
        },
        {
            "key": "discrepancy",
            "title": "Dữ liệu cần đối soát",
            "tone": "info",
            "description": "Có chênh lệch giữa tồn snapshot và tồn bán được từ các lô.",
            "items": discrepancy_alerts[:6],
            "count": len(discrepancy_alerts),
        },
    ]

    return {
        "medicine_options": medicine_options,
        "summary_cards": summary_cards,
        "alert_groups": [group for group in alert_groups if group["count"] > 0],
        "mode_label": dict(StockExportBatch.EXPORT_SCOPE_CHOICES).get(export_scope, "Xuất kho"),
    }


def build_inventory_alert_center(queryset, request=None):
    today = timezone.localdate()
    warning_deadline = today + timedelta(days=183)
    alert_queryset = queryset.filter(remaining_quantity__gt=0).select_related("medicine", "pharmacy")
    can_create_stock_export = bool(request is not None and can_create_admin_model(request.user, "stock_export"))

    def build_item_payload(lot, *, group_key):
        export_scope = StockExportBatch.EXPORT_SCOPE_EXPIRED if group_key == "expired" else StockExportBatch.EXPORT_SCOPE_RECONCILE
        return {
            "medicine_name": lot.medicine.name if lot.medicine else "Sản phẩm đã xóa",
            "quantity": int(lot.remaining_quantity or 0),
            "expiry_label": lot.expiry_date.strftime("%d/%m/%Y") if lot.expiry_date else "Chưa có HSD",
            "pharmacy_name": lot.pharmacy.name if lot.pharmacy else "Chưa gán chi nhánh",
            "source_label": lot.source_label or f"Lô #{lot.pk}",
            "cta_url": (
                f"{reverse('custom_admin_create', kwargs={'model_key': 'stock_export'})}"
                f"?pharmacy={lot.pharmacy_id}&export_scope={export_scope}"
            ) if lot.pharmacy_id and can_create_stock_export else "",
        }

    section_configs = [
        {
            "key": "expired",
            "title": "Lô đã hết hạn",
            "tone": "danger",
            "description": "Nên tạo phiếu xuất xử lý để loại khỏi kho theo đúng quy trình.",
            "queryset": alert_queryset.filter(expiry_date__isnull=False, expiry_date__lt=today).order_by("expiry_date", "pharmacy__name", "medicine__name"),
            "page_param": "expired_page",
        },
        {
            "key": "warning",
            "title": "Lô cận hạn ≤ 6 tháng",
            "tone": "warning",
            "description": "Nên theo dõi sát và ưu tiên xử lý trước khi thành hàng hết hạn.",
            "queryset": alert_queryset.filter(expiry_date__isnull=False, expiry_date__gte=today, expiry_date__lte=warning_deadline).order_by("expiry_date", "pharmacy__name", "medicine__name"),
            "page_param": "warning_page",
        },
    ]

    sections = []
    total_count = 0
    request_get = request.GET if request is not None else {}
    keep_open = False

    for config in section_configs:
        paginator = Paginator(config["queryset"], 3)
        page_number = request_get.get(config["page_param"], 1) or 1
        if str(page_number) not in {"", "1"}:
            keep_open = True
        page_obj = paginator.get_page(page_number)
        total_count += paginator.count
        if paginator.count == 0:
            continue

        query_pairs = []
        if request is not None:
            for key, value in request.GET.items():
                if key == config["page_param"]:
                    continue
                query_pairs.append((key, value))

        sections.append(
            {
                "key": config["key"],
                "title": config["title"],
                "tone": config["tone"],
                "count": paginator.count,
                "description": config["description"],
                "items": [build_item_payload(lot, group_key=config["key"]) for lot in page_obj.object_list],
                "page_obj": page_obj,
                "page_param": config["page_param"],
                "query_pairs": query_pairs,
            }
        )

    if request is not None and str(request.GET.get("inventory_alert_open", "")).strip() == "1":
        keep_open = True

    return {
        "total_count": total_count,
        "sections": sections,
        "keep_open": keep_open,
    }


def build_admin_permission_sections():
    action_labels = {
        "view": "Xem",
        "create": "Tạo",
        "update": "Cập nhật",
        "delete": "Xóa",
    }
    return [
        {
            "key": item["key"],
            "label": item["label"],
            "description": item["description"],
            "actions": [
                {
                    "key": action,
                    "label": action_labels.get(action, action.title()),
                }
                for action in item["actions"]
            ],
        }
        for item in ADMIN_PERMISSION_DEFINITIONS
    ]


def build_permission_matrix_rows(permission_sections, selected_permissions):
    action_order = ("view", "create", "update", "delete")
    rows = []
    for section in permission_sections:
        action_map = {item["key"]: item["label"] for item in section["actions"]}
        rows.append(
            {
                "key": section["key"],
                "label": section["label"],
                "description": section["description"],
                "actions": [
                    {
                        "key": action_key,
                        "label": action_map.get(action_key, ""),
                        "supported": action_key in action_map,
                        "checked": bool(selected_permissions.get(section["key"], {}).get(action_key)),
                    }
                    for action_key in action_order
                ],
            }
        )
    return rows


def extract_admin_permissions_from_request(request, *, target_user):
    payload = {}
    for item in ADMIN_PERMISSION_DEFINITIONS:
        module_key = item["key"]
        module_payload = {}
        for action in item["actions"]:
            module_payload[action] = request.POST.get(f"permission__{module_key}__{action}") == "1"
        payload[module_key] = module_payload
    return normalize_admin_permissions_payload(payload, user=target_user)


def update_profile_from_checkout(user, order):
    if not user.is_authenticated:
        return

    if not is_customer_user(user):
        return

    profile = get_or_create_user_profile(user)
    profile.address_text = order.address_text or ''
    profile.address_lat = order.delivery_lat
    profile.address_lng = order.delivery_lng
    profile.save(update_fields=['address_text', 'address_lat', 'address_lng', 'updated_at'])


def get_payment_channel_settings(payment_method):
    if payment_method == Order.PAYMENT_MOMO:
        return {
            'payment_method': Order.PAYMENT_MOMO,
            'display_name': 'Ví MoMo',
            'recipient_name': getattr(settings, 'PAYMENT_MOMO_NAME', 'GIS Pharma'),
            'account_number': getattr(settings, 'PAYMENT_MOMO_PHONE', '0901234567'),
            'provider_name': 'MoMo',
            'helper_text': 'Quét mã để mở nhanh màn hình thanh toán với sẵn số tiền và nội dung đối soát của đơn.',
        }
    return {
        'payment_method': Order.PAYMENT_BANK,
        'display_name': 'Chuyển khoản ngân hàng',
        'recipient_name': getattr(settings, 'PAYMENT_BANK_ACCOUNT_NAME', 'GIS PHARMA'),
        'account_number': getattr(settings, 'PAYMENT_BANK_ACCOUNT_NUMBER', '1029384756'),
        'provider_name': getattr(settings, 'PAYMENT_BANK_NAME', 'Vietcombank'),
        'helper_text': 'Quét mã để nhập nhanh tài khoản nhận, số tiền cần chuyển và nội dung đối soát.',
    }


def build_order_invoice_code(order):
    if order.pk and order.created_at:
        return f"HD{timezone.localtime(order.created_at).strftime('%Y%m%d')}-{order.pk:06d}"
    return 'HD-TAM'


def build_order_payment_reference(order):
    return f"{order.order_code}-{timezone.localtime(order.created_at).strftime('%d%m')}" if order.pk and order.created_at else 'DH-TAM'


def get_invoice_staff_name_for_pharmacy(pharmacy):
    if not pharmacy:
        return 'Nhân viên quầy thuốc'

    staff_profile = pharmacy.managed_staff_profiles.select_related('user').order_by('id').first()
    if staff_profile and staff_profile.user:
        display_name = get_user_display_name(staff_profile.user)
        if display_name:
            return display_name

    return f'Nhân viên phụ trách {pharmacy.name}'


def determine_initial_payment_status(payment_method):
    if payment_method == Order.PAYMENT_COD:
        return Order.PAYMENT_STATUS_COD_WAITING
    return Order.PAYMENT_STATUS_AWAITING_TRANSFER


def sanitize_qr_text(value, *, uppercase=False, max_length=None):
    raw_value = str(value or '').strip()
    if not raw_value:
        return ''

    normalized = unicodedata.normalize('NFKD', raw_value)
    ascii_value = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    cleaned = re.sub(r'[^0-9A-Za-z\-._ ]+', '', ascii_value)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    if uppercase:
        cleaned = cleaned.upper()
    if max_length is not None:
        cleaned = cleaned[:max_length].strip()
    return cleaned


def build_static_asset_url(relative_path):
    cleaned_path = str(relative_path or '').strip().replace('\\', '/').lstrip('/')
    if not cleaned_path:
        return ''

    static_url = str(getattr(settings, 'STATIC_URL', '/static/') or '/static/')
    if not static_url.endswith('/'):
        static_url += '/'
    return static_url + cleaned_path


def resolve_payment_qr_image_url(setting_name, default_relative_paths=None):
    configured_value = str(getattr(settings, setting_name, '') or '').strip()
    if configured_value:
        if configured_value.startswith(('http://', 'https://', 'data:', '/')):
            return configured_value
        return build_static_asset_url(configured_value)

    for relative_path in default_relative_paths or []:
        asset_path = settings.BASE_DIR / 'myapp' / 'static' / relative_path
        if asset_path.exists():
            return build_static_asset_url(relative_path)
    return ''


def build_payment_qr_payload(payment_method, amount, reference, recipient_name='', account_number='', provider_name=''):
    safe_amount = max(int(amount or 0), 0)
    payload_lines = [
        'GIS PHARMA PAYMENT',
        f'PHUONG THUC: {payment_method.upper()}',
    ]
    if provider_name:
        payload_lines.append(f'DON VI: {provider_name}')
    if recipient_name:
        payload_lines.append(f'NGUOI NHAN: {recipient_name}')
    if account_number:
        payload_lines.append(f'TAI KHOAN/VI: {account_number}')
    payload_lines.append(f'SO TIEN: {safe_amount} VND')
    payload_lines.append(f'NOI DUNG: {reference}')
    return '\n'.join(payload_lines)


def build_vietqr_quicklink(amount, reference, recipient_name='', account_number='', provider_name=''):
    bank_id = sanitize_qr_text(
        getattr(settings, 'PAYMENT_BANK_QR_BANK_ID', provider_name or ''),
        uppercase=False,
    ).lower().replace(' ', '')
    account_no = sanitize_qr_text(account_number)
    if not bank_id or not account_no:
        return ''

    account_name = sanitize_qr_text(recipient_name, uppercase=True, max_length=50)
    add_info = sanitize_qr_text(reference, uppercase=False, max_length=25)
    amount_text = str(max(int(amount or 0), 0))
    template = sanitize_qr_text(getattr(settings, 'PAYMENT_BANK_QR_TEMPLATE', 'compact2'), uppercase=False) or 'compact2'
    base_url = f'https://img.vietqr.io/image/{bank_id}-{account_no}-{template}.png'
    query_params = [f'amount={quote(amount_text)}']
    if add_info:
        query_params.append(f'addInfo={quote(add_info)}')
    if account_name:
        query_params.append(f'accountName={quote(account_name)}')
    return base_url + ('?' + '&'.join(query_params) if query_params else '')


def build_momo_qr_image_url(amount, reference):
    explicit_image = resolve_payment_qr_image_url(
        'PAYMENT_MOMO_QR_IMAGE_URL',
        default_relative_paths=[
            'images/payment-momo-qr.png',
            'images/payment-momo-qr.jpg',
            'images/payment-momo-qr.jpeg',
            'images/payment-momo-qr.webp',
        ],
    )
    if explicit_image:
        return explicit_image

    payment_link = (getattr(settings, 'PAYMENT_MOMO_PAYMENT_URL', '') or '').strip()
    if payment_link:
        return generate_qr_data_url(payment_link)

    return ''


def generate_qr_data_url(payload_text):
    if qrcode is None:
        svg_markup = (
            "<svg xmlns='http://www.w3.org/2000/svg' width='220' height='220' viewBox='0 0 220 220'>"
            "<rect width='220' height='220' fill='white'/>"
            "<rect x='18' y='18' width='184' height='184' rx='18' fill='white' stroke='#1146b8' stroke-width='6'/>"
            "<text x='110' y='86' text-anchor='middle' font-size='34' font-family='Arial' fill='#1146b8'>QR</text>"
            "<text x='110' y='118' text-anchor='middle' font-size='13' font-family='Arial' fill='#163057'>Thanh toán</text>"
            "<text x='110' y='146' text-anchor='middle' font-size='11' font-family='Arial' fill='#60708f'>Cài qrcode để sinh QR thật</text>"
            "</svg>"
        )
        encoded_svg = base64.b64encode(svg_markup.encode('utf-8')).decode('ascii')
        return f'data:image/svg+xml;base64,{encoded_svg}'

    qr = qrcode.QRCode(version=None, box_size=8, border=2)
    qr.add_data(payload_text or 'QR PAYMENT')
    qr.make(fit=True)
    image = qr.make_image(fill_color='black', back_color='white')
    buffer = BytesIO()
    image.save(buffer, format='PNG')
    encoded_image = base64.b64encode(buffer.getvalue()).decode('ascii')
    return f'data:image/png;base64,{encoded_image}'


def resolve_payment_qr_image(payment_method, amount, reference, recipient_name='', account_number='', provider_name=''):
    if payment_method == Order.PAYMENT_BANK:
        bank_qr_image = resolve_payment_qr_image_url(
            'PAYMENT_BANK_QR_IMAGE_URL',
            default_relative_paths=[
                'images/payment-bank-qr.png',
                'images/payment-bank-qr.jpg',
                'images/payment-bank-qr.jpeg',
                'images/payment-bank-qr.webp',
            ],
        )
        if bank_qr_image:
            return bank_qr_image

        quicklink = build_vietqr_quicklink(
            amount=amount,
            reference=reference,
            recipient_name=recipient_name,
            account_number=account_number,
            provider_name=provider_name,
        )
        if quicklink:
            return quicklink

    if payment_method == Order.PAYMENT_MOMO:
        momo_qr = build_momo_qr_image_url(amount, reference)
        if momo_qr:
            return momo_qr

    qr_payload = build_payment_qr_payload(
        payment_method=payment_method,
        amount=amount,
        reference=reference,
        recipient_name=recipient_name,
        account_number=account_number,
        provider_name=provider_name,
    )
    return generate_qr_data_url(qr_payload)


def build_payment_preview_payload(payment_method, amount, reference, pharmacy=None):
    safe_method = payment_method if payment_method in {Order.PAYMENT_COD, Order.PAYMENT_MOMO, Order.PAYMENT_BANK} else Order.PAYMENT_COD
    safe_amount = max(int(amount or 0), 0)

    if safe_method == Order.PAYMENT_COD:
        return {
            'payment_method': Order.PAYMENT_COD,
            'payment_label': 'Thanh toán khi nhận hàng (COD)',
            'amount_value': safe_amount,
            'amount_text': format_vnd(safe_amount),
            'qr_image': '',
            'show_qr': False,
            'recipient_name': '',
            'account_number': '',
            'provider_name': '',
            'transfer_note': reference,
            'helper_text': 'Khách hàng thanh toán trực tiếp cho nhân viên giao hàng khi nhận đơn.',
            'branch_name': pharmacy.name if pharmacy else '',
            'branch_address': pharmacy.address if pharmacy else '',
        }

    channel = get_payment_channel_settings(safe_method)
    return {
        'payment_method': safe_method,
        'payment_label': channel['display_name'],
        'amount_value': safe_amount,
        'amount_text': format_vnd(safe_amount),
        'qr_image': resolve_payment_qr_image(
            payment_method=safe_method,
            amount=safe_amount,
            reference=reference,
            recipient_name=channel['recipient_name'],
            account_number=channel['account_number'],
            provider_name=channel['provider_name'],
        ),
        'show_qr': True,
        'recipient_name': channel['recipient_name'],
        'account_number': channel['account_number'],
        'provider_name': channel['provider_name'],
        'transfer_note': reference,
        'helper_text': channel['helper_text'],
        'branch_name': pharmacy.name if pharmacy else '',
        'branch_address': pharmacy.address if pharmacy else '',
    }


def build_order_payment_preview(order):
    return build_payment_preview_payload(
        payment_method=order.payment_method,
        amount=order.final_total_price,
        reference=order.resolved_payment_reference,
        pharmacy=order.pharmacy,
    )


def normalize_medicine_key(name, unit, manufacturer='', origin=''):
    return (
        fold_text_for_match(name),
        fold_text_for_match(unit),
        fold_text_for_match(manufacturer),
        fold_text_for_match(origin),
    )


def normalize_catalog_key(name, unit, manufacturer=''):
    return build_medicine_catalog_key(name, unit, manufacturer)


def build_cart_requirements(cart):
    grouped_requirements = {}

    for cart_item in cart.items.select_related('medicine').all():
        key = normalize_medicine_key(
            cart_item.medicine.name,
            cart_item.medicine.unit,
            cart_item.medicine.manufacturer,
            cart_item.medicine.origin,
        )
        if key not in grouped_requirements:
            grouped_requirements[key] = {
                'name': cart_item.medicine.name,
                'unit': cart_item.medicine.unit,
                'manufacturer': cart_item.medicine.manufacturer,
                'origin': cart_item.medicine.origin,
                'quantity': 0,
            }
        grouped_requirements[key]['quantity'] += cart_item.quantity

    return list(grouped_requirements.values())


def build_cart_prescription_context(cart):
    prescription_items = []
    for cart_item in cart.items.select_related('medicine', 'medicine__pharmacy').order_by('id'):
        medicine = getattr(cart_item, 'medicine', None)
        if medicine and medicine.prescription_required:
            prescription_items.append(cart_item)
    return {
        'requires_prescription': bool(prescription_items),
        'items': prescription_items,
    }


def build_order_prescription_proof_cards(order):
    proof_cards = []
    if order and getattr(order, "pk", None):
        try:
            related_images = list(order.prescription_proof_images.all())
        except Exception:
            related_images = []
        for index, proof in enumerate(related_images, start=1):
            if getattr(proof, "image", None):
                proof_cards.append({
                    "url": proof.image.url,
                    "label": f"Ảnh đơn thuốc {index}",
                })

    legacy_image = getattr(order, "prescription_proof_image", None)
    if legacy_image and not proof_cards:
        proof_cards.append({
            "url": legacy_image.url,
            "label": "Ảnh đơn thuốc",
        })
    return proof_cards


def allocate_requirements_to_medicines(requirements, medicines):
    medicines_by_key = {}

    for medicine in medicines:
        key = normalize_medicine_key(
            medicine.name,
            medicine.unit,
            medicine.manufacturer,
            medicine.origin,
        )
        medicines_by_key.setdefault(key, medicine)

    allocations = []

    for requirement in requirements:
        key = normalize_medicine_key(
            requirement['name'],
            requirement['unit'],
            requirement.get('manufacturer', ''),
            requirement.get('origin', ''),
        )
        matched_medicine = medicines_by_key.get(key)
        if matched_medicine is None or matched_medicine.quantity < requirement['quantity']:
            return None

        allocations.append(
            {
                'medicine': matched_medicine,
                'quantity': requirement['quantity'],
            }
        )

    return allocations


def allocate_cart_to_pharmacy(cart, pharmacy):
    requirements = build_cart_requirements(cart)
    pharmacy_medicines = getattr(pharmacy, '_inventory_candidates', None)
    if pharmacy_medicines is None:
        pharmacy_medicines = list(
            Medicine.objects.filter(pharmacy=pharmacy).order_by('-quantity', 'id')
        )
    return allocate_requirements_to_medicines(requirements, pharmacy_medicines)


def get_checkout_candidate_pharmacy_ids(cart):
    requirements = build_cart_requirements(cart)
    if not requirements:
        return []

    candidate_ids = []
    pharmacy_queryset = get_available_pharmacies().prefetch_related(
        Prefetch(
            'medicines',
            queryset=Medicine.objects.order_by('-quantity', 'id'),
            to_attr='_inventory_candidates',
        )
    )

    for pharmacy in pharmacy_queryset:
        if allocate_requirements_to_medicines(
            requirements,
            getattr(pharmacy, '_inventory_candidates', []),
        ) is not None:
            candidate_ids.append(pharmacy.id)

    return candidate_ids


def sync_inventory_for_order_status_transition(order, previous_status, next_status=None):
    next_status = next_status or order.status
    if previous_status == next_status:
        return

    inventory_released_statuses = set(Order.INVENTORY_RELEASED_STATUSES)
    should_restore_inventory = previous_status not in inventory_released_statuses and next_status in inventory_released_statuses
    should_deduct_inventory_again = previous_status in inventory_released_statuses and next_status not in inventory_released_statuses

    if not (should_restore_inventory or should_deduct_inventory_again):
        return

    with transaction.atomic():
        locked_order = Order.objects.select_for_update().get(pk=order.pk)
        if should_restore_inventory:
            restore_inventory_for_order(locked_order)
            return

        if should_deduct_inventory_again:
            deduct_inventory_for_order(locked_order, prefer_existing_allocations=True)


def get_user_role_label(user):
    if user.is_superuser:
        return 'Quản trị viên hệ thống'
    if user.is_staff:
        return 'Nhân viên chi nhánh'
    return 'Khách hàng'


def get_medicine_search_queryset(search_keyword=''):
    medicine_queryset = Medicine.objects.select_related('pharmacy').all()

    if search_keyword:
        medicine_queryset = medicine_queryset.filter(
            Q(name__icontains=search_keyword)
            | Q(description__icontains=search_keyword)
            | Q(category__icontains=search_keyword)
            | Q(manufacturer__icontains=search_keyword)
            | Q(usage__icontains=search_keyword)
            | Q(ingredients__icontains=search_keyword)
            | Q(pharmacy__name__icontains=search_keyword)
        )

    return medicine_queryset


def get_popular_categories(limit=8):
    return list(
        Medicine.objects.exclude(category='')
        .values('category')
        .annotate(total=Count('id'))
        .order_by('-total', 'category')[:limit]
    )


def get_available_medicine_filter():
    today = timezone.localdate()
    return Q(medicines__quantity__gt=0) | (
        Q(medicines__lots__remaining_quantity__gt=0)
        & (Q(medicines__lots__expiry_date__isnull=True) | Q(medicines__lots__expiry_date__gte=today))
    )


def get_featured_pharmacies(limit=3):
    featured_items = list(
        Pharmacy.objects.annotate(
            available_total=Count('medicines', filter=get_available_medicine_filter(), distinct=True)
        )
        .filter(available_total__gt=0)
        .order_by('-id')[:limit]
    )

    for pharmacy in featured_items:
        pharmacy.open_state = build_opening_status(pharmacy.opening_hours)

    return featured_items


def ensure_home_page_defaults(content):
    if not content.hero_slides.exists():
        HomeHeroSlide.objects.bulk_create(
            [
                HomeHeroSlide(
                    content=content,
                    legacy_static_path=item["legacy_static_path"],
                    alt_text=item["alt_text"],
                    link_url=item["link_url"],
                    sort_order=index,
                    is_active=True,
                )
                for index, item in enumerate(DEFAULT_HOME_SLIDES)
            ]
        )

    if not content.category_spotlights.exists():
        HomeCategorySpotlightItem.objects.bulk_create(
            [
                HomeCategorySpotlightItem(
                    content=content,
                    title=item["title"],
                    subtitle=item["subtitle"],
                    icon_class=item["icon_class"],
                    link_url=item["link_url"],
                    sort_order=index,
                    is_active=True,
                )
                for index, item in enumerate(DEFAULT_HOME_CATEGORY_ITEMS)
            ]
        )

    if not content.service_commitments.exists():
        HomeServiceCommitmentItem.objects.bulk_create(
            [
                HomeServiceCommitmentItem(
                    content=content,
                    title=item["title"],
                    body=item["body"],
                    icon_class=item["icon_class"],
                    sort_order=index,
                    is_active=True,
                )
                for index, item in enumerate(DEFAULT_HOME_COMMITMENT_ITEMS)
            ]
        )


def ensure_about_page_slide_defaults(content):
    if content.hero_slides.exists():
        return

    AboutPageSlide.objects.bulk_create(
        [
            AboutPageSlide(
                content=content,
                legacy_static_path=item["legacy_static_path"],
                alt_text=item["alt_text"],
                link_url=item["link_url"],
                sort_order=index,
                is_active=True,
            )
            for index, item in enumerate(DEFAULT_ABOUT_SLIDES)
        ]
    )


def build_default_about_featured_branch_items(content):
    default_pharmacies = list(Pharmacy.objects.order_by("name")[:4])
    if not default_pharmacies:
        return []

    return [
        AboutFeaturedBranchItem(
            content=content,
            pharmacy=pharmacy,
            title=pharmacy.name,
            summary="",
            address=pharmacy.address or "",
            hours=pharmacy.opening_hours or "",
            badge=content.branch_showcase_badge,
            map_note=content.branch_showcase_map_note,
            icon_class="fas fa-clinic-medical",
            link_url=reverse("pharmacy_detail", kwargs={"pharmacy_id": pharmacy.pk}),
            link_label="Xem chi nhánh",
            image=pharmacy.image.name if getattr(pharmacy, "image", None) else None,
            sort_order=index,
            is_active=True,
        )
        for index, pharmacy in enumerate(default_pharmacies)
    ]


def ensure_about_featured_branch_defaults(content):
    existing_items = list(content.featured_branch_items.select_related("pharmacy").order_by("sort_order", "id"))
    if not existing_items:
        defaults = build_default_about_featured_branch_items(content)
        if defaults:
            AboutFeaturedBranchItem.objects.bulk_create(defaults)
        return

    pharmacies = list(Pharmacy.objects.order_by("name"))
    if not pharmacies:
        return

    available_by_name = {pharmacy.name.casefold(): pharmacy for pharmacy in pharmacies}
    unused_pharmacies = [pharmacy for pharmacy in pharmacies]

    for item in existing_items:
        if item.pharmacy_id:
            unused_pharmacies = [pharmacy for pharmacy in unused_pharmacies if pharmacy.pk != item.pharmacy_id]

    items_to_update = []
    for item in existing_items:
        if item.pharmacy_id:
            continue
        candidate = available_by_name.get((item.title or "").strip().casefold())
        if candidate is None and unused_pharmacies:
            candidate = unused_pharmacies.pop(0)
        if candidate is None:
            continue
        item.pharmacy = candidate
        item.title = candidate.name
        item.address = candidate.address or ""
        item.hours = candidate.opening_hours or ""
        item.image = candidate.image.name if getattr(candidate, "image", None) else item.image
        if not item.link_url:
            item.link_url = reverse("pharmacy_detail", kwargs={"pharmacy_id": candidate.pk})
        items_to_update.append(item)

    if items_to_update:
        AboutFeaturedBranchItem.objects.bulk_update(
            items_to_update,
            ["pharmacy", "title", "address", "hours", "image", "link_url"],
        )


def build_about_featured_branch_payload(item, content):
    pharmacy = getattr(item, "pharmacy", None)
    summary_text = str(getattr(item, "summary", "") or "").strip()
    if not summary_text and pharmacy is not None:
        summary_text = strip_tags(getattr(pharmacy, "desc", "") or "").strip()
    if summary_text:
        summary_text = summary_text[:180].rsplit(" ", 1)[0] + "…" if len(summary_text) > 180 else summary_text

    return {
        "title": (pharmacy.name if pharmacy is not None else item.title) or "Chi nhánh đang cập nhật",
        "summary": summary_text,
        "address": (pharmacy.address if pharmacy is not None else item.address) or "",
        "hours": (pharmacy.opening_hours if pharmacy is not None else item.hours) or "",
        "badge": item.badge or content.branch_showcase_badge,
        "map_note": item.map_note or content.branch_showcase_map_note,
        "icon_class": item.icon_class or "fas fa-clinic-medical",
        "link_url": item.link_url or (reverse("pharmacy_detail", kwargs={"pharmacy_id": pharmacy.pk}) if pharmacy is not None else ""),
        "link_label": item.link_label or "Xem chi nhánh",
        "image_url": (pharmacy.primary_image_url if pharmacy is not None else "") or (item.image.url if getattr(item, "image", None) else ""),
        "pharmacy_id": getattr(pharmacy, "pk", None),
    }


def build_about_section_layout(content, *, custom_blocks=None):
    custom_blocks = list(custom_blocks or [])
    layout_items = []

    for section in content.builtin_sections.order_by("sort_order", "id"):
        layout_items.append(
            {
                "kind": "builtin",
                "key": section.section_key,
                "sort_group": 0,
                "sort_order": section.sort_order,
                "object_id": section.pk or 0,
            }
        )

    for block in custom_blocks:
        layout_items.append(
            {
                "kind": "custom",
                "block": block,
                "sort_group": 1,
                "sort_order": block.sort_order,
                "object_id": block.pk or 0,
            }
        )

    layout_items.sort(
        key=lambda item: (
            item["sort_group"],
            item["sort_order"],
            item["object_id"],
        )
    )
    return layout_items


def build_admin_quick_link_choices():
    choices = [
        (reverse("home"), "Trang chủ"),
        (reverse("about"), "Trang giới thiệu"),
        (reverse("news_list"), "Trang tin tức"),
        (reverse("product_list"), "Danh sách sản phẩm"),
        (reverse("map_view"), "Bản đồ chi nhánh"),
        (reverse("cart_detail"), "Giỏ hàng"),
        (reverse("account"), "Tài khoản"),
        (reverse("order_history"), "Lịch sử đơn hàng"),
        (reverse("login"), "Đăng nhập"),
        (reverse("register"), "Đăng ký"),
    ]

    for category in (
        Medicine.objects.exclude(category="")
        .values_list("category", flat=True)
        .distinct()
        .order_by("category")[:20]
    ):
        choices.append((f"{reverse('product_list')}?category={quote(category)}", f"Danh mục sản phẩm: {category}"))

    for pharmacy in Pharmacy.objects.order_by("name")[:20]:
        choices.append(
            (
                reverse("pharmacy_detail", kwargs={"pharmacy_id": pharmacy.pk}),
                f"Chi nhánh: {pharmacy.name}",
            )
        )

    deduplicated = []
    seen_urls = set()
    for url, label in choices:
        if url in seen_urls:
            continue
        seen_urls.add(url)
        deduplicated.append((url, label))
    return deduplicated


def build_singleton_field_groups(form, field_groups):
    groups = []
    for title, field_names in field_groups:
        fields = []
        for field_name in field_names:
            if field_name not in form.fields:
                continue
            bound_field = form[field_name]
            fields.append(
                {
                    "field": bound_field,
                    "is_rich_text": bound_field.field.widget.attrs.get("data-rich-editor") == "1",
                    "is_textarea": isinstance(bound_field.field.widget, forms.Textarea),
                }
            )
        if fields:
            groups.append({"title": title, "fields": fields})
    return groups


def form_has_rich_editors(form):
    return any(field.widget.attrs.get("data-rich-editor") == "1" for field in form.fields.values())


def was_formset_submitted(request, prefix):
    return f"{prefix}-TOTAL_FORMS" in request.POST


def get_pharmacy_search_queryset(search_keyword=''):
    pharmacy_queryset = Pharmacy.objects.annotate(
        available_total=Count('medicines', filter=get_available_medicine_filter(), distinct=True)
    )

    if search_keyword:
        pharmacy_queryset = pharmacy_queryset.filter(
            Q(name__icontains=search_keyword)
            | Q(address__icontains=search_keyword)
            | Q(desc__icontains=search_keyword)
            | Q(phone__icontains=search_keyword)
            | Q(medicines__name__icontains=search_keyword)
            | Q(medicines__category__icontains=search_keyword)
        ).distinct()

    return pharmacy_queryset


def get_available_pharmacies():
    """
    Lấy danh sách nhà thuốc còn ít nhất một loại thuốc còn hàng để bán.
    """
    return Pharmacy.objects.filter(get_available_medicine_filter()).distinct()


def build_catalog_search_payload(keyword, medicine_limit=6, pharmacy_limit=6):
    medicines = attach_discount_payloads(list(
        get_medicine_search_queryset(keyword)
        .order_by('-quantity', Lower('name'), 'id')[:medicine_limit]
    ))
    pharmacies = (
        get_pharmacy_search_queryset(keyword)
        .order_by('-available_total', Lower('name'), 'id')[:pharmacy_limit]
    )

    return {
        'query': keyword,
        'products': [
            {
                'id': medicine.id,
                'name': medicine.name,
                'category': medicine.category or 'Thuốc / Dược phẩm',
                'manufacturer': medicine.manufacturer or 'Đang cập nhật',
                'pharmacy_name': medicine.pharmacy.name,
                'price_value': getattr(medicine, 'current_price_value', medicine.current_price),
                'price_text': f"{getattr(medicine, 'current_price_value', medicine.current_price):,} đ".replace(',', '.'),
                'is_in_stock': medicine.is_in_stock,
                'image': medicine.primary_image_url or MEDICINE_FALLBACK_IMAGE,
                'detail_url': reverse('medicine_detail', kwargs={'medicine_id': medicine.id}),
            }
            for medicine in medicines
        ],
        'pharmacies': [
            {
                'id': pharmacy.id,
                'name': pharmacy.name,
                'address': pharmacy.address,
                'phone': pharmacy.phone,
                'available_total': pharmacy.available_total,
                'image': pharmacy.primary_image_url or PHARMACY_FALLBACK_IMAGE,
                'detail_url': reverse('pharmacy_detail', kwargs={'pharmacy_id': pharmacy.id}),
                'map_url': f"{reverse('map_view')}?pharmacy_id={pharmacy.id}",
            }
            for pharmacy in pharmacies
        ],
    }


def build_nearby_pharmacy_payload(lat, lng, radius_km=0, keyword=''):
    try:
        base_lat = float(lat)
        base_lng = float(lng)
    except (TypeError, ValueError):
        raise ValueError('Tọa độ không hợp lệ.')

    try:
        radius_limit = float(radius_km)
    except (TypeError, ValueError):
        radius_limit = 0

    pharmacy_queryset = get_pharmacy_search_queryset(keyword).order_by(Lower('name'), 'id')
    nearby_items = []

    for pharmacy in pharmacy_queryset:
        if pharmacy.lat is None or pharmacy.lng is None:
            continue

        air_distance_km = calculate_air_distance_km(base_lat, base_lng, pharmacy.lat, pharmacy.lng)
        estimated_distance_km = round(estimate_road_distance_km(air_distance_km, 'motorbike'), 2)

        if radius_limit > 0 and estimated_distance_km > radius_limit:
            continue

        nearby_items.append(
            {
                'id': pharmacy.id,
                'name': pharmacy.name,
                'address': pharmacy.address,
                'phone': pharmacy.phone,
                'hours': pharmacy.opening_hours,
                'distance_km': estimated_distance_km,
                'available_total': pharmacy.available_total,
                'image': pharmacy.primary_image_url or PHARMACY_FALLBACK_IMAGE,
                'detail_url': reverse('pharmacy_detail', kwargs={'pharmacy_id': pharmacy.id}),
            }
        )

    nearby_items.sort(key=lambda item: (item['distance_km'], item['name'].casefold(), item['id']))
    return nearby_items


def build_checkout_pharmacy_payload(pharmacy_queryset):
    """
    Chuẩn bị dữ liệu nhà thuốc cho trang checkout.
    """
    checkout_payload = []

    for pharmacy in pharmacy_queryset:
        gallery_images = get_entity_gallery_urls(
            pharmacy,
            PHARMACY_FALLBACK_IMAGE,
        )

        checkout_payload.append(
            {
                'id': pharmacy.id,
                'name': pharmacy.name,
                'address': pharmacy.address,
                'phone': pharmacy.phone,
                'hours': pharmacy.opening_hours,
                'image': gallery_images[0],
                'gallery_images': gallery_images,
                'lat': pharmacy.lat,
                'lng': pharmacy.lng,
            }
        )

    return checkout_payload


def build_map_pharmacy_payload(pharmacy_queryset):
    """
    Chuẩn bị dữ liệu nhà thuốc cho trang bản đồ.
    """
    pharmacy_payload = []

    for pharmacy in pharmacy_queryset:
        gallery_images = get_entity_gallery_urls(
            pharmacy,
            PHARMACY_FALLBACK_IMAGE,
        )

        pharmacy_payload.append(
            {
                'id': pharmacy.id,
                'name': pharmacy.name,
                'address': pharmacy.address,
                'phone': pharmacy.phone,
                'hours': pharmacy.opening_hours,
                'desc': pharmacy.desc,
                'image': gallery_images[0],
                'gallery_images': gallery_images,
                'lat': pharmacy.lat,
                'lng': pharmacy.lng,
            }
        )

    return pharmacy_payload


def estimate_pharmacy_delivery_distance(pharmacy, delivery_lat, delivery_lng, delivery_mode='motorbike'):
    if pharmacy.lat is None or pharmacy.lng is None:
        return None

    air_distance_km = calculate_air_distance_km(
        pharmacy.lat,
        pharmacy.lng,
        delivery_lat,
        delivery_lng,
    )
    return round(estimate_road_distance_km(air_distance_km, delivery_mode), 2)


def rank_pharmacies_by_distance(pharmacies, delivery_lat, delivery_lng, delivery_mode='motorbike'):
    ranked_pharmacies = []

    for pharmacy in pharmacies:
        estimated_distance = estimate_pharmacy_delivery_distance(
            pharmacy,
            delivery_lat,
            delivery_lng,
            delivery_mode,
        )
        if estimated_distance is None:
            continue

        ranked_pharmacies.append((pharmacy, estimated_distance))

    ranked_pharmacies.sort(key=lambda item: (item[1], item[0].name.casefold(), item[0].id))
    return ranked_pharmacies


def choose_checkout_pharmacy(cart, pharmacies, delivery_lat, delivery_lng, delivery_mode='motorbike', departure_time_str=None):
    ranked_pharmacies = rank_pharmacies_by_distance(
        pharmacies,
        delivery_lat,
        delivery_lng,
        delivery_mode,
    )

    if not ranked_pharmacies:
        return {'error': 'Không tìm được chi nhánh phù hợp cho vị trí đã chọn.'}

    nearest_pharmacy, nearest_distance = ranked_pharmacies[0]

    selected_pharmacy = None
    selected_distance = None
    selected_allocations = []

    if cart.items.exists():
        for pharmacy, distance in ranked_pharmacies:
            allocations = allocate_cart_to_pharmacy(cart, pharmacy)
            if allocations is None:
                continue

            selected_pharmacy = pharmacy
            selected_distance = distance
            selected_allocations = allocations
            break

        if selected_pharmacy is None:
            return {'error': 'Không có chi nhánh nào đủ tồn kho để xử lý toàn bộ giỏ hàng hiện tại.'}
    else:
        selected_pharmacy = nearest_pharmacy
        selected_distance = nearest_distance

    route_result = delivery_service.estimate_route(
        start_lat=selected_pharmacy.lat,
        start_lng=selected_pharmacy.lng,
        end_lat=delivery_lat,
        end_lng=delivery_lng,
        delivery_mode=delivery_mode,
        departure_time_str=departure_time_str,
    )
    if 'routes' not in route_result or not route_result['routes']:
        return {'error': 'Không thể tính được tuyến giao hàng cho chi nhánh đã chọn.'}

    notice = ''
    if selected_pharmacy.id != nearest_pharmacy.id:
        notice = (
            f"Chi nhánh gần nhất là {nearest_pharmacy.name} "
            f"({nearest_distance:.1f} km) nhưng chưa đủ thuốc cho toàn bộ đơn hàng. "
            f"Hệ thống đã chuyển sang {selected_pharmacy.name} "
            f"({selected_distance:.1f} km)."
        )

    return {
        'pharmacy': selected_pharmacy,
        'route': route_result['routes'][0],
        'mode': route_result.get('mode', delivery_mode),
        'notice': notice,
        'allocations': selected_allocations,
        'nearest_pharmacy': nearest_pharmacy,
        'nearest_distance_km': nearest_distance,
    }


def find_pharmacy_in_list(pharmacies, pharmacy_id):
    pharmacy_id_text = str(pharmacy_id).strip()
    if not pharmacy_id_text:
        return None

    for pharmacy in pharmacies:
        if str(pharmacy.id) == pharmacy_id_text:
            return pharmacy

    return None


def verify_cart_stock(cart):
    """
    Kiểm tra toàn bộ sản phẩm trong giỏ có đủ tồn kho hay không.
    """
    for cart_item in cart.items.select_related('medicine').all():
        if cart_item.quantity > cart_item.medicine.quantity:
            return (
                False,
                f"Xin lỗi, thuốc '{cart_item.medicine.name}' chỉ còn {cart_item.medicine.quantity} sản phẩm.",
            )

    return True, ''


def home(request):
    """
    Trang chủ hiển thị danh sách thuốc và hỗ trợ tìm kiếm.
    """
    home_content = HomePageContent.get_solo()
    ensure_home_page_defaults(home_content)
    search_keyword = request.GET.get('q', '').strip()
    featured_discount_promotions = list(
        build_active_promotion_queryset(
            MedicinePromotion.objects.select_related('medicine__pharmacy').order_by('-discount_percent', '-id')
        )
    )
    sale_medicines = []
    seen_sale_keys = set()
    hidden_sale_keys = set()
    for promotion in featured_discount_promotions:
        medicine = promotion.medicine
        if medicine is None or medicine.quantity <= 0:
            continue
        sale_key = normalize_catalog_key(medicine.name, medicine.unit, medicine.manufacturer)
        hidden_sale_keys.add(sale_key)
        if sale_key in seen_sale_keys:
            continue
        payload = build_medicine_discount_payload(medicine, promotion)
        if not payload['has_discount']:
            continue
        medicine.active_discount_payload = payload
        medicine._prefetched_active_promotions = [payload['promotion']] if payload['promotion'] else []
        medicine.current_price_value = payload['discounted_price']
        medicine.original_price_value = payload['original_price']
        medicine.discount_percent_value = payload['discount_percent']
        medicine.has_discount_value = True
        sale_medicines.append(medicine)
        seen_sale_keys.add(sale_key)
        if len(sale_medicines) >= 8:
            break

    medicines_queryset = get_medicine_search_queryset(search_keyword).order_by('-id')
    medicines_source = list(medicines_queryset[:18] if not search_keyword else medicines_queryset)
    if hidden_sale_keys:
        medicines_source = [
            medicine for medicine in medicines_source
            if normalize_catalog_key(medicine.name, medicine.unit, medicine.manufacturer) not in hidden_sale_keys
        ]
    medicines = deduplicate_catalog_medicines(medicines_source, limit=(8 if not search_keyword else None))

    popular_categories = get_popular_categories(limit=6)
    featured_pharmacies = get_featured_pharmacies(limit=8)

    context = {
        'home_content': home_content,
        'home_hero_slides': list(home_content.hero_slides.filter(is_active=True).order_by('sort_order', 'id')),
        'home_category_items': list(home_content.category_spotlights.filter(is_active=True).order_by('sort_order', 'id')),
        'home_commitment_items': list(home_content.service_commitments.filter(is_active=True).order_by('sort_order', 'id')),
        'medicines': medicines,
        'sale_medicines': sale_medicines,
        'query': search_keyword,
        'popular_categories': popular_categories,
        'featured_pharmacies': featured_pharmacies,
        'featured_medicine_total': medicines_queryset.count() if search_keyword else Medicine.objects.count(),
        'available_pharmacy_total': get_available_pharmacies().count(),
    }
    return render(request, 'pages/home.html', context)


def about_view(request):
    about_content = AboutPageContent.get_solo()
    ensure_about_page_slide_defaults(about_content)
    ensure_about_featured_branch_defaults(about_content)
    featured_categories = list(
        Medicine.objects.exclude(category="")
        .values_list("category", flat=True)
        .distinct()
        .order_by("category")[:8]
    )
    active_branch_items = list(
        about_content.featured_branch_items.select_related("pharmacy").filter(is_active=True).order_by("sort_order", "id")
    )
    branch_items_source = active_branch_items or build_default_about_featured_branch_items(about_content)
    custom_blocks = list(about_content.custom_blocks.filter(is_active=True).order_by("sort_order", "id"))
    product_type_summary = {
        'medicine': Medicine.objects.filter(product_type=MEDICINE_PRODUCT_TYPE_MEDICINE).count(),
        'supplement': Medicine.objects.filter(product_type=MEDICINE_PRODUCT_TYPE_SUPPLEMENT).count(),
    }
    context = {
        "about_sections": build_about_section_layout(about_content, custom_blocks=custom_blocks),
        "featured_categories": featured_categories,
        "about_branch_items": [build_about_featured_branch_payload(item, about_content) for item in branch_items_source],
        "product_type_summary": product_type_summary,
        "about_content": about_content,
        "about_slides": list(about_content.hero_slides.filter(is_active=True).order_by("sort_order", "id")),
        "about_story_items": [
            {"icon": "fas fa-map-marked-alt", "title": about_content.story_item_1_title, "body": about_content.story_item_1_body},
            {"icon": "fas fa-database", "title": about_content.story_item_2_title, "body": about_content.story_item_2_body},
            {"icon": "fas fa-shipping-fast", "title": about_content.story_item_3_title, "body": about_content.story_item_3_body},
        ],
        "about_problem_items": [
            {"icon": "fas fa-search", "title": about_content.problem_item_1_title, "body": about_content.problem_item_1_body},
            {"icon": "fas fa-eye", "title": about_content.problem_item_2_title, "body": about_content.problem_item_2_body},
            {"icon": "fas fa-map-pin", "title": about_content.problem_item_3_title, "body": about_content.problem_item_3_body},
        ],
        "about_value_cards": [
            {"icon": "fas fa-layer-group", "title": about_content.value_card_1_title, "body": about_content.value_card_1_body},
            {"icon": "fas fa-hand-holding-medical", "title": about_content.value_card_2_title, "body": about_content.value_card_2_body},
            {"icon": "fas fa-map", "title": about_content.value_card_3_title, "body": about_content.value_card_3_body},
        ],
        "about_journey_steps": [
            {"number": "01", "title": about_content.step_1_title, "body": about_content.step_1_body},
            {"number": "02", "title": about_content.step_2_title, "body": about_content.step_2_body},
            {"number": "03", "title": about_content.step_3_title, "body": about_content.step_3_body},
            {"number": "04", "title": about_content.step_4_title, "body": about_content.step_4_body},
        ],
        "about_branch_role_items": [
            {"icon": "fas fa-clock", "title": about_content.branch_role_item_1_title, "body": about_content.branch_role_item_1_body},
            {"icon": "fas fa-image", "title": about_content.branch_role_item_2_title, "body": about_content.branch_role_item_2_body},
            {"icon": "fas fa-map-pin", "title": about_content.branch_role_item_3_title, "body": about_content.branch_role_item_3_body},
        ],
        "about_custom_blocks": custom_blocks,
        "pharmacy_total": Pharmacy.objects.count(),
        "medicine_total": Medicine.objects.count(),
        "order_total": Order.objects.count(),
        "review_total": MedicineReview.objects.count() + PharmacyReview.objects.count(),
    }
    return render(request, "pages/about.html", context)


def news_list_view(request):
    article_queryset = NewsArticle.objects.filter(is_published=True).order_by("-published_at", "-created_at", "-id")
    page_obj = Paginator(article_queryset, 6).get_page(request.GET.get("page"))
    context = {
        "page_title": "Tin tức",
        "featured_article": page_obj.object_list[0] if page_obj.number == 1 and page_obj.object_list else None,
        "page_obj": page_obj,
        "latest_articles": list(article_queryset[:5]),
    }
    return render(request, "news/list.html", context)


def news_detail_view(request, slug):
    article = get_object_or_404(NewsArticle.objects.filter(is_published=True), slug=slug)
    related_articles = list(
        NewsArticle.objects.filter(is_published=True).exclude(pk=article.pk).order_by("-published_at", "-created_at", "-id")[:4]
    )
    return render(
        request,
        "news/detail.html",
        {
            "article": article,
            "related_articles": related_articles,
        },
    )


def product_list(request):
    """
    Trang danh sách toàn bộ sản phẩm.
    """
    search_keyword = request.GET.get('q', '').strip()
    active_product_type = request.GET.get('product_type', '').strip()
    active_category = request.GET.get('category', '').strip()
    active_pharmacy = request.GET.get('pharmacy', '').strip()
    availability = request.GET.get('availability', 'in_stock').strip() or 'in_stock'
    sort_key = request.GET.get('sort', 'popular').strip() or 'popular'
    sale_filter = request.GET.get('sale', '').strip()

    medicines_queryset = get_medicine_search_queryset(search_keyword)
    if active_product_type in {choice[0] for choice in MEDICINE_PRODUCT_TYPE_CHOICES}:
        medicines_queryset = medicines_queryset.filter(product_type=active_product_type)
    else:
        active_product_type = ''
    if active_category:
        medicines_queryset = medicines_queryset.filter(category__iexact=active_category)
    if active_pharmacy.isdigit():
        medicines_queryset = medicines_queryset.filter(pharmacy_id=int(active_pharmacy))
    else:
        active_pharmacy = ''

    if availability == 'in_stock':
        medicines_queryset = medicines_queryset.filter(quantity__gt=0)
    elif availability != 'all':
        availability = 'in_stock'
        medicines_queryset = medicines_queryset.filter(quantity__gt=0)

    if sale_filter not in {'', 'yes'}:
        sale_filter = ''

    medicines = list(medicines_queryset.order_by('name', 'unit', 'manufacturer', 'origin', 'price', '-quantity', 'id'))
    promotion_map = build_active_promotion_map(medicines)

    grouped_products = {}

    for medicine in medicines:
        discount_payload = build_medicine_discount_payload(medicine, promotion_map.get(medicine.id))
        current_price = discount_payload['discounted_price']
        original_price = discount_payload['original_price']
        key = normalize_catalog_key(
            medicine.name,
            medicine.unit,
            medicine.manufacturer,
        )
        branch_entry = {
            'pharmacy_id': medicine.pharmacy_id,
            'pharmacy_name': medicine.pharmacy.name,
            'medicine_id': medicine.id,
            'quantity': medicine.quantity,
            'price_value': current_price,
            'price_text': f"{current_price:,} đ".replace(',', '.'),
            'original_price_value': original_price,
            'discount_percent': discount_payload['discount_percent'],
            'has_discount': discount_payload['has_discount'],
        }

        if key not in grouped_products:
            grouped_products[key] = {
                'id': medicine.id,
                'detail_id': medicine.id,
                'name': medicine.name,
                'product_type': medicine.product_type,
                'product_type_label': dict(MEDICINE_PRODUCT_TYPE_CHOICES).get(medicine.product_type, 'Sản phẩm'),
                'category': medicine.category or 'Thuốc / Dược phẩm',
                'unit': medicine.unit or 'Hộp',
                'manufacturer': medicine.manufacturer or 'Đang cập nhật',
                'origin': medicine.origin or 'Đang cập nhật',
                'description': build_medicine_short_description(
                    getattr(medicine, 'short_description', ''),
                    medicine.description,
                    medicine.usage,
                    'Thông tin sản phẩm đang được cập nhật.',
                ),
                'primary_image_url': medicine.primary_image_url or MEDICINE_FALLBACK_IMAGE,
                'is_in_stock': medicine.quantity > 0,
                'min_price_value': current_price,
                'max_price_value': current_price,
                'min_original_price_value': original_price,
                'branch_count': 0,
                'in_stock_branch_count': 0,
                'total_stock': 0,
                'availability_entries': [],
                'has_discount': discount_payload['has_discount'],
                'discount_percent': discount_payload['discount_percent'],
                '_representative_rank': (
                    0 if medicine.quantity > 0 else 1,
                    0 if discount_payload['has_discount'] else 1,
                    current_price,
                    medicine.id,
                ),
            }

        product = grouped_products[key]
        product['branch_count'] += 1
        product['min_price_value'] = min(product['min_price_value'], current_price)
        product['max_price_value'] = max(product['max_price_value'], current_price)
        product['min_original_price_value'] = min(product['min_original_price_value'], original_price)
        product['has_discount'] = product['has_discount'] or discount_payload['has_discount']
        product['discount_percent'] = max(product['discount_percent'], discount_payload['discount_percent'])

        if medicine.quantity > 0:
            product['is_in_stock'] = True
            product['in_stock_branch_count'] += 1
            product['total_stock'] += medicine.quantity
            product['availability_entries'].append(branch_entry)

        representative_rank = (
            0 if medicine.quantity > 0 else 1,
            0 if discount_payload['has_discount'] else 1,
            current_price,
            medicine.id,
        )
        if representative_rank < product['_representative_rank']:
            product['_representative_rank'] = representative_rank
            product['id'] = medicine.id
            product['detail_id'] = medicine.id
            product['description'] = build_medicine_short_description(
                getattr(medicine, 'short_description', ''),
                medicine.description,
                medicine.usage,
                product['description'],
            )
            product['primary_image_url'] = medicine.primary_image_url or product['primary_image_url']
            product['product_type'] = medicine.product_type
            product['product_type_label'] = dict(MEDICINE_PRODUCT_TYPE_CHOICES).get(medicine.product_type, 'Sản phẩm')

    products = list(grouped_products.values())

    for product in products:
        product['availability_entries'].sort(
            key=lambda item: (-item['quantity'], item['price_value'], item['pharmacy_name'].casefold())
        )
        product['availability_preview'] = product['availability_entries'][:3]
        product['extra_branch_count'] = max(product['in_stock_branch_count'] - len(product['availability_preview']), 0)
        product['min_price_text'] = f"{product['min_price_value']:,} đ".replace(',', '.')
        product['max_price_text'] = f"{product['max_price_value']:,} đ".replace(',', '.')
        product['min_original_price_text'] = f"{product['min_original_price_value']:,} đ".replace(',', '.')
        del product['_representative_rank']

    if sale_filter == 'yes':
        products = [product for product in products if product['has_discount']]

    if sort_key == 'name':
        products.sort(key=lambda item: (item['name'].casefold(), item['manufacturer'].casefold(), item['id']))
    elif sort_key == 'price_low':
        products.sort(key=lambda item: (item['min_price_value'], item['name'].casefold(), item['id']))
    elif sort_key == 'price_high':
        products.sort(key=lambda item: (-item['max_price_value'], item['name'].casefold(), item['id']))
    elif sort_key == 'stock':
        products.sort(key=lambda item: (-item['total_stock'], -item['in_stock_branch_count'], item['name'].casefold(), item['id']))
    else:
        sort_key = 'popular'
        products.sort(
            key=lambda item: (
                -int(item['has_discount']),
                -item['in_stock_branch_count'],
                -item['total_stock'],
                item['min_price_value'],
                item['name'].casefold(),
                item['id'],
            )
        )

    paginator = Paginator(products, 6)
    page_obj = paginator.get_page(request.GET.get('page'))
    query_params = request.GET.copy()
    query_params.pop('page', None)
    pharmacy_options = list(get_available_pharmacies().order_by('name').values('id', 'name'))
    active_pharmacy_name = ''
    if active_pharmacy:
        active_pharmacy_name = next(
            (
                pharmacy['name']
                for pharmacy in pharmacy_options
                if str(pharmacy['id']) == str(active_pharmacy)
            ),
            '',
        )

    return render(
        request,
        'catalog/products.html',
        {
            'page_obj': page_obj,
            'products': page_obj.object_list,
            'query': search_keyword,
            'active_product_type': active_product_type,
            'active_category': active_category,
            'active_pharmacy': active_pharmacy,
            'availability': availability,
            'sale_filter': sale_filter,
            'sort_key': sort_key,
            'popular_categories': get_popular_categories(limit=8),
            'product_type_options': list(MEDICINE_PRODUCT_TYPE_CHOICES),
            'pharmacy_options': pharmacy_options,
            'active_pharmacy_name': active_pharmacy_name,
            'available_pharmacy_total': len(pharmacy_options),
            'product_total': paginator.count,
            'current_page_count': len(page_obj.object_list),
            'query_string': query_params.urlencode(),
        }
    )


def build_medicine_detail_context(request, medicine, review_form=None):
    related_queryset = Medicine.objects.select_related('pharmacy').exclude(pk=medicine.pk)
    if medicine.category:
        related_queryset = related_queryset.filter(
            Q(category__iexact=medicine.category) | Q(pharmacy=medicine.pharmacy)
        )
    else:
        related_queryset = related_queryset.filter(pharmacy=medicine.pharmacy)

    related_medicines = attach_discount_payloads(list(related_queryset.order_by('-id')[:6]))

    target_key = normalize_catalog_key(medicine.name, medicine.unit, medicine.manufacturer)
    same_product_candidates = [
        item for item in Medicine.objects.select_related('pharmacy').order_by('-quantity', 'pharmacy__name', 'id')
        if normalize_catalog_key(item.name, item.unit, item.manufacturer) == target_key
    ]
    attach_discount_payloads(same_product_candidates)
    grouped_branch_products = {}
    for item in same_product_candidates:
        rank = (
            0 if item.quantity > 0 else 1,
            getattr(item, 'current_price_value', item.current_price),
            -int(item.quantity or 0),
            item.id,
        )
        current = grouped_branch_products.get(item.pharmacy_id)
        if current is None or rank < current[0]:
            grouped_branch_products[item.pharmacy_id] = (rank, item)
    same_product_branches = [entry[1] for entry in grouped_branch_products.values()]
    same_product_branches.sort(key=lambda item: (-int(item.quantity or 0), item.pharmacy.name.casefold(), item.id))
    attach_discount_payloads([medicine])
    availability_rows = [
        {
            'pharmacy_id': item.pharmacy_id,
            'pharmacy_name': item.pharmacy.name,
            'quantity': item.quantity,
            'price': getattr(item, 'current_price_value', item.current_price),
            'original_price': getattr(item, 'original_price_value', item.price),
            'has_discount': getattr(item, 'has_discount_value', item.has_active_discount),
            'discount_percent': getattr(item, 'discount_percent_value', item.discount_percent),
        }
        for item in same_product_branches
        if item.quantity > 0
    ]
    shared_gallery = []
    for item in same_product_branches:
        for image_url in get_entity_gallery_urls(item, MEDICINE_FALLBACK_IMAGE):
            if image_url not in shared_gallery:
                shared_gallery.append(image_url)

    if not shared_gallery:
        shared_gallery = get_entity_gallery_urls(medicine, MEDICINE_FALLBACK_IMAGE)

    system_prices = sorted({(getattr(item, 'current_price_value', item.current_price)) for item in same_product_branches})
    original_prices = sorted({item.price for item in same_product_branches})
    availability_summary = {
        'branch_total': len(same_product_branches),
        'in_stock_branch_total': len(availability_rows),
        'total_stock': sum(item['quantity'] for item in availability_rows),
        'rows': availability_rows[:3],
        'extra_branch_count': max(len(availability_rows) - 3, 0),
        'is_price_consistent': len(system_prices) <= 1,
        'system_price_value': system_prices[0] if system_prices else medicine.current_price,
        'price_min_value': system_prices[0] if system_prices else medicine.current_price,
        'price_max_value': system_prices[-1] if system_prices else medicine.current_price,
        'original_price_value': original_prices[0] if original_prices else medicine.price,
        'has_discount': getattr(medicine, 'has_discount_value', medicine.has_active_discount),
        'discount_percent': getattr(medicine, 'discount_percent_value', medicine.discount_percent),
    }

    cart = get_or_create_cart(request)
    cart_item = cart.items.filter(medicine=medicine).first()

    review_queryset = medicine.reviews.select_related('user').order_by('-updated_at', '-id')
    existing_review = None
    if request.user.is_authenticated:
        existing_review = review_queryset.filter(user=request.user).first()
    if review_form is None:
        review_form = MedicineReviewForm(instance=existing_review)

    review_summary = build_review_summary(review_queryset)
    review_panel = build_review_panel(review_queryset, per_page=5)
    can_review = can_user_review_medicine(request.user, medicine, existing_review)

    return {
        'medicine': medicine,
        'medicine_gallery': shared_gallery,
        'medicine_fallback_image': MEDICINE_FALLBACK_IMAGE,
        'related_medicines': related_medicines,
        'availability_summary': availability_summary,
        'cart_item': cart_item,
        'medicine_reviews': review_panel['items'],
        'medicine_review_summary': review_summary,
        'medicine_review_panel': review_panel,
        'medicine_review_form': review_form,
        'medicine_user_review': existing_review,
        'medicine_review_access': build_review_access_context(
            request.user.is_authenticated,
            can_review,
            'sản phẩm này',
        ),
        'medicine_reviews_api_url': reverse('medicine_reviews_api', args=[medicine.id]),
    }


def medicine_detail(request, medicine_id):
    """
    Trang chi tiết sản phẩm thuốc.
    """
    medicine = get_object_or_404(
        Medicine.objects.select_related('pharmacy'),
        pk=medicine_id,
    )
    return render(request, 'catalog/medicine_detail.html', build_medicine_detail_context(request, medicine))


@login_required(login_url='/login/')
def submit_medicine_review(request, medicine_id):
    if request.method != 'POST':
        return redirect('medicine_detail', medicine_id=medicine_id)

    medicine = get_object_or_404(Medicine.objects.select_related('pharmacy'), pk=medicine_id)
    existing_review = MedicineReview.objects.filter(user=request.user, medicine=medicine).first()

    if not can_user_review_medicine(request.user, medicine, existing_review):
        messages.error(request, 'Chỉ khách hàng đã mua và có đơn hàng hoàn thành mới được đánh giá sản phẩm này.')
        return redirect('medicine_detail', medicine_id=medicine.id)

    form = MedicineReviewForm(request.POST, instance=existing_review)

    if form.is_valid():
        review = form.save(commit=False)
        review.user = request.user
        review.medicine = medicine
        if existing_review:
            review.is_edited = True
        review.save()
        if existing_review:
            messages.success(request, 'Đã cập nhật đánh giá sản phẩm của bạn.')
        else:
            messages.success(request, 'Cảm ơn bạn đã gửi đánh giá cho sản phẩm.')
        return redirect('medicine_detail', medicine_id=medicine.id)

    messages.error(request, 'Vui lòng chọn số sao hợp lệ trước khi gửi đánh giá.')
    return render(request, 'catalog/medicine_detail.html', build_medicine_detail_context(request, medicine, form))


def build_pharmacy_detail_context(request, pharmacy, review_form=None):
    medicines = Medicine.objects.filter(pharmacy=pharmacy).order_by('name')
    has_stock = medicines.filter(quantity__gt=0).exists()
    review_queryset = pharmacy.reviews.select_related('user').order_by('-updated_at', '-id')
    existing_review = None
    if request.user.is_authenticated:
        existing_review = review_queryset.filter(user=request.user).first()
    if review_form is None:
        review_form = PharmacyReviewForm(instance=existing_review)

    review_summary = build_review_summary(review_queryset)
    review_panel = build_review_panel(review_queryset, per_page=5)
    can_review = can_user_review_pharmacy(request.user, pharmacy, existing_review)

    return {
        'pharmacy': pharmacy,
        'pharmacy_gallery': get_entity_gallery_urls(
            pharmacy,
            PHARMACY_FALLBACK_IMAGE,
        ),
        'pharmacy_open_state': build_opening_status(pharmacy.opening_hours),
        'medicines': medicines,
        'has_stock': has_stock,
        'popular_categories': get_popular_categories(limit=6),
        'pharmacy_reviews': review_panel['items'],
        'pharmacy_review_summary': review_summary,
        'pharmacy_review_panel': review_panel,
        'pharmacy_review_form': review_form,
        'pharmacy_user_review': existing_review,
        'pharmacy_review_access': build_review_access_context(
            request.user.is_authenticated,
            can_review,
            'chi nhánh này',
        ),
        'pharmacy_reviews_api_url': reverse('pharmacy_reviews_api', args=[pharmacy.id]),
    }


def pharmacy_detail(request, pharmacy_id):
    """
    Trang chi tiết một nhà thuốc.
    """
    pharmacy = get_object_or_404(Pharmacy, pk=pharmacy_id)
    return render(request, 'locations/pharmacy_detail.html', build_pharmacy_detail_context(request, pharmacy))


@login_required(login_url='/login/')
def submit_pharmacy_review(request, pharmacy_id):
    if request.method != 'POST':
        return redirect('pharmacy_detail', pharmacy_id=pharmacy_id)

    pharmacy = get_object_or_404(Pharmacy, pk=pharmacy_id)
    existing_review = PharmacyReview.objects.filter(user=request.user, pharmacy=pharmacy).first()

    if not can_user_review_pharmacy(request.user, pharmacy, existing_review):
        messages.error(request, 'Chỉ khách hàng đã mua và có đơn hàng hoàn thành mới được đánh giá chi nhánh này.')
        return redirect('pharmacy_detail', pharmacy_id=pharmacy.id)

    form = PharmacyReviewForm(request.POST, instance=existing_review)

    if form.is_valid():
        review = form.save(commit=False)
        review.user = request.user
        review.pharmacy = pharmacy
        if existing_review:
            review.is_edited = True
        review.save()
        if existing_review:
            messages.success(request, 'Đã cập nhật đánh giá chi nhánh của bạn.')
        else:
            messages.success(request, 'Cảm ơn bạn đã gửi đánh giá cho chi nhánh.')
        return redirect('pharmacy_detail', pharmacy_id=pharmacy.id)

    messages.error(request, 'Vui lòng chọn số sao hợp lệ trước khi gửi đánh giá.')
    return render(request, 'locations/pharmacy_detail.html', build_pharmacy_detail_context(request, pharmacy, form))


def medicine_reviews_api(request, medicine_id):
    medicine = get_object_or_404(Medicine, pk=medicine_id)
    active_rating = parse_review_rating(request.GET.get('rating'))
    page_number = request.GET.get('page') or 1
    review_queryset = medicine.reviews.select_related('user').order_by('-updated_at', '-id')
    review_panel = build_review_panel(review_queryset, active_rating=active_rating, page_number=page_number, per_page=5)
    return JsonResponse(build_review_api_payload(review_panel))


def pharmacy_reviews_api(request, pharmacy_id):
    pharmacy = get_object_or_404(Pharmacy, pk=pharmacy_id)
    active_rating = parse_review_rating(request.GET.get('rating'))
    page_number = request.GET.get('page') or 1
    review_queryset = pharmacy.reviews.select_related('user').order_by('-updated_at', '-id')
    review_panel = build_review_panel(review_queryset, active_rating=active_rating, page_number=page_number, per_page=5)
    return JsonResponse(build_review_api_payload(review_panel))


def add_to_cart(request, medicine_id):
    """
    Thêm sản phẩm vào giỏ hàng.
    - GET: thêm nhanh 1 sản phẩm
    - POST: thêm theo số lượng từ trang chi tiết
    """
    medicine = get_object_or_404(Medicine, id=medicine_id)
    cart = get_or_create_cart(request)

    quantity_raw = request.POST.get('quantity') or request.GET.get('quantity') or '1'
    buy_now = request.POST.get('buy_now') or request.GET.get('buy_now')

    try:
        quantity_to_add = int(quantity_raw)
    except (TypeError, ValueError):
        quantity_to_add = 1

    if quantity_to_add < 1:
        quantity_to_add = 1

    fallback_url = request.META.get('HTTP_REFERER')
    expects_json = request_expects_json(request)

    if medicine.quantity <= 0:
        message_text = f"Thuốc '{medicine.name}' hiện đã hết hàng."
        if expects_json:
            return JsonResponse({'message': message_text, 'cart_items_count': get_cart_items_count(cart)}, status=400)
        messages.error(request, message_text)
        if fallback_url:
            return redirect(fallback_url)
        return redirect('medicine_detail', medicine_id=medicine.id)

    cart_item, created = CartItem.objects.get_or_create(cart=cart, medicine=medicine)
    current_quantity = 0 if created else cart_item.quantity
    new_quantity = current_quantity + quantity_to_add

    if new_quantity > medicine.quantity:
        message_text = f"Thuốc '{medicine.name}' chỉ còn {medicine.quantity} sản phẩm trong kho."
        if expects_json:
            return JsonResponse({'message': message_text, 'cart_items_count': get_cart_items_count(cart)}, status=400)
        messages.error(request, message_text)
        if fallback_url:
            return redirect(fallback_url)
        return redirect('medicine_detail', medicine_id=medicine.id)

    cart_item.quantity = new_quantity
    cart_item.save()

    if buy_now:
        message_text = f"Đã thêm '{medicine.name}' vào giỏ và chuyển sang thanh toán."
        if expects_json:
            return JsonResponse(
                {
                    'message': message_text,
                    'cart_items_count': get_cart_items_count(cart),
                    'redirect_url': reverse('checkout'),
                }
            )
        messages.success(request, message_text)
        return redirect('checkout')

    message_text = f"Đã thêm '{medicine.name}' vào giỏ hàng."
    if expects_json:
        return JsonResponse(
            {
                'message': message_text,
                'cart_items_count': get_cart_items_count(cart),
            }
        )
    messages.success(request, message_text)
    if fallback_url:
        return redirect(fallback_url)
    return redirect('medicine_detail', medicine_id=medicine.id)


@ensure_csrf_cookie
def cart_detail(request):
    """
    Trang xem giỏ hàng.
    """
    cart = get_or_create_cart(request)
    cart_pricing = build_cart_pricing_snapshot(cart, request.user)
    return render(
        request,
        'shop/cart.html',
        {
            'cart': cart,
            'cart_items': cart_pricing['items'],
            'cart_pricing': cart_pricing,
        },
    )


def update_cart_item(request, item_id):
    """
    Cập nhật số lượng sản phẩm trong giỏ hàng.
    Cho phép giảm số lượng ngay cả khi giỏ đang vượt tồn kho hiện tại.
    """
    if request.method != 'POST':
        return redirect('cart_detail')

    cart = get_or_create_cart(request)
    cart_item = get_object_or_404(CartItem.objects.select_related('medicine'), id=item_id, cart=cart)
    medicine = cart_item.medicine
    expects_json = request_expects_json(request)

    focus_item_id = (request.POST.get('focus_item_id') or '').strip()
    if not focus_item_id.isdigit():
        focus_item_id = str(item_id)

    def build_cart_redirect():
        redirect_url = reverse('cart_detail')
        if focus_item_id:
            redirect_url += f'#cart-item-{focus_item_id}'
        return redirect(redirect_url)

    def build_cart_response(message_text, *, status=200, removed=False):
        cart.refresh_from_db()
        cart_pricing = build_cart_pricing_snapshot(cart, request.user)
        remaining_items = cart_pricing['items']
        line_item_total = cart_pricing['line_item_count']
        total_quantity = cart_pricing['total_quantity']
        total_price = cart_pricing['final_product_total']
        payload = {
            'message': message_text,
            'item_id': item_id,
            'removed': removed,
            'line_item_count': line_item_total,
            'cart_items_count': total_quantity,
            'cart_total_text': format_vnd(total_price),
            'cart_is_empty': line_item_total == 0,
            'customer_discount_text': format_vnd(cart_pricing['loyalty_discount_total']),
        }
        if not removed:
            current_item = next((item for item in remaining_items if item.id == item_id), None)
            if current_item:
                payload.update(
                    {
                        'quantity': current_item.quantity,
                        'row_total_text': format_vnd(current_item.line_final_total),
                        'stock_quantity': current_item.medicine.quantity,
                        'stock_note': f'Còn {current_item.medicine.quantity} sản phẩm trong kho',
                        'is_over_stock': current_item.quantity > current_item.medicine.quantity,
                    }
                )
        return JsonResponse(payload, status=status)

    raw_quantity = request.POST.get('quantity', '').strip()
    action = (request.POST.get('action') or '').strip()

    try:
        quantity = int(raw_quantity)
    except (TypeError, ValueError):
        quantity = cart_item.quantity

    if action == 'increase':
        quantity = cart_item.quantity + 1
    elif action == 'decrease':
        quantity = cart_item.quantity - 1

    if quantity <= 0:
        cart_item.delete()
        message_text = f"Đã xóa '{medicine.name}' khỏi giỏ hàng."
        if expects_json:
            return build_cart_response(message_text, removed=True)
        messages.success(request, message_text)
        return build_cart_redirect()

    if medicine.quantity <= 0:
        message_text = f"Thuốc '{medicine.name}' hiện đã hết hàng. Vui lòng xóa khỏi giỏ hàng."
        if expects_json:
            return build_cart_response(message_text, status=400)
        messages.error(request, message_text)
        return build_cart_redirect()

    if quantity > medicine.quantity:
        message_text = (
            f"Thuốc '{medicine.name}' hiện chỉ còn {medicine.quantity} sản phẩm. "
            "Bạn vẫn có thể giảm số lượng trong giỏ hàng."
        )
        if expects_json:
            return build_cart_response(message_text, status=400)
        messages.error(request, message_text)
        return build_cart_redirect()

    cart_item.quantity = quantity
    cart_item.save(update_fields=['quantity'])
    message_text = f"Đã cập nhật số lượng '{medicine.name}' trong giỏ hàng."
    if expects_json:
        return build_cart_response(message_text)
    messages.success(request, message_text)
    return build_cart_redirect()


def remove_from_cart(request, item_id):
    """
    Xóa một sản phẩm khỏi giỏ hàng.
    """
    cart = get_or_create_cart(request)
    cart_item = get_object_or_404(CartItem, id=item_id, cart=cart)
    cart_item.delete()
    return redirect('cart_detail')


def checkout(request):
    """
    Xử lý trang thanh toán.
    """
    cart = get_or_create_cart(request)

    if not cart.items.exists():
        messages.warning(request, 'Bạn cần có sản phẩm trong giỏ hàng trước khi thanh toán.')
        return redirect('cart_detail')

    available_pharmacies = list(
        get_available_pharmacies().prefetch_related(
            Prefetch(
                'medicines',
                queryset=Medicine.objects.order_by('-quantity', 'id'),
                to_attr='_inventory_candidates',
            )
        )
    )
    if not available_pharmacies:
        messages.error(request, 'Hiện chưa có chi nhánh khả dụng để xử lý đơn hàng.')
        return redirect('cart_detail')

    has_fulfillment_branch = any(
        allocate_cart_to_pharmacy(cart, pharmacy) is not None
        for pharmacy in available_pharmacies
    )
    if not has_fulfillment_branch:
        messages.error(request, 'Không có chi nhánh nào đủ tồn kho để xử lý toàn bộ giỏ hàng hiện tại.')
        return redirect('cart_detail')

    pharmacy_payload = build_checkout_pharmacy_payload(available_pharmacies)
    profile = get_or_create_user_profile(request.user) if request.user.is_authenticated else None
    saved_address = build_saved_address_payload(profile) if profile and is_customer_user(request.user) else None
    cart_pricing = build_cart_pricing_snapshot(cart, request.user)
    loyalty_context = cart_pricing['loyalty']
    prescription_context = build_cart_prescription_context(cart)
    requested_address = (request.POST.get('address_text') or request.GET.get('address') or '').strip()
    requested_pharmacy_id = (request.POST.get('pharmacy_id') or request.GET.get('pharmacy_id') or '').strip()
    requested_lat = (request.POST.get('delivery_lat') or request.GET.get('delivery_lat') or '').strip()
    requested_lng = (request.POST.get('delivery_lng') or request.GET.get('delivery_lng') or '').strip()
    requested_departure_time = get_default_departure_time_value(request.POST.get('departure_time') or '')

    def build_checkout_context(form_instance):
        preselected_delivery = None

        if requested_lat and requested_lng:
            try:
                preselected_delivery = {
                    'lat': float(requested_lat),
                    'lng': float(requested_lng),
                    'address_text': requested_address,
                }
            except (TypeError, ValueError):
                preselected_delivery = None
        elif saved_address and saved_address.get('lat') is not None and saved_address.get('lng') is not None:
            preselected_delivery = saved_address

        return {
            'form': form_instance,
            'cart': cart,
            'cart_pricing': cart_pricing,
            'loyalty_context': loyalty_context,
            'is_customer': is_customer_user(request.user),
            'pharmacies': pharmacy_payload,
            'saved_address': saved_address,
            'preselected_delivery': preselected_delivery,
            'preselected_pharmacy_id': requested_pharmacy_id,
            'departure_time_value': requested_departure_time,
            'prescription_context': prescription_context,
        }

    if request.method == 'POST':
        form = CheckoutForm(request.POST, request.FILES)

        if not form.is_valid():
            messages.error(request, 'Thông tin chưa hợp lệ. Vui lòng kiểm tra lại biểu mẫu.')
            return render(request, 'shop/checkout.html', build_checkout_context(form))

        uploaded_prescription_images = form.cleaned_data.get('prescription_proof_image') or []
        if prescription_context['requires_prescription'] and not uploaded_prescription_images:
            form.add_error('prescription_proof_image', 'Giỏ hàng có thuốc cần kê đơn. Vui lòng tải lên ảnh đơn thuốc.')
            messages.error(request, 'Giỏ hàng có thuốc cần kê đơn. Vui lòng bổ sung ảnh đơn thuốc trước khi đặt hàng.')
            return render(request, 'shop/checkout.html', build_checkout_context(form))

        delivery_lat = request.POST.get('delivery_lat', '').strip()
        delivery_lng = request.POST.get('delivery_lng', '').strip()
        selected_pharmacy_id = request.POST.get('pharmacy_id', '').strip()
        departure_time = get_default_departure_time_value(request.POST.get('departure_time', ''))
        payment_method = form.cleaned_data.get('payment_method') or Order.PAYMENT_COD
        invoice_requested = bool(form.cleaned_data.get('invoice_requested'))

        if not delivery_lat or not delivery_lng:
            messages.error(request, 'Vui lòng chọn vị trí giao hàng trên bản đồ để hệ thống tính chi nhánh và phí ship.')
            return render(request, 'shop/checkout.html', build_checkout_context(form))

        best_delivery_result = None

        if selected_pharmacy_id:
            selected_pharmacy = find_pharmacy_in_list(available_pharmacies, selected_pharmacy_id)
            selected_allocations = allocate_cart_to_pharmacy(cart, selected_pharmacy) if selected_pharmacy else None

            if selected_pharmacy and selected_allocations is not None:
                route_result = delivery_service.estimate_route(
                    start_lat=selected_pharmacy.lat,
                    start_lng=selected_pharmacy.lng,
                    end_lat=delivery_lat,
                    end_lng=delivery_lng,
                    delivery_mode='motorbike',
                    departure_time_str=departure_time,
                )

                if 'routes' in route_result and route_result['routes']:
                    best_delivery_result = {
                        'pharmacy': selected_pharmacy,
                        'route': route_result['routes'][0],
                        'allocations': selected_allocations,
                        'notice': '',
                    }

        if best_delivery_result is None:
            best_delivery_result = choose_checkout_pharmacy(
                cart,
                available_pharmacies,
                delivery_lat,
                delivery_lng,
                'motorbike',
                departure_time,
            )

        if 'error' in best_delivery_result:
            messages.error(request, best_delivery_result['error'])
            return render(request, 'shop/checkout.html', build_checkout_context(form))

        selected_pharmacy = best_delivery_result['pharmacy']
        selected_route = best_delivery_result['route']
        selected_allocations = best_delivery_result.get('allocations')
        if selected_allocations is None:
            selected_allocations = allocate_cart_to_pharmacy(cart, selected_pharmacy)

        if selected_allocations is None:
            messages.error(request, 'Chi nhánh được chọn không còn đủ tồn kho cho giỏ hàng hiện tại.')
            return redirect('cart_detail')

        try:
            with transaction.atomic():
                locked_medicines = {
                    medicine.id: medicine
                    for medicine in Medicine.objects.select_for_update().select_related('pharmacy').filter(
                        id__in=[allocation['medicine'].id for allocation in selected_allocations]
                    )
                }
                finalized_allocations = []

                for allocation in selected_allocations:
                    locked_medicine = locked_medicines.get(allocation['medicine'].id)
                    quantity = allocation['quantity']

                    if locked_medicine is None:
                        raise ValueError('Một số sản phẩm trong giỏ không còn khả dụng.')

                    if quantity > locked_medicine.quantity:
                        raise ValueError(
                            f"Thuốc '{locked_medicine.name}' tại {locked_medicine.pharmacy.name} chỉ còn {locked_medicine.quantity} {locked_medicine.unit.lower()}."
                        )

                    finalized_allocations.append({
                        'medicine': locked_medicine,
                        'quantity': quantity,
                    })

                allocated_requires_prescription = any(
                    allocation['medicine'].prescription_required
                    for allocation in finalized_allocations
                )
                requires_prescription_for_order = (
                    prescription_context['requires_prescription']
                    or allocated_requires_prescription
                )
                if requires_prescription_for_order and not uploaded_prescription_images:
                    form.add_error('prescription_proof_image', 'Đơn có thuốc cần kê đơn. Vui lòng tải lên ảnh đơn thuốc.')
                    messages.error(request, 'Đơn có thuốc cần kê đơn. Vui lòng bổ sung ảnh đơn thuốc trước khi đặt hàng.')
                    return render(request, 'shop/checkout.html', build_checkout_context(form))

                selected_total_product_price_before_loyalty = sum(
                    allocation['medicine'].current_price * allocation['quantity']
                    for allocation in finalized_allocations
                )
                selected_loyalty_discount_total = calculate_loyalty_discount_total(
                    selected_total_product_price_before_loyalty,
                    loyalty_context['discount_percent'],
                )
                selected_total_product_price = max(
                    selected_total_product_price_before_loyalty - selected_loyalty_discount_total,
                    0,
                )

                order = form.save(commit=False)

                if request.user.is_authenticated:
                    order.user = request.user

                order.delivery_lat = float(delivery_lat)
                order.delivery_lng = float(delivery_lng)
                order.pharmacy = selected_pharmacy
                order.distance_km = selected_route['distance_km']
                order.shipping_fee = selected_route['shipping_fee_value']
                order.total_product_price = selected_total_product_price
                order.final_total_price = selected_total_product_price + order.shipping_fee
                order.customer_tier_name = loyalty_context['label'] if loyalty_context['discount_percent'] > 0 else ''
                order.customer_tier_discount_percent = loyalty_context['discount_percent']
                order.customer_tier_discount_total = selected_loyalty_discount_total
                order.payment_method = payment_method
                order.payment_status = determine_initial_payment_status(payment_method)
                order.prescription_status = (
                    Order.PRESCRIPTION_STATUS_PENDING
                    if requires_prescription_for_order
                    else Order.PRESCRIPTION_STATUS_NOT_REQUIRED
                )
                order.invoice_requested = invoice_requested
                order.invoice_staff_name = get_invoice_staff_name_for_pharmacy(selected_pharmacy)
                order.estimated_delivery_at = build_order_estimated_delivery_at(
                    timezone.now(),
                    selected_route.get('departure_time') or departure_time,
                    selected_route.get('duration_min'),
                )
                order.save()
                for uploaded_image in uploaded_prescription_images[:3]:
                    OrderPrescriptionProof.objects.create(order=order, image=uploaded_image)
                order.invoice_code = build_order_invoice_code(order)
                order.payment_reference = build_order_payment_reference(order)
                order.save(update_fields=['invoice_code', 'payment_reference'])
                for allocation in finalized_allocations:
                    medicine = allocation['medicine']
                    quantity = allocation['quantity']
                    final_unit_price = medicine.current_price

                    OrderItem.objects.create(
                        order=order,
                        medicine=medicine,
                        medicine_name=medicine.name,
                        price=final_unit_price,
                        quantity=quantity,
                    )

                deduct_inventory_for_order(order, prefer_existing_allocations=False)

                cart.items.all().delete()

                transaction.on_commit(
                    lambda confirmed_order=order, current_request=request: (
                        send_order_confirmation_email(confirmed_order, request=current_request),
                        send_order_invoice_email(confirmed_order, request=current_request),
                    )
                )
        except ValueError as exc:
            messages.error(request, f"{exc} Vui lòng kiểm tra lại giỏ hàng.")
            return redirect('checkout')

        messages.success(request, 'Đặt hàng thành công.')
        if payment_method in {Order.PAYMENT_BANK, Order.PAYMENT_MOMO}:
            messages.info(request, 'Vui lòng thanh toán theo mã đơn thật và gửi ảnh chứng từ trong chi tiết đơn hàng để nhân viên đối soát.')
        if order.requires_prescription_review:
            messages.info(request, 'Đơn có thuốc kê đơn sẽ được dược sĩ duyệt trước khi chuẩn bị giao.')
        if best_delivery_result.get('notice'):
            messages.info(request, best_delivery_result['notice'])
        return redirect('order_history_detail', order_id=order.pk)

    initial_data = {}
    if profile:
        initial_data = {
            'full_name': profile.full_name or request.user.get_full_name() or request.user.username,
            'phone': profile.phone,
            'address_text': requested_address or profile.address_text,
        }
    elif requested_address:
        initial_data['address_text'] = requested_address

    form = CheckoutForm(initial=initial_data)
    return render(request, 'shop/checkout.html', build_checkout_context(form))


def checkout_page(request):
    return checkout(request)


def register_view(request):
    """
    Xử lý đăng ký tài khoản.
    """
    if request.method == 'POST':
        form = RegisterForm(request.POST)

        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.email = form.cleaned_data['email']
            user.is_active = False
            user.save()
            get_or_create_user_profile(user)

            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = email_activation_token.make_token(user)
            email_sent = send_registration_confirmation_email(
                user=user,
                uid=uid,
                token=token,
                request=request,
            )
            if not email_sent:
                user.delete()
                form.add_error(None, "Không thể gửi email xác nhận lúc này. Vui lòng thử lại sau ít phút.")
            else:
                messages.success(
                    request,
                    "Hệ thống đã gửi email xác nhận đăng ký. Vui lòng mở Mailtrap/hộp thư và bấm link xác nhận trước khi đăng nhập.",
                )
                return redirect('login')
    else:
        form = RegisterForm()

    return render(request, 'account/register.html', {'form': form})


def activate_account_view(request, uidb64, token):
    """
    Kích hoạt tài khoản sau khi người dùng bấm link xác nhận email.
    """
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and user.is_active:
        messages.info(request, "Tài khoản này đã được xác nhận trước đó. Bạn có thể đăng nhập ngay.")
        return redirect('login')

    if user is not None and email_activation_token.check_token(user, token):
        user.is_active = True
        user.save(update_fields=['is_active'])
        messages.success(request, "Xác nhận email thành công. Tài khoản đã được kích hoạt, bạn có thể đăng nhập.")
        return redirect('login')

    messages.error(request, "Liên kết xác nhận không hợp lệ hoặc đã hết hạn. Vui lòng đăng ký lại hoặc liên hệ hỗ trợ.")
    return redirect('register')


@ensure_csrf_cookie
def login_view(request):
    """
    Xử lý đăng nhập.
    """
    if request.method == 'POST':
        form = LoginForm(request.POST)

        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)
                return redirect('home')

            inactive_user = User.objects.filter(username__iexact=username, is_active=False).first()
            if inactive_user and inactive_user.check_password(password):
                message = 'Tài khoản này chưa xác nhận email. Vui lòng mở email xác nhận trong Mailtrap/hộp thư trước khi đăng nhập.'
            else:
                message = 'Tên đăng nhập hoặc mật khẩu không đúng.'
            form.add_error(None, message)
            messages.error(request, message)
    else:
        form = LoginForm()

    return render(request, 'account/login.html', {'form': form})


def logout_view(request):
    """
    Đăng xuất tài khoản.
    """
    logout(request)
    return redirect('home')


@login_required(login_url='/login/')
def account_view(request):
    profile = get_or_create_user_profile(request.user)
    is_customer = is_customer_user(request.user)
    user_orders = Order.objects.filter(user=request.user)
    loyalty_context = build_customer_loyalty_context(request.user)
    auto_complete_overdue_shipping_orders(user_orders)
    user_orders = Order.objects.filter(user=request.user)
    recent_orders = list(
        user_orders.select_related('pharmacy').prefetch_related('items__medicine', 'return_request').order_by('-created_at', '-id')[:3]
    )
    recent_orders = [decorate_order_for_customer_display(order) for order in recent_orders]
    cart = get_or_create_cart(request)
    show_password_modal = False
    account_form = AccountProfileForm(
        user=request.user,
        profile=profile,
        is_customer=is_customer,
    )
    password_change_form = ProfilePasswordChangeForm(user=request.user)

    if request.method == 'POST':
        form_action = (request.POST.get('form_action') or 'profile').strip()
        if form_action == 'change_password':
            password_change_form = ProfilePasswordChangeForm(user=request.user, data=request.POST)
            if password_change_form.is_valid():
                updated_user = password_change_form.save()
                update_session_auth_hash(request, updated_user)
                send_password_changed_email(updated_user, request=request, change_source='account')
                messages.success(request, 'Đã cập nhật mật khẩu tài khoản.')
                return redirect('account')
            show_password_modal = True
            messages.error(request, 'Biểu mẫu đổi mật khẩu chưa hợp lệ. Vui lòng kiểm tra lại.')
        else:
            account_form = AccountProfileForm(
                request.POST,
                user=request.user,
                profile=profile,
                is_customer=is_customer,
            )
            if account_form.is_valid():
                full_name = (account_form.cleaned_data.get('full_name') or '').strip()
                new_email = (account_form.cleaned_data.get('email') or '').strip()
                new_phone = (account_form.cleaned_data.get('phone') or '').strip()
                previous_email = (request.user.email or '').strip()
                previous_full_name = (profile.full_name or request.user.get_full_name() or '').strip()
                previous_phone = (profile.phone or '').strip()
                previous_address_text = (profile.address_text or '').strip()
                changed_fields = []

                if full_name != previous_full_name:
                    changed_fields.append('Họ tên')
                if new_email != previous_email:
                    changed_fields.append('Email')
                if new_phone != previous_phone:
                    changed_fields.append('Số điện thoại')

                request.user.email = new_email
                request.user.first_name = full_name
                request.user.last_name = ''
                request.user.save(update_fields=['email', 'first_name', 'last_name'])

                profile.full_name = full_name
                profile.phone = new_phone

                if is_customer:
                    profile.address_text = (account_form.cleaned_data.get('address_text') or '').strip()
                    profile.address_lat = account_form.cleaned_data.get('address_lat')
                    profile.address_lng = account_form.cleaned_data.get('address_lng')
                    if profile.address_text != previous_address_text:
                        changed_fields.append('Địa chỉ giao hàng mặc định')

                    if profile.address_text and (profile.address_lat is None or profile.address_lng is None):
                        try:
                            first_match = search_address_candidates(profile.address_text, limit=1)
                        except Exception:
                            first_match = []

                        if first_match:
                            profile.address_lat = first_match[0]['lat']
                            profile.address_lng = first_match[0]['lng']
                    elif not profile.address_text:
                        profile.address_lat = None
                        profile.address_lng = None

                profile.save()
                if changed_fields:
                    send_account_profile_updated_email(
                        request.user,
                        previous_email=previous_email,
                        changed_fields=changed_fields,
                        request=request,
                    )
                messages.success(request, 'Đã cập nhật thông tin tài khoản.')
                return redirect('account')
            messages.error(request, 'Biểu mẫu thông tin tài khoản chưa hợp lệ. Vui lòng kiểm tra lại.')

    context = {
        'account_role': get_user_role_label(request.user),
        'orders_total': user_orders.count(),
        'orders_pending': user_orders.filter(status__in=[Order.STATUS_PENDING, Order.STATUS_CONFIRMED, Order.STATUS_PACKING]).count(),
        'orders_completed': user_orders.filter(status=Order.STATUS_COMPLETED).count(),
        'recent_orders': recent_orders,
        'cart_items_total': get_cart_items_count(cart),
        'account_form': account_form,
        'password_change_form': password_change_form,
        'profile': profile,
        'is_customer': is_customer,
        'saved_address': build_saved_address_payload(profile),
        'managed_work_pharmacy': profile.managed_pharmacy if request.user.is_staff and not request.user.is_superuser else None,
        'loyalty_context': loyalty_context,
        'show_password_modal': show_password_modal,
    }
    return render(request, 'account/profile.html', context)


@login_required(login_url='/login/')
def order_history(request):
    """
    Hiển thị lịch sử đơn hàng của người dùng đang đăng nhập.
    """
    filter_state = get_order_history_filter_state(request)
    orders = build_order_history_queryset_for_user(request.user, filter_state=filter_state)
    page_obj = Paginator(orders, 3).get_page(request.GET.get('page'))
    query_params = request.GET.copy()
    query_params.pop('page', None)
    context = {
        'orders': page_obj.object_list,
        'page_obj': page_obj,
        'query_string': query_params.urlencode(),
        'filter_state': filter_state,
        'status_choices': Order.STATUS_CHOICES,
        'payment_method_choices': Order.PAYMENT_METHOD_CHOICES,
        'refund_status_choices': [("none", "Chưa tạo yêu cầu"), (ReturnRefundRequest.STATUS_PROCESSING, "Đang xử lý"), (ReturnRefundRequest.STATUS_APPROVED, "Chấp nhận hoàn tiền"), (ReturnRefundRequest.STATUS_REJECTED, "Từ chối hoàn tiền")],
    }
    return render(request, 'account/orders/history.html', context)


@login_required(login_url='/login/')
def order_history_detail(request, order_id):
    order = get_order_for_customer_or_404(request.user, order_id)
    auto_complete_order_if_due(order)
    order.refresh_from_db()
    return_request = getattr(order, 'return_request', None)
    decorate_order_for_customer_display(order)
    order_items = list(
        order.items.select_related('medicine').prefetch_related('lot_allocations__lot').all()
    )
    context = {
        'order': order,
        'order_items': order_items,
        'return_request': return_request,
        'prescription_proof_cards': build_order_prescription_proof_cards(order),
    }
    return render(request, 'account/orders/detail.html', context)


@login_required(login_url='/login/')
def upload_payment_proof(request, order_id):
    order = get_order_for_customer_or_404(request.user, order_id)
    auto_complete_order_if_due(order)
    order.refresh_from_db()

    if not order.can_upload_payment_proof:
        messages.error(request, 'Đơn hàng này không còn ở trạng thái cần bổ sung chứng từ thanh toán.')
        return redirect('order_history_detail', order_id=order.pk)

    if request.method == 'POST':
        form = PaymentProofUploadForm(request.POST, request.FILES, instance=order)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã gửi chứng từ thanh toán. Nhân viên sẽ đối soát và xác nhận sớm.')
            return redirect('order_history_detail', order_id=order.pk)
        messages.error(request, 'Chứng từ thanh toán chưa hợp lệ. Vui lòng kiểm tra lại.')
    else:
        form = PaymentProofUploadForm(instance=order)

    context = {
        'order': order,
        'form': form,
        'payment_preview': build_order_payment_preview(order),
    }
    return render(request, 'account/orders/payment_proof_form.html', context)


@login_required(login_url='/login/')
def order_item_review_redirect(request, order_id, item_id):
    order = get_order_for_customer_or_404(request.user, order_id)
    auto_complete_order_if_due(order)
    order.refresh_from_db()

    order_item = get_object_or_404(
        OrderItem.objects.select_related('medicine', 'order'),
        pk=item_id,
        order=order,
    )

    if order.status != Order.STATUS_COMPLETED:
        messages.error(request, 'Chỉ đơn hàng đã hoàn thành mới có thể đánh giá sản phẩm.')
        return redirect('order_history_detail', order_id=order.pk)

    if not order_item.medicine_id:
        messages.error(request, 'Sản phẩm gốc của dòng hàng này không còn tồn tại nên chưa thể mở trang đánh giá.')
        return redirect('order_history_detail', order_id=order.pk)

    return redirect(f"{reverse('medicine_detail', args=[order_item.medicine_id])}#review-form-section")


@login_required(login_url='/login/')
def cancel_order(request, order_id):
    if request.method != 'POST':
        raise Http404('Yêu cầu không hợp lệ.')

    order = get_order_for_customer_or_404(request.user, order_id)
    auto_complete_order_if_due(order)
    order.refresh_from_db()

    if not order.can_customer_cancel:
        messages.error(request, 'Chỉ đơn chưa giao mới có thể hủy.')
        return redirect('order_history_detail', order_id=order.pk)

    try:
        with transaction.atomic():
            locked_order = Order.objects.select_for_update().get(pk=order.pk, user=request.user)
            if not locked_order.can_customer_cancel:
                raise ValueError('Đơn hàng không còn ở trạng thái có thể hủy.')
            locked_order.status = Order.STATUS_CANCELLED
            locked_order.cancelled_at = timezone.now()
            locked_order.save()
            transaction.on_commit(
                lambda cancelled_order=locked_order, current_request=request: send_order_cancelled_email(
                    cancelled_order,
                    request=current_request,
                )
            )
        messages.success(request, f'Đã hủy đơn hàng {order.order_code}.')
    except ValueError as exc:
        messages.error(request, str(exc))

    return redirect('order_history')


@login_required(login_url='/login/')
def confirm_order_received(request, order_id):
    if request.method != 'POST':
        raise Http404('Yêu cầu không hợp lệ.')

    order = get_order_for_customer_or_404(request.user, order_id)
    auto_complete_order_if_due(order)
    order.refresh_from_db()

    if order.status == Order.STATUS_COMPLETED:
        messages.info(request, f'Đơn hàng {order.order_code} đã ở trạng thái hoàn thành.')
        return redirect('order_history_detail', order_id=order.pk)

    if order.status != Order.STATUS_SHIPPING:
        messages.error(request, 'Chỉ đơn đang giao mới có thể xác nhận đã nhận hàng.')
        return redirect('order_history')

    if order.requires_payment_confirmation and order.payment_status != Order.PAYMENT_STATUS_PAID:
        messages.error(request, 'Đơn chuyển khoản/MoMo cần được xác nhận thanh toán trước khi hoàn thành.')
        return redirect('order_history_detail', order_id=order.pk)

    with transaction.atomic():
        locked_order = Order.objects.select_for_update().get(pk=order.pk, user=request.user)
        if locked_order.status != Order.STATUS_SHIPPING:
            messages.error(request, 'Đơn hàng không còn ở trạng thái đang giao.')
            return redirect('order_history')
        if locked_order.requires_payment_confirmation and locked_order.payment_status != Order.PAYMENT_STATUS_PAID:
            messages.error(request, 'Đơn chuyển khoản/MoMo cần được xác nhận thanh toán trước khi hoàn thành.')
            return redirect('order_history_detail', order_id=order.pk)
        complete_order_workflow(locked_order, completed_by_customer=True)

    messages.success(request, f'Đã xác nhận nhận hàng cho đơn {order.order_code}.')
    return redirect('order_history_detail', order_id=order.pk)


@login_required(login_url='/login/')
def create_or_update_return_request(request, order_id):
    order = get_order_for_customer_or_404(request.user, order_id)
    auto_complete_order_if_due(order)
    order.refresh_from_db()

    if order.status != Order.STATUS_COMPLETED:
        messages.error(request, 'Chỉ đơn hàng đã hoàn thành mới có thể gửi yêu cầu trả hàng / hoàn tiền.')
        return redirect('order_history_detail', order_id=order.pk)

    existing_request = getattr(order, 'return_request', None)
    if existing_request and existing_request.is_finalized and request.method == 'POST':
        messages.error(request, 'Yêu cầu này đã được xử lý xong và không thể chỉnh sửa thêm.')
        return redirect('order_history_detail', order_id=order.pk)

    if request.method == 'POST':
        is_existing_request = existing_request is not None
        form = ReturnRefundRequestForm(request.POST, request.FILES, instance=existing_request, order=order)
        if form.is_valid():
            with transaction.atomic():
                return_request = form.save(commit=False)
                return_request.order = order
                return_request.save()

                existing_evidences = list(return_request.evidences.all())
                for evidence in existing_evidences:
                    keep_value = str(request.POST.get(f'evidence_keep_{evidence.pk}', '1') or '1').strip().lower()
                    if keep_value not in {'1', 'true', 'yes', 'on'}:
                        evidence.delete()
                        continue

                    replacement = request.FILES.get(f'evidence_replace_{evidence.pk}')
                    if replacement:
                        evidence.image = replacement
                        evidence.save(update_fields=['image'])

                for uploaded_image in form.cleaned_data.get('proof_images') or []:
                    ReturnRefundEvidence.objects.create(request=return_request, image=uploaded_image)
                transaction.on_commit(
                    lambda saved_request=return_request, current_request=request, was_update=is_existing_request: send_return_request_received_email(
                        saved_request,
                        request=current_request,
                        is_update=was_update,
                    )
                )
            messages.success(request, 'Đã lưu yêu cầu trả hàng / hoàn tiền. Bộ phận xử lý sẽ kiểm tra và phản hồi sớm.')
            return redirect('order_history_detail', order_id=order.pk)
    else:
        form = ReturnRefundRequestForm(instance=existing_request, order=order, initial=build_return_request_initial(order, existing_request))

    context = {
        'order': order,
        'return_request': existing_request,
        'form': form,
        'evidence_images': existing_request.evidences.all() if existing_request else [],
    }
    return render(request, 'account/returns/request_form.html', context)


@login_required(login_url='/login/')
def order_invoice_view(request, order_id):
    order_queryset = Order.objects.select_related('pharmacy', 'user').prefetch_related('items__medicine')
    order = get_object_or_404(order_queryset, pk=order_id)
    auto_complete_order_if_due(order)
    order.refresh_from_db()

    if not request.user.is_superuser and not request.user.is_staff and order.user_id != request.user.id:
        raise Http404('Không tìm thấy hóa đơn phù hợp.')
    if request.user.is_staff and not request.user.is_superuser:
        ensure_object_is_within_admin_scope(request.user, 'order', order)

    payment_preview = build_order_payment_preview(order)
    context = {
        'order': order,
        'order_items': order.items.all(),
        'payment_preview': payment_preview,
        'invoice_printed_at': timezone.localtime(),
        'invoice_staff_display_name': get_invoice_staff_display_name(order),
    }
    return render(request, 'account/orders/invoice_detail.html', context)


def payment_preview_api(request):
    payment_method = (request.GET.get('payment_method') or Order.PAYMENT_COD).strip().lower()
    amount = request.GET.get('amount', '0').strip()
    pharmacy_id = (request.GET.get('pharmacy_id') or '').strip()
    reference = (request.GET.get('reference') or '').strip()

    try:
        amount_value = max(int(float(amount or 0)), 0)
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Số tiền thanh toán không hợp lệ.'}, status=400)

    pharmacy = None
    if pharmacy_id:
        pharmacy = Pharmacy.objects.filter(pk=pharmacy_id).first()

    if not reference:
        if pharmacy is not None:
            reference = f"TAM-{pharmacy.pk}-{amount_value}"
        else:
            reference = f"TAM-{amount_value}"

    preview_payload = build_payment_preview_payload(
        payment_method=payment_method,
        amount=amount_value,
        reference=reference,
        pharmacy=pharmacy,
    )
    return JsonResponse(preview_payload)


@csrf_exempt
@admin_panel_required
def editor_upload_api(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Phương thức không hợp lệ.'}, status=405)

    uploaded = request.FILES.get('upload') or request.FILES.get('file')
    if not uploaded:
        return JsonResponse({'error': 'Chưa nhận được ảnh tải lên.'}, status=400)

    content_type = str(getattr(uploaded, 'content_type', '') or '').lower()
    if not content_type.startswith('image/'):
        return JsonResponse({'error': 'Chỉ chấp nhận tệp ảnh.'}, status=400)

    original_name = Path(getattr(uploaded, 'name', 'image')).name or 'image'
    extension = Path(original_name).suffix or '.jpg'
    safe_extension = extension if len(extension) <= 10 else '.jpg'
    saved_name = default_storage.save(
        f"editor/{timezone.now().strftime('%Y%m%d')}/{uuid4().hex}{safe_extension}",
        uploaded,
    )
    uploaded_url = default_storage.url(saved_name)
    ckeditor_func_num = (request.GET.get("CKEditorFuncNum") or request.POST.get("CKEditorFuncNum") or "").strip()
    if ckeditor_func_num:
        response_script = (
            "<script>"
            f"window.parent.CKEDITOR.tools.callFunction({ckeditor_func_num}, '{escapejs(uploaded_url)}', '');"
            "</script>"
        )
        return HttpResponse(response_script)
    return JsonResponse({'uploaded': 1, 'fileName': original_name, 'url': uploaded_url})


def get_route_api(request):
    """
    API lấy tuyến đường giữa 2 điểm.
    """
    start_lat = request.GET.get('start_lat')
    start_lng = request.GET.get('start_lng')
    end_lat = request.GET.get('end_lat')
    end_lng = request.GET.get('end_lng')
    delivery_mode = request.GET.get('mode', 'motorbike')
    departure_time = get_default_departure_time_value(request.GET.get('departure_time', ''))

    if not all([start_lat, start_lng, end_lat, end_lng]):
        return JsonResponse({'error': 'Thiếu tọa độ.'}, status=400)

    route_result = delivery_service.estimate_route(
        start_lat=start_lat,
        start_lng=start_lng,
        end_lat=end_lat,
        end_lng=end_lng,
        delivery_mode=delivery_mode,
        departure_time_str=departure_time,
    )

    status_code = 200 if 'routes' in route_result else 400
    return JsonResponse(route_result, status=status_code)


def find_best_pharmacy_api(request):
    """
    API tìm chi nhánh phù hợp nhất cho vị trí giao hàng.
    """
    delivery_lat = request.GET.get('delivery_lat')
    delivery_lng = request.GET.get('delivery_lng')
    delivery_mode = request.GET.get('mode', 'motorbike')
    departure_time = get_default_departure_time_value(request.GET.get('departure_time', ''))
    selected_pharmacy_id = request.GET.get('pharmacy_id', '').strip()

    if not all([delivery_lat, delivery_lng]):
        return JsonResponse({'error': 'Thiếu tọa độ giao hàng.'}, status=400)

    available_pharmacies = get_available_pharmacies()
    cart = get_or_create_cart(request)

    if selected_pharmacy_id:
        selected_pharmacy = find_pharmacy_in_list(available_pharmacies, selected_pharmacy_id)
        if selected_pharmacy:
            route_result = delivery_service.estimate_route(
                start_lat=selected_pharmacy.lat,
                start_lng=selected_pharmacy.lng,
                end_lat=delivery_lat,
                end_lng=delivery_lng,
                delivery_mode=delivery_mode,
                departure_time_str=departure_time,
            )
            if 'routes' in route_result and route_result['routes']:
                best_delivery_result = {
                    'pharmacy': selected_pharmacy,
                    'route': route_result['routes'][0],
                    'mode': route_result.get('mode', delivery_mode),
                }
            else:
                best_delivery_result = {'error': 'Không thể tính được đường đi cho chi nhánh đã chọn.'}
        else:
            best_delivery_result = {'error': 'Không tìm thấy chi nhánh đã chọn.'}
    else:
        best_delivery_result = choose_checkout_pharmacy(
            cart,
            available_pharmacies,
            delivery_lat,
            delivery_lng,
            delivery_mode,
            departure_time,
        )

    if 'error' in best_delivery_result:
        return JsonResponse(best_delivery_result, status=400)

    selected_pharmacy = best_delivery_result['pharmacy']

    response_data = {
        'pharmacy': {
            'id': selected_pharmacy.id,
            'name': selected_pharmacy.name,
            'address': selected_pharmacy.address,
            'phone': selected_pharmacy.phone,
            'hours': selected_pharmacy.opening_hours,
            'image': get_entity_gallery_urls(
                selected_pharmacy,
                PHARMACY_FALLBACK_IMAGE,
            )[0],
            'lat': selected_pharmacy.lat,
            'lng': selected_pharmacy.lng,
        },
        'route': best_delivery_result['route'],
        'mode': best_delivery_result.get('mode', delivery_mode),
    }

    return JsonResponse(response_data)


def find_best_pharmacy_api_v2(request):
    delivery_lat = request.GET.get('delivery_lat')
    delivery_lng = request.GET.get('delivery_lng')
    delivery_mode = request.GET.get('mode', 'motorbike')
    departure_time = get_default_departure_time_value(request.GET.get('departure_time', ''))
    selected_pharmacy_id = request.GET.get('pharmacy_id', '').strip()

    if not all([delivery_lat, delivery_lng]):
        return JsonResponse({'error': 'Thiếu tọa độ giao hàng.'}, status=400)

    cart = get_or_create_cart(request)
    loyalty_context = build_customer_loyalty_context(request.user)
    available_pharmacies = list(
        get_available_pharmacies().prefetch_related(
            Prefetch(
                'medicines',
                queryset=Medicine.objects.order_by('-quantity', 'id'),
                to_attr='_inventory_candidates',
            )
        )
    )

    if not available_pharmacies:
        return JsonResponse({'error': 'Không có chi nhánh phù hợp cho giỏ hàng hiện tại.'}, status=400)

    if selected_pharmacy_id:
        selected_pharmacy = find_pharmacy_in_list(available_pharmacies, selected_pharmacy_id)
        if not selected_pharmacy:
            return JsonResponse({'error': 'Không tìm thấy chi nhánh đã chọn.'}, status=400)
        if cart.items.exists() and allocate_cart_to_pharmacy(cart, selected_pharmacy) is None:
            return JsonResponse(
                {'error': 'Chi nhánh đã chọn chưa đủ thuốc cho toàn bộ đơn hàng.'},
                status=400,
            )

        route_result = delivery_service.estimate_route(
            start_lat=selected_pharmacy.lat,
            start_lng=selected_pharmacy.lng,
            end_lat=delivery_lat,
            end_lng=delivery_lng,
            delivery_mode=delivery_mode,
            departure_time_str=departure_time,
        )
        if 'routes' not in route_result or not route_result['routes']:
            return JsonResponse({'error': 'Không thể tính được đường đi cho chi nhánh đã chọn.'}, status=400)

        best_delivery_result = {
            'pharmacy': selected_pharmacy,
            'route': route_result['routes'][0],
            'mode': route_result.get('mode', delivery_mode),
            'notice': '',
        }
    else:
        best_delivery_result = choose_checkout_pharmacy(
            cart,
            available_pharmacies,
            delivery_lat,
            delivery_lng,
            delivery_mode,
            departure_time,
        )

    if 'error' in best_delivery_result:
        return JsonResponse(best_delivery_result, status=400)

    selected_pharmacy = best_delivery_result['pharmacy']
    selected_allocations = best_delivery_result.get('allocations')
    if selected_allocations is None:
        selected_allocations = allocate_cart_to_pharmacy(cart, selected_pharmacy) if cart.items.exists() else []
    product_total_before_loyalty = sum(
        allocation['medicine'].current_price * allocation['quantity']
        for allocation in selected_allocations
    )
    loyalty_discount_total = calculate_loyalty_discount_total(
        product_total_before_loyalty,
        loyalty_context['discount_percent'],
    )
    product_total_value = max(product_total_before_loyalty - loyalty_discount_total, 0)
    response_data = {
        'pharmacy': {
            'id': selected_pharmacy.id,
            'name': selected_pharmacy.name,
            'address': selected_pharmacy.address,
            'phone': selected_pharmacy.phone,
            'hours': selected_pharmacy.opening_hours,
            'image': get_entity_gallery_urls(selected_pharmacy, PHARMACY_FALLBACK_IMAGE)[0],
            'lat': selected_pharmacy.lat,
            'lng': selected_pharmacy.lng,
        },
        'route': best_delivery_result['route'],
        'mode': best_delivery_result.get('mode', delivery_mode),
        'product_total_value': product_total_value,
        'product_total_before_loyalty': product_total_before_loyalty,
        'loyalty_discount_total': loyalty_discount_total,
        'loyalty_discount_percent': loyalty_context['discount_percent'],
        'loyalty_tier_label': loyalty_context['label'],
        'notice': best_delivery_result.get('notice', ''),
    }
    return JsonResponse(response_data)


def catalog_search_api(request):
    keyword = request.GET.get('q', '').strip()

    if not keyword:
        return JsonResponse(
            {
                'query': '',
                'products': [],
                'pharmacies': [],
            }
        )

    payload = build_catalog_search_payload(keyword)
    payload['product_total'] = len(payload['products'])
    payload['pharmacy_total'] = len(payload['pharmacies'])
    return JsonResponse(payload)


def nearby_pharmacies_api(request):
    lat = request.GET.get('lat')
    lng = request.GET.get('lng')
    radius = request.GET.get('radius', '0')
    keyword = request.GET.get('q', '').strip()

    if not lat or not lng:
        return JsonResponse({'error': 'Thiếu tọa độ để lọc bán kính.'}, status=400)

    try:
        nearby_items = build_nearby_pharmacy_payload(lat, lng, radius_km=radius, keyword=keyword)
    except ValueError as exc:
        return JsonResponse({'error': str(exc)}, status=400)

    try:
        radius_value = float(radius or 0)
    except (TypeError, ValueError):
        radius_value = 0

    return JsonResponse(
        {
            'lat': float(lat),
            'lng': float(lng),
            'radius_km': radius_value,
            'query': keyword,
            'pharmacies': nearby_items,
            'total': len(nearby_items),
        }
    )


def search_address_api(request):
    """
    API tìm kiếm địa chỉ cho các trang bản đồ.

    Nhận thêm `lat`, `lng` (tuỳ chọn) — tâm bản đồ user đang xem — để
    geocoder ưu tiên kết quả gần khu vực đó, tránh trả về địa chỉ trùng
    tên ở thành phố khác.
    """
    keyword = request.GET.get('q', '').strip()

    if not keyword:
        return JsonResponse({'error': 'Vui lòng nhập địa chỉ cần tìm.'}, status=400)

    bias_lat = request.GET.get('lat')
    bias_lng = request.GET.get('lng')

    try:
        results = search_address_candidates(keyword, bias_lat=bias_lat, bias_lng=bias_lng)
    except Exception:
        return JsonResponse({'error': 'Không kết nối được dịch vụ tìm địa chỉ.'}, status=502)

    return JsonResponse({
        'query': keyword,
        'results': results,
    })


def reverse_address_api(request):
    """
    API lấy địa chỉ từ tọa độ trên bản đồ.
    """
    lat = request.GET.get('lat')
    lng = request.GET.get('lng')

    if not lat or not lng:
        return JsonResponse({'error': 'Thiếu tọa độ để tìm địa chỉ.'}, status=400)

    try:
        result = reverse_geocode_coordinates(lat, lng)
    except Exception:
        return JsonResponse({'error': 'Không kết nối được dịch vụ tra cứu địa chỉ.'}, status=502)

    return JsonResponse({
        'address': result.get('display_name', ''),
        'lat': result.get('lat'),
        'lng': result.get('lng'),
    })


@login_required(login_url='/login/')
def save_profile_address_api(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Phương thức không hợp lệ.'}, status=405)

    if not is_customer_user(request.user):
        return JsonResponse({'error': 'Chỉ tài khoản khách hàng mới có thể lưu địa chỉ giao hàng.'}, status=403)

    profile = get_or_create_user_profile(request.user)
    address_text = (request.POST.get('address_text') or '').strip()
    lat = request.POST.get('lat')
    lng = request.POST.get('lng')

    profile.address_text = address_text

    if address_text and lat and lng:
        try:
            profile.address_lat = float(lat)
            profile.address_lng = float(lng)
        except (TypeError, ValueError):
            return JsonResponse({'error': 'Tọa độ không hợp lệ.'}, status=400)
    elif not address_text:
        profile.address_lat = None
        profile.address_lng = None

    profile.save()
    return JsonResponse({'saved': True})


def map_view(request):
    """
    Trang bản đồ hiển thị danh sách các nhà thuốc còn hàng.
    """
    pharmacies = get_available_pharmacies()
    pharmacy_payload = build_map_pharmacy_payload(pharmacies)
    saved_address = None
    selected_pharmacy_id = request.GET.get('pharmacy_id', '').strip() or request.GET.get('pharmacy', '').strip()
    if is_customer_user(request.user):
        saved_address = build_saved_address_payload(get_or_create_user_profile(request.user))
    return render(
        request,
        'locations/map.html',
        {
            'pharmacies': pharmacy_payload,
            'saved_address': saved_address,
            'selected_pharmacy_id': selected_pharmacy_id,
            'map_departure_time': get_default_departure_time_value(request.GET.get('departure_time', '')),
        },
    )


# =========================================================
# CUSTOM ADMIN PANEL
# =========================================================

LOW_STOCK_THRESHOLD = 10
ADMIN_PAGE_SIZE = 5


def staff_check(user):
    return user.is_authenticated and user.is_staff


def get_object_label(obj):
    if hasattr(obj, 'name') and obj.name:
        return obj.name
    if hasattr(obj, 'username') and obj.username:
        return obj.username
    if hasattr(obj, 'full_name') and obj.full_name:
        return obj.full_name
    return f'#{obj.pk}'


def build_search_query(keyword, fields):
    query = Q()
    for field in fields:
        query |= Q(**{f'{field}__icontains': keyword})
    return query


def format_money(value):
    try:
        return f"{int(value):,}".replace(',', '.') + ' đ'
    except (TypeError, ValueError):
        return '0 đ'


def render_badge(label, tone='secondary'):
    return format_html('<span class="admin-badge admin-badge-{}">{}</span>', tone, label)


def render_image_thumb(image_field, alt_text, empty_text='Chưa có ảnh'):
    if image_field:
        return format_html(
            '<div class="table-thumb-wrap"><img src="{}" alt="{}" class="table-thumb"></div>',
            image_field.url,
            alt_text,
        )
    return format_html(
        '<div class="table-thumb table-thumb-empty"><i class="fas fa-image"></i><span>{}</span></div>',
        empty_text,
    )


def render_stock_badge(quantity):
    if quantity <= 0:
        return render_badge('Hết hàng', 'danger')
    if quantity <= LOW_STOCK_THRESHOLD:
        return render_badge('Sắp hết', 'warning')
    return render_badge('Còn hàng', 'success')


def render_prescription_badge(required):
    return render_badge('Cần kê đơn', 'info') if required else render_badge('Không kê đơn', 'success')


def render_order_status_badge(status):
    mapping = {
        Order.STATUS_PENDING: ('Chờ xử lý', 'warning'),
        Order.STATUS_CONFIRMED: ('Đã xác nhận', 'primary'),
        Order.STATUS_PACKING: ('Đang chuẩn bị', 'secondary'),
        Order.STATUS_SHIPPING: ('Đang giao', 'info'),
        Order.STATUS_COMPLETED: ('Hoàn thành', 'success'),
        Order.STATUS_CANCELLED: ('Đã hủy', 'danger'),
        Order.STATUS_FAILED_DELIVERY: ('Giao không thành công', 'danger'),
    }
    label, tone = mapping.get(status, ('Không xác định', 'secondary'))
    return render_badge(label, tone)


def render_user_role_badge(user):
    if user.is_superuser:
        return render_badge('Quản trị viên cấp cao', 'danger')
    if user.is_staff:
        return render_badge('Nhân viên', 'info')
    return render_badge('Khách hàng', 'secondary')


def can_delete_object(request_user, model_key, obj=None):
    if not user_has_admin_permission(request_user, model_key, "delete"):
        return False
    if model_key == 'user' and obj and obj == request_user:
        return False
    return True


def can_access_admin_model(user, model_key):
    return user_has_admin_permission(user, model_key, "view")


def require_admin_model_access(request, model_key):
    if not can_access_admin_model(request.user, model_key):
        raise PermissionDenied('Tài khoản hiện tại không có quyền truy cập chức năng quản trị này.')

    if request.user.is_staff and not request.user.is_superuser and model_key in BRANCH_SCOPED_ADMIN_MODELS:
        managed_pharmacy = get_admin_scope_pharmacy(request.user)
        if managed_pharmacy is None:
            raise PermissionDenied('Tài khoản nhân viên chưa được gán chi nhánh quản lý. Vui lòng liên hệ quản trị viên.')
    return None


def build_review_insight_context(request):
    managed_pharmacy = get_admin_scope_pharmacy(request.user)

    medicine_queryset = Medicine.objects.select_related('pharmacy').all()
    pharmacy_queryset = Pharmacy.objects.all()
    medicine_reviews_qs = MedicineReview.objects.select_related('medicine__pharmacy', 'user')
    pharmacy_reviews_qs = PharmacyReview.objects.select_related('pharmacy', 'user')
    completed_items_qs = OrderItem.objects.filter(order__status=Order.STATUS_COMPLETED).select_related('medicine', 'order__pharmacy')

    if managed_pharmacy is not None:
        medicine_queryset = medicine_queryset.filter(pharmacy=managed_pharmacy)
        pharmacy_queryset = pharmacy_queryset.filter(pk=managed_pharmacy.pk)
        medicine_reviews_qs = medicine_reviews_qs.filter(medicine__pharmacy=managed_pharmacy)
        pharmacy_reviews_qs = pharmacy_reviews_qs.filter(pharmacy=managed_pharmacy)
        completed_items_qs = completed_items_qs.filter(order__pharmacy=managed_pharmacy)

    medicine_rating_map = {
        row['medicine_id']: {'average': float(row['average'] or 0), 'count': int(row['count'] or 0)}
        for row in medicine_reviews_qs.values('medicine_id').annotate(average=Avg('rating'), count=Count('id'))
    }
    pharmacy_rating_map = {
        row['pharmacy_id']: {'average': float(row['average'] or 0), 'count': int(row['count'] or 0)}
        for row in pharmacy_reviews_qs.values('pharmacy_id').annotate(average=Avg('rating'), count=Count('id'))
    }
    sold_quantity_map = {
        row['medicine_id']: int(row['total_quantity'] or 0)
        for row in completed_items_qs.values('medicine_id').annotate(total_quantity=Sum('quantity'))
        if row['medicine_id']
    }

    product_insights = []
    for medicine in medicine_queryset:
        rating_info = medicine_rating_map.get(medicine.pk, {'average': 0, 'count': 0})
        avg_rating = round(rating_info['average'], 1) if rating_info['count'] else 0
        review_count = rating_info['count']
        sold_quantity = sold_quantity_map.get(medicine.pk, 0)

        if review_count >= 2 and avg_rating >= 4.3:
            status = 'high'
            recommendation = 'Được đánh giá rất tốt. Ưu tiên nhập thêm để giữ doanh số và tránh thiếu hàng.'
        elif review_count >= 2 and avg_rating <= 2.5 and sold_quantity <= 3:
            status = 'low'
            recommendation = 'Đánh giá thấp và bán chậm. Nên giảm nhập, rà soát chất lượng; nếu kéo dài có thể ngừng nhập.'
        elif review_count >= 2 and avg_rating <= 3.0:
            status = 'watch'
            recommendation = 'Đánh giá chưa tốt. Cần kiểm tra phản hồi khách hàng trước khi quyết định nhập thêm.'
        elif sold_quantity <= 1 and medicine.quantity > LOW_STOCK_THRESHOLD:
            status = 'watch'
            recommendation = 'Bán chậm so với tồn kho hiện tại. Nên theo dõi trước khi tiếp tục nhập thêm.'
        else:
            status = 'neutral'
            recommendation = 'Mức đánh giá và sức bán đang ổn định. Tiếp tục theo dõi thêm theo chu kỳ bán hàng.'

        product_insights.append({
            'medicine': medicine,
            'avg_rating': avg_rating,
            'review_count': review_count,
            'sold_quantity': sold_quantity,
            'stock_quantity': medicine.quantity,
            'status': status,
            'recommendation': recommendation,
        })

    top_products = sorted(
        [item for item in product_insights if item['review_count'] > 0],
        key=lambda item: (-item['avg_rating'], -item['review_count'], -item['sold_quantity'], item['medicine'].name.casefold()),
    )[:6]
    watch_products = sorted(
        [item for item in product_insights if item['status'] in {'low', 'watch'}],
        key=lambda item: (item['avg_rating'] if item['review_count'] else 99, item['sold_quantity'], -item['stock_quantity'], item['medicine'].name.casefold()),
    )[:6]

    pharmacy_insights = []
    reward_staff = []
    for pharmacy in pharmacy_queryset.prefetch_related('managed_staff_profiles__user'):
        rating_info = pharmacy_rating_map.get(pharmacy.pk, {'average': 0, 'count': 0})
        avg_rating = round(rating_info['average'], 1) if rating_info['count'] else 0
        review_count = rating_info['count']
        staff_names = [get_user_display_name(profile.user) for profile in pharmacy.managed_staff_profiles.select_related('user')]

        if review_count >= 2 and avg_rating >= 4.3:
            status = 'high'
            recommendation = 'Chi nhánh được khách hàng đánh giá rất tốt. Đề xuất đưa nhân viên quản lý vào danh sách khen thưởng.'
            for staff_name in staff_names:
                reward_staff.append({'pharmacy_name': pharmacy.name, 'staff_name': staff_name, 'avg_rating': avg_rating})
        elif review_count >= 2 and avg_rating <= 2.8:
            status = 'low'
            recommendation = 'Chi nhánh bị đánh giá thấp. Đưa vào diện giám sát chặt chẽ, ưu tiên rà soát quy trình phục vụ.'
        else:
            status = 'neutral'
            recommendation = 'Mức đánh giá hiện tại ổn định. Tiếp tục theo dõi thêm phản hồi khách hàng.'

        pharmacy_insights.append({
            'pharmacy': pharmacy,
            'avg_rating': avg_rating,
            'review_count': review_count,
            'staff_names': staff_names,
            'status': status,
            'recommendation': recommendation,
        })

    top_pharmacies = sorted(
        [item for item in pharmacy_insights if item['review_count'] > 0],
        key=lambda item: (-item['avg_rating'], -item['review_count'], item['pharmacy'].name.casefold()),
    )[:6]
    watch_pharmacies = sorted(
        [item for item in pharmacy_insights if item['status'] == 'low'],
        key=lambda item: (item['avg_rating'], -item['review_count'], item['pharmacy'].name.casefold()),
    )[:6]

    return {
        'page_title': 'Phân tích đánh giá',
        'current_model': 'review_insights',
        'scope_label': managed_pharmacy.name if managed_pharmacy else 'Toàn hệ thống',
        'product_review_total': medicine_reviews_qs.count(),
        'avg_product_rating': medicine_reviews_qs.aggregate(avg=Avg('rating')).get('avg') or 0,
        'pharmacy_review_total': pharmacy_reviews_qs.count(),
        'avg_pharmacy_rating': pharmacy_reviews_qs.aggregate(avg=Avg('rating')).get('avg') or 0,
        'top_products': top_products,
        'watch_products': watch_products,
        'top_pharmacies': top_pharmacies,
        'watch_pharmacies': watch_pharmacies,
        'reward_staff': reward_staff,
        'recent_product_reviews': medicine_reviews_qs.order_by('-updated_at', '-id')[:4],
        'recent_pharmacy_reviews': pharmacy_reviews_qs.order_by('-updated_at', '-id')[:4],
    }


def build_admin_reports_context(request, *, paginate=True):
    reports_page_size = 6
    managed_pharmacy = get_admin_scope_pharmacy(request.user)
    orders_base = filter_queryset_by_admin_scope(Order.objects.select_related('pharmacy'), request.user, 'order')
    medicines_base = filter_queryset_by_admin_scope(Medicine.objects.select_related('pharmacy'), request.user, 'medicine')
    return_requests_base = filter_queryset_by_admin_scope(
        ReturnRefundRequest.objects.select_related('order__pharmacy'),
        request.user,
        'return_request',
    )

    today = timezone.localdate()

    def parse_date_value(raw_value):
        raw_value = (raw_value or '').strip()
        if not raw_value:
            return None
        try:
            return datetime.strptime(raw_value, '%Y-%m-%d').date()
        except ValueError:
            return None

    def get_previous_month_bounds(anchor_date):
        first_day_this_month = anchor_date.replace(day=1)
        last_day_previous_month = first_day_this_month - timedelta(days=1)
        first_day_previous_month = last_day_previous_month.replace(day=1)
        return first_day_previous_month, last_day_previous_month

    def get_month_bounds(anchor_date):
        month_start = anchor_date.replace(day=1)
        month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        return month_start, month_end

    def parse_month_value(raw_value):
        raw_value = (raw_value or '').strip()
        if not raw_value:
            return None
        try:
            month_value = datetime.strptime(raw_value, '%Y-%m').date().replace(day=1)
        except ValueError:
            return None
        if month_value > today.replace(day=1):
            return None
        return month_value

    def parse_year_value(raw_value):
        raw_value = (raw_value or '').strip()
        if not raw_value:
            return None
        try:
            year_value = int(raw_value)
        except (TypeError, ValueError):
            return None
        if year_value < 2000 or year_value > today.year:
            return None
        return year_value

    def resolve_range(range_key, start_raw=None, end_raw=None, month_raw=None, year_raw=None):
        if range_key == '7d':
            return today - timedelta(days=6), today
        if range_key == '30d':
            return today - timedelta(days=29), today
        if range_key == '90d':
            return today - timedelta(days=89), today
        if range_key == 'last_month':
            return get_previous_month_bounds(today)
        if range_key == 'month':
            return get_month_bounds(parse_month_value(month_raw) or today.replace(day=1))
        if range_key == 'this_year':
            return today.replace(month=1, day=1), today
        if range_key == 'year':
            year_value = parse_year_value(year_raw) or today.year
            return date(year_value, 1, 1), date(year_value, 12, 31)
        if range_key == 'custom':
            custom_start = parse_date_value(start_raw) or today.replace(day=1)
            custom_end = parse_date_value(end_raw) or today
            return custom_start, custom_end
        return today.replace(day=1), today

    range_key = (request.GET.get('range') or 'this_month').strip()
    if range_key not in {'7d', '30d', '90d', 'this_month', 'last_month', 'month', 'this_year', 'year', 'custom'}:
        range_key = 'this_month'

    selected_month_value = (request.GET.get('month') or today.strftime('%Y-%m')).strip()
    if parse_month_value(selected_month_value) is None:
        selected_month_value = today.strftime('%Y-%m')
    selected_year_value = (request.GET.get('year') or str(today.year)).strip()
    if parse_year_value(selected_year_value) is None:
        selected_year_value = str(today.year)

    start_date, end_date = resolve_range(
        range_key,
        request.GET.get('start_date'),
        request.GET.get('end_date'),
        request.GET.get('month'),
        request.GET.get('year'),
    )
    if end_date > today:
        end_date = today
    if start_date > today:
        start_date = today
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    range_span_days = max((end_date - start_date).days + 1, 1)
    previous_end_date = start_date - timedelta(days=1)
    previous_start_date = previous_end_date - timedelta(days=range_span_days - 1)

    pharmacy_options = []
    selected_pharmacy = managed_pharmacy
    selected_pharmacy_id = ''
    if managed_pharmacy is None:
        pharmacy_options = [
            {'value': '', 'label': 'Toàn hệ thống'}
        ] + [
            {'value': str(pharmacy.pk), 'label': pharmacy.name}
            for pharmacy in Pharmacy.objects.order_by('name', 'id')
        ]
        selected_pharmacy_id = (request.GET.get('pharmacy') or '').strip()
        if selected_pharmacy_id.isdigit():
            selected_pharmacy = Pharmacy.objects.filter(pk=int(selected_pharmacy_id)).first()
            if selected_pharmacy is not None:
                orders_base = orders_base.filter(pharmacy=selected_pharmacy)
                medicines_base = medicines_base.filter(pharmacy=selected_pharmacy)
                return_requests_base = return_requests_base.filter(order__pharmacy=selected_pharmacy)
            else:
                selected_pharmacy_id = ''
        else:
            selected_pharmacy = None

    payment_choices = [(method, label) for method, label in Order.PAYMENT_METHOD_CHOICES]
    selected_payment_method = (request.GET.get('payment_method') or '').strip()
    valid_payment_methods = {method for method, _ in payment_choices}
    if selected_payment_method not in valid_payment_methods:
        selected_payment_method = ''
    if selected_payment_method:
        orders_base = orders_base.filter(payment_method=selected_payment_method)

    group_by_param = (request.GET.get('group_by') or 'auto').strip()
    if group_by_param not in {'auto', 'day', 'week', 'month'}:
        group_by_param = 'auto'
    if group_by_param == 'auto':
        if range_span_days <= 31:
            group_by = 'day'
        elif range_span_days <= 120:
            group_by = 'week'
        else:
            group_by = 'month'
    else:
        group_by = group_by_param

    def resolve_order_report_date(order):
        reference_dt = order.completed_at or order.created_at
        if not reference_dt:
            return today
        if timezone.is_aware(reference_dt):
            return timezone.localtime(reference_dt).date()
        return reference_dt.date()

    def resolve_request_report_date(return_request):
        reference_dt = return_request.created_at
        if not reference_dt:
            return today
        if timezone.is_aware(reference_dt):
            return timezone.localtime(reference_dt).date()
        return reference_dt.date()

    def build_delta_meta(current_value, previous_value):
        if current_value == previous_value:
            if current_value == 0:
                return {'text': 'Chưa phát sinh dữ liệu trong 2 kỳ', 'tone': 'neutral'}
            return {'text': 'Không đổi so với kỳ trước', 'tone': 'neutral'}
        if previous_value == 0:
            if current_value > 0:
                return {'text': 'Phát sinh mới so với kỳ trước', 'tone': 'up'}
            return {'text': 'Giảm về 0 so với kỳ trước', 'tone': 'down'}
        diff_value = current_value - previous_value
        percent_value = round(abs(diff_value) * 100 / previous_value)
        if diff_value > 0:
            return {'text': f'Tăng {percent_value}% so với kỳ trước', 'tone': 'up'}
        return {'text': f'Giảm {percent_value}% so với kỳ trước', 'tone': 'down'}

    def paginate_report_items(items, page_param):
        paginator = Paginator(items, reports_page_size)
        page_obj = paginator.get_page(request.GET.get(page_param))
        query_params = request.GET.copy()
        if page_param in query_params:
            del query_params[page_param]
        return page_obj, query_params.urlencode()

    def start_of_week(day_value):
        return day_value - timedelta(days=day_value.weekday())

    def start_of_month(day_value):
        return day_value.replace(day=1)

    def next_month(day_value):
        return (day_value.replace(day=28) + timedelta(days=4)).replace(day=1)

    def build_bucket_sequence(start_value, end_value, mode):
        sequence = []
        if mode == 'day':
            cursor = start_value
            while cursor <= end_value:
                sequence.append(cursor)
                cursor += timedelta(days=1)
            return sequence
        if mode == 'week':
            cursor = start_of_week(start_value)
            while cursor <= end_value:
                sequence.append(cursor)
                cursor += timedelta(days=7)
            return sequence
        cursor = start_of_month(start_value)
        while cursor <= end_value:
            sequence.append(cursor)
            cursor = next_month(cursor)
        return sequence

    def resolve_bucket_start(day_value, mode):
        if mode == 'day':
            return day_value
        if mode == 'week':
            return start_of_week(day_value)
        return start_of_month(day_value)

    def format_bucket_label(bucket_start, mode):
        if mode == 'day':
            return bucket_start.strftime('%d/%m')
        if mode == 'week':
            return bucket_start.strftime('%d/%m')
        return bucket_start.strftime('%m/%Y')

    def format_bucket_range(bucket_start, mode):
        if mode == 'day':
            return bucket_start.strftime('%d/%m/%Y')
        if mode == 'week':
            bucket_end = min(bucket_start + timedelta(days=6), end_date)
            return f"{bucket_start.strftime('%d/%m/%Y')} - {bucket_end.strftime('%d/%m/%Y')}"
        month_end = next_month(bucket_start) - timedelta(days=1)
        bucket_end = min(month_end, end_date)
        return f"{bucket_start.strftime('%m/%Y')} ({bucket_start.strftime('%d/%m')} - {bucket_end.strftime('%d/%m')})"

    all_orders = list(orders_base.order_by('-completed_at', '-created_at', '-id'))
    filtered_orders = []
    previous_period_orders = []
    for order in all_orders:
        report_date = resolve_order_report_date(order)
        order._report_date = report_date
        if start_date <= report_date <= end_date:
            filtered_orders.append(order)
        elif previous_start_date <= report_date <= previous_end_date:
            previous_period_orders.append(order)

    filtered_completed_orders = [order for order in filtered_orders if order.status == Order.STATUS_COMPLETED]
    previous_completed_orders = [order for order in previous_period_orders if order.status == Order.STATUS_COMPLETED]

    total_orders = len(filtered_orders)
    completed_order_count = len(filtered_completed_orders)
    status_counts = {
        status_key: sum(1 for order in filtered_orders if order.status == status_key)
        for status_key, _status_label in Order.STATUS_CHOICES
    }
    pending_order_count = status_counts.get(Order.STATUS_PENDING, 0)
    shipping_order_count = status_counts.get(Order.STATUS_SHIPPING, 0)
    cancelled_order_count = status_counts.get(Order.STATUS_CANCELLED, 0)

    total_completed_revenue = sum(int(order.final_total_price or 0) for order in filtered_completed_orders)
    previous_completed_revenue = sum(int(order.final_total_price or 0) for order in previous_completed_orders)
    total_product_revenue = sum(int(order.total_product_price or 0) for order in filtered_completed_orders)
    previous_product_revenue = sum(int(order.total_product_price or 0) for order in previous_completed_orders)
    total_loyalty_discount = sum(int(order.customer_tier_discount_total or 0) for order in filtered_completed_orders)
    previous_loyalty_discount = sum(int(order.customer_tier_discount_total or 0) for order in previous_completed_orders)

    def resolve_order_gross_product_total(order):
        snapshot_value = getattr(order, 'product_subtotal_before_tier_discount', None)
        if snapshot_value is not None:
            return int(snapshot_value or 0)
        return int(order.total_product_price or 0) + int(order.customer_tier_discount_total or 0)

    gross_product_revenue = sum(resolve_order_gross_product_total(order) for order in filtered_completed_orders)
    shipping_fee_revenue = sum(max(int(order.final_total_price or 0) - int(order.total_product_price or 0), 0) for order in filtered_completed_orders)
    previous_shipping_fee_revenue = sum(
        max(int(order.final_total_price or 0) - int(order.total_product_price or 0), 0)
        for order in previous_completed_orders
    )
    average_order_value = round(total_completed_revenue / completed_order_count) if completed_order_count else 0
    previous_average_order_value = round(previous_completed_revenue / len(previous_completed_orders)) if previous_completed_orders else 0
    completion_rate = round(completed_order_count * 100 / total_orders) if total_orders else 0
    previous_total_orders = len(previous_period_orders)
    previous_completion_rate = round(len(previous_completed_orders) * 100 / previous_total_orders) if previous_total_orders else 0

    timeline_buckets = build_bucket_sequence(start_date, end_date, group_by)
    timeline_map = {
        bucket_start: {
            'label': format_bucket_label(bucket_start, group_by),
            'range_label': format_bucket_range(bucket_start, group_by),
            'value': 0,
            'order_count': 0,
        }
        for bucket_start in timeline_buckets
    }
    for order in filtered_completed_orders:
        bucket_key = resolve_bucket_start(order._report_date, group_by)
        if bucket_key in timeline_map:
            timeline_map[bucket_key]['value'] += int(order.final_total_price or 0)
            timeline_map[bucket_key]['order_count'] += 1
    timeline_chart = list(timeline_map.values())
    max_timeline_value = max((item['value'] for item in timeline_chart), default=0)
    for item in timeline_chart:
        item['height_percent'] = max(10, round(item['value'] * 100 / max_timeline_value)) if item['value'] and max_timeline_value else 0
    timeline_chart_min_width = max(720, len(timeline_chart) * 86)

    payment_method_map = {
        method: {
            'key': method,
            'label': label,
            'order_count': 0,
            'revenue': 0,
        }
        for method, label in Order.PAYMENT_METHOD_CHOICES
    }
    for order in filtered_completed_orders:
        method_key = order.payment_method or Order.PAYMENT_COD
        if method_key not in payment_method_map:
            payment_method_map[method_key] = {
                'key': method_key,
                'label': method_key,
                'order_count': 0,
                'revenue': 0,
            }
        payment_method_map[method_key]['order_count'] += 1
        payment_method_map[method_key]['revenue'] += int(order.final_total_price or 0)
    payment_method_breakdown = [
        item for item in sorted(
            payment_method_map.values(),
            key=lambda current_item: (-current_item['revenue'], -current_item['order_count'], current_item['label'])
        )
        if item['order_count'] > 0
    ]
    max_payment_revenue = max((item['revenue'] for item in payment_method_breakdown), default=0)
    for item in payment_method_breakdown:
        item['share_percent'] = round(item['revenue'] * 100 / total_completed_revenue) if total_completed_revenue else 0
        item['bar_percent'] = max(12, round(item['revenue'] * 100 / max_payment_revenue)) if item['revenue'] and max_payment_revenue else 0

    status_meta = {
        Order.STATUS_PENDING: {'label': 'Chờ xử lý', 'tone': 'warning', 'icon': 'fas fa-hourglass-half'},
        Order.STATUS_CONFIRMED: {'label': 'Đã xác nhận', 'tone': 'primary', 'icon': 'fas fa-clipboard-check'},
        Order.STATUS_PACKING: {'label': 'Đang chuẩn bị', 'tone': 'secondary', 'icon': 'fas fa-box-open'},
        Order.STATUS_SHIPPING: {'label': 'Đang giao', 'tone': 'info', 'icon': 'fas fa-shipping-fast'},
        Order.STATUS_COMPLETED: {'label': 'Hoàn thành', 'tone': 'success', 'icon': 'fas fa-check-circle'},
        Order.STATUS_CANCELLED: {'label': 'Đã hủy', 'tone': 'danger', 'icon': 'fas fa-times-circle'},
        Order.STATUS_FAILED_DELIVERY: {'label': 'Giao không thành công', 'tone': 'danger', 'icon': 'fas fa-exclamation-circle'},
    }
    max_status_count = max(status_counts.values(), default=0)
    status_breakdown = []
    for status_key, _status_label in Order.STATUS_CHOICES:
        meta = status_meta.get(status_key, {'label': _status_label, 'tone': 'secondary', 'icon': 'fas fa-layer-group'})
        count_value = int(status_counts.get(status_key, 0))
        status_breakdown.append({
            'key': status_key,
            'label': meta['label'],
            'icon': meta['icon'],
            'tone': meta['tone'],
            'count': count_value,
            'share_percent': round(count_value * 100 / total_orders) if total_orders else 0,
            'bar_percent': max(12, round(count_value * 100 / max_status_count)) if count_value and max_status_count else 0,
        })

    completed_order_ids = [order.pk for order in filtered_completed_orders if order.pk]
    completed_items_qs = OrderItem.objects.select_related('medicine', 'order__pharmacy').none()
    if completed_order_ids:
        completed_items_qs = OrderItem.objects.filter(order_id__in=completed_order_ids).select_related('medicine', 'order__pharmacy')

    product_map = {}
    for item in completed_items_qs:
        product_key = item.medicine_id or f"archived-{(item.medicine_name or '').casefold()}"
        if product_key not in product_map:
            product_map[product_key] = {
                'name': item.medicine_name,
                'quantity': 0,
                'revenue': 0,
                'order_count': 0,
                'unit': item.medicine.unit if item.medicine else '',
                'current_stock': item.medicine.quantity if item.medicine else 0,
            }
        product_map[product_key]['quantity'] += int(item.quantity or 0)
        product_map[product_key]['revenue'] += int(item.price or 0) * int(item.quantity or 0)
        product_map[product_key]['order_count'] += 1
        if item.medicine:
            product_map[product_key]['current_stock'] = item.medicine.quantity
            product_map[product_key]['unit'] = item.medicine.unit or product_map[product_key]['unit']
    top_products = sorted(
        product_map.values(),
        key=lambda current_item: (
            -current_item['quantity'],
            -current_item['revenue'],
            (current_item['name'] or 'Sản phẩm không xác định').casefold(),
        )
    )

    branch_map = {}
    for order in filtered_completed_orders:
        branch_label = order.pharmacy.name if order.pharmacy else 'Chưa gán chi nhánh'
        if branch_label not in branch_map:
            branch_map[branch_label] = {
                'label': branch_label,
                'order_count': 0,
                'revenue': 0,
                'product_revenue': 0,
            }
        branch_map[branch_label]['order_count'] += 1
        branch_map[branch_label]['revenue'] += int(order.final_total_price or 0)
        branch_map[branch_label]['product_revenue'] += int(order.total_product_price or 0)
    branch_performance = sorted(
        branch_map.values(),
        key=lambda current_item: (-current_item['revenue'], -current_item['order_count'], current_item['label'].casefold())
    )
    max_branch_revenue = max((item['revenue'] for item in branch_performance), default=0)
    for item in branch_performance:
        item['average_order_value'] = round(item['revenue'] / item['order_count']) if item['order_count'] else 0
        item['share_percent'] = round(item['revenue'] * 100 / total_completed_revenue) if total_completed_revenue else 0
        item['bar_percent'] = max(12, round(item['revenue'] * 100 / max_branch_revenue)) if item['revenue'] and max_branch_revenue else 0

    low_stock_queryset = medicines_base.filter(quantity__gt=0, quantity__lte=LOW_STOCK_THRESHOLD).order_by('quantity', 'name')
    out_of_stock_queryset = medicines_base.filter(quantity__lte=0).order_by('name', 'id')
    out_of_stock_count = out_of_stock_queryset.count()
    expiring_soon_qs = get_expiring_soon_medicines_queryset(medicines_base)
    expiring_soon_queryset = expiring_soon_qs.order_by('expiry_date', 'name')
    low_stock_count = low_stock_queryset.count()
    expiring_soon_count = expiring_soon_qs.count()

    return_requests = list(return_requests_base.order_by('-created_at', '-id'))
    filtered_return_requests = [
        item for item in return_requests
        if start_date <= resolve_request_report_date(item) <= end_date
    ]
    return_status_meta = {
        ReturnRefundRequest.STATUS_PROCESSING: {'label': 'Đang xử lý', 'tone': 'warning'},
        ReturnRefundRequest.STATUS_APPROVED: {'label': 'Đã duyệt', 'tone': 'success'},
        ReturnRefundRequest.STATUS_REJECTED: {'label': 'Từ chối', 'tone': 'danger'},
    }
    return_status_breakdown = []
    return_total = len(filtered_return_requests)
    max_return_count = 0
    return_status_counts = {}
    for status_key, _status_label in ReturnRefundRequest.STATUS_CHOICES:
        count_value = sum(1 for item in filtered_return_requests if item.status == status_key)
        return_status_counts[status_key] = count_value
        max_return_count = max(max_return_count, count_value)
    for status_key, _status_label in ReturnRefundRequest.STATUS_CHOICES:
        meta = return_status_meta.get(status_key, {'label': _status_label, 'tone': 'secondary'})
        count_value = return_status_counts.get(status_key, 0)
        return_status_breakdown.append({
            'key': status_key,
            'label': meta['label'],
            'tone': meta['tone'],
            'count': count_value,
            'share_percent': round(count_value * 100 / return_total) if return_total else 0,
            'bar_percent': max(12, round(count_value * 100 / max_return_count)) if count_value and max_return_count else 0,
        })

    dominant_payment = payment_method_breakdown[0] if payment_method_breakdown else None
    leading_product = top_products[0] if top_products else None
    leading_branch = branch_performance[0] if branch_performance else None
    strongest_bucket = max(timeline_chart, key=lambda item: (item['value'], item['order_count']), default=None)

    insight_cards = [
        {
            'label': 'Đỉnh doanh thu trong kỳ',
            'value': strongest_bucket['value'] if strongest_bucket else 0,
            'format': 'currency',
            'note': (
                f"{strongest_bucket['range_label']} • {strongest_bucket['order_count']} đơn hoàn thành"
                if strongest_bucket and strongest_bucket['value'] > 0
                else 'Chưa có mốc doanh thu nổi bật trong phạm vi đang lọc.'
            ),
        },
        {
            'label': 'Phương thức thanh toán chủ đạo',
            'value': dominant_payment['revenue'] if dominant_payment else 0,
            'format': 'currency',
            'note': (
                f"{dominant_payment['label']} • {dominant_payment['share_percent']}% doanh thu hoàn thành"
                if dominant_payment
                else 'Chưa có đơn hoàn thành để xác định phương thức chiếm ưu thế.'
            ),
        },
        {
            'label': 'Sản phẩm dẫn đầu',
            'value': leading_product['quantity'] if leading_product else 0,
            'format': 'number',
            'note': (
                f"{leading_product['name']} • {leading_product['revenue']:,} đ doanh thu gộp".replace(',', '.')
                if leading_product
                else 'Chưa có sản phẩm nào phát sinh bán trong phạm vi báo cáo.'
            ),
        },
        {
            'label': 'Chi nhánh nổi bật',
            'value': leading_branch['revenue'] if leading_branch else 0,
            'format': 'currency',
            'note': (
                f"{leading_branch['label']} • {leading_branch['order_count']} đơn hoàn thành"
                if leading_branch
                else 'Chưa có chi nhánh nào phát sinh doanh thu trong phạm vi hiện tại.'
            ),
        },
    ]

    summary_cards = [
        {
            'label': 'Doanh thu hoàn thành',
            'value': total_completed_revenue,
            'helper': 'Tổng thanh toán của các đơn đã hoàn thành trong kỳ lọc.',
            'tone': 'primary',
            'format': 'currency',
            'delta': build_delta_meta(total_completed_revenue, previous_completed_revenue),
        },
        {
            'label': 'Tiền hàng thuần',
            'value': total_product_revenue,
            'helper': 'Đã trừ ưu đãi hạng khách hàng, chưa cộng phí giao hàng.',
            'tone': 'success',
            'format': 'currency',
            'delta': build_delta_meta(total_product_revenue, previous_product_revenue),
        },
        {
            'label': 'Ưu đãi hạng khách hàng',
            'value': total_loyalty_discount,
            'helper': 'Tổng mức giảm trên phần tiền hàng của các đơn hoàn thành.',
            'tone': 'warning',
            'format': 'currency',
            'delta': build_delta_meta(total_loyalty_discount, previous_loyalty_discount),
        },
        {
            'label': 'Phí giao hàng thu được',
            'value': shipping_fee_revenue,
            'helper': 'Phần chênh giữa tổng thanh toán và tiền hàng thuần.',
            'tone': 'info',
            'format': 'currency',
            'delta': build_delta_meta(shipping_fee_revenue, previous_shipping_fee_revenue),
        },
        {
            'label': 'Đơn phát sinh',
            'value': total_orders,
            'helper': 'Tổng số đơn có ngày ghi nhận nằm trong kỳ báo cáo.',
            'tone': 'secondary',
            'format': 'number',
            'delta': build_delta_meta(total_orders, previous_total_orders),
        },
        {
            'label': 'Đơn hoàn thành',
            'value': completed_order_count,
            'helper': 'Dùng để tính các chỉ số doanh thu và bán hàng.',
            'tone': 'primary',
            'format': 'number',
            'delta': build_delta_meta(completed_order_count, len(previous_completed_orders)),
        },
        {
            'label': 'Tỷ lệ hoàn thành',
            'value': completion_rate,
            'helper': 'Tỷ trọng đơn đã hoàn tất trên tổng đơn phát sinh.',
            'tone': 'success',
            'format': 'percent',
            'delta': build_delta_meta(completion_rate, previous_completion_rate),
        },
        {
            'label': 'Giá trị đơn trung bình',
            'value': average_order_value,
            'helper': 'Trung bình trên tổng thanh toán của các đơn hoàn thành.',
            'tone': 'warning',
            'format': 'currency',
            'delta': build_delta_meta(average_order_value, previous_average_order_value),
        },
    ]

    if paginate:
        top_products_page_obj, top_products_query_string = paginate_report_items(top_products, 'top_page')
        branch_performance_page_obj, branch_performance_query_string = paginate_report_items(branch_performance, 'branch_page')
        low_stock_page_obj, low_stock_query_string = paginate_report_items(low_stock_queryset, 'stock_page')
        expiring_soon_page_obj, expiring_soon_query_string = paginate_report_items(expiring_soon_queryset, 'expiry_page')

        top_products = list(top_products_page_obj.object_list)
        branch_performance = list(branch_performance_page_obj.object_list)
        low_stock_medicines = list(low_stock_page_obj.object_list)
        expiring_soon_medicines = list(expiring_soon_page_obj.object_list)
    else:
        top_products_page_obj = None
        branch_performance_page_obj = None
        low_stock_page_obj = None
        expiring_soon_page_obj = None
        top_products_query_string = ''
        branch_performance_query_string = ''
        low_stock_query_string = ''
        expiring_soon_query_string = ''
        low_stock_medicines = list(low_stock_queryset)
        expiring_soon_medicines = list(expiring_soon_queryset)
    out_of_stock_medicines = list(out_of_stock_queryset) if not paginate else []

    filter_chips = [
        {'icon': 'fas fa-calendar-alt', 'label': f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"},
        {'icon': 'fas fa-layer-group', 'label': selected_pharmacy.name if selected_pharmacy else 'Toàn hệ thống'},
        {'icon': 'fas fa-wallet', 'label': dict(payment_choices).get(selected_payment_method, 'Tất cả phương thức')},
        {'icon': 'fas fa-chart-bar', 'label': {'day': 'Nhóm theo ngày', 'week': 'Nhóm theo tuần', 'month': 'Nhóm theo tháng'}.get(group_by, 'Tự động')},
    ]

    period_label_map = {
        '7d': '7 ngày gần nhất',
        '30d': '30 ngày gần nhất',
        '90d': '90 ngày gần nhất',
        'this_month': 'Tháng hiện tại',
        'last_month': 'Tháng trước',
        'month': 'Theo tháng',
        'this_year': 'Năm hiện tại',
        'year': 'Theo năm',
        'custom': 'Tự chọn',
    }

    filter_form = {
        'range_key': range_key,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'selected_month': selected_month_value,
        'selected_year': selected_year_value,
        'selected_pharmacy_id': selected_pharmacy_id,
        'selected_payment_method': selected_payment_method,
        'group_by': group_by_param,
        'range_options': [
            ('7d', '7 ngày gần nhất'),
            ('30d', '30 ngày gần nhất'),
            ('90d', '90 ngày gần nhất'),
            ('this_month', 'Tháng hiện tại'),
            ('last_month', 'Tháng trước'),
            ('month', 'Theo tháng'),
            ('this_year', 'Năm hiện tại'),
            ('year', 'Theo năm'),
            ('custom', 'Tự chọn'),
        ],
        'group_by_options': [
            ('auto', 'Tự động'),
            ('day', 'Theo ngày'),
            ('week', 'Theo tuần'),
            ('month', 'Theo tháng'),
        ],
        'payment_options': [('', 'Tất cả phương thức')] + payment_choices,
        'pharmacy_options': pharmacy_options,
    }

    return {
        'page_title': 'Báo cáo & thống kê',
        'current_model': 'reports',
        'managed_pharmacy': managed_pharmacy,
        'selected_pharmacy': selected_pharmacy,
        'scope_label': selected_pharmacy.name if selected_pharmacy else 'Toàn hệ thống',
        'summary_cards': summary_cards,
        'timeline_chart': timeline_chart,
        'timeline_chart_min_width': timeline_chart_min_width,
        'timeline_group_label': {'day': 'ngày', 'week': 'tuần', 'month': 'tháng'}.get(group_by, 'ngày'),
        'payment_method_breakdown': payment_method_breakdown,
        'status_breakdown': status_breakdown,
        'top_products': top_products,
        'top_products_page_obj': top_products_page_obj,
        'top_products_query_string': top_products_query_string,
        'branch_performance': branch_performance,
        'branch_performance_page_obj': branch_performance_page_obj,
        'branch_performance_query_string': branch_performance_query_string,
        'low_stock_medicines': low_stock_medicines,
        'low_stock_page_obj': low_stock_page_obj,
        'low_stock_query_string': low_stock_query_string,
        'expiring_soon_medicines': expiring_soon_medicines,
        'expiring_soon_page_obj': expiring_soon_page_obj,
        'expiring_soon_query_string': expiring_soon_query_string,
        'out_of_stock_medicines': out_of_stock_medicines,
        'low_stock_count': low_stock_count,
        'expiring_soon_count': expiring_soon_count,
        'out_of_stock_count': out_of_stock_count,
        'return_status_breakdown': return_status_breakdown,
        'return_total': return_total,
        'filter_form': filter_form,
        'filter_chips': filter_chips,
        'range_label': period_label_map.get(range_key, 'Tháng hiện tại'),
        'previous_period_label': f"{previous_start_date.strftime('%d/%m/%Y')} - {previous_end_date.strftime('%d/%m/%Y')}",
        'report_period_label': f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
        'total_completed_revenue': total_completed_revenue,
        'total_product_revenue': total_product_revenue,
        'gross_product_revenue': gross_product_revenue,
        'total_loyalty_discount': total_loyalty_discount,
        'completed_order_count': completed_order_count,
        'total_orders': total_orders,
        'pending_order_count': pending_order_count,
        'shipping_order_count': shipping_order_count,
        'cancelled_order_count': cancelled_order_count,
        'average_order_value': average_order_value,
        'completion_rate': completion_rate,
        'insight_cards': insight_cards,
        'filtered_orders': filtered_orders,
        'filtered_completed_orders': filtered_completed_orders,
        'filtered_return_requests': filtered_return_requests,
        'report_start_date': start_date,
        'report_end_date': end_date,
        'group_by': group_by,
    }


ADMIN_REPORT_EXCEL_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _build_admin_reports_workbook_legacy(report_context):
    Workbook = get_excel_workbook_builder()
    try:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Máy đang thiếu thư viện openpyxl nên chưa thể định dạng file Excel .xlsx."
        ) from exc

    workbook = Workbook()
    title_fill = PatternFill('solid', fgColor='1F64E0')
    title_font = Font(color='FFFFFF', bold=True, size=14)
    header_fill = PatternFill('solid', fgColor='EAF2FF')
    header_font = Font(color='163057', bold=True)
    thin_side = Side(style='thin', color='D7E2F0')
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    currency_keywords = ('doanh thu', 'tiền', 'phí', 'tổng', 'giảm', 'giá', 'thanh toán')

    def to_excel_datetime(value):
        if value is None:
            return ''
        if isinstance(value, datetime):
            if timezone.is_aware(value):
                return timezone.localtime(value).replace(tzinfo=None)
            return value
        return value

    def display_date(value):
        value = to_excel_datetime(value)
        return value

    def bool_label(value):
        return 'Có' if value else 'Không'

    def prepare_cell_value(value):
        if isinstance(value, Decimal):
            return float(value)
        return to_excel_datetime(value)

    def autosize_columns(worksheet):
        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)

    def style_table(worksheet, header_row, headers):
        worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)
        last_column = max(len(headers), 1)
        last_row = max(worksheet.max_row, header_row)
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(last_column)}{last_row}"

        for cell in worksheet[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border

        for row in worksheet.iter_rows(min_row=header_row + 1, max_row=worksheet.max_row, max_col=last_column):
            for cell in row:
                cell.border = cell_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if isinstance(cell.value, datetime):
                    cell.number_format = 'dd/mm/yyyy hh:mm'
                elif isinstance(cell.value, date):
                    cell.number_format = 'dd/mm/yyyy'
                else:
                    header = str(headers[cell.column - 1]).casefold() if cell.column <= len(headers) else ''
                    if isinstance(cell.value, (int, float)) and any(keyword in header for keyword in currency_keywords):
                        cell.number_format = '#,##0" đ"'

        autosize_columns(worksheet)

    def write_table(title, headers, rows, *, worksheet=None):
        worksheet = worksheet or workbook.create_sheet(title=title[:31])
        worksheet.append([title])
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        worksheet.row_dimensions[1].height = 24
        worksheet.append(headers)
        for row in rows:
            worksheet.append([prepare_cell_value(value) for value in row])
        style_table(worksheet, 2, headers)
        return worksheet

    summary_sheet = workbook.active
    summary_sheet.title = 'Tong quan'

    summary_rows = [
        ('Kỳ báo cáo', report_context['range_label']),
        ('Thời gian', report_context['report_period_label']),
        ('Phạm vi', report_context['scope_label']),
        ('So sánh kỳ trước', report_context['previous_period_label']),
    ]
    for card in report_context['summary_cards']:
        summary_rows.append((card['label'], card['value'], card.get('helper', ''), card.get('delta', {}).get('text', '')))
    write_table('Tổng quan báo cáo', ['Chỉ tiêu', 'Giá trị', 'Ghi chú', 'So sánh'], summary_rows, worksheet=summary_sheet)

    orders = report_context.get('filtered_orders') or []
    order_rows = []
    for order in orders:
        order_rows.append([
            order.order_code,
            display_date(order.created_at),
            getattr(order, '_report_date', ''),
            order.full_name,
            order.phone,
            order.pharmacy.name if order.pharmacy else '',
            order.get_status_display(),
            order.get_payment_method_display(),
            order.get_payment_status_display(),
            int(order.total_product_price or 0),
            int(order.customer_tier_discount_total or 0),
            int(order.shipping_fee or 0),
            int(order.final_total_price or 0),
            bool_label(order.invoice_requested),
            order.resolved_invoice_code,
            order.resolved_payment_reference,
        ])
    write_table(
        'Don hang',
        ['Mã đơn', 'Ngày tạo', 'Ngày báo cáo', 'Khách hàng', 'SĐT', 'Chi nhánh', 'Trạng thái', 'Phương thức', 'Thanh toán', 'Tiền hàng', 'Giảm hạng KH', 'Phí giao', 'Tổng thanh toán', 'Yêu cầu hóa đơn', 'Mã hóa đơn', 'Mã tham chiếu'],
        order_rows,
    )

    order_ids = [order.pk for order in orders if order.pk]
    order_item_rows = []
    if order_ids:
        order_items = (
            OrderItem.objects
            .filter(order_id__in=order_ids)
            .select_related('order__pharmacy', 'medicine')
            .order_by('-order__created_at', '-order_id', 'id')
        )
        for item in order_items:
            order_item_rows.append([
                item.order.order_code,
                item.order.pharmacy.name if item.order.pharmacy else '',
                item.medicine_name,
                item.quantity,
                int(item.price or 0),
                int(item.line_total or 0),
                item.medicine.unit if item.medicine else '',
                item.medicine.quantity if item.medicine else '',
            ])
    write_table(
        'Chi tiet san pham',
        ['Mã đơn', 'Chi nhánh', 'Sản phẩm', 'Số lượng', 'Đơn giá', 'Thành tiền', 'Đơn vị hiện tại', 'Tồn hiện tại'],
        order_item_rows,
    )

    write_table(
        'Doanh thu theo ky',
        ['Mốc thời gian', 'Khoảng thời gian', 'Doanh thu hoàn thành', 'Số đơn hoàn thành'],
        [
            [item['label'], item['range_label'], item['value'], item['order_count']]
            for item in report_context.get('timeline_chart', [])
        ],
    )

    write_table(
        'Top san pham',
        ['Sản phẩm', 'Số lượng bán', 'Doanh thu', 'Số dòng đơn', 'Đơn vị', 'Tồn hiện tại'],
        [
            [item['name'], item['quantity'], item['revenue'], item['order_count'], item.get('unit', ''), item.get('current_stock', '')]
            for item in report_context.get('top_products', [])
        ],
    )

    write_table(
        'Chi nhanh',
        ['Chi nhánh', 'Số đơn hoàn thành', 'Doanh thu', 'Tiền hàng', 'Giá trị đơn TB', 'Tỷ trọng doanh thu (%)'],
        [
            [item['label'], item['order_count'], item['revenue'], item.get('product_revenue', 0), item['average_order_value'], item['share_percent']]
            for item in report_context.get('branch_performance', [])
        ],
    )

    return_rows = []
    for item in report_context.get('filtered_return_requests', []):
        order = item.order
        return_rows.append([
            order.order_code,
            display_date(item.created_at),
            item.get_status_display(),
            order.full_name,
            item.contact_email,
            item.contact_phone,
            item.reason,
            item.admin_note,
            item.processed_by_display_name,
            display_date(item.processed_at),
        ])
    write_table(
        'Tra hang hoan tien',
        ['Mã đơn', 'Ngày tạo yêu cầu', 'Trạng thái', 'Khách hàng', 'Email liên hệ', 'SĐT liên hệ', 'Lý do', 'Ghi chú xử lý', 'Người xử lý', 'Ngày xử lý'],
        return_rows,
    )

    stock_rows = []
    for medicine in report_context.get('low_stock_medicines', []):
        stock_rows.append(['Sắp hết hàng', medicine.name, medicine.pharmacy.name if medicine.pharmacy else '', medicine.quantity, medicine.unit, display_date(medicine.expiry_date)])
    for medicine in report_context.get('out_of_stock_medicines', []):
        stock_rows.append(['Hết hàng', medicine.name, medicine.pharmacy.name if medicine.pharmacy else '', medicine.quantity, medicine.unit, display_date(medicine.expiry_date)])
    for medicine in report_context.get('expiring_soon_medicines', []):
        stock_rows.append(['Sắp hết hạn', medicine.name, medicine.pharmacy.name if medicine.pharmacy else '', medicine.quantity, medicine.unit, display_date(medicine.expiry_date)])
    write_table(
        'Canh bao kho',
        ['Loại cảnh báo', 'Sản phẩm', 'Chi nhánh', 'Tồn kho', 'Đơn vị', 'Hạn sử dụng'],
        stock_rows,
    )

    workbook.properties.title = 'Báo cáo GIS Pharma'
    workbook.properties.subject = report_context['report_period_label']
    workbook.properties.creator = 'GIS Pharma Admin'
    workbook.active = 0
    return workbook


def _load_excel_builder_tools():
    Workbook = get_excel_workbook_builder()
    try:
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table as ExcelTable
        from openpyxl.worksheet.table import TableStyleInfo
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Máy đang thiếu thư viện openpyxl nên chưa thể định dạng file Excel .xlsx."
        ) from exc

    return {
        "Workbook": Workbook,
        "BarChart": BarChart,
        "LineChart": LineChart,
        "PieChart": PieChart,
        "Reference": Reference,
        "DataLabelList": DataLabelList,
        "CellIsRule": CellIsRule,
        "Alignment": Alignment,
        "Border": Border,
        "Font": Font,
        "PatternFill": PatternFill,
        "Side": Side,
        "get_column_letter": get_column_letter,
        "ExcelTable": ExcelTable,
        "TableStyleInfo": TableStyleInfo,
    }


def _excel_local_value(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            return timezone.localtime(value).replace(tzinfo=None)
        return value
    return value


def _excel_unique_sheet_title(workbook, raw_title):
    cleaned = re.sub(r"[\[\]\*:/\\?]+", " ", str(raw_title or "Sheet")).strip() or "Sheet"
    base = cleaned[:31]
    existing = {name.casefold() for name in workbook.sheetnames}
    if base.casefold() not in existing:
        return base
    for index in range(2, 100):
        suffix = f" {index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate.casefold() not in existing:
            return candidate
    return base[:27] + " 99"


def _excel_table_name(raw_name, used_names):
    base = re.sub(r"[^A-Za-z0-9_]+", "_", str(raw_name or "Table")).strip("_") or "Table"
    if base[0].isdigit():
        base = f"T_{base}"
    candidate = base[:240]
    index = 1
    while candidate in used_names:
        suffix = f"_{index}"
        candidate = f"{base[:240 - len(suffix)]}{suffix}"
        index += 1
    used_names.add(candidate)
    return candidate


def _excel_style_title(worksheet, title, tools, *, last_column=8, subtitle=""):
    Alignment = tools["Alignment"]
    Font = tools["Font"]
    PatternFill = tools["PatternFill"]

    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(last_column, 1))
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.fill = PatternFill("solid", fgColor="17315D")
    title_cell.font = Font(color="FFFFFF", bold=True, size=15)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 28
    if subtitle:
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(last_column, 1))
        subtitle_cell = worksheet.cell(row=2, column=1, value=subtitle)
        subtitle_cell.fill = PatternFill("solid", fgColor="F4F8FF")
        subtitle_cell.font = Font(color="5F7290", italic=True)
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        worksheet.row_dimensions[2].height = 24


def _excel_autosize_columns(worksheet, tools, *, min_width=10, max_width=48):
    get_column_letter = tools["get_column_letter"]
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)


def _excel_apply_table_style(worksheet, headers, header_row, tools, table_name, used_table_names):
    Alignment = tools["Alignment"]
    Border = tools["Border"]
    CellIsRule = tools["CellIsRule"]
    ExcelTable = tools["ExcelTable"]
    Font = tools["Font"]
    PatternFill = tools["PatternFill"]
    Side = tools["Side"]
    TableStyleInfo = tools["TableStyleInfo"]
    get_column_letter = tools["get_column_letter"]

    thin_side = Side(style="thin", color="D7E2F0")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill("solid", fgColor="DDEBFF")
    header_font = Font(color="17315D", bold=True)
    currency_keywords = (
        "doanh thu", "tiền", "tien", "phí", "phi", "tổng", "tong",
        "giá", "gia", "thanh toán", "thanh toan", "ưu đãi", "uu dai",
    )
    percent_keywords = ("%", "tỷ lệ", "ty le", "tỷ trọng", "ty trong")

    last_column = max(len(headers), 1)
    last_row = max(worksheet.max_row, header_row)
    ref = f"A{header_row}:{get_column_letter(last_column)}{last_row}"
    has_data_rows = last_row > header_row
    worksheet.freeze_panes = f"A{header_row + 1}"

    for cell in worksheet[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border

    if has_data_rows:
        for row in worksheet.iter_rows(min_row=header_row + 1, max_row=last_row, max_col=last_column):
            for cell in row:
                cell.border = cell_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                header_text = str(headers[cell.column - 1]).casefold() if cell.column <= len(headers) else ""
                if isinstance(cell.value, datetime):
                    cell.number_format = "dd/mm/yyyy hh:mm"
                elif isinstance(cell.value, date):
                    cell.number_format = "dd/mm/yyyy"
                elif isinstance(cell.value, (int, float)) and any(keyword in header_text for keyword in currency_keywords):
                    cell.number_format = '#,##0" đ";[Red]-#,##0" đ"'
                elif isinstance(cell.value, (int, float)) and any(keyword in header_text for keyword in percent_keywords):
                    cell.number_format = '0.0"%"'
                elif isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0;[Red]-#,##0"

        display_name = _excel_table_name(table_name, used_table_names)
        table = ExcelTable(displayName=display_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    else:
        worksheet.auto_filter.ref = ref

    _excel_autosize_columns(worksheet, tools)
    if has_data_rows:
        quantity_column = None
        for index, header in enumerate(headers, start=1):
            header_text = str(header).casefold()
            if "tồn" in header_text or "ton" in header_text:
                quantity_column = index
                break
        if quantity_column:
            col_letter = get_column_letter(quantity_column)
            data_ref = f"{col_letter}{header_row + 1}:{col_letter}{worksheet.max_row}"
            worksheet.conditional_formatting.add(
                data_ref,
                CellIsRule(operator="lessThanOrEqual", formula=[str(LOW_STOCK_THRESHOLD)], fill=PatternFill("solid", fgColor="FFF2CC")),
            )


def _excel_write_table_sheet(workbook, title, headers, rows, tools, used_table_names, *, table_name, note=""):
    worksheet = workbook.create_sheet(_excel_unique_sheet_title(workbook, title))
    last_column = max(len(headers), 1)
    _excel_style_title(worksheet, title, tools, last_column=last_column, subtitle=note)
    header_row = 4
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=header_row, column=column, value=str(header))
    for raw_row in rows:
        row_values = [_excel_local_value(value) for value in raw_row]
        if len(row_values) < len(headers):
            row_values.extend([""] * (len(headers) - len(row_values)))
        worksheet.append(row_values[:len(headers)])
    _excel_apply_table_style(worksheet, headers, header_row, tools, table_name, used_table_names)
    return worksheet, header_row, worksheet.max_row


def _excel_value_for_display(value, value_format):
    if value_format == "percent":
        return f"{value}%"
    return value


def _excel_add_report_chart(workbook, dashboard, source_sheet, header_row, last_row, tools, *, chart_type, title, data_col, label_col, anchor, height=7, width=13):
    if last_row <= header_row:
        return
    Reference = tools["Reference"]
    if chart_type == "line":
        chart = tools["LineChart"]()
        chart.style = 13
    elif chart_type == "pie":
        chart = tools["PieChart"]()
        chart.dataLabels = tools["DataLabelList"]()
        chart.dataLabels.showPercent = True
    else:
        chart = tools["BarChart"]()
        chart.style = 10
    chart.title = title
    chart.height = height
    chart.width = width
    data = Reference(source_sheet, min_col=data_col, min_row=header_row, max_row=last_row)
    labels = Reference(source_sheet, min_col=label_col, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    if hasattr(chart, "legend"):
        chart.legend = None
    if hasattr(chart, "y_axis"):
        chart.y_axis.numFmt = '#,##0'
    dashboard.add_chart(chart, anchor)


def sanitize_excel_filename(value, prefix):
    raw_value = re.sub(r"[^A-Za-z0-9._-]+", "-", str(value or "").strip()).strip("-") or prefix
    if raw_value.lower().endswith(".xlsx"):
        return raw_value
    return f"{raw_value}.xlsx"


def build_admin_reports_workbook(report_context):
    tools = _load_excel_builder_tools()
    Workbook = tools["Workbook"]
    Alignment = tools["Alignment"]
    Font = tools["Font"]
    PatternFill = tools["PatternFill"]
    Border = tools["Border"]
    Side = tools["Side"]
    get_column_letter = tools["get_column_letter"]
    workbook = Workbook()
    used_table_names = set()

    dashboard = workbook.active
    dashboard.title = _excel_unique_sheet_title(workbook, "Dashboard")
    dashboard.sheet_view.showGridLines = False
    _excel_style_title(
        dashboard,
        "Báo cáo doanh thu GIS Pharma",
        tools,
        last_column=13,
        subtitle="File gồm sheet tổng quan, dữ liệu chi tiết, biểu đồ và sheet đối soát để kiểm tra số liệu sau khi mở bằng Excel.",
    )
    dashboard.column_dimensions["A"].width = 26
    dashboard.column_dimensions["B"].width = 18
    dashboard.column_dimensions["C"].width = 24
    dashboard.column_dimensions["D"].width = 42
    dashboard.column_dimensions["E"].width = 20

    label_font = Font(color="5F7290", bold=True)
    value_font = Font(color="17315D", bold=True)
    info_fill = PatternFill("solid", fgColor="F4F8FF")
    thin_side = Side(style="thin", color="D7E2F0")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    meta_rows = [
        ("Kỳ báo cáo", report_context["range_label"]),
        ("Thời gian", report_context["report_period_label"]),
        ("Phạm vi", report_context["scope_label"]),
        ("So sánh kỳ trước", report_context["previous_period_label"]),
        ("Kiểu nhóm thời gian", report_context.get("timeline_group_label", "")),
    ]
    for row_index, (label, value) in enumerate(meta_rows, start=4):
        dashboard.cell(row=row_index, column=1, value=label).font = label_font
        dashboard.cell(row=row_index, column=2, value=_excel_local_value(value)).font = value_font
        for column in range(1, 5):
            cell = dashboard.cell(row=row_index, column=column)
            cell.fill = info_fill
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    summary_headers = ["Chỉ tiêu", "Giá trị", "So sánh", "Ghi chú"]
    summary_start = 10
    for col, header in enumerate(summary_headers, start=1):
        cell = dashboard.cell(row=summary_start, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="DDEBFF")
        cell.font = Font(color="17315D", bold=True)
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_offset, card in enumerate(report_context.get("summary_cards", []), start=1):
        row_index = summary_start + row_offset
        dashboard.cell(row=row_index, column=1, value=card["label"])
        dashboard.cell(row=row_index, column=2, value=_excel_value_for_display(card["value"], card.get("format")))
        dashboard.cell(row=row_index, column=3, value=card.get("delta", {}).get("text", ""))
        dashboard.cell(row=row_index, column=4, value=card.get("helper", ""))
        for column in range(1, 5):
            cell = dashboard.cell(row=row_index, column=column)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column == 1:
                cell.font = value_font
            if column == 2 and isinstance(cell.value, (int, float)) and card.get("format") == "currency":
                cell.number_format = '#,##0" đ";[Red]-#,##0" đ"'
            elif column == 2 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0;[Red]-#,##0"

    insight_start = summary_start + len(report_context.get("summary_cards", [])) + 3
    dashboard.cell(row=insight_start, column=1, value="Điểm cần chú ý").font = Font(color="17315D", bold=True, size=12)
    dashboard.merge_cells(start_row=insight_start, start_column=1, end_row=insight_start, end_column=4)
    for row_offset, card in enumerate(report_context.get("insight_cards", []), start=1):
        row_index = insight_start + row_offset
        dashboard.cell(row=row_index, column=1, value=card["label"]).font = value_font
        dashboard.cell(row=row_index, column=2, value=_excel_value_for_display(card["value"], card.get("format")))
        dashboard.cell(row=row_index, column=3, value=card.get("note", ""))
        dashboard.merge_cells(start_row=row_index, start_column=3, end_row=row_index, end_column=4)
        for column in range(1, 5):
            cell = dashboard.cell(row=row_index, column=column)
            cell.fill = info_fill
            cell.border = cell_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column == 2 and isinstance(cell.value, (int, float)) and card.get("format") == "currency":
                cell.number_format = '#,##0" đ";[Red]-#,##0" đ"'
            elif column == 2 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0;[Red]-#,##0"

    dashboard.freeze_panes = "A10"

    orders = report_context.get("filtered_orders") or []
    order_rows = []
    for order in orders:
        order_rows.append([
            order.order_code,
            order.created_at,
            getattr(order, "_report_date", ""),
            order.full_name,
            order.phone,
            order.pharmacy.name if order.pharmacy else "",
            order.get_status_display(),
            order.status,
            order.get_payment_method_display(),
            order.get_payment_status_display(),
            int(order.total_product_price or 0),
            int(order.customer_tier_discount_total or 0),
            int(order.shipping_fee or 0),
            int(order.final_total_price or 0),
            "Có" if order.invoice_requested else "Không",
            order.resolved_invoice_code,
            order.resolved_payment_reference,
        ])
    orders_sheet, orders_header_row, orders_last_row = _excel_write_table_sheet(
        workbook,
        "Don hang",
        [
            "Mã đơn", "Ngày tạo", "Ngày báo cáo", "Khách hàng", "SĐT", "Chi nhánh",
            "Trạng thái", "Mã trạng thái", "Phương thức", "Thanh toán", "Tiền hàng",
            "Giảm hạng KH", "Phí giao", "Tổng thanh toán", "Yêu cầu hóa đơn",
            "Mã hóa đơn", "Mã tham chiếu",
        ],
        order_rows,
        tools,
        used_table_names,
        table_name="tblOrders",
        note="Chi tiết đơn hàng trong phạm vi lọc. Doanh thu chính chỉ tính các đơn có mã trạng thái completed.",
    )

    order_ids = [order.pk for order in orders if getattr(order, "pk", None)]
    order_item_rows = []
    if order_ids:
        order_items = (
            OrderItem.objects
            .filter(order_id__in=order_ids)
            .select_related("order__pharmacy", "medicine")
            .order_by("-order__created_at", "-order_id", "id")
        )
        for item in order_items:
            order_item_rows.append([
                item.order.order_code,
                item.order.pharmacy.name if item.order.pharmacy else "",
                item.medicine_name,
                item.quantity,
                int(item.price or 0),
                int(item.line_total or 0),
                item.medicine.unit if item.medicine else "",
                item.medicine.quantity if item.medicine else "",
            ])
    _excel_write_table_sheet(
        workbook,
        "Chi tiet san pham",
        ["Mã đơn", "Chi nhánh", "Sản phẩm", "Số lượng", "Đơn giá", "Thành tiền", "Đơn vị hiện tại", "Tồn hiện tại"],
        order_item_rows,
        tools,
        used_table_names,
        table_name="tblOrderItems",
        note="Dòng sản phẩm của các đơn trong phạm vi báo cáo, dùng để kiểm tra doanh thu theo sản phẩm.",
    )

    timeline_sheet, timeline_header_row, timeline_last_row = _excel_write_table_sheet(
        workbook,
        "Doanh thu theo ky",
        ["Mốc thời gian", "Khoảng thời gian", "Doanh thu hoàn thành", "Số đơn hoàn thành"],
        [
            [item["label"], item["range_label"], item["value"], item["order_count"]]
            for item in report_context.get("timeline_chart", [])
        ],
        tools,
        used_table_names,
        table_name="tblRevenueTimeline",
        note="Chuỗi thời gian được nhóm theo ngày, tuần hoặc tháng tùy khoảng lọc.",
    )

    payment_sheet, payment_header_row, payment_last_row = _excel_write_table_sheet(
        workbook,
        "Thanh toan",
        ["Phương thức", "Số đơn hoàn thành", "Doanh thu", "Tỷ trọng doanh thu (%)"],
        [
            [item["label"], item["order_count"], item["revenue"], item["share_percent"]]
            for item in report_context.get("payment_method_breakdown", [])
        ],
        tools,
        used_table_names,
        table_name="tblPaymentBreakdown",
        note="Cơ cấu doanh thu theo phương thức thanh toán, chỉ lấy đơn đã hoàn thành.",
    )

    _excel_write_table_sheet(
        workbook,
        "Trang thai don",
        ["Trạng thái", "Số đơn", "Tỷ trọng (%)"],
        [
            [item["label"], item["count"], item["share_percent"]]
            for item in report_context.get("status_breakdown", [])
        ],
        tools,
        used_table_names,
        table_name="tblOrderStatus",
        note="Cơ cấu trạng thái của toàn bộ đơn phát sinh trong kỳ.",
    )

    branch_sheet, branch_header_row, branch_last_row = _excel_write_table_sheet(
        workbook,
        "Chi nhanh",
        ["Chi nhánh", "Số đơn hoàn thành", "Doanh thu", "Tiền hàng", "Giá trị đơn TB", "Tỷ trọng doanh thu (%)"],
        [
            [item["label"], item["order_count"], item["revenue"], item.get("product_revenue", 0), item["average_order_value"], item["share_percent"]]
            for item in report_context.get("branch_performance", [])
        ],
        tools,
        used_table_names,
        table_name="tblBranchRevenue",
        note="Xếp hạng doanh thu theo chi nhánh trong phạm vi lọc.",
    )

    _excel_write_table_sheet(
        workbook,
        "Top san pham",
        ["Sản phẩm", "Số lượng bán", "Doanh thu", "Số dòng đơn", "Đơn vị", "Tồn hiện tại"],
        [
            [item["name"], item["quantity"], item["revenue"], item["order_count"], item.get("unit", ""), item.get("current_stock", "")]
            for item in report_context.get("top_products", [])
        ],
        tools,
        used_table_names,
        table_name="tblTopProducts",
        note="Sản phẩm bán chạy theo số lượng trong các đơn đã hoàn thành.",
    )

    return_rows = []
    for item in report_context.get("filtered_return_requests", []):
        order = item.order
        return_rows.append([
            order.order_code,
            item.created_at,
            item.get_status_display(),
            order.full_name,
            item.contact_email,
            item.contact_phone,
            item.reason,
            item.admin_note,
            item.processed_by_display_name,
            item.processed_at,
        ])
    _excel_write_table_sheet(
        workbook,
        "Tra hang hoan tien",
        ["Mã đơn", "Ngày tạo yêu cầu", "Trạng thái", "Khách hàng", "Email liên hệ", "SĐT liên hệ", "Lý do", "Ghi chú xử lý", "Người xử lý", "Ngày xử lý"],
        return_rows,
        tools,
        used_table_names,
        table_name="tblReturns",
        note="Yêu cầu trả hàng/hoàn tiền phát sinh trong kỳ báo cáo.",
    )

    stock_rows = []
    for medicine in report_context.get("low_stock_medicines", []):
        stock_rows.append(["Sắp hết hàng", medicine.name, medicine.pharmacy.name if medicine.pharmacy else "", medicine.quantity, medicine.unit, medicine.expiry_date])
    for medicine in report_context.get("out_of_stock_medicines", []):
        stock_rows.append(["Hết hàng", medicine.name, medicine.pharmacy.name if medicine.pharmacy else "", medicine.quantity, medicine.unit, medicine.expiry_date])
    for medicine in report_context.get("expiring_soon_medicines", []):
        stock_rows.append(["Sắp hết hạn", medicine.name, medicine.pharmacy.name if medicine.pharmacy else "", medicine.quantity, medicine.unit, medicine.expiry_date])
    _excel_write_table_sheet(
        workbook,
        "Canh bao kho",
        ["Loại cảnh báo", "Sản phẩm", "Chi nhánh", "Tồn kho", "Đơn vị", "Hạn sử dụng"],
        stock_rows,
        tools,
        used_table_names,
        table_name="tblStockAlerts",
        note="Các cảnh báo tồn kho liên quan đến kỳ báo cáo hiện tại.",
    )

    audit_sheet = workbook.create_sheet(_excel_unique_sheet_title(workbook, "Doi soat"))
    audit_sheet.sheet_view.showGridLines = False
    _excel_style_title(audit_sheet, "Đối soát số liệu", tools, last_column=5, subtitle="Mở file bằng Excel để công thức SUMIFS/COUNTIF tự tính lại nếu cần.")
    audit_sheet.append([])
    orders_sheet_ref = f"'{orders_sheet.title.replace(chr(39), chr(39) * 2)}'"
    audit_rows = [
        ["Chỉ tiêu", "Giá trị từ hệ thống", "Công thức kiểm tra trong file", "Chênh lệch", "Ghi chú"],
        [
            "Doanh thu đơn hoàn thành",
            int(report_context.get("total_completed_revenue") or 0),
            f'=SUMIFS({orders_sheet_ref}!N:N,{orders_sheet_ref}!H:H,"{Order.STATUS_COMPLETED}")',
            "=B5-C5",
            f"Phải bằng 0 nếu dữ liệu đơn hàng trong sheet {orders_sheet.title} khớp context.",
        ],
        [
            "Số đơn hoàn thành",
            int(report_context.get("completed_order_count") or 0),
            f'=COUNTIF({orders_sheet_ref}!H:H,"{Order.STATUS_COMPLETED}")',
            "=B6-C6",
            "Dùng mã trạng thái để tránh sai lệch do ngôn ngữ hiển thị.",
        ],
        [
            "Tổng số đơn phát sinh",
            int(report_context.get("total_orders") or 0),
            f"=COUNTA({orders_sheet_ref}!A5:A1048576)",
            "=B7-C7",
            f"Tổng dòng đơn trong sheet {orders_sheet.title}.",
        ],
    ]
    for row in audit_rows:
        audit_sheet.append(row)
    _excel_apply_table_style(audit_sheet, audit_rows[0], 4, tools, "tblAuditChecks", used_table_names)
    for row_index in range(5, audit_sheet.max_row + 1):
        audit_sheet.cell(row=row_index, column=2).number_format = '#,##0'
        audit_sheet.cell(row=row_index, column=3).number_format = '#,##0'
        audit_sheet.cell(row=row_index, column=4).number_format = '#,##0'

    info_sheet = workbook.create_sheet(_excel_unique_sheet_title(workbook, "Thong tin"))
    info_sheet.sheet_view.showGridLines = False
    _excel_style_title(info_sheet, "Thông tin file", tools, last_column=4, subtitle="Mô tả bộ lọc và phạm vi dữ liệu được dùng để tạo file.")
    info_sheet.append([])
    info_rows = [
        ["Thuộc tính", "Giá trị"],
        ["Thời điểm xuất", timezone.localtime().replace(tzinfo=None)],
        ["Kỳ báo cáo", report_context["range_label"]],
        ["Thời gian", report_context["report_period_label"]],
        ["Phạm vi", report_context["scope_label"]],
        ["Phương thức thanh toán", next((chip["label"] for chip in report_context.get("filter_chips", []) if chip.get("icon") == "fas fa-wallet"), "")],
    ]
    for row in info_rows:
        info_sheet.append(row)
    _excel_apply_table_style(info_sheet, info_rows[0], 4, tools, "tblReportInfo", used_table_names)

    _excel_add_report_chart(
        workbook,
        dashboard,
        timeline_sheet,
        timeline_header_row,
        timeline_last_row,
        tools,
        chart_type="line",
        title="Doanh thu theo kỳ",
        data_col=3,
        label_col=1,
        anchor="G4",
        height=7,
        width=14,
    )
    _excel_add_report_chart(
        workbook,
        dashboard,
        branch_sheet,
        branch_header_row,
        branch_last_row,
        tools,
        chart_type="bar",
        title="Doanh thu theo chi nhánh",
        data_col=3,
        label_col=1,
        anchor="G21",
        height=7,
        width=14,
    )
    _excel_add_report_chart(
        workbook,
        dashboard,
        payment_sheet,
        payment_header_row,
        payment_last_row,
        tools,
        chart_type="pie",
        title="Cơ cấu thanh toán",
        data_col=3,
        label_col=1,
        anchor="A31",
        height=7,
        width=10,
    )

    for worksheet in workbook.worksheets:
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.5
        worksheet.page_margins.bottom = 0.5
        worksheet.print_options.horizontalCentered = True
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.alignment is None:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except AttributeError:
        pass
    workbook.properties.title = "Báo cáo doanh thu GIS Pharma"
    workbook.properties.subject = report_context["report_period_label"]
    workbook.properties.creator = "GIS Pharma Admin"
    workbook.active = 0
    return workbook


def _related_items(value):
    if value is None:
        return []
    if hasattr(value, "all"):
        return list(value.all())
    return list(value)


def build_stock_export_workbook(export_batch, export_items=None):
    tools = _load_excel_builder_tools()
    Workbook = tools["Workbook"]
    Alignment = tools["Alignment"]
    Font = tools["Font"]
    PatternFill = tools["PatternFill"]
    Border = tools["Border"]
    Side = tools["Side"]
    workbook = Workbook()
    used_table_names = set()
    export_items = list(export_items if export_items is not None else export_batch.items.all())

    scope_labels = dict(StockExportBatch.EXPORT_SCOPE_CHOICES)
    export_scope_label = (
        export_batch.get_export_scope_display()
        if hasattr(export_batch, "get_export_scope_display")
        else scope_labels.get(getattr(export_batch, "export_scope", ""), getattr(export_batch, "export_scope", ""))
    )
    export_code = getattr(export_batch, "resolved_export_code", "") or getattr(export_batch, "export_code", "") or "PX-TAM"
    created_at = getattr(export_batch, "created_at", None)

    summary_sheet = workbook.active
    summary_sheet.title = _excel_unique_sheet_title(workbook, "Phieu xuat")
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet_ref = f"'{summary_sheet.title.replace(chr(39), chr(39) * 2)}'"
    _excel_style_title(summary_sheet, "Phiếu xuất kho", tools, last_column=8, subtitle="Bản Excel chỉ đọc, tạo khi tải xuống và không lưu thêm dữ liệu vào database.")

    label_font = Font(color="5F7290", bold=True)
    value_font = Font(color="17315D", bold=True)
    info_fill = PatternFill("solid", fgColor="F4F8FF")
    thin_side = Side(style="thin", color="D7E2F0")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    pharmacy = getattr(export_batch, "pharmacy", None)
    meta_rows = [
        ("Mã phiếu", export_code),
        ("Chi nhánh", getattr(pharmacy, "name", "")),
        ("Loại phiếu", export_scope_label),
        ("Ngày lập", created_at),
        ("Người lập", getattr(export_batch, "resolved_exported_by_name", "") or getattr(export_batch, "exported_by_name", "")),
        ("Email", getattr(export_batch, "exported_by_email", "")),
        ("Chức vụ", getattr(export_batch, "exported_by_role", "")),
        ("Nơi nhận / mục đích", getattr(export_batch, "destination_name", "")),
        ("Ghi chú", getattr(export_batch, "note", "")),
    ]
    for row_index, (label, value) in enumerate(meta_rows, start=4):
        summary_sheet.cell(row=row_index, column=1, value=label).font = label_font
        summary_sheet.cell(row=row_index, column=2, value=_excel_local_value(value)).font = value_font
        summary_sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=4)
        for column in range(1, 5):
            cell = summary_sheet.cell(row=row_index, column=column)
            cell.fill = info_fill
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(cell.value, datetime):
                cell.number_format = "dd/mm/yyyy hh:mm"

    item_headers = ["STT", "Sản phẩm", "Nhà sản xuất", "Đơn vị", "Tồn trước", "Số lượng xuất", "Tồn sau", "Ghi chú"]
    item_rows = []
    for index, item in enumerate(export_items, start=1):
        item_rows.append([
            index,
            getattr(item, "medicine_name", ""),
            getattr(item, "manufacturer", ""),
            getattr(item, "unit", ""),
            int(getattr(item, "previous_quantity", 0) or 0),
            int(getattr(item, "exported_quantity", 0) or 0),
            int(getattr(item, "remaining_quantity", 0) or 0),
            getattr(item, "note", ""),
        ])
    item_header_row = 16
    for column, header in enumerate(item_headers, start=1):
        summary_sheet.cell(row=item_header_row, column=column, value=header)
    for row in item_rows:
        summary_sheet.append([_excel_local_value(value) for value in row])
    _excel_apply_table_style(summary_sheet, item_headers, item_header_row, tools, "tblStockExportItems", used_table_names)

    allocation_rows = []
    for item in export_items:
        for allocation in _related_items(getattr(item, "lot_allocations", [])):
            allocation_rows.append([
                getattr(item, "medicine_name", ""),
                getattr(item, "unit", ""),
                getattr(allocation, "lot_source_label", "") or (str(getattr(allocation, "lot", "")) if getattr(allocation, "lot", None) else ""),
                getattr(allocation, "lot_expiry_date", None),
                int(getattr(allocation, "lot_import_price", 0) or 0),
                int(getattr(allocation, "quantity", 0) or 0),
            ])
    lots_sheet, _, _ = _excel_write_table_sheet(
        workbook,
        "Lo xuat",
        ["Sản phẩm", "Đơn vị", "Lô xuất", "Hạn sử dụng", "Giá nhập snapshot", "Số lượng phân bổ"],
        allocation_rows,
        tools,
        used_table_names,
        table_name="tblStockExportLots",
        note="Chi tiết lô FEFO thực tế đã được trừ khỏi kho khi lập phiếu.",
    )
    lots_sheet_ref = f"'{lots_sheet.title.replace(chr(39), chr(39) * 2)}'"

    audit_sheet = workbook.create_sheet(_excel_unique_sheet_title(workbook, "Doi soat"))
    audit_sheet.sheet_view.showGridLines = False
    _excel_style_title(audit_sheet, "Đối soát phiếu xuất", tools, last_column=5, subtitle="Các công thức giúp đối chiếu tổng dòng và tổng số lượng sau khi mở file bằng Excel.")
    audit_sheet.append([])
    expected_lines = int(getattr(export_batch, "total_lines", 0) or len(export_items))
    expected_quantity = int(getattr(export_batch, "total_quantity", 0) or sum(int(getattr(item, "exported_quantity", 0) or 0) for item in export_items))
    audit_rows = [
        ["Chỉ tiêu", "Giá trị lưu trên phiếu", "Công thức kiểm tra trong file", "Chênh lệch", "Ghi chú"],
        ["Tổng dòng hàng", expected_lines, f"=COUNTA({summary_sheet_ref}!B17:B1048576)", "=B5-C5", "Chênh lệch cần bằng 0."],
        ["Tổng số lượng xuất", expected_quantity, f"=SUM({summary_sheet_ref}!F:F)", "=B6-C6", "Đối chiếu với total_quantity của phiếu."],
        ["Tổng phân bổ theo lô", expected_quantity, f"=SUM({lots_sheet_ref}!F:F)", "=B7-C7", "Nếu có lô xuất, tổng phân bổ phải bằng tổng số lượng xuất."],
    ]
    for row in audit_rows:
        audit_sheet.append(row)
    _excel_apply_table_style(audit_sheet, audit_rows[0], 4, tools, "tblStockExportAudit", used_table_names)

    for worksheet in workbook.worksheets:
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.5
        worksheet.page_margins.bottom = 0.5
        worksheet.print_options.horizontalCentered = True

    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except AttributeError:
        pass
    workbook.properties.title = f"Phiếu xuất kho {export_code}"
    workbook.properties.subject = export_scope_label
    workbook.properties.creator = "GIS Pharma Admin"
    workbook.active = 0
    return workbook


ADMIN_MODELS = {
    'pharmacy': {
        'model': Pharmacy,
        'title': 'Chi nhánh',
        'title_plural': 'Chi nhánh',
        'form_create': PharmacyAdminForm,
        'form_update': PharmacyAdminForm,
        'search_fields': ['name', 'address', 'phone', 'desc'],
    },
    'medicine': {
        'model': Medicine,
        'title': 'Sản phẩm thuốc',
        'title_plural': 'Sản phẩm thuốc',
        'form_create': MedicineAdminForm,
        'form_update': MedicineAdminForm,
        'search_fields': ['name', 'category', 'manufacturer', 'origin', 'pharmacy__name'],
    },
    'order': {
        'model': Order,
        'title': 'Đơn hàng',
        'title_plural': 'Đơn hàng',
        'search_fields': ['full_name', 'phone', 'address_text', 'pharmacy__name'],
    },
    'return_request': {
        'model': ReturnRefundRequest,
        'title': 'Yêu cầu trả hàng / hoàn tiền',
        'title_plural': 'Yêu cầu trả hàng / hoàn tiền',
        'search_fields': ['order__full_name', 'order__phone', 'order__address_text', 'order__pharmacy__name', 'reason', 'contact_phone', 'contact_email'],
    },
    'purchase_import': {
        'model': PurchaseImportBatch,
        'title': 'Phiếu nhập hàng',
        'title_plural': 'Nhập hàng bằng Excel',
        'search_fields': ['invoice_code', 'pharmacy__name', 'imported_by_name', 'note'],
    },
    'stock_export': {
        'model': StockExportBatch,
        'title': 'Phiếu xuất kho',
        'title_plural': 'Phiếu xuất kho',
        'search_fields': ['export_code', 'pharmacy__name', 'exported_by_name', 'destination_name', 'note'],
    },
    'promotion': {
        'model': MedicinePromotion,
        'title': 'Khuyến mãi sản phẩm',
        'title_plural': 'Khuyến mãi sản phẩm',
        'form_create': PromotionAdminForm,
        'form_update': PromotionAdminForm,
        'search_fields': ['title', 'medicine__name', 'medicine__manufacturer', 'medicine__pharmacy__name', 'note'],
    },
    'inventory_lot': {
        'model': MedicineLot,
        'title': 'Lô tồn kho',
        'title_plural': 'Lô tồn kho FEFO',
        'search_fields': ['source_label', 'medicine__name', 'medicine__manufacturer', 'pharmacy__name', 'note'],
    },
    'news': {
        'model': NewsArticle,
        'title': 'Tin tức',
        'title_plural': 'Tin tức',
        'form_create': NewsArticleAdminForm,
        'form_update': NewsArticleAdminForm,
        'search_fields': ['title', 'slug', 'summary', 'content'],
    },
    'user': {
        'model': User,
        'title': 'Tài khoản',
        'title_plural': 'Tài khoản',
        'form_create': CustomUserCreateForm,
        'form_update': CustomUserUpdateForm,
        'search_fields': ['username', 'email', 'first_name', 'last_name'],
    },
}


def get_admin_config(model_key):
    config = ADMIN_MODELS.get(model_key)
    if not config:
        raise Http404('Không tìm thấy module quản trị')
    return config


def get_admin_form_sections(form, model_key):
    layouts = {
        'pharmacy': [
            {
                'title': 'Thông tin chi nhánh',
                'icon': 'fas fa-clinic-medical',
                'fields': ['name', 'address', 'phone', 'desc'],
            },
            {
                'title': 'Thời gian hoạt động và bản đồ',
                'icon': 'fas fa-clock',
                'fields': ['open_time', 'close_time', 'lat', 'lng'],
            },
            {
                'title': 'Hình ảnh đại diện',
                'icon': 'fas fa-image',
                'fields': ['image', 'gallery_images'],
            },
        ],
        'medicine': [
            {
                'title': 'Thông tin cơ bản',
                'icon': 'fas fa-pills',
                'fields': ['name', 'pharmacy', 'product_type', 'category', 'unit', 'short_description', 'description', 'image', 'gallery_images'],
            },
            {
                'title': 'Thông tin bán hàng',
                'icon': 'fas fa-cash-register',
                'fields': ['price', 'quantity', 'expiry_date', 'prescription_required'],
            },
            {
                'title': 'Nguồn gốc sản phẩm',
                'icon': 'fas fa-industry',
                'fields': ['manufacturer', 'origin'],
            },
        ],
        'promotion': [
            {
                'title': 'Thiết lập khuyến mãi',
                'icon': 'fas fa-tags',
                'fields': ['medicine', 'title', 'discount_percent', 'start_date', 'end_date', 'is_active', 'note'],
            },
        ],
        'news': [
            {
                'title': 'Thong tin bai viet',
                'icon': 'fas fa-newspaper',
                'fields': ['title', 'slug', 'cover_image'],
            },
            {
                'title': 'Tom tat va noi dung',
                'icon': 'fas fa-align-left',
                'fields': ['summary', 'content'],
            },
            {
                'title': 'Xuat ban',
                'icon': 'fas fa-upload',
                'fields': ['published_at', 'is_published'],
            },
        ],
        'user': [
            {
                'title': 'Thông tin đăng nhập',
                'icon': 'fas fa-user-shield',
                'fields': ['username', 'email'],
            },
            {
                'title': 'Thông tin cá nhân',
                'icon': 'fas fa-id-card',
                'fields': ['first_name', 'last_name'],
            },
            {
                'title': 'Bảo mật tài khoản',
                'icon': 'fas fa-lock',
                'fields': ['password1', 'password2', 'new_password', 'confirm_new_password'],
            },
            {
                'title': 'Vai trò và phạm vi quản lý',
                'icon': 'fas fa-user-cog',
                'fields': ['role', 'managed_pharmacy', 'is_active'],
            },
        ],
    }

    sections = []
    for section in layouts.get(model_key, []):
        normal_fields = []
        checkbox_fields = []
        for field_name in section['fields']:
            if field_name not in form.fields:
                continue
            bound_field = form[field_name]
            input_type = getattr(bound_field.field.widget, 'input_type', '')
            is_rich_text = bound_field.field.widget.attrs.get('data-rich-editor') == '1'
            field_info = {
                'field': bound_field,
                'is_rich_text': is_rich_text,
                'full_width': is_rich_text or input_type in {'textarea', 'file'} or field_name in {'desc', 'short_description', 'description', 'usage', 'ingredients', 'dosage'},
            }
            if input_type == 'checkbox':
                checkbox_fields.append(field_info)
            else:
                normal_fields.append(field_info)
        if normal_fields or checkbox_fields:
            sections.append({
                'title': section['title'],
                'icon': section['icon'],
                'normal_fields': normal_fields,
                'checkbox_fields': checkbox_fields,
            })
    return sections


def apply_admin_sort(queryset, model_key, sort_key):
    sort_map = {
        'pharmacy': {
            'newest': '-id',
            'name_asc': 'name',
            'name_desc': '-name',
            'medicine_desc': '-medicine_total',
        },
        'medicine': {
            'newest': '-id',
            'name_asc': 'name',
            'price_low': 'price',
            'price_high': '-price',
            'stock_low': 'quantity',
            'stock_high': '-quantity',
        },
        'order': {
            'newest': ('-created_at', '-id'),
            'oldest': ('created_at', 'id'),
            'total_high': '-final_total_price',
            'total_low': 'final_total_price',
        },
        'return_request': {
            'newest': ('-created_at', '-id'),
            'oldest': ('created_at', 'id'),
        },
        'purchase_import': {
            'newest': ('-created_at', '-id'),
            'oldest': ('created_at', 'id'),
            'quantity_high': '-total_quantity',
            'quantity_low': 'total_quantity',
        },
        'stock_export': {
            'newest': ('-created_at', '-id'),
            'oldest': ('created_at', 'id'),
            'quantity_high': '-total_quantity',
            'quantity_low': 'total_quantity',
        },
        'promotion': {
            'newest': ('-created_at', '-id'),
            'discount_high': '-discount_percent',
            'discount_low': 'discount_percent',
        },
        'inventory_lot': {
            'expiry_soon': ('expiry_date', 'id'),
            'newest': ('-created_at', '-id'),
            'remaining_high': '-remaining_quantity',
            'remaining_low': 'remaining_quantity',
        },
        'news': {
            'newest': ('-published_at', '-created_at', '-id'),
            'oldest': ('published_at', 'created_at', 'id'),
            'title_asc': 'title',
            'title_desc': '-title',
        },
        'user': {
            'newest': '-date_joined',
            'username_asc': 'username',
            'username_desc': '-username',
        },
    }
    selected = sort_map.get(model_key, {}).get(sort_key)
    if isinstance(selected, (list, tuple)):
        return queryset.order_by(*selected)
    return queryset.order_by(selected or '-id')


def get_pharmacy_filter_options(request):
    return [
        {
            'name': 'stock_state',
            'label': 'Tình trạng',
            'value': request.GET.get('stock_state', ''),
            'options': [
                ('', 'Tất cả'),
                ('available', 'Có hàng'),
                ('empty', 'Không có hàng'),
            ],
        },
        {
            'name': 'sort',
            'label': 'Sắp xếp',
            'value': request.GET.get('sort', 'name_asc'),
            'options': [
                ('name_asc', 'Tên A-Z'),
                ('name_desc', 'Tên Z-A'),
                ('medicine_desc', 'Nhiều sản phẩm nhất'),
                ('newest', 'Mới cập nhật'),
            ],
        },
    ]


def get_medicine_filter_options(request, pharmacy_queryset=None, selected_pharmacy_id=''):
    pharmacy_queryset = pharmacy_queryset if pharmacy_queryset is not None else Pharmacy.objects.order_by('name')
    pharmacy_options = [(str(pharmacy.pk), pharmacy.name) for pharmacy in pharmacy_queryset]
    if len(pharmacy_options) > 1:
        pharmacy_options = [('', 'Tất cả')] + pharmacy_options
    return [
        {
            'name': 'pharmacy',
            'label': 'Chi nhánh',
            'value': selected_pharmacy_id,
            'options': pharmacy_options,
        },
        {
            'name': 'product_type',
            'label': 'Loại sản phẩm',
            'value': request.GET.get('product_type', ''),
            'options': [('', 'Tất cả')] + [(value, label) for value, label in MEDICINE_PRODUCT_TYPE_CHOICES],
        },
        {
            'name': 'stock',
            'label': 'Tồn kho',
            'value': request.GET.get('stock', ''),
            'options': [
                ('', 'Tất cả'),
                ('in', 'Còn hàng'),
                ('low', 'Sắp hết'),
                ('out', 'Hết hàng'),
            ],
        },
        {
            'name': 'rx',
            'label': 'Kê đơn',
            'value': request.GET.get('rx', ''),
            'options': [
                ('', 'Tất cả'),
                ('yes', 'Cần kê đơn'),
                ('no', 'Không kê đơn'),
            ],
        },
        {
            'name': 'sort',
            'label': 'Sắp xếp',
            'value': request.GET.get('sort', 'newest'),
            'options': [
                ('newest', 'Mới nhất'),
                ('name_asc', 'Tên A-Z'),
                ('price_low', 'Giá tăng dần'),
                ('price_high', 'Giá giảm dần'),
                ('stock_low', 'Tồn kho thấp nhất'),
                ('stock_high', 'Tồn kho cao nhất'),
            ],
        },
    ]


def get_order_filter_options(request, pharmacy_queryset=None, selected_pharmacy_id=''):
    pharmacy_queryset = pharmacy_queryset if pharmacy_queryset is not None else Pharmacy.objects.order_by('name')
    pharmacy_options = [(str(pharmacy.pk), pharmacy.name) for pharmacy in pharmacy_queryset]
    if len(pharmacy_options) > 1:
        pharmacy_options = [('', 'Tất cả')] + pharmacy_options
    return [
        {
            'name': 'status',
            'label': 'Trạng thái',
            'value': request.GET.get('status', ''),
            'options': [('', 'Tất cả')] + [(value, label) for value, label in Order.STATUS_CHOICES],
        },
        {
            'name': 'pharmacy',
            'label': 'Chi nhánh',
            'value': selected_pharmacy_id,
            'options': pharmacy_options,
        },
        {
            'name': 'sort',
            'label': 'Sắp xếp',
            'value': request.GET.get('sort', 'newest'),
            'options': [
                ('newest', 'Mới nhất'),
                ('oldest', 'Cũ nhất'),
                ('total_high', 'Tổng tiền giảm dần'),
                ('total_low', 'Tổng tiền tăng dần'),
            ],
        },
    ]




def get_return_request_filter_options(request, pharmacy_queryset=None, selected_pharmacy_id=''):
    pharmacy_queryset = pharmacy_queryset if pharmacy_queryset is not None else Pharmacy.objects.order_by('name')
    pharmacy_options = [(str(pharmacy.pk), pharmacy.name) for pharmacy in pharmacy_queryset]
    if len(pharmacy_options) > 1:
        pharmacy_options = [('', 'Tất cả')] + pharmacy_options
    return [
        {
            'name': 'status',
            'label': 'Trạng thái',
            'value': request.GET.get('status', ''),
            'options': [('', 'Tất cả'), (ReturnRefundRequest.STATUS_PROCESSING, 'Đang xử lý'), (ReturnRefundRequest.STATUS_APPROVED, 'Chấp nhận hoàn tiền'), (ReturnRefundRequest.STATUS_REJECTED, 'Từ chối hoàn tiền')],
        },
        {
            'name': 'pharmacy',
            'label': 'Chi nhánh',
            'value': selected_pharmacy_id,
            'options': pharmacy_options,
        },
        {
            'name': 'sort',
            'label': 'Sắp xếp',
            'value': request.GET.get('sort', 'newest'),
            'options': [
                ('newest', 'Mới nhất'),
                ('oldest', 'Cũ nhất'),
            ],
        },
    ]

def get_purchase_import_filter_options(request, pharmacy_queryset=None, selected_pharmacy_id=''):
    pharmacy_queryset = pharmacy_queryset if pharmacy_queryset is not None else Pharmacy.objects.order_by('name')
    pharmacy_options = [(str(pharmacy.pk), pharmacy.name) for pharmacy in pharmacy_queryset]
    if len(pharmacy_options) > 1:
        pharmacy_options = [('', 'Tất cả')] + pharmacy_options
    return [
        {
            'name': 'pharmacy',
            'label': 'Chi nhánh',
            'value': selected_pharmacy_id,
            'options': pharmacy_options,
            'type': 'select',
        },
        {
            'name': 'imported_from',
            'label': 'Từ ngày nhập',
            'value': request.GET.get('imported_from', ''),
            'type': 'date',
        },
        {
            'name': 'imported_to',
            'label': 'Đến ngày nhập',
            'value': request.GET.get('imported_to', ''),
            'type': 'date',
        },
        {
            'name': 'sort',
            'label': 'Sắp xếp',
            'value': request.GET.get('sort', 'newest'),
            'options': [
                ('newest', 'Mới nhất'),
                ('oldest', 'Cũ nhất'),
                ('quantity_high', 'Số lượng nhập giảm dần'),
                ('quantity_low', 'Số lượng nhập tăng dần'),
            ],
            'type': 'select',
        },
    ]


def get_stock_export_filter_options(request, pharmacy_queryset=None, selected_pharmacy_id=''):
    pharmacy_queryset = pharmacy_queryset if pharmacy_queryset is not None else Pharmacy.objects.order_by('name')
    pharmacy_options = [(str(pharmacy.pk), pharmacy.name) for pharmacy in pharmacy_queryset]
    if len(pharmacy_options) > 1:
        pharmacy_options = [('', 'Tất cả')] + pharmacy_options
    return [
        {
            'name': 'pharmacy',
            'label': 'Chi nhánh',
            'value': selected_pharmacy_id,
            'options': pharmacy_options,
            'type': 'select',
        },
        {
            'name': 'export_scope',
            'label': 'Loại phiếu xuất',
            'value': request.GET.get('export_scope', ''),
            'options': [
                ('', 'Tất cả'),
                *StockExportBatch.EXPORT_SCOPE_CHOICES,
            ],
            'type': 'select',
        },
        {
            'name': 'exported_from',
            'label': 'Từ ngày xuất',
            'value': request.GET.get('exported_from', ''),
            'type': 'date',
        },
        {
            'name': 'exported_to',
            'label': 'Đến ngày xuất',
            'value': request.GET.get('exported_to', ''),
            'type': 'date',
        },
        {
            'name': 'sort',
            'label': 'Sắp xếp',
            'value': request.GET.get('sort', 'newest'),
            'options': [
                ('newest', 'Mới nhất'),
                ('oldest', 'Cũ nhất'),
                ('quantity_high', 'Số lượng xuất giảm dần'),
                ('quantity_low', 'Số lượng xuất tăng dần'),
            ],
            'type': 'select',
        },
    ]


def get_promotion_filter_options(request, pharmacy_queryset=None, selected_pharmacy_id=''):
    pharmacy_queryset = pharmacy_queryset if pharmacy_queryset is not None else Pharmacy.objects.order_by('name')
    pharmacy_options = [(str(pharmacy.pk), pharmacy.name) for pharmacy in pharmacy_queryset]
    if len(pharmacy_options) > 1:
        pharmacy_options = [('', 'Tất cả')] + pharmacy_options
    return [
        {
            'name': 'pharmacy',
            'label': 'Chi nhánh',
            'value': selected_pharmacy_id,
            'options': pharmacy_options,
            'type': 'select',
        },
        {
            'name': 'promo_state',
            'label': 'Trạng thái chương trình',
            'value': request.GET.get('promo_state', ''),
            'options': [
                ('', 'Tất cả'),
                ('active', 'Đang áp dụng'),
                ('upcoming', 'Sắp bắt đầu'),
                ('ended', 'Đã kết thúc'),
                ('inactive', 'Tắt thủ công'),
            ],
            'type': 'select',
        },
        {
            'name': 'medicine_kind',
            'label': 'Loại sản phẩm',
            'value': request.GET.get('medicine_kind', ''),
            'options': [
                ('', 'Tất cả'),
                ('expiring', 'HSD ≤ 6 tháng'),
                ('normal', 'Sản phẩm thường'),
            ],
            'type': 'select',
        },
        {
            'name': 'sort',
            'label': 'Sắp xếp',
            'value': request.GET.get('sort', 'newest'),
            'options': [
                ('newest', 'Mới nhất'),
                ('discount_high', 'Giảm cao nhất'),
                ('discount_low', 'Giảm thấp nhất'),
            ],
            'type': 'select',
        },
    ]


def build_grouped_promotion_candidates(medicines, *, active_group_keys=None):
    today = timezone.localdate()
    warning_deadline = today + timedelta(days=183)
    active_group_keys = set(active_group_keys or set())
    grouped = {}

    for medicine in medicines:
        group_key = normalize_catalog_key(medicine.name, medicine.unit, medicine.manufacturer)
        expiry_date = medicine.expiry_date
        is_warning = bool(expiry_date and today <= expiry_date <= warning_deadline)
        is_expired = bool(expiry_date and expiry_date < today)
        entry = grouped.setdefault(group_key, {
            'medicine': medicine,
            'name': medicine.name,
            'unit': medicine.unit or 'Hộp',
            'manufacturer': medicine.manufacturer or '',
            'total_stock': 0,
            'warning_stock': 0,
            'nearest_expiry': None,
            'is_warning': False,
            'is_expired': False,
            'is_active': group_key in active_group_keys,
            'group_key': group_key,
        })

        entry['total_stock'] += int(medicine.quantity or 0)
        if is_warning:
            entry['warning_stock'] += int(medicine.quantity or 0)
            entry['is_warning'] = True
        if is_expired:
            entry['is_expired'] = True
        if expiry_date:
            if entry['nearest_expiry'] is None or expiry_date < entry['nearest_expiry']:
                entry['nearest_expiry'] = expiry_date

        representative_rank = (
            0 if medicine.quantity > 0 else 1,
            -int(medicine.quantity or 0),
            medicine.id,
        )
        current_rank = getattr(entry['medicine'], '_promo_representative_rank', None)
        if current_rank is None or representative_rank < current_rank:
            medicine._promo_representative_rank = representative_rank
            entry['medicine'] = medicine

    return list(grouped.values())


def build_promotion_admin_suggestion_groups(user, current_promotion=None):
    scoped_medicines = list(
        filter_queryset_by_admin_scope(
            Medicine.objects.select_related('pharmacy'),
            user,
            'medicine',
        ).filter(quantity__gt=0)
    )

    active_group_keys = {
        normalize_catalog_key(promotion.medicine.name, promotion.medicine.unit, promotion.medicine.manufacturer)
        for promotion in build_active_promotion_queryset(MedicinePromotion.objects.select_related('medicine'))
        if getattr(promotion, 'medicine', None)
    }
    if current_promotion and getattr(current_promotion, 'medicine', None):
        active_group_keys.discard(
            normalize_catalog_key(
                current_promotion.medicine.name,
                current_promotion.medicine.unit,
                current_promotion.medicine.manufacturer,
            )
        )

    grouped_candidates = build_grouped_promotion_candidates(scoped_medicines, active_group_keys=active_group_keys)
    candidate_items = [item for item in grouped_candidates if not item['is_active']]
    suggested_items = sorted(
        [item for item in candidate_items if item['is_warning']],
        key=lambda item: (-item['warning_stock'], item['nearest_expiry'] or timezone.localdate(), item['name'].casefold()),
    )[:8]
    regular_items = sorted(
        [item for item in candidate_items if not item['is_warning']],
        key=lambda item: (-item['total_stock'], item['name'].casefold()),
    )[:8]
    active_items = sorted(
        [item for item in grouped_candidates if item['is_active']],
        key=lambda item: (item['name'].casefold(), -(item['total_stock'] or 0)),
    )[:8]

    def decorate(items, *, tone, kind_label):
        decorated = []
        for entry in items:
            decorated.append({
                'medicine': entry['medicine'],
                'tone': tone,
                'kind_label': kind_label,
                'expiry_label': entry['nearest_expiry'].strftime('%d/%m/%Y') if entry['nearest_expiry'] else 'Chưa có HSD',
                'total_stock': entry['total_stock'],
                'warning_stock': entry['warning_stock'],
                'unit': entry['unit'],
            })
        return decorated

    return {
        'suggested': decorate(suggested_items, tone='warning', kind_label='Ưu tiên cận hạn ≤ 6 tháng'),
        'regular': decorate(regular_items, tone='info', kind_label='Sản phẩm thường'),
        'active': decorate(active_items, tone='success', kind_label='Đang có khuyến mãi'),
    }


def get_inventory_lot_filter_options(request, pharmacy_queryset=None, selected_pharmacy_id=''):
    pharmacy_queryset = pharmacy_queryset if pharmacy_queryset is not None else Pharmacy.objects.order_by('name')
    pharmacy_options = [(str(pharmacy.pk), pharmacy.name) for pharmacy in pharmacy_queryset]
    if len(pharmacy_options) > 1:
        pharmacy_options = [('', 'Tất cả')] + pharmacy_options
    effective_expiry_state = (request.GET.get('expiry_state', '') or 'sellable').strip()
    return [
        {
            'name': 'pharmacy',
            'label': 'Chi nhánh',
            'value': selected_pharmacy_id,
            'options': pharmacy_options,
            'type': 'select',
        },
        {
            'name': 'expiry_state',
            'label': 'HSD',
            'value': effective_expiry_state,
            'options': [
                ('sellable', 'Đang bán được'),
                ('warning', 'HSD ≤ 6 tháng'),
                ('expired', 'Đã hết hạn'),
                ('', 'Tất cả'),
            ],
            'type': 'select',
        },
        {
            'name': 'lot_state',
            'label': 'Tồn lô',
            'value': request.GET.get('lot_state', ''),
            'options': [
                ('', 'Tất cả'),
                ('available', 'Còn tồn'),
                ('empty', 'Đã xuất hết'),
            ],
            'type': 'select',
        },
        {
            'name': 'sort',
            'label': 'Sắp xếp',
            'value': request.GET.get('sort', 'expiry_soon'),
            'options': [
                ('expiry_soon', 'HSD gần nhất'),
                ('newest', 'Lô mới nhất'),
                ('remaining_high', 'Tồn lô nhiều nhất'),
                ('remaining_low', 'Tồn lô ít nhất'),
            ],
            'type': 'select',
        },
    ]


def get_news_filter_options(request):
    return [
        {
            'name': 'publish_state',
            'label': 'Trang thai',
            'value': request.GET.get('publish_state', ''),
            'options': [
                ('', 'Tat ca'),
                ('published', 'Da xuat ban'),
                ('draft', 'Nhap'),
            ],
            'type': 'select',
        },
        {
            'name': 'published_from',
            'label': 'Tu ngay xuat ban',
            'value': request.GET.get('published_from', ''),
            'type': 'date',
        },
        {
            'name': 'published_to',
            'label': 'Den ngay xuat ban',
            'value': request.GET.get('published_to', ''),
            'type': 'date',
        },
        {
            'name': 'sort',
            'label': 'Sap xep',
            'value': request.GET.get('sort', 'newest'),
            'options': [
                ('newest', 'Moi nhat'),
                ('oldest', 'Cu nhat'),
                ('title_asc', 'Tieu de A-Z'),
                ('title_desc', 'Tieu de Z-A'),
            ],
            'type': 'select',
        },
    ]


def get_user_filter_options(request):
    return [
        {
            'name': 'role',
            'label': 'Vai trò',
            'value': request.GET.get('role', ''),
            'options': [
                ('', 'Tất cả'),
                ('customer', 'Khách hàng'),
                ('staff', 'Nhân viên'),
                ('superuser', 'Quản trị viên cấp cao'),
            ],
        },
        {
            'name': 'active',
            'label': 'Trạng thái',
            'value': request.GET.get('active', ''),
            'options': [
                ('', 'Tất cả'),
                ('yes', 'Đang hoạt động'),
                ('no', 'Đã khóa'),
            ],
        },
        {
            'name': 'sort',
            'label': 'Sắp xếp',
            'value': request.GET.get('sort', 'newest'),
            'options': [
                ('newest', 'Mới nhất'),
                ('username_asc', 'Tên đăng nhập A-Z'),
                ('username_desc', 'Tên đăng nhập Z-A'),
            ],
        },
    ]


def build_list_data(model_key, request):
    keyword = request.GET.get('q', '').strip()
    managed_pharmacy = get_admin_scope_pharmacy(request.user)

    if model_key == 'pharmacy':
        queryset = Pharmacy.objects.annotate(
            medicine_total=Count('medicines', distinct=True),
            available_total=Count('medicines', filter=Q(medicines__quantity__gt=0), distinct=True),
        )
        if keyword:
            queryset = queryset.filter(build_search_query(keyword, ADMIN_MODELS[model_key]['search_fields']))

        stock_state = request.GET.get('stock_state', '')
        if stock_state == 'available':
            queryset = queryset.filter(available_total__gt=0)
        elif stock_state == 'empty':
            queryset = queryset.filter(available_total=0)

        queryset = apply_admin_sort(queryset, model_key, request.GET.get('sort', 'name_asc'))
        columns = ['Ảnh', 'Chi nhánh', 'Liên hệ', 'Giờ hoạt động', 'Sản phẩm', 'Tình trạng']
        summary_cards = [
            {'label': 'Tổng chi nhánh', 'value': queryset.count(), 'tone': 'primary'},
            {'label': 'Chi nhánh có hàng', 'value': queryset.filter(available_total__gt=0).count(), 'tone': 'success'},
            {'label': 'Chi nhánh chưa có hàng', 'value': queryset.filter(available_total=0).count(), 'tone': 'warning'},
        ]
        filter_options = get_pharmacy_filter_options(request)

        rows = []
        for obj in queryset:
            status_badge = render_badge('Có hàng', 'success') if obj.available_total > 0 else render_badge('Không có hàng', 'danger')
            actions = []
            if can_update_admin_model(request.user, model_key):
                actions.append({'url': reverse('custom_admin_update', kwargs={'model_key': 'pharmacy', 'pk': obj.pk}), 'label': 'Cập nhật', 'icon': 'fas fa-pen', 'class': 'btn-primary'})
            if can_delete_object(request.user, model_key, obj):
                actions.append({'url': reverse('custom_admin_delete', kwargs={'model_key': 'pharmacy', 'pk': obj.pk}), 'label': 'Xóa', 'icon': 'fas fa-trash', 'class': 'btn-danger'})
            rows.append({
                'cells': [
                    render_image_thumb(obj.image, obj.name, 'Ảnh chi nhánh'),
                    format_html('<div class="cell-title">{}</div><div class="cell-sub">ID: #{}</div>', obj.name, obj.pk),
                    format_html('<div>{}</div><div class="cell-sub">{}</div>', obj.address, obj.phone or '-'),
                    obj.opening_hours or '-',
                    format_html('<strong>{}</strong><div class="cell-sub">{} loại đang có hàng</div>', obj.medicine_total, obj.available_total),
                    status_badge,
                ],
                'actions': actions,
            })
        return queryset, columns, rows, filter_options, summary_cards, keyword

    if model_key == 'medicine':
        base_queryset = filter_queryset_by_admin_scope(Medicine.objects.select_related('pharmacy').annotate(active_lot_total=Count('lots', filter=Q(lots__remaining_quantity__gt=0), distinct=True)), request.user, 'medicine')
        queryset = base_queryset
        if keyword:
            queryset = queryset.filter(build_search_query(keyword, ADMIN_MODELS[model_key]['search_fields']))

        pharmacy_queryset = Pharmacy.objects.order_by('name')
        if managed_pharmacy is not None:
            pharmacy_queryset = Pharmacy.objects.filter(pk=managed_pharmacy.pk)
        pharmacy_id = request.GET.get('pharmacy', '')
        if pharmacy_id:
            queryset = queryset.filter(pharmacy_id=pharmacy_id)

        product_type = request.GET.get('product_type', '')
        if product_type in {choice[0] for choice in MEDICINE_PRODUCT_TYPE_CHOICES}:
            queryset = queryset.filter(product_type=product_type)

        stock = request.GET.get('stock', '')
        if stock == 'in':
            queryset = queryset.filter(quantity__gt=LOW_STOCK_THRESHOLD)
        elif stock == 'low':
            queryset = queryset.filter(quantity__gt=0, quantity__lte=LOW_STOCK_THRESHOLD)
        elif stock == 'out':
            queryset = queryset.filter(quantity__lte=0)

        rx = request.GET.get('rx', '')
        if rx == 'yes':
            queryset = queryset.filter(prescription_required=True)
        elif rx == 'no':
            queryset = queryset.filter(prescription_required=False)

        queryset = apply_admin_sort(queryset, model_key, request.GET.get('sort', 'newest'))
        columns = ['Ảnh', 'Sản phẩm', 'Chi nhánh', 'Giá bán', 'Tồn kho', 'Kê đơn']
        summary_cards = [
            {'label': 'Tổng sản phẩm', 'value': base_queryset.count(), 'tone': 'primary'},
            {'label': 'Sắp hết hàng', 'value': base_queryset.filter(quantity__gt=0, quantity__lte=LOW_STOCK_THRESHOLD).count(), 'tone': 'warning'},
            {'label': 'Hết hàng', 'value': base_queryset.filter(quantity__lte=0).count(), 'tone': 'danger'},
            {'label': 'HSD ≤ 6 tháng', 'value': get_expiring_soon_medicines_queryset(base_queryset).count(), 'tone': 'info'},
        ]
        filter_options = get_medicine_filter_options(request, pharmacy_queryset=pharmacy_queryset, selected_pharmacy_id=pharmacy_id)

        rows = []
        for obj in queryset:
            actions = []
            if can_update_admin_model(request.user, model_key):
                actions.append({'url': reverse('custom_admin_update', kwargs={'model_key': 'medicine', 'pk': obj.pk}), 'label': 'Cập nhật', 'icon': 'fas fa-pen', 'class': 'btn-primary'})
            if can_delete_object(request.user, model_key, obj):
                actions.append({'url': reverse('custom_admin_delete', kwargs={'model_key': 'medicine', 'pk': obj.pk}), 'label': 'Xóa', 'icon': 'fas fa-trash', 'class': 'btn-danger'})
            rows.append({
                'cells': [
                    render_image_thumb(obj.image, obj.name, 'Ảnh sản phẩm'),
                    format_html(
                        '<div class="cell-title">{}</div><div class="cell-sub">{} • {} • {} • HSD cảnh báo: {} • {} lô còn tồn</div>',
                        obj.name,
                        obj.get_product_type_display(),
                        obj.category or 'Chưa phân loại',
                        obj.unit or '-',
                        obj.expiry_date.strftime('%d/%m/%Y') if obj.expiry_date else 'Chưa cập nhật',
                        getattr(obj, 'active_lot_total', 0),
                    ),
                    obj.pharmacy.name if obj.pharmacy else '-',
                    format_money(obj.price),
                    format_html('<div>{}</div><div class="mt-1">{}</div>', obj.quantity, render_stock_badge(obj.quantity)),
                    render_prescription_badge(obj.prescription_required),
                ],
                'actions': actions,
            })
        return queryset, columns, rows, filter_options, summary_cards, keyword

    if model_key == 'order':
        base_queryset = filter_queryset_by_admin_scope(
            Order.objects.select_related('pharmacy', 'user').annotate(item_total=Count('items')),
            request.user,
            'order',
        )
        auto_complete_overdue_shipping_orders(base_queryset)
        base_queryset = filter_queryset_by_admin_scope(
            Order.objects.select_related('pharmacy', 'user').annotate(item_total=Count('items')),
            request.user,
            'order',
        )
        queryset = base_queryset
        if keyword:
            queryset = queryset.filter(build_search_query(keyword, ADMIN_MODELS[model_key]['search_fields']))

        status = request.GET.get('status', '')
        if status:
            queryset = queryset.filter(status=status)

        pharmacy_queryset = Pharmacy.objects.order_by('name')
        if managed_pharmacy is not None:
            pharmacy_queryset = Pharmacy.objects.filter(pk=managed_pharmacy.pk)
        pharmacy_id = request.GET.get('pharmacy', '')
        if pharmacy_id:
            queryset = queryset.filter(pharmacy_id=pharmacy_id)

        queryset = apply_admin_sort(queryset, model_key, request.GET.get('sort', 'newest'))
        columns = ['Mã đơn', 'Khách hàng', 'Chi nhánh xử lý', 'Tổng tiền', 'Trạng thái']
        summary_cards = [
            {'label': 'Đơn cần xử lý', 'value': base_queryset.filter(status__in=[Order.STATUS_PENDING, Order.STATUS_CONFIRMED, Order.STATUS_PACKING]).count(), 'tone': 'warning'},
            {'label': 'Đang giao', 'value': base_queryset.filter(status=Order.STATUS_SHIPPING).count(), 'tone': 'info'},
            {'label': 'Hoàn thành', 'value': base_queryset.filter(status=Order.STATUS_COMPLETED).count(), 'tone': 'success'},
            {'label': 'Đã hủy / giao lỗi', 'value': base_queryset.filter(status__in=[Order.STATUS_CANCELLED, Order.STATUS_FAILED_DELIVERY]).count(), 'tone': 'danger'},
        ]
        filter_options = get_order_filter_options(request, pharmacy_queryset=pharmacy_queryset, selected_pharmacy_id=pharmacy_id)

        rows = []
        for obj in queryset:
            rows.append({
                'cells': [
                    format_html(
                        '<div class="cell-title">#{}</div><div class="cell-sub">{} • {} sản phẩm</div>',
                        obj.pk,
                        obj.created_at.strftime('%d/%m/%Y %H:%M'),
                        obj.item_total,
                    ),
                    format_html('<div>{}</div><div class="cell-sub">{}</div>', obj.full_name, obj.phone),
                    obj.pharmacy.name if obj.pharmacy else render_badge('Chưa gán', 'secondary'),
                    format_money(obj.final_total_price),
                    render_order_status_badge(obj.status),
                ],
                'actions': ([
                    {'url': reverse('custom_admin_order_detail', kwargs={'pk': obj.pk}), 'label': 'Xem chi tiết', 'icon': 'fas fa-eye', 'class': 'btn-info'},
                ] + ([
                    {'url': reverse('custom_admin_delete', kwargs={'model_key': 'order', 'pk': obj.pk}), 'label': 'Xóa', 'icon': 'fas fa-trash', 'class': 'btn-danger'},
                ] if can_delete_object(request.user, 'order', obj) else [])),
            })
        return queryset, columns, rows, filter_options, summary_cards, keyword

    if model_key == 'return_request':
        base_queryset = filter_queryset_by_admin_scope(
            ReturnRefundRequest.objects.select_related('order__pharmacy', 'order__user', 'processed_by').annotate(proof_total=Count('evidences')),
            request.user,
            'return_request',
        )
        queryset = base_queryset
        if keyword:
            queryset = queryset.filter(build_search_query(keyword, ADMIN_MODELS[model_key]['search_fields']))

        status = request.GET.get('status', '')
        if status:
            queryset = queryset.filter(status=status)

        pharmacy_queryset = Pharmacy.objects.order_by('name')
        if managed_pharmacy is not None:
            pharmacy_queryset = Pharmacy.objects.filter(pk=managed_pharmacy.pk)
        pharmacy_id = request.GET.get('pharmacy', '')
        if pharmacy_id:
            queryset = queryset.filter(order__pharmacy_id=pharmacy_id)

        queryset = apply_admin_sort(queryset, model_key, request.GET.get('sort', 'newest'))
        columns = ['Yêu cầu', 'Đơn hàng', 'Khách hàng', 'Hoàn tiền', 'Trạng thái']
        summary_cards = [
            {'label': 'Tổng yêu cầu', 'value': base_queryset.count(), 'tone': 'primary'},
            {'label': 'Đang xử lý', 'value': base_queryset.filter(status=ReturnRefundRequest.STATUS_PROCESSING).count(), 'tone': 'warning'},
            {'label': 'Chấp nhận hoàn tiền', 'value': base_queryset.filter(status=ReturnRefundRequest.STATUS_APPROVED).count(), 'tone': 'success'},
            {'label': 'Từ chối hoàn tiền', 'value': base_queryset.filter(status=ReturnRefundRequest.STATUS_REJECTED).count(), 'tone': 'danger'},
        ]
        filter_options = get_return_request_filter_options(request, pharmacy_queryset=pharmacy_queryset, selected_pharmacy_id=pharmacy_id)

        rows = []
        for obj in queryset:
            refund_target = obj.bank_account_number or obj.momo_account_number or '-'
            rows.append({
                'cells': [
                    format_html('<div class="cell-title">Yêu cầu #{}</div><div class="cell-sub">{} ảnh chứng minh</div>', obj.pk, obj.proof_total),
                    format_html('<div>{}</div><div class="cell-sub">{}</div>', obj.order.order_code, obj.order.pharmacy.name if obj.order.pharmacy else 'Chưa gán chi nhánh'),
                    format_html('<div>{}</div><div class="cell-sub">{}</div>', obj.order.full_name, obj.contact_phone or obj.order.phone),
                    format_html('<div>{}</div><div class="cell-sub">{}</div>', refund_target, obj.contact_email or '-'),
                    render_badge('Đang xử lý', 'warning') if obj.status == ReturnRefundRequest.STATUS_PROCESSING else (render_badge('Chấp nhận hoàn tiền', 'success') if obj.status == ReturnRefundRequest.STATUS_APPROVED else render_badge('Từ chối hoàn tiền', 'danger')),
                ],
                'actions': ([
                    {'url': reverse('custom_admin_return_request_detail', kwargs={'pk': obj.pk}), 'label': 'Xem chi tiết', 'icon': 'fas fa-eye', 'class': 'btn-info'},
                ] + ([
                    {'url': reverse('custom_admin_delete', kwargs={'model_key': 'return_request', 'pk': obj.pk}), 'label': 'Xóa', 'icon': 'fas fa-trash', 'class': 'btn-danger'},
                ] if can_delete_object(request.user, 'return_request', obj) else [])),
            })
        return queryset, columns, rows, filter_options, summary_cards, keyword

    if model_key == 'purchase_import':
        base_queryset = filter_queryset_by_admin_scope(
            PurchaseImportBatch.objects.select_related('pharmacy', 'imported_by').prefetch_related('items'),
            request.user,
            'purchase_import',
        )
        queryset = base_queryset
        if keyword:
            queryset = queryset.filter(build_search_query(keyword, ADMIN_MODELS[model_key]['search_fields']))

        pharmacy_queryset = Pharmacy.objects.order_by('name')
        if managed_pharmacy is not None:
            pharmacy_queryset = Pharmacy.objects.filter(pk=managed_pharmacy.pk)
        pharmacy_id = request.GET.get('pharmacy', '')
        if pharmacy_id:
            queryset = queryset.filter(pharmacy_id=pharmacy_id)

        imported_from = (request.GET.get('imported_from') or '').strip()
        if imported_from:
            queryset = queryset.filter(created_at__date__gte=imported_from)
        imported_to = (request.GET.get('imported_to') or '').strip()
        if imported_to:
            queryset = queryset.filter(created_at__date__lte=imported_to)

        queryset = apply_admin_sort(queryset, model_key, request.GET.get('sort', 'newest'))
        columns = ['Phiếu nhập', 'Chi nhánh', 'Người phụ trách', 'Tổng nhập', 'Tệp Excel']
        summary_cards = [
            {'label': 'Tổng phiếu nhập', 'value': base_queryset.count(), 'tone': 'primary'},
            {'label': 'Tổng số lượng nhập', 'value': base_queryset.aggregate(total=Sum('total_quantity')).get('total') or 0, 'tone': 'success'},
            {'label': 'Dòng nhập hợp lệ', 'value': base_queryset.aggregate(total=Sum('total_lines')).get('total') or 0, 'tone': 'info'},
        ]
        filter_options = get_purchase_import_filter_options(request, pharmacy_queryset=pharmacy_queryset, selected_pharmacy_id=pharmacy_id)

        rows = []
        for obj in queryset:
            rows.append({
                'cells': [
                    format_html('<div class="cell-title">{}</div><div class="cell-sub">{} • {} dòng</div>', obj.resolved_invoice_code, obj.created_at.strftime('%d/%m/%Y %H:%M'), obj.total_lines),
                    obj.pharmacy.name if obj.pharmacy else '-',
                    format_html('<div>{}</div><div class="cell-sub">{}</div>', obj.resolved_imported_by_name or 'Chưa cập nhật', obj.imported_by.email if obj.imported_by else '-'),
                    format_html('<div><strong>{}</strong></div><div class="cell-sub">{} đơn vị sản phẩm</div>', obj.total_quantity, obj.total_lines),
                    format_html('<a href="{}" target="_blank" rel="noopener">Mở file</a>', obj.source_file.url) if obj.source_file else '-',
                ],
                'actions': ([
                    {'url': reverse('custom_admin_purchase_import_detail', kwargs={'pk': obj.pk}), 'label': 'Xem chi tiết', 'icon': 'fas fa-eye', 'class': 'btn-info'},
                    {'url': reverse('custom_admin_purchase_import_receipt', kwargs={'pk': obj.pk}), 'label': 'Xuất phiếu', 'icon': 'fas fa-file-export', 'class': 'btn-success'},
                ] + ([
                    {'url': reverse('custom_admin_delete', kwargs={'model_key': 'purchase_import', 'pk': obj.pk}), 'label': 'Xóa', 'icon': 'fas fa-trash', 'class': 'btn-danger'},
                ] if can_delete_object(request.user, 'purchase_import', obj) else [])),
            })
        return queryset, columns, rows, filter_options, summary_cards, keyword

    if model_key == 'stock_export':
        base_queryset = filter_queryset_by_admin_scope(
            StockExportBatch.objects.select_related('pharmacy', 'exported_by').prefetch_related('items'),
            request.user,
            'stock_export',
        )
        queryset = base_queryset
        if keyword:
            queryset = queryset.filter(build_search_query(keyword, ADMIN_MODELS[model_key]['search_fields']))

        pharmacy_queryset = Pharmacy.objects.order_by('name')
        if managed_pharmacy is not None:
            pharmacy_queryset = Pharmacy.objects.filter(pk=managed_pharmacy.pk)
        pharmacy_id = request.GET.get('pharmacy', '')
        if pharmacy_id:
            queryset = queryset.filter(pharmacy_id=pharmacy_id)
        export_scope = (request.GET.get('export_scope') or '').strip()
        if export_scope in dict(StockExportBatch.EXPORT_SCOPE_CHOICES):
            queryset = queryset.filter(export_scope=export_scope)

        exported_from = (request.GET.get('exported_from') or '').strip()
        if exported_from:
            queryset = queryset.filter(created_at__date__gte=exported_from)
        exported_to = (request.GET.get('exported_to') or '').strip()
        if exported_to:
            queryset = queryset.filter(created_at__date__lte=exported_to)

        queryset = apply_admin_sort(queryset, model_key, request.GET.get('sort', 'newest'))
        columns = ['Phiếu xuất', 'Chi nhánh', 'Người lập phiếu', 'Loại phiếu', 'Mục đích xuất', 'Tổng xuất']
        summary_cards = [
            {'label': 'Tổng phiếu xuất', 'value': base_queryset.count(), 'tone': 'primary'},
            {'label': 'Tổng số lượng xuất', 'value': base_queryset.aggregate(total=Sum('total_quantity')).get('total') or 0, 'tone': 'warning'},
            {'label': 'Dòng xuất kho', 'value': base_queryset.aggregate(total=Sum('total_lines')).get('total') or 0, 'tone': 'info'},
        ]
        filter_options = get_stock_export_filter_options(request, pharmacy_queryset=pharmacy_queryset, selected_pharmacy_id=pharmacy_id)

        rows = []
        for obj in queryset:
            rows.append({
                'cells': [
                    format_html('<div class="cell-title">{}</div><div class="cell-sub">{} • {} dòng</div>', obj.resolved_export_code, obj.created_at.strftime('%d/%m/%Y %H:%M'), obj.total_lines),
                    obj.pharmacy.name if obj.pharmacy else '-',
                    format_html('<div>{}</div><div class="cell-sub">{}</div>', obj.resolved_exported_by_name or 'Chưa cập nhật', obj.exported_by_email or '-'),
                    render_badge(obj.get_export_scope_display(), 'danger' if obj.export_scope == StockExportBatch.EXPORT_SCOPE_EXPIRED else ('info' if obj.export_scope == StockExportBatch.EXPORT_SCOPE_RECONCILE else 'success')),
                    format_html('<div>{}</div><div class="cell-sub">{}</div>', obj.destination_name or 'Xuất nội bộ', obj.note or 'Không có ghi chú'),
                    format_html('<div><strong>{}</strong></div><div class="cell-sub">{} dòng sản phẩm</div>', obj.total_quantity, obj.total_lines),
                ],
                'actions': ([
                    {'url': reverse('custom_admin_stock_export_detail', kwargs={'pk': obj.pk}), 'label': 'Xem chi tiết', 'icon': 'fas fa-eye', 'class': 'btn-info'},
                    {'url': reverse('custom_admin_stock_export_receipt', kwargs={'pk': obj.pk}), 'label': 'Xuất phiếu', 'icon': 'fas fa-file-export', 'class': 'btn-success'},
                ] + ([
                    {'url': reverse('custom_admin_delete', kwargs={'model_key': 'stock_export', 'pk': obj.pk}), 'label': 'Xóa', 'icon': 'fas fa-trash', 'class': 'btn-danger'},
                ] if can_delete_object(request.user, 'stock_export', obj) else [])),
            })
        return queryset, columns, rows, filter_options, summary_cards, keyword

    if model_key == 'promotion':
        base_queryset = filter_queryset_by_admin_scope(
            MedicinePromotion.objects.select_related('medicine__pharmacy', 'created_by'),
            request.user,
            'promotion',
        )
        queryset = base_queryset
        if keyword:
            queryset = queryset.filter(build_search_query(keyword, ADMIN_MODELS[model_key]['search_fields']))

        pharmacy_queryset = Pharmacy.objects.order_by('name')
        if managed_pharmacy is not None:
            pharmacy_queryset = Pharmacy.objects.filter(pk=managed_pharmacy.pk)
        pharmacy_id = request.GET.get('pharmacy', '')
        if pharmacy_id:
            queryset = queryset.filter(medicine__pharmacy_id=pharmacy_id)

        today = timezone.localdate()
        warning_deadline = today + timedelta(days=183)
        promo_state = request.GET.get('promo_state', '')
        if promo_state == 'active':
            queryset = queryset.filter(is_active=True).filter(Q(start_date__isnull=True) | Q(start_date__lte=today), Q(end_date__isnull=True) | Q(end_date__gte=today))
        elif promo_state == 'upcoming':
            queryset = queryset.filter(is_active=True, start_date__gt=today)
        elif promo_state == 'ended':
            queryset = queryset.filter(end_date__isnull=False, end_date__lt=today)
        elif promo_state == 'inactive':
            queryset = queryset.filter(is_active=False)

        medicine_kind = request.GET.get('medicine_kind', '')
        if medicine_kind == 'expiring':
            queryset = queryset.filter(medicine__expiry_date__isnull=False, medicine__expiry_date__gte=today, medicine__expiry_date__lte=warning_deadline)
        elif medicine_kind == 'normal':
            queryset = queryset.exclude(medicine__expiry_date__isnull=False, medicine__expiry_date__gte=today, medicine__expiry_date__lte=warning_deadline)

        queryset = apply_admin_sort(queryset, model_key, request.GET.get('sort', 'newest'))
        columns = ['Chương trình', 'Sản phẩm', 'Chi nhánh', 'Mức giảm', 'Trạng thái']
        summary_cards = [
            {'label': 'Tổng khuyến mãi', 'value': base_queryset.count(), 'tone': 'primary'},
            {'label': 'Đang áp dụng', 'value': build_active_promotion_queryset(base_queryset).count(), 'tone': 'success'},
            {'label': 'Thuốc cận hạn ≤ 6 tháng', 'value': base_queryset.filter(medicine__expiry_date__isnull=False, medicine__expiry_date__gte=today, medicine__expiry_date__lte=warning_deadline).count(), 'tone': 'warning'},
            {'label': 'Sản phẩm thường', 'value': base_queryset.exclude(medicine__expiry_date__isnull=False, medicine__expiry_date__gte=today, medicine__expiry_date__lte=warning_deadline).count(), 'tone': 'info'},
        ]
        filter_options = get_promotion_filter_options(request, pharmacy_queryset=pharmacy_queryset, selected_pharmacy_id=pharmacy_id)
        rows = []
        for obj in queryset:
            if obj.is_currently_active:
                status_badge = render_badge('Đang áp dụng', 'success')
            elif not obj.is_active:
                status_badge = render_badge('Đã tắt', 'secondary')
            elif obj.start_date and obj.start_date > today:
                status_badge = render_badge('Sắp bắt đầu', 'info')
            else:
                status_badge = render_badge('Đã kết thúc', 'danger')
            if obj.medicine.expiry_date and today <= obj.medicine.expiry_date <= warning_deadline:
                medicine_kind_badge = render_badge('HSD ≤ 6 tháng', 'warning')
            elif obj.medicine.expiry_date and obj.medicine.expiry_date < today:
                medicine_kind_badge = render_badge('Đã hết hạn', 'danger')
            else:
                medicine_kind_badge = render_badge('Sản phẩm thường', 'info')
            rows.append({
                'cells': [
                    format_html('<div class="cell-title">{}</div><div class="cell-sub">{}</div>', obj.resolved_title, obj.note or 'Không có ghi chú'),
                    format_html('<div>{}</div><div class="cell-sub">{} · {} · HSD gần nhất: {}</div>', obj.medicine.name, obj.medicine.unit or 'Hộp', medicine_kind_badge, obj.medicine.expiry_date.strftime('%d/%m/%Y') if obj.medicine.expiry_date else 'Chưa cập nhật'),
                    'Toàn hệ thống',
                    format_html('<strong>{}%</strong><div class="cell-sub">{} → {}</div>', obj.discount_percent, format_vnd(obj.medicine.price), format_vnd(get_discounted_price(obj.medicine.price, obj.discount_percent))),
                    status_badge,
                ],
                'actions': ([{
                    'url': reverse('custom_admin_update', kwargs={'model_key': 'promotion', 'pk': obj.pk}),
                    'label': 'Cập nhật',
                    'icon': 'fas fa-pen',
                    'class': 'btn-primary',
                }] if can_update_admin_model(request.user, model_key) else []) + ([
                    {'url': reverse('custom_admin_delete', kwargs={'model_key': 'promotion', 'pk': obj.pk}), 'label': 'Xóa', 'icon': 'fas fa-trash', 'class': 'btn-danger'},
                ] if can_delete_object(request.user, 'promotion', obj) else []),
            })
        return queryset, columns, rows, filter_options, summary_cards, keyword

    if model_key == 'inventory_lot':
        base_queryset = filter_queryset_by_admin_scope(
            MedicineLot.objects.select_related('medicine__pharmacy', 'purchase_batch', 'purchase_item'),
            request.user,
            'inventory_lot',
        )
        queryset = base_queryset
        if keyword:
            queryset = queryset.filter(build_search_query(keyword, ADMIN_MODELS[model_key]['search_fields']))
        pharmacy_queryset = Pharmacy.objects.order_by('name')
        if managed_pharmacy is not None:
            pharmacy_queryset = Pharmacy.objects.filter(pk=managed_pharmacy.pk)
        pharmacy_id = request.GET.get('pharmacy', '')
        if pharmacy_id:
            queryset = queryset.filter(pharmacy_id=pharmacy_id)
        today = timezone.localdate()
        expiry_state = (request.GET.get('expiry_state', '') or 'sellable').strip()
        if expiry_state == 'sellable':
            queryset = queryset.filter(remaining_quantity__gt=0).filter(Q(expiry_date__isnull=True) | Q(expiry_date__gte=today + timedelta(days=184)))
        elif expiry_state == 'warning':
            queryset = queryset.filter(expiry_date__isnull=False, expiry_date__gte=today, expiry_date__lte=today + timedelta(days=183), remaining_quantity__gt=0)
        elif expiry_state == 'expired':
            queryset = queryset.filter(expiry_date__isnull=False, expiry_date__lt=today)
        lot_state = request.GET.get('lot_state', '')
        if lot_state == 'available':
            queryset = queryset.filter(remaining_quantity__gt=0)
        elif lot_state == 'empty':
            queryset = queryset.filter(remaining_quantity__lte=0)
        queryset = apply_admin_sort(queryset, model_key, request.GET.get('sort', 'expiry_soon'))
        columns = ['Lô', 'Sản phẩm', 'Chi nhánh', 'Hạn sử dụng', 'Tồn lô', 'Trạng thái']
        summary_cards = [
            {'label': 'Tổng lô', 'value': base_queryset.count(), 'tone': 'primary'},
            {'label': 'Lô còn tồn', 'value': base_queryset.filter(remaining_quantity__gt=0).count(), 'tone': 'success'},
            {'label': 'Lô HSD ≤ 6 tháng', 'value': base_queryset.filter(expiry_date__isnull=False, expiry_date__gte=today, expiry_date__lte=today + timedelta(days=183), remaining_quantity__gt=0).count(), 'tone': 'warning'},
            {'label': 'Lô hết hạn', 'value': base_queryset.filter(expiry_date__isnull=False, expiry_date__lt=today).count(), 'tone': 'danger'},
        ]
        filter_options = get_inventory_lot_filter_options(request, pharmacy_queryset=pharmacy_queryset, selected_pharmacy_id=pharmacy_id)
        rows = []
        for obj in queryset:
            if obj.remaining_quantity <= 0:
                lot_badge = render_badge('Đã xuất hết', 'secondary')
            elif obj.expiry_date and obj.expiry_date < today:
                lot_badge = render_badge('Đã hết hạn', 'danger')
            elif obj.expiry_date and obj.expiry_date <= today + timedelta(days=183):
                lot_badge = render_badge('HSD ≤ 6 tháng', 'warning')
            else:
                lot_badge = render_badge('Đang bán được', 'success')
            action_list = []
            if can_update_admin_model(request.user, 'medicine'):
                action_list.append({'url': reverse('custom_admin_update', kwargs={'model_key': 'medicine', 'pk': obj.medicine_id}), 'label': 'Mở thuốc', 'icon': 'fas fa-pills', 'class': 'btn-info'})
            if obj.purchase_batch_id:
                action_list.append({'url': reverse('custom_admin_purchase_import_detail', kwargs={'pk': obj.purchase_batch_id}), 'label': 'Phiếu nhập', 'icon': 'fas fa-file-alt', 'class': 'btn-primary'})
            if obj.pharmacy_id and obj.remaining_quantity > 0 and obj.expiry_date and can_create_admin_model(request.user, 'stock_export'):
                export_scope = StockExportBatch.EXPORT_SCOPE_EXPIRED if obj.expiry_date < today else StockExportBatch.EXPORT_SCOPE_RECONCILE
                action_list.append({
                    'url': f"{reverse('custom_admin_create', kwargs={'model_key': 'stock_export'})}?pharmacy={obj.pharmacy_id}&export_scope={export_scope}",
                    'label': 'Đi xử lý',
                    'icon': 'fas fa-file-export',
                    'class': 'btn-warning',
                })
            rows.append({
                'cells': [
                    format_html('<div class="cell-title">{}</div><div class="cell-sub">{} • {}</div>', obj.source_label or f'Lô #{obj.pk}', obj.get_source_type_display(), obj.created_at.strftime('%d/%m/%Y %H:%M')),
                    format_html('<div>{}</div><div class="cell-sub">NSX: {}</div>', obj.medicine.name if obj.medicine else '-', obj.medicine.manufacturer if obj.medicine else '-'),
                    obj.pharmacy.name if obj.pharmacy else '-',
                    obj.expiry_date.strftime('%d/%m/%Y') if obj.expiry_date else 'Chưa có',
                    format_html('<div><strong>{}</strong></div><div class="cell-sub">Nhập {} • còn {}</div>', obj.remaining_quantity, obj.received_quantity, obj.remaining_quantity),
                    lot_badge,
                ],
                'actions': action_list,
            })
        return queryset, columns, rows, filter_options, summary_cards, keyword

    if model_key == 'news':
        base_queryset = NewsArticle.objects.select_related('created_by', 'updated_by')
        queryset = base_queryset
        if keyword:
            queryset = queryset.filter(build_search_query(keyword, ADMIN_MODELS[model_key]['search_fields']))

        publish_state = request.GET.get('publish_state', '')
        if publish_state == 'published':
            queryset = queryset.filter(is_published=True)
        elif publish_state == 'draft':
            queryset = queryset.filter(is_published=False)

        published_from = (request.GET.get('published_from') or '').strip()
        if published_from:
            queryset = queryset.filter(published_at__date__gte=published_from)
        published_to = (request.GET.get('published_to') or '').strip()
        if published_to:
            queryset = queryset.filter(published_at__date__lte=published_to)

        queryset = apply_admin_sort(queryset, model_key, request.GET.get('sort', 'newest'))
        columns = ['Ảnh', 'Bài viết', 'Người cập nhật', 'Xuất bản', 'Trạng thái']
        summary_cards = [
            {'label': 'Tổng bài viết', 'value': base_queryset.count(), 'tone': 'primary'},
            {'label': 'Đã xuất bản', 'value': base_queryset.filter(is_published=True).count(), 'tone': 'success'},
            {'label': 'Nháp', 'value': base_queryset.filter(is_published=False).count(), 'tone': 'warning'},
            {'label': 'Có ảnh đại diện', 'value': base_queryset.exclude(cover_image='').count(), 'tone': 'info'},
        ]
        filter_options = get_news_filter_options(request)
        rows = []
        for obj in queryset:
            published_label = obj.published_at.strftime('%d/%m/%Y %H:%M') if obj.published_at else 'Chưa hẹn giờ'
            rows.append({
                'cells': [
                    render_image_thumb(obj.cover_image, obj.title, 'Ảnh bài viết'),
                    format_html('<div class="cell-title">{}</div><div class="cell-sub">/{}</div>', obj.title, obj.slug),
                    format_html('<div>{}</div><div class="cell-sub">{}</div>', obj.author_display_name, obj.updated_at.strftime('%d/%m/%Y %H:%M')),
                    published_label,
                    render_badge('Đã xuất bản', 'success') if obj.is_published else render_badge('Nháp', 'warning'),
                ],
                'actions': (
                    ([{
                        'url': reverse('news_detail', kwargs={'slug': obj.slug}),
                        'label': 'Xem ngoài site',
                        'icon': 'fas fa-external-link-alt',
                        'class': 'btn-info',
                    }] if obj.is_published else [])
                    + ([{
                        'url': reverse('custom_admin_update', kwargs={'model_key': 'news', 'pk': obj.pk}),
                        'label': 'Cập nhật',
                        'icon': 'fas fa-pen',
                        'class': 'btn-primary',
                    }] if can_update_admin_model(request.user, 'news') else [])
                    + ([{
                        'url': reverse('custom_admin_delete', kwargs={'model_key': 'news', 'pk': obj.pk}),
                        'label': 'Xóa',
                        'icon': 'fas fa-trash',
                        'class': 'btn-danger',
                    }] if can_delete_object(request.user, 'news', obj) else [])
                ),
            })
        return queryset, columns, rows, filter_options, summary_cards, keyword

    queryset = User.objects.all()
    if keyword:
        queryset = queryset.filter(build_search_query(keyword, ADMIN_MODELS[model_key]['search_fields']))

    role = request.GET.get('role', '')
    if role == 'customer':
        queryset = queryset.filter(is_staff=False, is_superuser=False)
    elif role == 'staff':
        queryset = queryset.filter(is_staff=True, is_superuser=False)
    elif role == 'superuser':
        queryset = queryset.filter(is_superuser=True)

    active = request.GET.get('active', '')
    if active == 'yes':
        queryset = queryset.filter(is_active=True)
    elif active == 'no':
        queryset = queryset.filter(is_active=False)

    queryset = apply_admin_sort(queryset, model_key, request.GET.get('sort', 'newest'))
    columns = ['Tài khoản', 'Thông tin liên hệ', 'Vai trò', 'Trạng thái']
    summary_cards = [
        {'label': 'Tổng tài khoản', 'value': User.objects.count(), 'tone': 'primary'},
        {'label': 'Nhân viên', 'value': User.objects.filter(is_staff=True, is_superuser=False).count(), 'tone': 'info'},
        {'label': 'Khách hàng', 'value': User.objects.filter(is_staff=False, is_superuser=False).count(), 'tone': 'secondary'},
        {'label': 'Bị khóa', 'value': User.objects.filter(is_active=False).count(), 'tone': 'danger'},
    ]
    filter_options = get_user_filter_options(request)
    profile_map = {
        profile.user_id: profile
        for profile in UserProfile.objects.select_related('managed_pharmacy').filter(user__in=queryset)
    }

    rows = []
    for obj in queryset:
        profile = profile_map.get(obj.pk)
        managed_branch_label = profile.managed_pharmacy.name if profile and profile.managed_pharmacy else 'Chưa gán chi nhánh'
        actions = []
        if can_update_admin_model(request.user, model_key) and (not obj.is_superuser or request.user.is_superuser):
            actions.append({'url': reverse('custom_admin_update', kwargs={'model_key': 'user', 'pk': obj.pk}), 'label': 'Cập nhật', 'icon': 'fas fa-pen', 'class': 'btn-primary'})
        if can_delete_object(request.user, model_key, obj):
            actions.append({'url': reverse('custom_admin_delete', kwargs={'model_key': 'user', 'pk': obj.pk}), 'label': 'Xóa', 'icon': 'fas fa-trash', 'class': 'btn-danger'})
        rows.append({
            'cells': [
                format_html('<div class="cell-title">{}</div><div class="cell-sub">ID: #{}</div>', obj.username, obj.pk),
                format_html(
                    '<div>{}</div><div class="cell-sub">{} {}{}</div>',
                    obj.email or '-',
                    obj.last_name or '',
                    obj.first_name or '',
                    format_html(' • {}', managed_branch_label) if obj.is_staff and not obj.is_superuser else '',
                ),
                render_user_role_badge(obj),
                render_badge('Đang hoạt động', 'success') if obj.is_active else render_badge('Đã khóa', 'danger'),
            ],
            'actions': actions,
        })
    return queryset, columns, rows, filter_options, summary_cards, keyword


@admin_panel_required
def custom_admin_dashboard(request):
    denied_response = require_admin_model_access(request, 'dashboard')
    if denied_response:
        return denied_response

    try:
        managed_pharmacy = get_admin_scope_pharmacy(request.user)
        if request.user.is_staff and not request.user.is_superuser and managed_pharmacy is None:
            raise PermissionDenied('Tài khoản nhân viên chưa được gán chi nhánh quản lý. Vui lòng liên hệ quản trị viên.')

        orders_base = filter_queryset_by_admin_scope(Order.objects.select_related('pharmacy'), request.user, 'order')
        medicines_base = filter_queryset_by_admin_scope(Medicine.objects.select_related('pharmacy'), request.user, 'medicine')

        recent_orders_queryset = orders_base.order_by('-created_at', '-id')
        low_stock_queryset = medicines_base.filter(quantity__gt=0, quantity__lte=LOW_STOCK_THRESHOLD).order_by('quantity', 'name')
        expiring_soon_queryset = get_expiring_soon_medicines_queryset(medicines_base).order_by('expiry_date', 'name')
        expired_queryset = get_expired_medicines_queryset(medicines_base).order_by('expiry_date', 'name')

        if managed_pharmacy is not None:
            branch_overview_queryset = Pharmacy.objects.filter(pk=managed_pharmacy.pk).annotate(
                medicine_total=Count('medicines', distinct=True),
                available_total=Count('medicines', filter=Q(medicines__quantity__gt=0), distinct=True),
            ).order_by('name')
            pharmacy_count = 1
        else:
            branch_overview_queryset = Pharmacy.objects.annotate(
                medicine_total=Count('medicines', distinct=True),
                available_total=Count('medicines', filter=Q(medicines__quantity__gt=0), distinct=True),
            ).order_by('-available_total', 'name')
            pharmacy_count = Pharmacy.objects.count()

        recent_orders_page_obj = Paginator(recent_orders_queryset, ADMIN_PAGE_SIZE).get_page(request.GET.get('orders_page'))
        low_stock_page_obj = Paginator(low_stock_queryset, 5).get_page(request.GET.get('stock_page'))
        expiring_soon_page_obj = Paginator(expiring_soon_queryset, 5).get_page(request.GET.get('expiry_page'))
        expired_page_obj = Paginator(expired_queryset, 5).get_page(request.GET.get('expired_page'))
        branch_overview_page_obj = Paginator(branch_overview_queryset, ADMIN_PAGE_SIZE).get_page(request.GET.get('branch_page'))

        context = {
            'page_title': 'Dashboard quản trị hệ thống',
            'current_model': 'dashboard',
            'pharmacy_count': pharmacy_count,
            'medicine_count': medicines_base.count(),
            'pending_order_count': orders_base.filter(status__in=[Order.STATUS_PENDING, Order.STATUS_CONFIRMED, Order.STATUS_PACKING]).count(),
            'shipping_order_count': orders_base.filter(status=Order.STATUS_SHIPPING).count(),
            'low_stock_count': medicines_base.filter(quantity__gt=0, quantity__lte=LOW_STOCK_THRESHOLD).count(),
            'out_of_stock_count': medicines_base.filter(quantity__lte=0).count(),
            'expiring_soon_count': expiring_soon_queryset.count(),
            'expired_medicine_count': expired_queryset.count(),
            'recent_orders': recent_orders_page_obj.object_list,
            'recent_orders_page_obj': recent_orders_page_obj,
            'low_stock_medicines': low_stock_page_obj.object_list,
            'low_stock_page_obj': low_stock_page_obj,
            'expiring_soon_medicines': expiring_soon_page_obj.object_list,
            'expiring_soon_page_obj': expiring_soon_page_obj,
            'expired_medicines': expired_page_obj.object_list,
            'expired_page_obj': expired_page_obj,
            'branch_overview': branch_overview_page_obj.object_list,
            'branch_overview_page_obj': branch_overview_page_obj,
            'can_manage_pharmacy': can_access_admin_model(request.user, 'pharmacy'),
            'can_manage_medicine': can_access_admin_model(request.user, 'medicine'),
            'can_manage_order': can_access_admin_model(request.user, 'order'),
            'can_manage_user': can_access_admin_model(request.user, 'user'),
            'can_manage_purchase_import': can_access_admin_model(request.user, 'purchase_import'),
            'can_create_pharmacy': request.user.is_superuser,
            'can_create_medicine': request.user.is_staff,
            'can_create_user': request.user.is_superuser,
            'managed_pharmacy': managed_pharmacy,
        }
        return render(request, 'admin_panel/dashboard.html', context)
    except PermissionDenied:
        raise
    except Exception:
        logger.exception('Admin dashboard fallback activated for user %s', getattr(request.user, 'pk', None))
        messages.warning(
            request,
            'Trang tổng quan đang được chuyển sang chế độ an toàn. Bạn vẫn có thể tiếp tục quản lý đơn hàng, trả hàng / hoàn tiền, nhập hàng và sản phẩm ngay bên dưới.'
        )
        return custom_admin_list(request, 'order')


@admin_panel_required
def custom_admin_list(request, model_key):
    denied_response = require_admin_model_access(request, model_key)
    if denied_response:
        return denied_response

    config = get_admin_config(model_key)
    queryset, columns, rows, filter_options, summary_cards, keyword = build_list_data(model_key, request)

    paginator = Paginator(queryset, ADMIN_PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get('page'))
    page_rows = rows[(page_obj.start_index() - 1):page_obj.end_index()] if page_obj.paginator.count else []

    create_allowed = (
        model_key in {'pharmacy', 'medicine', 'user', 'purchase_import', 'stock_export', 'promotion', 'news'}
        and can_create_admin_model(request.user, model_key)
    )
    create_label = {
        'pharmacy': 'Thêm chi nhánh',
        'medicine': 'Thêm sản phẩm thuốc',
        'user': 'Thêm tài khoản',
        'purchase_import': 'Nhập hàng bằng Excel',
        'stock_export': 'Tạo phiếu xuất kho',
        'promotion': 'Tạo khuyến mãi',
        'news': 'Thêm bài viết',
    }.get(model_key, '')

    context = {
        'page_title': f'Quản lý {config["title_plural"]}',
        'title': config['title_plural'],
        'model_key': model_key,
        'current_model': model_key,
        'columns': columns,
        'rows': page_rows,
        'page_obj': page_obj,
        'keyword': keyword,
        'filter_options': filter_options,
        'summary_cards': summary_cards,
        'create_allowed': create_allowed,
        'create_label': create_label,
    }
    if model_key == 'news':
        context.update(
            {
                'news_cards': [
                    {
                        'article': obj,
                        'preview_url': reverse('news_detail', kwargs={'slug': obj.slug}) if obj.is_published else '',
                        'edit_url': (
                            reverse('custom_admin_update', kwargs={'model_key': 'news', 'pk': obj.pk})
                            if can_update_admin_model(request.user, 'news')
                            else ''
                        ),
                        'delete_url': (
                            reverse('custom_admin_delete', kwargs={'model_key': 'news', 'pk': obj.pk})
                            if can_delete_object(request.user, 'news', obj)
                            else ''
                        ),
                        'status_label': 'Đã xuất bản' if obj.is_published else 'Nháp',
                        'status_tone': 'success' if obj.is_published else 'warning',
                    }
                    for obj in page_obj.object_list
                ],
                'news_filter_state': {
                    item['name']: item.get('value', '')
                    for item in filter_options
                },
                'news_sort_options': next(
                    (item.get('options', []) for item in filter_options if item.get('name') == 'sort'),
                    [],
                ),
            }
        )
        return render(request, 'admin_panel/content/news_list.html', context)
    if model_key == 'inventory_lot':
        context['inventory_alert_center'] = build_inventory_alert_center(
            filter_queryset_by_admin_scope(
                MedicineLot.objects.select_related('medicine', 'pharmacy'),
                request.user,
                'inventory_lot',
            ),
            request=request,
        )
    return render(request, 'admin_panel/shared/list.html', context)


@admin_panel_required
def custom_admin_home_page(request):
    denied_response = require_admin_model_access(request, 'home_page')
    if denied_response:
        return denied_response

    home_content = HomePageContent.get_solo()
    ensure_home_page_defaults(home_content)
    can_update_home_page = user_has_admin_permission(request.user, 'home_page', 'update')
    quick_link_choices = build_admin_quick_link_choices()

    if request.method == 'POST':
        if not can_update_home_page:
            raise PermissionDenied('Tài khoản hiện tại không có quyền cập nhật trang chủ.')
        form = HomePageContentForm(request.POST, instance=home_content)
        hero_slide_formset = HomeHeroSlideFormSet(
            request.POST,
            request.FILES,
            instance=home_content,
            prefix='hero_slides',
            form_kwargs={'link_choices': quick_link_choices},
        )
        category_item_formset = HomeCategorySpotlightItemFormSet(
            request.POST,
            instance=home_content,
            prefix='category_items',
            form_kwargs={'link_choices': quick_link_choices},
        )
        commitment_item_formset = HomeServiceCommitmentItemFormSet(
            request.POST,
            instance=home_content,
            prefix='commitment_items',
        )
        if (
            form.is_valid()
            and hero_slide_formset.is_valid()
            and category_item_formset.is_valid()
            and commitment_item_formset.is_valid()
        ):
            form.save()
            hero_slide_formset.save()
            category_item_formset.save()
            commitment_item_formset.save()
            messages.success(request, 'Đã cập nhật nội dung trang chủ.')
            return redirect('custom_admin_home_page')
    else:
        form = HomePageContentForm(instance=home_content)
        hero_slide_formset = HomeHeroSlideFormSet(
            instance=home_content,
            prefix='hero_slides',
            form_kwargs={'link_choices': quick_link_choices},
        )
        category_item_formset = HomeCategorySpotlightItemFormSet(
            instance=home_content,
            prefix='category_items',
            form_kwargs={'link_choices': quick_link_choices},
        )
        commitment_item_formset = HomeServiceCommitmentItemFormSet(
            instance=home_content,
            prefix='commitment_items',
        )

    context = {
        'page_title': 'Quản lý trang chủ',
        'current_model': 'home_page',
        'form': form,
        'field_groups': build_singleton_field_groups(form, HomePageContentForm.FIELD_GROUPS),
        'home_content': home_content,
        'hero_slide_formset': hero_slide_formset,
        'category_item_formset': category_item_formset,
        'commitment_item_formset': commitment_item_formset,
        'can_update_home_page': can_update_home_page,
    }
    return render(request, 'admin_panel/content/home_page_form.html', context)


@admin_panel_required
def custom_admin_permissions_center(request):
    denied_response = require_admin_model_access(request, 'permission')
    if denied_response:
        return denied_response

    if request.method == 'POST' and not user_has_admin_permission(request.user, 'permission', 'update'):
        raise PermissionDenied('Tài khoản hiện tại không có quyền cập nhật phân quyền.')

    user_queryset = User.objects.filter(is_staff=True).order_by('-is_superuser', 'username')
    selected_user_id = (request.POST.get('user_id') or request.GET.get('user') or '').strip()
    selected_user = None
    if selected_user_id.isdigit():
        selected_user = user_queryset.filter(pk=int(selected_user_id)).first()
    if selected_user is None:
        selected_user = user_queryset.first()

    if request.method == 'POST':
        if selected_user is None:
            messages.error(request, 'Không tìm thấy tài khoản cần phân quyền.')
            return redirect('custom_admin_permissions_center')
        if selected_user.is_superuser:
            messages.info(request, 'Tài khoản quản trị viên cấp cao luôn có toàn quyền và không cần cấu hình chi tiết.')
            return redirect(f"{reverse('custom_admin_permissions_center')}?user={selected_user.pk}")

        target_profile = get_or_create_user_profile(selected_user)
        target_profile.admin_permissions = extract_admin_permissions_from_request(request, target_user=selected_user)
        target_profile.save(update_fields=['admin_permissions', 'updated_at'])
        messages.success(request, f"Đã cập nhật phân quyền cho tài khoản '{selected_user.username}'.")
        return redirect(f"{reverse('custom_admin_permissions_center')}?user={selected_user.pk}")

    permission_sections = build_admin_permission_sections()
    selected_permissions = get_user_admin_permissions(selected_user) if selected_user else {}
    permission_matrix_rows = build_permission_matrix_rows(permission_sections, selected_permissions)
    profile_map = {
        profile.user_id: profile
        for profile in UserProfile.objects.select_related('managed_pharmacy').filter(user__in=user_queryset)
    }
    permission_user_cards = []
    for candidate in user_queryset:
        profile = profile_map.get(candidate.pk)
        candidate_permissions = get_user_admin_permissions(candidate)
        enabled_count = sum(
            1
            for item in ADMIN_PERMISSION_DEFINITIONS
            for action in item['actions']
            if candidate_permissions.get(item['key'], {}).get(action)
        )
        permission_user_cards.append(
            {
                'user': candidate,
                'profile': profile,
                'role_label': get_user_role_label(candidate),
                'managed_branch_label': profile.managed_pharmacy.name if profile and profile.managed_pharmacy else 'Chưa gán chi nhánh',
                'enabled_count': enabled_count,
            }
        )

    selected_enabled_count = sum(
        1
        for row in permission_matrix_rows
        for action in row['actions']
        if action['supported'] and action['checked']
    )
    selected_module_count = sum(
        1
        for row in permission_matrix_rows
        if any(action['supported'] and action['checked'] for action in row['actions'])
    )

    for row in permission_matrix_rows:
        row['enabled_count'] = sum(
            1 for action in row['actions'] if action['supported'] and action['checked']
        )
    row_map = {row['key']: row for row in permission_matrix_rows}

    PERMISSION_GROUP_DEFINITIONS = [
        {
            'label': 'Nội dung website',
            'description': 'Trang chủ, giới thiệu và tin tức hiển thị ngoài site.',
            'icon': 'fas fa-palette',
            'keys': ['home_page', 'about_page', 'news'],
        },
        {
            'label': 'Vận hành hàng ngày',
            'description': 'Tổng quan, đơn hàng, đổi trả - công việc thường xuyên.',
            'icon': 'fas fa-tasks',
            'keys': ['dashboard', 'order', 'return_request'],
        },
        {
            'label': 'Kho & sản phẩm',
            'description': 'Danh mục thuốc, nhập xuất, tồn kho, khuyến mãi và báo cáo.',
            'icon': 'fas fa-boxes',
            'keys': ['medicine', 'purchase_import', 'stock_export', 'inventory_lot', 'promotion', 'reports'],
        },
        {
            'label': 'Hệ thống & tài khoản',
            'description': 'Chi nhánh, tài khoản người dùng và quản lý phân quyền.',
            'icon': 'fas fa-shield-alt',
            'keys': ['pharmacy', 'user', 'permission'],
        },
    ]
    grouped_keys = {key for group in PERMISSION_GROUP_DEFINITIONS for key in group['keys']}
    permission_groups = []
    for group in PERMISSION_GROUP_DEFINITIONS:
        rows = [row_map[key] for key in group['keys'] if key in row_map]
        if rows:
            permission_groups.append({**group, 'rows': rows})
    leftover_rows = [row for row in permission_matrix_rows if row['key'] not in grouped_keys]
    if leftover_rows:
        permission_groups.append({
            'label': 'Khác',
            'description': 'Các quyền chưa được phân nhóm.',
            'icon': 'fas fa-ellipsis-h',
            'rows': leftover_rows,
        })

    context = {
        'page_title': 'Phân quyền tài khoản quản trị',
        'current_model': 'permission',
        'permission_sections': permission_sections,
        'permission_matrix_rows': permission_matrix_rows,
        'permission_groups': permission_groups,
        'permission_user_cards': permission_user_cards,
        'selected_permission_user': selected_user,
        'selected_permissions': selected_permissions,
        'selected_profile': profile_map.get(selected_user.pk) if selected_user else None,
        'selected_enabled_count': selected_enabled_count,
        'selected_module_count': selected_module_count,
        'can_update_permissions': user_has_admin_permission(request.user, 'permission', 'update'),
    }
    return render(request, 'admin_panel/system/permissions.html', context)


@admin_panel_required
def custom_admin_about_page(request):
    denied_response = require_admin_model_access(request, 'about_page')
    if denied_response:
        return denied_response

    about_content = AboutPageContent.get_solo()
    ensure_about_page_slide_defaults(about_content)
    ensure_about_featured_branch_defaults(about_content)
    can_update_about_page = user_has_admin_permission(request.user, 'about_page', 'update')
    builtin_sections_submitted = was_formset_submitted(request, 'builtin_sections')
    slides_submitted = was_formset_submitted(request, 'slides')
    featured_branches_submitted = was_formset_submitted(request, 'featured_branches')
    custom_blocks_submitted = was_formset_submitted(request, 'custom_blocks')
    quick_link_choices = build_admin_quick_link_choices()

    if request.method == 'POST':
        if not can_update_about_page:
            raise PermissionDenied('Tài khoản hiện tại không có quyền cập nhật trang giới thiệu.')
        form = AboutPageContentForm(request.POST, instance=about_content)
        if builtin_sections_submitted:
            builtin_section_formset = AboutBuiltinSectionFormSet(
                request.POST,
                request.FILES,
                instance=about_content,
                prefix='builtin_sections',
            )
            builtin_section_formset_valid = builtin_section_formset.is_valid()
        else:
            builtin_section_formset = AboutBuiltinSectionFormSet(
                instance=about_content,
                prefix='builtin_sections',
            )
            builtin_section_formset_valid = True
        if slides_submitted:
            slide_formset = AboutPageSlideFormSet(
                request.POST,
                request.FILES,
                instance=about_content,
                prefix='slides',
                form_kwargs={'link_choices': quick_link_choices},
            )
            slide_formset_valid = slide_formset.is_valid()
        else:
            slide_formset = AboutPageSlideFormSet(
                instance=about_content,
                prefix='slides',
                form_kwargs={'link_choices': quick_link_choices},
            )
            slide_formset_valid = True
        if featured_branches_submitted:
            featured_branch_formset = AboutFeaturedBranchItemFormSet(
                request.POST,
                request.FILES,
                instance=about_content,
                prefix='featured_branches',
                form_kwargs={'link_choices': quick_link_choices},
            )
            featured_branch_formset_valid = featured_branch_formset.is_valid()
        else:
            featured_branch_formset = AboutFeaturedBranchItemFormSet(
                instance=about_content,
                prefix='featured_branches',
                form_kwargs={'link_choices': quick_link_choices},
            )
            featured_branch_formset_valid = True
        if custom_blocks_submitted:
            custom_block_formset = AboutCustomBlockFormSet(
                request.POST,
                request.FILES,
                instance=about_content,
                prefix='custom_blocks',
                form_kwargs={'link_choices': quick_link_choices},
            )
            custom_block_formset_valid = custom_block_formset.is_valid()
        else:
            custom_block_formset = AboutCustomBlockFormSet(
                instance=about_content,
                prefix='custom_blocks',
                form_kwargs={'link_choices': quick_link_choices},
            )
            custom_block_formset_valid = True
        if (
            form.is_valid()
            and builtin_section_formset_valid
            and slide_formset_valid
            and featured_branch_formset_valid
            and custom_block_formset_valid
        ):
            form.save()
            if builtin_sections_submitted:
                builtin_section_formset.save()
            if slides_submitted:
                slide_formset.save()
            if featured_branches_submitted:
                featured_branch_formset.save()
            if custom_blocks_submitted:
                custom_block_formset.save()
            messages.success(request, 'Đã cập nhật nội dung trang Giới thiệu.')
            return redirect('custom_admin_about_page')
    else:
        form = AboutPageContentForm(instance=about_content)
        builtin_section_formset = AboutBuiltinSectionFormSet(
            instance=about_content,
            prefix='builtin_sections',
        )
        slide_formset = AboutPageSlideFormSet(
            instance=about_content,
            prefix='slides',
            form_kwargs={'link_choices': quick_link_choices},
        )
        featured_branch_formset = AboutFeaturedBranchItemFormSet(
            instance=about_content,
            prefix='featured_branches',
            form_kwargs={'link_choices': quick_link_choices},
        )
        custom_block_formset = AboutCustomBlockFormSet(
            instance=about_content,
            prefix='custom_blocks',
            form_kwargs={'link_choices': quick_link_choices},
        )

    about_field_groups = build_singleton_field_groups(form, AboutPageContentForm.FIELD_GROUPS)
    hero_field_group = about_field_groups[0] if about_field_groups else None
    remaining_field_groups = about_field_groups[1:] if len(about_field_groups) > 1 else []

    context = {
        'page_title': 'Quản lý trang Giới thiệu',
        'current_model': 'about_page',
        'form': form,
        'field_groups': remaining_field_groups,
        'hero_field_group': hero_field_group,
        'about_content': about_content,
        'builtin_section_formset': builtin_section_formset,
        'slide_formset': slide_formset,
        'featured_branch_formset': featured_branch_formset,
        'custom_block_formset': custom_block_formset,
        'can_update_about_page': can_update_about_page,
        'has_rich_editors': form_has_rich_editors(form),
    }
    return render(request, 'admin_panel/content/about_page_form.html', context)


@admin_panel_required
def custom_admin_create(request, model_key):
    denied_response = require_admin_model_access(request, model_key)
    if denied_response:
        return denied_response
    if not can_create_admin_model(request.user, model_key):
        raise PermissionDenied('Tài khoản hiện tại không có quyền tạo dữ liệu ở chức năng quản trị này.')

    if model_key == 'purchase_import':
        if request.method == 'POST':
            form = PurchaseImportExcelForm(request.POST, request.FILES, admin_user=request.user)
            if form.is_valid():
                try:
                    created_batches = process_purchase_import_excel(form=form, admin_user=request.user)
                except (RuntimeError, ValueError) as exc:
                    form.add_error(None, str(exc))
                else:
                    if len(created_batches) == 1:
                        batch = created_batches[0]
                        messages.success(request, f"Đã nhập hàng thành công cho phiếu {batch.resolved_invoice_code}.")
                        return redirect('custom_admin_purchase_import_detail', pk=batch.pk)
                    messages.success(
                        request,
                        f"Đã nhập hàng đồng loạt cho {len(created_batches)} chi nhánh. Hệ thống đã tạo {len(created_batches)} phiếu nhập riêng để dễ đối soát."
                    )
                    return redirect('custom_admin_list', model_key='purchase_import')
        else:
            form = PurchaseImportExcelForm(admin_user=request.user)

        context = {
            'page_title': 'Nhập hàng bằng Excel',
            'current_model': 'purchase_import',
            'form': form,
            'sample_columns': ['medicine_id', 'medicine_name', 'manufacturer', 'unit', 'quantity', 'expiry_date', 'import_price', 'sale_price', 'category', 'origin', 'note'],
            'purchase_import_preview_payload': build_purchase_import_preview_payload(request.user),
        }
        return render(request, 'admin_panel/inventory/purchase_import_form.html', context)

    if model_key == 'stock_export':
        if request.method == 'POST':
            batch_form = StockExportBatchForm(request.POST, admin_user=request.user)
            selected_pharmacy = get_requested_stock_export_pharmacy(request, batch_form=batch_form)
            selected_export_scope = get_requested_stock_export_scope(request, batch_form=batch_form)
            item_formset = StockExportItemFormSet(
                request.POST,
                prefix='items',
                form_kwargs={
                    'pharmacy': selected_pharmacy,
                    'admin_user': request.user,
                    'allocation_mode': selected_export_scope,
                },
            )
            if batch_form.is_valid() and item_formset.is_valid():
                try:
                    with transaction.atomic():
                        batch = create_stock_export_batch(
                            batch_form=batch_form,
                            item_formset=item_formset,
                            admin_user=request.user,
                        )
                except ValueError as exc:
                    batch_form.add_error(None, str(exc))
                else:
                    messages.success(request, f"Đã tạo phiếu xuất kho {batch.resolved_export_code}.")
                    return redirect('custom_admin_stock_export_detail', pk=batch.pk)
        else:
            selected_pharmacy = get_requested_stock_export_pharmacy(request)
            selected_export_scope = get_requested_stock_export_scope(request)
            initial = {}
            if selected_pharmacy:
                initial['pharmacy'] = selected_pharmacy.pk
            if selected_export_scope:
                initial['export_scope'] = selected_export_scope
            batch_form = StockExportBatchForm(admin_user=request.user, initial=initial)
            item_formset = StockExportItemFormSet(
                prefix='items',
                form_kwargs={
                    'pharmacy': selected_pharmacy,
                    'admin_user': request.user,
                    'allocation_mode': selected_export_scope,
                },
            )

        selected_export_scope = get_requested_stock_export_scope(request, batch_form=batch_form if request.method == 'POST' else None)
        stock_export_insights = build_stock_export_medicine_insights(
            get_requested_stock_export_pharmacy(request, batch_form=batch_form if request.method == 'POST' else None),
            export_scope=selected_export_scope,
        )
        context = {
            'page_title': 'Tạo phiếu xuất kho',
            'current_model': 'stock_export',
            'batch_form': batch_form,
            'item_formset': item_formset,
            'selected_pharmacy': get_requested_stock_export_pharmacy(request, batch_form=batch_form if request.method == 'POST' else None),
            'selected_export_scope': selected_export_scope,
            'stock_export_insights': stock_export_insights,
        }
        return render(request, 'admin_panel/inventory/stock_export_form.html', context)

    if model_key not in {'pharmacy', 'medicine', 'user', 'promotion', 'news'}:
        raise Http404('Không tìm thấy trang thêm dữ liệu')

    config = get_admin_config(model_key)
    form_class = config['form_create']

    if request.method == 'POST':
        form = form_class(request.POST, request.FILES, admin_user=request.user)
        if form.is_valid():
            obj = form.save()
            messages.success(request, f"Đã thêm {config['title'].lower()} '{get_object_label(obj)}' thành công.")
            return redirect('custom_admin_list', model_key=model_key)
    else:
        form = form_class(admin_user=request.user)

    context = {
        'page_title': f'Thêm {config["title"]}',
        'title': config['title'],
        'model_key': model_key,
        'current_model': model_key,
        'form': form,
        'form_sections': get_admin_form_sections(form, model_key),
        'has_rich_editors': form_has_rich_editors(form),
        'is_create': True,
        'object': None,
        'promotion_suggestion_groups': build_promotion_admin_suggestion_groups(request.user) if model_key == 'promotion' else None,
    }
    if model_key == 'news':
        context.update(
            {
                'news_preview_url': '',
                'can_delete_current': False,
            }
        )
        return render(request, 'admin_panel/content/news_form.html', context)
    return render(request, 'admin_panel/shared/form.html', context)


@admin_panel_required
def custom_admin_update(request, model_key, pk):
    denied_response = require_admin_model_access(request, model_key)
    if denied_response:
        return denied_response
    if not can_update_admin_model(request.user, model_key):
        raise PermissionDenied('Tài khoản hiện tại không có quyền cập nhật dữ liệu ở chức năng quản trị này.')

    if model_key == 'order':
        return redirect('custom_admin_order_detail', pk=pk)
    if model_key == 'return_request':
        return redirect('custom_admin_return_request_detail', pk=pk)
    if model_key == 'purchase_import':
        return redirect('custom_admin_purchase_import_detail', pk=pk)
    if model_key == 'stock_export':
        return redirect('custom_admin_stock_export_detail', pk=pk)
    if model_key == 'inventory_lot':
        obj = get_object_or_404(MedicineLot.objects.select_related('medicine'), pk=pk)
        return redirect('custom_admin_update', model_key='medicine', pk=obj.medicine_id)

    config = get_admin_config(model_key)
    model = config['model']
    obj = get_object_or_404(model, pk=pk)
    ensure_object_is_within_admin_scope(request.user, model_key, obj)

    if model_key == 'user' and obj.is_superuser and not request.user.is_superuser:
        raise PermissionDenied('Bạn không có quyền chỉnh sửa tài khoản quản trị viên cấp cao.')

    form_class = config['form_update']

    if request.method == 'POST':
        post_data = request.POST.copy()
        if request.POST.get('delete_image_action') == '1' and 'delete_image' in form_class(instance=obj, admin_user=request.user).fields:
            post_data['delete_image'] = 'on'
        form = form_class(post_data, request.FILES, instance=obj, admin_user=request.user)
        if form.is_valid():
            updated_obj = form.save()
            messages.success(request, f"Đã cập nhật {config['title'].lower()} '{get_object_label(updated_obj)}'.")
            return redirect('custom_admin_list', model_key=model_key)
    else:
        form = form_class(instance=obj, admin_user=request.user)

    inventory_lots = []
    if model_key == 'medicine':
        inventory_lots = list(
            MedicineLot.objects.filter(medicine=obj)
            .order_by(F('expiry_date').asc(nulls_last=True), '-remaining_quantity', 'id')
        )

    context = {
        'page_title': f'Cập nhật {config["title"]}',
        'title': config['title'],
        'model_key': model_key,
        'current_model': model_key,
        'form': form,
        'form_sections': get_admin_form_sections(form, model_key),
        'has_rich_editors': form_has_rich_editors(form),
        'object': obj,
        'inventory_lots': inventory_lots,
        'is_create': False,
        'promotion_suggestion_groups': build_promotion_admin_suggestion_groups(request.user, current_promotion=obj) if model_key == 'promotion' else None,
    }
    if model_key == 'news':
        context.update(
            {
                'news_preview_url': reverse('news_detail', kwargs={'slug': obj.slug}) if obj.is_published else '',
                'can_delete_current': can_delete_object(request.user, 'news', obj),
            }
        )
        return render(request, 'admin_panel/content/news_form.html', context)
    return render(request, 'admin_panel/shared/form.html', context)


@admin_panel_required
def custom_admin_order_detail(request, pk):
    denied_response = require_admin_model_access(request, 'order')
    if denied_response:
        return denied_response

    order = get_object_or_404(
        Order.objects.select_related('pharmacy', 'user').prefetch_related('items__medicine', 'items__lot_allocations__lot', 'return_request', 'prescription_proof_images'),
        pk=pk,
    )
    auto_complete_order_if_due(order)
    order.refresh_from_db()
    ensure_object_is_within_admin_scope(request.user, 'order', order)
    can_update_order = can_update_admin_model(request.user, 'order')

    if request.method == 'POST':
        if not can_update_order:
            raise PermissionDenied('Tài khoản hiện tại không có quyền cập nhật đơn hàng.')
        previous_status = order.status
        previous_payment_status = order.payment_status
        form = OrderStatusUpdateForm(request.POST, request.FILES, instance=order, admin_user=request.user)
        if form.is_valid():
            try:
                with transaction.atomic():
                    updated_order = form.save(commit=False)
                    status_changed = previous_status != updated_order.status
                    payment_confirmed = (
                        previous_payment_status != Order.PAYMENT_STATUS_PAID
                        and updated_order.payment_status == Order.PAYMENT_STATUS_PAID
                    )
                    if status_changed and updated_order.status == Order.STATUS_COMPLETED:
                        if updated_order.payment_method == Order.PAYMENT_COD:
                            updated_order.payment_status = Order.PAYMENT_STATUS_PAID
                            payment_confirmed = previous_payment_status != Order.PAYMENT_STATUS_PAID
                            if updated_order.payment_confirmed_at is None:
                                updated_order.payment_confirmed_at = timezone.now()
                                updated_order.payment_confirmed_by = request.user
                        updated_order.completed_at = timezone.now()
                    if status_changed and updated_order.status in {Order.STATUS_CANCELLED, Order.STATUS_FAILED_DELIVERY}:
                        updated_order.cancelled_at = timezone.now()
                    updated_order.save()
                    if status_changed:
                        transaction.on_commit(
                            lambda changed_order=updated_order, old_status=previous_status, current_request=request: send_order_status_update_email(
                                changed_order,
                                old_status,
                                request=current_request,
                            )
                        )
                    if payment_confirmed and updated_order.payment_method in {Order.PAYMENT_BANK, Order.PAYMENT_MOMO}:
                        transaction.on_commit(
                            lambda paid_order=updated_order, current_request=request: send_order_payment_confirmed_email(
                                paid_order,
                                request=current_request,
                            )
                        )
                messages.success(request, f'Đã cập nhật đơn hàng #{order.pk}.')
                return redirect('custom_admin_order_detail', pk=order.pk)
            except ValueError as exc:
                form.add_error(None, str(exc))
    else:
        form = OrderStatusUpdateForm(instance=order, admin_user=request.user)
        if not can_update_order:
            disable_form_fields(form)

    order.user_return_request = getattr(order, 'return_request', None)
    order_items = list(order.items.select_related('medicine').prefetch_related('lot_allocations__lot').all())
    context = {
        'page_title': f'Chi tiết đơn hàng #{order.pk}',
        'current_model': 'order',
        'order': order,
        'update_form': form,
        'order_items': order_items,
        'return_request': getattr(order, 'return_request', None),
        'prescription_proof_cards': build_order_prescription_proof_cards(order),
        'can_delete_order': can_delete_object(request.user, 'order', order),
        'can_update_order': can_update_order,
    }
    return render(request, 'admin_panel/orders/order_detail.html', context)


def apply_return_request_status_change(*, updated_request, new_status, actor):
    now = timezone.now()
    previous_status = updated_request.status

    if previous_status == new_status:
        updated_request.admin_note = updated_request.admin_note or ""
        if new_status == ReturnRefundRequest.STATUS_PROCESSING:
            updated_request.processed_at = None
            updated_request.processed_by = None
        else:
            updated_request.processed_at = now
            updated_request.processed_by = actor
        updated_request.save(update_fields=["admin_note", "processed_at", "processed_by", "updated_at"])
        return updated_request

    updated_request.status = new_status
    if new_status == ReturnRefundRequest.STATUS_PROCESSING:
        updated_request.processed_at = None
        updated_request.processed_by = None
    else:
        updated_request.processed_at = now
        updated_request.processed_by = actor

    updated_request.save()
    return updated_request




@admin_panel_required
def custom_admin_return_request_detail(request, pk):
    denied_response = require_admin_model_access(request, 'return_request')
    if denied_response:
        return denied_response

    return_request = get_object_or_404(
        ReturnRefundRequest.objects.select_related('order__pharmacy', 'order__user', 'processed_by').prefetch_related('evidences'),
        pk=pk,
    )
    ensure_object_is_within_admin_scope(request.user, 'return_request', return_request)
    can_update_return_request = can_update_admin_model(request.user, 'return_request')

    if request.method == 'POST':
        if not can_update_return_request:
            raise PermissionDenied('Tài khoản hiện tại không có quyền cập nhật yêu cầu trả hàng / hoàn tiền.')
        previous_status = return_request.status
        form = ReturnRefundRequestAdminUpdateForm(request.POST, instance=return_request)
        if form.is_valid():
            try:
                with transaction.atomic():
                    updated_request = ReturnRefundRequest.objects.select_for_update().select_related('order').get(pk=return_request.pk)
                    updated_request.admin_note = form.cleaned_data.get('admin_note') or ''
                    apply_return_request_status_change(
                        updated_request=updated_request,
                        new_status=form.cleaned_data['status'],
                        actor=request.user,
                    )
                    if previous_status != updated_request.status:
                        transaction.on_commit(
                            lambda changed_request=updated_request, old_status=previous_status, current_request=request: send_return_request_status_update_email(
                                changed_request,
                                old_status,
                                request=current_request,
                            )
                        )
                messages.success(request, f'Đã cập nhật yêu cầu #{return_request.pk}.')
                return redirect('custom_admin_return_request_detail', pk=return_request.pk)
            except ValueError as exc:
                form.add_error(None, str(exc))
    else:
        form = ReturnRefundRequestAdminUpdateForm(instance=return_request)
        if not can_update_return_request:
            disable_form_fields(form)

    context = {
        'page_title': f'Yêu cầu trả hàng / hoàn tiền #{return_request.pk}',
        'current_model': 'return_request',
        'return_request': return_request,
        'update_form': form,
        'evidence_images': return_request.evidences.all(),
        'can_delete_return_request': can_delete_object(request.user, 'return_request', return_request),
        'can_update_return_request': can_update_return_request,
    }
    return render(request, 'admin_panel/returns/return_request_detail.html', context)


@admin_panel_required
def custom_admin_purchase_import_detail(request, pk):
    denied_response = require_admin_model_access(request, 'purchase_import')
    if denied_response:
        return denied_response

    purchase_batch = get_object_or_404(
        PurchaseImportBatch.objects.select_related('pharmacy', 'imported_by').prefetch_related('items__medicine'),
        pk=pk,
    )
    ensure_object_is_within_admin_scope(request.user, 'purchase_import', purchase_batch)

    context = {
        'page_title': f'Phiếu nhập {purchase_batch.resolved_invoice_code}',
        'current_model': 'purchase_import',
        'purchase_batch': purchase_batch,
        'purchase_items': purchase_batch.items.all(),
        'can_delete_purchase_import': can_delete_object(request.user, 'purchase_import', purchase_batch),
    }
    return render(request, 'admin_panel/inventory/purchase_import_detail.html', context)


@admin_panel_required
def custom_admin_purchase_import_receipt(request, pk):
    denied_response = require_admin_model_access(request, 'purchase_import')
    if denied_response:
        return denied_response

    purchase_batch = get_object_or_404(
        PurchaseImportBatch.objects.select_related('pharmacy', 'imported_by').prefetch_related('items__medicine'),
        pk=pk,
    )
    ensure_object_is_within_admin_scope(request.user, 'purchase_import', purchase_batch)
    ensure_purchase_import_receipt_pdf(purchase_batch, force=True)
    purchase_batch.receipt_pdf.open('rb')
    response = FileResponse(purchase_batch.receipt_pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{sanitize_receipt_filename(purchase_batch.resolved_invoice_code, "phieu-nhap-kho")}"'
    return response


@admin_panel_required
def custom_admin_stock_export_detail(request, pk):
    denied_response = require_admin_model_access(request, 'stock_export')
    if denied_response:
        return denied_response

    export_batch = get_object_or_404(
        StockExportBatch.objects.select_related('pharmacy', 'exported_by').prefetch_related('items__medicine', 'items__lot_allocations__lot'),
        pk=pk,
    )
    ensure_object_is_within_admin_scope(request.user, 'stock_export', export_batch)

    context = {
        'page_title': f'Phiếu xuất {export_batch.resolved_export_code}',
        'current_model': 'stock_export',
        'export_batch': export_batch,
        'export_items': export_batch.items.all(),
        'can_delete_stock_export': can_delete_object(request.user, 'stock_export', export_batch),
    }
    return render(request, 'admin_panel/inventory/stock_export_detail.html', context)


@admin_panel_required
def custom_admin_stock_export_receipt(request, pk):
    denied_response = require_admin_model_access(request, 'stock_export')
    if denied_response:
        return denied_response

    export_batch = get_object_or_404(
        StockExportBatch.objects.select_related('pharmacy', 'exported_by').prefetch_related('items__medicine'),
        pk=pk,
    )
    ensure_object_is_within_admin_scope(request.user, 'stock_export', export_batch)
    ensure_stock_export_receipt_pdf(export_batch, force=True)
    export_batch.receipt_pdf.open('rb')
    response = FileResponse(export_batch.receipt_pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{sanitize_receipt_filename(export_batch.resolved_export_code, "phieu-xuat-kho")}"'
    return response


@admin_panel_required
def custom_admin_stock_export_excel(request, pk):
    denied_response = require_admin_model_access(request, 'stock_export')
    if denied_response:
        return denied_response

    export_batch = get_object_or_404(
        StockExportBatch.objects.select_related('pharmacy', 'exported_by').prefetch_related('items__medicine', 'items__lot_allocations__lot'),
        pk=pk,
    )
    ensure_object_is_within_admin_scope(request.user, 'stock_export', export_batch)
    try:
        workbook = build_stock_export_workbook(export_batch, list(export_batch.items.all()))
    except RuntimeError as exc:
        messages.error(request, str(exc))
        return redirect('custom_admin_stock_export_detail', pk=export_batch.pk)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = sanitize_excel_filename(export_batch.resolved_export_code, "phieu-xuat-kho")
    response = FileResponse(
        output,
        as_attachment=True,
        filename=filename,
        content_type=ADMIN_REPORT_EXCEL_CONTENT_TYPE,
    )
    response['Content-Length'] = str(output.getbuffer().nbytes)
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['X-Content-Type-Options'] = 'nosniff'
    return response



def rollback_purchase_import_batch_inventory(batch):
    lots = list(
        MedicineLot.objects.select_for_update()
        .filter(purchase_batch=batch)
        .select_related('medicine')
        .order_by('id')
    )
    if not lots:
        return

    touched_lot_ids = []
    for lot in lots:
        if lot.remaining_quantity != lot.received_quantity:
            raise ValueError(
                f"Không thể xóa phiếu nhập {batch.resolved_invoice_code} vì lô của thuốc '{lot.medicine.name}' đã được xuất bán một phần hoặc đã điều chỉnh tồn kho."
            )
        if lot.order_allocations.exists():
            raise ValueError(
                f"Không thể xóa phiếu nhập {batch.resolved_invoice_code} vì lô của thuốc '{lot.medicine.name}' đã từng được dùng để xử lý đơn hàng."
            )
        touched_lot_ids.append(lot.id)

    affected_medicines = {lot.medicine for lot in lots if lot.medicine_id}
    MedicineLot.objects.filter(id__in=touched_lot_ids).delete()
    for medicine in affected_medicines:
        recalculate_medicine_inventory_snapshot(medicine)


def restore_stock_export_batch_inventory(batch):
    export_items = list(
        batch.items.select_related("medicine").prefetch_related("lot_allocations__lot").order_by("id")
    )
    affected_medicines = set()
    locked_lot_ids = [
        allocation.lot_id
        for item in export_items
        for allocation in item.lot_allocations.all()
        if allocation.lot_id
    ]
    locked_lots = {
        lot.id: lot
        for lot in MedicineLot.objects.select_for_update().filter(id__in=locked_lot_ids)
    }

    for item in export_items:
        medicine = getattr(item, "medicine", None)
        if medicine is None:
            continue
        affected_medicines.add(medicine)
        allocations = list(item.lot_allocations.all())
        if not allocations:
            MedicineLot.objects.create(
                medicine=medicine,
                pharmacy=batch.pharmacy or medicine.pharmacy,
                source_type=MedicineLot.SOURCE_MANUAL,
                source_label=f"Phục hồi {batch.resolved_export_code}",
                import_price=0,
                expiry_date=medicine.expiry_date,
                received_quantity=item.exported_quantity,
                remaining_quantity=item.exported_quantity,
                note="Phục hồi tồn kho khi hủy phiếu xuất không có dữ liệu lô chi tiết.",
            )
            continue

        for allocation in allocations:
            locked_lot = locked_lots.get(allocation.lot_id)
            if locked_lot is None:
                MedicineLot.objects.create(
                    medicine=medicine,
                    pharmacy=batch.pharmacy or medicine.pharmacy,
                    source_type=MedicineLot.SOURCE_MANUAL,
                    source_label=allocation.lot_source_label or f"Phục hồi {batch.resolved_export_code}",
                    import_price=allocation.lot_import_price or 0,
                    expiry_date=allocation.lot_expiry_date,
                    received_quantity=allocation.quantity,
                    remaining_quantity=allocation.quantity,
                    note="Phục hồi tồn kho khi lô cũ không còn tồn tại trong hệ thống.",
                )
                continue
            locked_lot.remaining_quantity += allocation.quantity
            locked_lot.save(update_fields=["remaining_quantity"])

    for medicine in affected_medicines:
        recalculate_medicine_inventory_snapshot(medicine)


@admin_panel_required
def custom_admin_delete(request, model_key, pk):
    denied_response = require_admin_model_access(request, model_key)
    if denied_response:
        return denied_response

    config = get_admin_config(model_key)
    model = config['model']
    obj = get_object_or_404(model, pk=pk)
    ensure_object_is_within_admin_scope(request.user, model_key, obj)

    if not can_delete_object(request.user, model_key, obj):
        raise PermissionDenied('Bạn không có quyền xóa dữ liệu này.')

    if request.method == 'POST':
        object_name = get_object_label(obj)
        try:
            with transaction.atomic():
                if model_key == 'purchase_import':
                    locked_batch = PurchaseImportBatch.objects.select_for_update().prefetch_related('items__medicine').get(pk=obj.pk)
                    rollback_purchase_import_batch_inventory(locked_batch)
                    locked_batch.delete()
                elif model_key == 'stock_export':
                    locked_batch = StockExportBatch.objects.select_for_update().prefetch_related('items__lot_allocations__lot', 'items__medicine').get(pk=obj.pk)
                    restore_stock_export_batch_inventory(locked_batch)
                    locked_batch.delete()
                else:
                    obj.delete()
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect('custom_admin_list', model_key=model_key)

        messages.success(request, f"Đã xóa {config['title'].lower()} '{object_name}'.")
        return redirect('custom_admin_list', model_key=model_key)

    context = {
        'page_title': f'Xóa {config["title"]}',
        'title': config['title'],
        'model_key': model_key,
        'current_model': model_key,
        'object': obj,
        'object_name': get_object_label(obj),
    }
    return render(request, 'admin_panel/shared/delete.html', context)



@admin_panel_required
def custom_admin_reports(request):
    denied_response = require_admin_model_access(request, 'reports')
    if denied_response:
        return denied_response
    return render(request, 'admin_panel/analytics/reports.html', build_admin_reports_context(request))


@admin_panel_required
def custom_admin_reports_export_excel(request):
    denied_response = require_admin_model_access(request, 'reports')
    if denied_response:
        return denied_response

    report_context = build_admin_reports_context(request, paginate=False)
    try:
        workbook = build_admin_reports_workbook(report_context)
    except RuntimeError as exc:
        messages.error(request, str(exc))
        return redirect('custom_admin_reports')

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = (
        f"bao-cao-gis-pharma-"
        f"{report_context['report_start_date'].strftime('%Y%m%d')}-"
        f"{report_context['report_end_date'].strftime('%Y%m%d')}.xlsx"
    )
    response = FileResponse(
        output,
        as_attachment=True,
        filename=filename,
        content_type=ADMIN_REPORT_EXCEL_CONTENT_TYPE,
    )
    response['Content-Length'] = str(output.getbuffer().nbytes)
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['X-Content-Type-Options'] = 'nosniff'
    return response


@admin_panel_required
def custom_admin_review_insights(request):
    messages.info(request, 'Mục Phân tích đánh giá đã được loại khỏi khu vực quản trị để giao diện gọn và dễ dùng hơn.')
    return redirect('custom_admin_reports')


def custom_csrf_failure_view(request, reason=''):
    message_text = 'Phiên làm việc đã hết hạn hoặc yêu cầu không hợp lệ. Vui lòng tải lại trang rồi thử lại.'
    if request_expects_json(request):
        return JsonResponse({'message': message_text}, status=403)
    return render(
        request,
        'errors/403.html',
        {
            'requested_path': request.get_full_path(),
            'error_message': message_text if not reason else f"{message_text} ({reason})",
        },
        status=403,
    )


def custom_400_view(request, exception=None):
    return render(
        request,
        'errors/400.html',
        {
            'requested_path': request.get_full_path(),
            'error_message': str(exception) if exception else '',
        },
        status=400,
    )


def custom_403_view(request, exception=None):
    return render(
        request,
        'errors/403.html',
        {
            'requested_path': request.get_full_path(),
            'error_message': str(exception) if exception else '',
        },
        status=403,
    )


def custom_404_view(request, exception=None):
    return render(
        request,
        'errors/404.html',
        {
            'requested_path': request.get_full_path(),
        },
        status=404,
    )


def custom_500_view(request):
    return render(
        request,
        'errors/500.html',
        {
            'requested_path': request.get_full_path(),
        },
        status=500,
    )


def error_preview(request, status_code):
    template_map = {
        400: ('errors/400.html', {'requested_path': request.get_full_path(), 'error_message': 'Dữ liệu gửi lên không hợp lệ.'}),
        403: ('errors/403.html', {'requested_path': request.get_full_path(), 'error_message': 'Đây là trang xem trước lỗi 403.'}),
        404: ('errors/404.html', {'requested_path': '/duong-dan-khong-ton-tai/'}),
        405: ('errors/405.html', {'requested_path': request.get_full_path()}),
        500: ('errors/500.html', {'requested_path': request.get_full_path()}),
    }
    payload = template_map.get(status_code)
    if not payload:
        raise Http404('Không hỗ trợ xem trước mã lỗi này.')

    template_name, context = payload
    context['is_preview'] = True
    return render(request, template_name, context)
